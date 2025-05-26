#!/usr/bin/env python3
"""
Check Excel file for cladding data in the expected locations.
"""

import sys
from openpyxl import load_workbook

def check_excel_cladding(excel_path: str):
    """Check Excel file for cladding data."""
    print(f"🔍 CHECKING EXCEL FILE: {excel_path}")
    print("=" * 50)
    
    wb = load_workbook(excel_path)
    print(f"Sheets: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        if 'CANOPY' in sheet_name:
            sheet = wb[sheet_name]
            print(f"\n📋 {sheet_name}:")
            print(f"  C19 (cladding indicator): {repr(sheet['C19'].value)}")
            
            # Check rows 19-24 for cladding data
            cladding_found = False
            for row in range(19, 25):
                p_val = sheet[f'P{row}'].value  # Dimensions
                q_val = sheet[f'Q{row}'].value  # Position
                n_val = sheet[f'N{row}'].value  # Price
                
                if p_val or q_val or n_val:
                    cladding_found = True
                    print(f"  Row {row}: P={repr(p_val)}, Q={repr(q_val)}, N={repr(n_val)}")
            
            if not cladding_found:
                print("  ❌ No cladding data found in rows 19-24")
            
            # Also check if there's any data in the N column around those rows
            print(f"\n  📊 N column data around rows 19-24:")
            for row in range(15, 30):
                n_val = sheet[f'N{row}'].value
                if n_val:
                    print(f"    N{row}: {repr(n_val)}")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python check_excel_cladding.py <excel_file>")
        sys.exit(1)
    
    check_excel_cladding(sys.argv[1]) 