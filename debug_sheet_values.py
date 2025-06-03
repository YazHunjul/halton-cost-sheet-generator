#!/usr/bin/env python3
"""
Debug UV canopy sheet values.
"""

import sys
import os
sys.path.append('src')

from openpyxl import load_workbook

def debug_sheet_values():
    """Debug the actual values in canopy sheets."""
    
    excel_files = []
    if os.path.exists('output'):
        for file in os.listdir('output'):
            if file.endswith('.xlsx'):
                excel_files.append(os.path.join('output', file))
    
    latest_file = max(excel_files, key=os.path.getmtime)
    print(f"📁 Checking: {latest_file}")
    
    try:
        wb = load_workbook(latest_file, data_only=True)
        
        # Find UV and non-UV canopy sheets
        uv_sheets = [s for s in wb.sheetnames if 'CANOPY (UV)' in s]
        canopy_sheets = [s for s in wb.sheetnames if 'CANOPY - ' in s and 'CANOPY (UV)' not in s]
        
        print(f"🔍 UV sheets: {uv_sheets}")
        print(f"🔍 Regular canopy sheets: {canopy_sheets}")
        
        for uv_sheet_name in uv_sheets:
            print(f"\n📊 Checking {uv_sheet_name}:")
            uv_sheet = wb[uv_sheet_name]
            
            # Check N9 value (total price)
            n9_value = uv_sheet['N9'].value
            print(f"   N9 value: {n9_value} (type: {type(n9_value)})")
            
            # Find corresponding non-UV sheet
            area_part = uv_sheet_name.replace('CANOPY (UV) - ', '')
            non_uv_sheet_name = f"CANOPY - {area_part}"
            
            if non_uv_sheet_name in wb.sheetnames:
                print(f"📊 Checking {non_uv_sheet_name}:")
                non_uv_sheet = wb[non_uv_sheet_name]
                
                non_uv_n9_value = non_uv_sheet['N9'].value
                print(f"   N9 value: {non_uv_n9_value} (type: {type(non_uv_n9_value)})")
                
                # Calculate difference
                try:
                    if n9_value and non_uv_n9_value:
                        uv_cost = float(n9_value) - float(non_uv_n9_value)
                        print(f"   UV Extra Over Cost: £{uv_cost:.2f}")
                    else:
                        print(f"   Cannot calculate: UV={n9_value}, Non-UV={non_uv_n9_value}")
                except Exception as e:
                    print(f"   Error calculating: {e}")
            else:
                print(f"❌ Corresponding non-UV sheet not found: {non_uv_sheet_name}")
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    debug_sheet_values() 