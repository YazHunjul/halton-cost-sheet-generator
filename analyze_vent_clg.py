#!/usr/bin/env python3
"""Analyze the ventilated ceiling-only Excel file to understand its structure."""

import openpyxl
import json

# Load the Excel file
excel_path = '/Users/yazan/Downloads/36398 Cost Sheet 11082025.xlsx'
workbook = openpyxl.load_workbook(excel_path, data_only=True)

# Analyze the structure
analysis = {
    'sheet_names': workbook.sheetnames,
    'has_canopy_sheets': any('CANOPY' in sheet for sheet in workbook.sheetnames),
    'has_vent_clg_sheets': any('VENT CLG' in sheet for sheet in workbook.sheetnames),
    'vent_clg_data': {}
}

# Examine VENT CLG sheets
for sheet_name in workbook.sheetnames:
    if 'VENT CLG' in sheet_name:
        sheet = workbook[sheet_name]
        analysis['vent_clg_data'][sheet_name] = {
            'customer': sheet['H3'].value,
            'project_number': sheet['H4'].value,
            'date': sheet['H5'].value,
            'B12': sheet['B12'].value,
            'D12': sheet['D12'].value,
            'N12_price': sheet['N12'].value,
            'N182_delivery': sheet['N182'].value,
            'N193_commissioning': sheet['N193'].value,
            # Check for area name in the sheet title
            'sheet_title': sheet_name
        }

# Check JOB TOTAL sheet
if 'JOB TOTAL' in workbook.sheetnames:
    job_total_sheet = workbook['JOB TOTAL']
    analysis['job_total'] = {
        'has_sheet': True,
        'total_value': job_total_sheet['C24'].value if job_total_sheet['C24'].value else 'Empty'
    }

print(json.dumps(analysis, indent=2, default=str))