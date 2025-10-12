"""
Debug script to verify Reactaway option is saved in session state and project data.
"""
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src'))

from utils.excel import save_to_excel
from datetime import datetime
import json

# Simulate what happens in the app when user checks Reactaway
print("=" * 60)
print("DEBUG: Reactaway State Verification")
print("=" * 60)

# Test 1: Verify checkbox state handling
print("\n1. Testing checkbox state handling:")
test_options = {
    'uvc': False,
    'recoair': False,
    'marvel': False,
    'vent_clg': False,
    'pollustop': False,
    'aerolys': False,
    'xeu': False,
    'reactaway': True  # This should trigger sheet creation
}
print(f"   Test options: {json.dumps(test_options, indent=2)}")
print(f"   Reactaway value: {test_options.get('reactaway', False)}")
print(f"   ✅ Checkbox would be: {'CHECKED' if test_options['reactaway'] else 'UNCHECKED'}")

# Test 2: Verify the has_reactaway check
print("\n2. Testing has_reactaway detection:")
test_area = {
    'name': 'Kitchen Area',
    'options': test_options,
    'canopies': []
}
has_reactaway = test_area.get("options", {}).get("reactaway", False)
print(f"   has_reactaway = {has_reactaway}")
print(f"   Expected: True")
print(f"   ✅ PASS" if has_reactaway else "   ❌ FAIL")

# Test 3: Create minimal project and check Excel generation
print("\n3. Testing Excel generation with Reactaway:")
project_data = {
    "project_name": "Debug Reactaway Project",
    "project_number": "DEBUG-002",
    "customer": "Debug Customer",
    "company": "Debug Company",
    "address": "Debug Address",
    "project_location": "Debug Location",
    "delivery_location": "Debug Delivery",
    "sales_contact": "Debug Sales",
    "estimator": "DS",
    "date": datetime.now().strftime("%Y-%m-%d"),
    "revision": "A",
    "levels": [
        {
            "level_number": 1,
            "level_name": "Ground Floor",
            "areas": [
                {
                    "name": "Kitchen with Reactaway",
                    "options": {
                        "uvc": False,
                        "recoair": False,
                        "marvel": False,
                        "vent_clg": False,
                        "pollustop": False,
                        "aerolys": False,
                        "xeu": False,
                        "reactaway": True  # ← This should create REACTAWAY sheet
                    },
                    "canopies": [
                        {
                            "reference_number": "C001",
                            "model": "KVH-6",
                            "configuration": "Island",
                            "length": 2000,
                            "width": 1000,
                            "height": 555,
                            "wall_cladding": {
                                "type": "None",
                                "width": 0,
                                "height": 0,
                                "position": []
                            },
                            "options": {
                                "fire_suppression": False,
                                "sdu": False
                            }
                        }
                    ]
                }
            ]
        }
    ]
}

print(f"   Project: {project_data['project_name']}")
print(f"   Area: {project_data['levels'][0]['areas'][0]['name']}")
print(f"   Reactaway in area options: {project_data['levels'][0]['areas'][0]['options']['reactaway']}")

# Check the exact logic used in excel.py
area = project_data['levels'][0]['areas'][0]
has_reactaway_check = area.get("options", {}).get("reactaway", False)
print(f"   Excel.py would detect has_reactaway as: {has_reactaway_check}")

if has_reactaway_check:
    print(f"   ✅ REACTAWAY sheet SHOULD be created")
else:
    print(f"   ❌ REACTAWAY sheet WOULD NOT be created")

# Test 4: Generate and verify
print("\n4. Generating Excel file...")
try:
    output_path = save_to_excel(project_data, "templates/excel/COST SHEET R19.2 SEPT2025ss.xlsx")
    print(f"   ✅ Excel generated: {output_path}")

    from openpyxl import load_workbook
    wb = load_workbook(output_path)

    # Check for REACTAWAY sheet
    reactaway_sheets = [s for s in wb.sheetnames if 'REACTAWAY' in s and 'Ground Floor' in s]

    if reactaway_sheets:
        sheet = wb[reactaway_sheets[0]]
        print(f"\n   ✅ SUCCESS! REACTAWAY sheet found:")
        print(f"      Name: {reactaway_sheets[0]}")
        print(f"      State: {sheet.sheet_state}")
        print(f"      Title: {sheet['B1'].value}")
    else:
        print(f"\n   ❌ FAILED! No REACTAWAY sheet found")
        print(f"      All sheets: {', '.join(wb.sheetnames[:10])}...")

except Exception as e:
    print(f"   ❌ Error: {str(e)}")
    import traceback
    traceback.print_exc()

print("\n" + "=" * 60)
print("DEBUG COMPLETE")
print("=" * 60)
