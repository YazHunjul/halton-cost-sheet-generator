#!/usr/bin/env python3
"""Debug RecoAir validation issues."""

import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src'))

from openpyxl import load_workbook
from utils.excel import validate_cell_data, clear_validation_errors

excel_path = '/Users/yazan/Downloads/36068 Cost Sheet 25032025 (1).xlsx'

print("=" * 80)
print("DEBUG RECOAIR VALIDATION ISSUES")
print("=" * 80)

# Load the workbook
wb = load_workbook(excel_path, data_only=True)

for sheet_name in wb.sheetnames:
    if 'RECOAIR' in sheet_name:
        sheet = wb[sheet_name]
        print(f"\n{sheet_name}:")
        print("-" * 40)
        
        # Clear any previous validation errors
        clear_validation_errors()
        
        # Test the validation on key cells first
        print("Testing key cell validations:")
        
        # Test N36 validation
        n36_valid, n36_value, n36_error = validate_cell_data(
            sheet_name, 'N36', sheet['N36'].value, 'number', 'Total Delivery and Installation (N36)'
        )
        print(f"  N36: {sheet['N36'].value} -> Valid: {n36_valid}, Value: {n36_value}, Error: {n36_error}")
        
        # Test N46 validation
        n46_valid, n46_value, n46_error = validate_cell_data(
            sheet_name, 'N46', sheet['N46'].value, 'number', 'Commissioning Price (N46)'
        )
        print(f"  N46: {sheet['N46'].value} -> Valid: {n46_valid}, Value: {n46_value}, Error: {n46_error}")
        
        # Test N40 validation
        n40_valid, n40_value, n40_error = validate_cell_data(
            sheet_name, 'N40', sheet['N40'].value, 'number', 'Flat Pack Price'
        )
        print(f"  N40: {sheet['N40'].value} -> Valid: {n40_valid}, Value: {n40_value}, Error: {n40_error}")
        
        print(f"\nTesting unit selection validations:")
        
        # Test rows 14-28 for selection validation
        for row in range(14, 29):
            e_value = sheet[f'E{row}'].value
            c_value = sheet[f'C{row}'].value
            
            if e_value is not None and str(e_value).strip():
                # Test selection validation
                selection_valid, selection_num, selection_error = validate_cell_data(
                    sheet_name, f'E{row}', e_value, 'integer', f'RecoAir Unit Quantity (Row {row})'
                )
                
                print(f"  Row {row}: E{row}={e_value} -> Valid: {selection_valid}, Value: {selection_num}, Error: {selection_error}")
                
                if selection_valid and selection_num >= 1:
                    # Test dimension validations for this selected unit
                    f_value = sheet[f'F{row}'].value
                    g_value = sheet[f'G{row}'].value
                    h_value = sheet[f'H{row}'].value
                    
                    width_valid, width, width_error = validate_cell_data(
                        sheet_name, f'F{row}', f_value, 'integer', f'RecoAir Unit Width (Row {row})'
                    )
                    length_valid, length, length_error = validate_cell_data(
                        sheet_name, f'G{row}', g_value, 'integer', f'RecoAir Unit Length (Row {row})'
                    )
                    height_valid, height, height_error = validate_cell_data(
                        sheet_name, f'H{row}', h_value, 'integer', f'RecoAir Unit Height (Row {row})'
                    )
                    
                    print(f"    F{row}={f_value} -> Valid: {width_valid}, Value: {width}, Error: {width_error}")
                    print(f"    G{row}={g_value} -> Valid: {length_valid}, Value: {length}, Error: {length_error}")
                    print(f"    H{row}={h_value} -> Valid: {height_valid}, Value: {height}, Error: {height_error}")
                    
                    # Test N12 validation (base price)
                    n12_valid, n12_value, n12_error = validate_cell_data(
                        sheet_name, 'N12', sheet['N12'].value, 'number', 'RecoAir Unit Base Price (N12)'
                    )
                    print(f"    N12={sheet['N12'].value} -> Valid: {n12_valid}, Value: {n12_value}, Error: {n12_error}")
                    
                    # Check if all validations would pass
                    all_valid = all([selection_valid, width_valid, length_valid, height_valid, n12_valid, n36_valid, n46_valid])
                    print(f"    ALL VALIDATIONS PASS: {'✅' if all_valid else '❌'}")

print("\n" + "=" * 80)