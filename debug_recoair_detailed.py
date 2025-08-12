#!/usr/bin/env python3
"""Debug RecoAir pricing in detail to understand the 148k vs 151k discrepancy."""

import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src'))

from openpyxl import load_workbook
from utils.excel import read_excel_project_data
from utils.word import collect_recoair_pricing_schedule_data

excel_path = '/Users/yazan/Downloads/36068 Cost Sheet 25032025 (1).xlsx'

print("=" * 80)
print("DETAILED RECOAIR PRICING DEBUG")
print("=" * 80)

# 1. Check Excel totals directly
wb = load_workbook(excel_path, data_only=True)

print("\n1. EXCEL JOB TOTAL VERIFICATION:")
print("-" * 40)
if 'JOB TOTAL' in wb.sheetnames:
    job_total_sheet = wb['JOB TOTAL']
    t28_value = job_total_sheet['T28'].value
    t24_value = job_total_sheet['T24'].value
    print(f"  T28 (Total including RecoAir): £{t28_value:,.2f}" if t28_value else "  T28: None")
    print(f"  T24 (Other systems): £{t24_value:,.2f}" if t24_value else "  T24: None")

# 2. Read project data and check what we get
print("\n2. PROJECT DATA ANALYSIS:")
print("-" * 40)

project_data = read_excel_project_data(excel_path)

total_recoair_from_project = 0
for level in project_data.get('levels', []):
    for area in level.get('areas', []):
        recoair_price = area.get('recoair_price', 0)
        flat_pack_data = area.get('recoair_flat_pack', {})
        flat_pack_price = flat_pack_data.get('price', 0) if flat_pack_data.get('has_flat_pack', False) else 0
        
        if recoair_price > 0:
            print(f"  Area: {area.get('name')}")
            print(f"    RecoAir Price (excluding flat pack): £{recoair_price:,.2f}")
            if flat_pack_price > 0:
                print(f"    Flat Pack Price: £{flat_pack_price:,.2f}")
            print(f"    Total with flat pack: £{recoair_price + flat_pack_price:,.2f}")
            
            # Check individual units
            for unit in area.get('recoair_units', []):
                print(f"    Unit {unit.get('item_reference', 'N/A')}: £{unit.get('unit_price', 0):,.2f}")
            
            total_recoair_from_project += recoair_price

print(f"\n  Total RecoAir (excluding flat pack): £{total_recoair_from_project:,.2f}")

# 3. Check what the Word generation function produces
print("\n3. WORD GENERATION PRICING:")
print("-" * 40)

recoair_data = collect_recoair_pricing_schedule_data(project_data)
job_totals = recoair_data['job_totals']

print(f"  Units total: £{job_totals['total_units_price']:,.2f}")
print(f"  Delivery total: £{job_totals['total_delivery_price']:,.2f}")
print(f"  Commissioning total: £{job_totals['total_commissioning_price']:,.2f}")
print(f"  Flat pack total: £{job_totals['total_flat_pack_price']:,.2f}")
print(f"  Job total (all): £{job_totals['job_total']:,.2f}")

# Calculate job total without flat pack
job_total_without_flat_pack = (
    job_totals['total_units_price'] + 
    job_totals['total_delivery_price'] + 
    job_totals['total_commissioning_price']
)
print(f"  Job total (excluding flat pack): £{job_total_without_flat_pack:,.2f}")

print("\n4. COMPARISON:")
print("-" * 40)
expected_total = 148355.41  # From Excel
actual_total_with_flat_pack = job_totals['job_total']
actual_total_without_flat_pack = job_total_without_flat_pack

print(f"  Expected (from Excel): £{expected_total:,.2f}")
print(f"  Actual (with flat pack): £{actual_total_with_flat_pack:,.2f}")
print(f"  Actual (without flat pack): £{actual_total_without_flat_pack:,.2f}")

print(f"\n  Difference (with flat pack): £{actual_total_with_flat_pack - expected_total:,.2f}")
print(f"  Difference (without flat pack): £{actual_total_without_flat_pack - expected_total:,.2f}")

# 5. Check individual RecoAir sheets in detail
print("\n5. DETAILED SHEET ANALYSIS:")
print("-" * 40)

for sheet_name in wb.sheetnames:
    if 'RECOAIR' in sheet_name:
        sheet = wb[sheet_name]
        print(f"\n  {sheet_name}:")
        
        # Base unit price (N12)
        n12 = sheet['N12'].value or 0
        print(f"    N12 (Base unit price): £{n12:,.2f}")
        
        # Flat pack (N40)
        n40 = sheet['N40'].value or 0
        print(f"    N40 (Flat pack): £{n40:,.2f}")
        
        # Delivery subtotal (N182)
        n182 = sheet['N182'].value or 0
        print(f"    N182 (Delivery subtotal): £{n182:,.2f}")
        
        # Commissioning (N193)
        n193 = sheet['N193'].value or 0
        print(f"    N193 (Commissioning): £{n193:,.2f}")
        
        # Net delivery (N182 - N193)
        net_delivery = n182 - n193
        print(f"    Net delivery (N182-N193): £{net_delivery:,.2f}")
        
        # What should be the area total (excluding flat pack)
        area_total_should_be = n12 + net_delivery
        print(f"    Area total (N12 + net delivery): £{area_total_should_be:,.2f}")
        
        # What would be total with flat pack
        area_total_with_flat_pack = area_total_should_be + n40
        print(f"    Area total (with flat pack): £{area_total_with_flat_pack:,.2f}")

print("\n" + "=" * 80)