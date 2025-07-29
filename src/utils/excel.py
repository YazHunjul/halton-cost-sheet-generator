"""
Excel generation utilities for Halton quotation system.
Handles creation and manipulation of Excel workbooks based on templates.
"""
from typing import Dict, List, Union, Optional, Any
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from config.business_data import VALID_CANOPY_MODELS
from config.constants import is_feature_enabled
from utils.date_utils import format_date_for_display, get_current_date

# Constants for Excel operations
TEMPLATE_PATHS = {
    "R19.1": "templates/excel/Cost Sheet R19.1 May 2025.xlsx",
    "R19.2": "templates/excel/Cost Sheet R19.2 Jun 2025.xlsx"
}
DEFAULT_TEMPLATE_PATH = TEMPLATE_PATHS["R19.2"]  # Default to latest version

# Other constants
BASE_SHEET_NAME = "CANOPY"  # The template sheet to copy from
RECOAIR_SHEET_NAME = "RECOAIR"
EDGE_BOX_SHEET_NAME = "EDGE BOX"
FIRE_SUPPRESSION_SHEET_NAME = "FIRE SUPPRESSION"  # Template sheet name
LISTS_SHEET_NAME = "Lists"

# Output sheet name mapping
OUTPUT_SHEET_NAMES = {
    FIRE_SUPPRESSION_SHEET_NAME: "FIRE SUPP"  # Map template name to output name
}

def safe_upper(value):
    """
    Safely convert a value to uppercase only if it contains letters.
    Numbers like "1.2" will be returned as-is.
    """
    if value is None:
        return value
    str_value = str(value)
    # Check if the string contains any letters
    if any(c.isalpha() for c in str_value):
        return str_value.upper()
    return str_value

# Cell mappings for different data points (CANOPY, FIRE SUPP, JOB TOTAL, etc.)
CELL_MAPPINGS = {
    "project_number": "C3",  # Job No
    "company": "C5",         # Company (changed from customer)
    "estimator": "C7",       # Sales Manager / Estimator Initials
    "project_name": "G3",    # Project Name (changed from F3)
    "project_location": "G5",  # Project Location (changed from F5)
    "date": "G7",             # Date (changed from F7)
    "revision": "K7",         # Revision
}

# Row spacing for canopy entries
CANOPY_ROW_SPACING = 17

# Starting row for canopy data
CANOPY_START_ROW = 14

# Tab color mapping for different levels
TAB_COLORS = [
    "FF92D050",  # Light green
    "FF00B0F0",  # Light blue
    "FFFF9900",  # Orange
    "FFFF00FF",  # Pink
    "FF7030A0",  # Purple
    "FFFF0000",  # Red
    "FF00FF00",  # Green
    "FF0070C0",  # Blue
    "FFFFC000",  # Gold
    "FF00FFFF",  # Cyan
]

def remove_external_links(wb: Workbook) -> None:
    """
    Remove external links from workbook to prevent 'unsafe external sources' warning.
    
    Args:
        wb (Workbook): Workbook to clean
    """
    try:
        # Remove external links if they exist
        if hasattr(wb, 'external_references') and wb.external_references:
            for ext_ref in wb.external_references:
                if hasattr(ext_ref, 'clear'):
                    ext_ref.clear()
        
        # Alternative approach: check for external links in defined names
        if hasattr(wb, 'defined_names') and wb.defined_names:
            names_to_remove = []
            for name in wb.defined_names:
                if hasattr(name, 'value') and name.value and ('[' in str(name.value) or '.xlsx' in str(name.value)):
                    names_to_remove.append(name.name)
            
            for name_to_remove in names_to_remove:
                try:
                    del wb.defined_names[name_to_remove]
                except:
                    pass
                    
    except Exception as e:
        print(f"Warning: Could not remove external links: {str(e)}")

def get_template_path(version: str = None) -> str:
    """
    Get the template path for a specific version.
    
    Args:
        version (str, optional): Version identifier (e.g. "R19.1" or "R19.2")
        
    Returns:
        str: Path to the template file
    """
    if version and version in TEMPLATE_PATHS:
        return TEMPLATE_PATHS[version]
    return DEFAULT_TEMPLATE_PATH

def load_template_workbook(template_path: str = None, version: str = None) -> Workbook:
    """
    Load the Excel template workbook and remove external links.
    
    Args:
        template_path (str, optional): Path to the template file. If None, uses default.
        version (str, optional): Version identifier (e.g. "R19.1" or "R19.2")
    
    Returns:
        Workbook: Template workbook with external links removed
    """
    try:
        # Use provided template path, version-specific path, or fall back to default
        if template_path is None:
            if version:
                template_path = get_template_path(version)
            else:
                template_path = DEFAULT_TEMPLATE_PATH
        
        # Try relative path from src directory first, then from project root
        template_paths = [
            f"../{template_path}",  # From src directory
            template_path           # From project root
        ]
        
        wb = None
        for path in template_paths:
            try:
                wb = load_workbook(path)
                print(f"✅ Successfully loaded template: {path}")
                break
            except FileNotFoundError:
                continue
        
        if wb is None:
            raise FileNotFoundError(f"Could not find template file '{template_path}' in any of the expected locations")
        
        # Remove external links to prevent "unsafe external sources" warning
        remove_external_links(wb)
        
        return wb
    except Exception as e:
        raise Exception(f"Failed to load template workbook: {str(e)}")

def copy_template_sheet(wb: Workbook, sheet_name: str, new_name: str) -> Worksheet:
    """
    Create a copy of a template sheet with a new name.
    
    Args:
        wb (Workbook): The workbook to modify
        sheet_name (str): Name of the template sheet to copy
        new_name (str): Name for the new sheet
    
    Returns:
        Worksheet: The newly created worksheet
    """
    try:
        template_sheet = wb[sheet_name]
        new_sheet = wb.copy_worksheet(template_sheet)
        new_sheet.title = new_name
        return new_sheet
    except KeyError:
        raise Exception(f"Template sheet '{sheet_name}' not found in workbook. Available sheets: {wb.sheetnames}")
    except Exception as e:
        raise Exception(f"Failed to copy template sheet: {str(e)}")

def get_sheet_type(project_type: str, canopy_data: Optional[Dict] = None) -> List[str]:
    """
    Determine which template sheets to use based on project and canopy type.
    
    Args:
        project_type (str): Type of project (Canopy or RecoAir)
        canopy_data (Dict, optional): Canopy configuration data
    
    Returns:
        List[str]: List of template sheet names to use
    """
    sheets = []
    
    if project_type == "RecoAir Project":
        sheets.append(RECOAIR_SHEET_NAME)
    elif project_type == "Canopy Project":
        # Check if canopy has specific requirements for EDGE BOX
        if canopy_data and canopy_data.get("configuration") == "Edge":
            sheets.append(EDGE_BOX_SHEET_NAME)
        else:
            sheets.append(BASE_SHEET_NAME)
        
        # Add Fire Suppression sheet if the option is enabled
        if canopy_data and canopy_data.get("options", {}).get("fire_suppression"):
            sheets.append(FIRE_SUPPRESSION_SHEET_NAME)
    else:
        raise ValueError(f"Unknown project type: {project_type}")
    
    return sheets

def get_output_sheet_name(template_sheet_name: str) -> str:
    """
    Get the output sheet name for a given template sheet name.
    
    Args:
        template_sheet_name (str): The name of the template sheet
    
    Returns:
        str: The name to use in the output file
    """
    return OUTPUT_SHEET_NAMES.get(template_sheet_name, template_sheet_name)

def get_initials(name_str: str) -> str:
    """
    Convert name string into initials. Handles multiple names separated by 'and' or '/'.
    Example: "Yazan Hunjul / Joe Salloum" -> "YH/JS"
    
    Args:
        name_str (str): Name string potentially containing multiple names
        
    Returns:
        str: Initials with slash separator
    """
    if not name_str:
        return ""
    
    # Split by common separators (and, /, &)
    names = [n.strip() for n in name_str.replace(" and ", "/").replace("&", "/").split("/")]
    
    # Get initials for each name
    initials_list = []
    for name in names:
        parts = name.strip().split()
        if parts:
            initials = ''.join(part[0].upper() for part in parts if part)
            initials_list.append(initials)
    
    # Join with slash
    return '/'.join(initials_list)

def _calculate_net_canopy_price(sheet: Worksheet, ref_row: int) -> float:
    """
    Calculate net canopy price by reading from P{ref_row} formula result, 
    or manually calculating if formula result is not available.
    
    Args:
        sheet (Worksheet): The worksheet to read from
        ref_row (int): Reference row (12, 29, 46, etc.)
    
    Returns:
        float: Net canopy price (canopy total minus cladding)
    """
    try:
        # Try to read the calculated value from P12, P29, P46, etc.
        p_cell_value = sheet[f'P{ref_row}'].value
        if p_cell_value and isinstance(p_cell_value, (int, float)):
            return float(p_cell_value)
        
        # If formula result not available, calculate manually
        # Sum the canopy-related prices from the subtotal range, EXCLUDING the cladding price
        # The Excel formula is P12=N12-N19, so we need: subtotal(N14:N27) - N19
        start_row = ref_row + 2  # N14, N31, N48, etc.
        cladding_row = ref_row + 7  # N19, N36, N53, etc.
        end_row = ref_row + 15   # N27, N44, N61, etc.
        
        # Calculate subtotal (sum all values in the range)
        subtotal = 0
        for row in range(start_row, end_row + 1):
            cell_value = sheet[f'N{row}'].value
            if cell_value and isinstance(cell_value, (int, float)):
                subtotal += float(cell_value)
        
        # Subtract cladding price (this matches the Excel formula P12=N12-N19)
        cladding_price = sheet[f'N{cladding_row}'].value or 0
        if isinstance(cladding_price, (int, float)):
            cladding_price = float(cladding_price)
        else:
            cladding_price = 0
        
        net_price = subtotal - cladding_price
        return net_price
        
    except Exception as e:
        print(f"Warning: Could not calculate net canopy price for ref_row {ref_row}: {str(e)}")
        return 0

def _calculate_net_delivery_price(sheet: Worksheet) -> float:
    """
    Calculate net delivery & installation price by reading from N182 and subtracting N193.
    
    Excel structure:
    - N182: =SUBTOTAL(9,N183:N197) (TOTAL delivery price including all components)
    - N193: Commissioning price
    - Net Delivery & Installation = N182 - N193
    
    Args:
        sheet (Worksheet): The worksheet to read from
    
    Returns:
        float: Net delivery & installation price (N182 minus N193)
    """
    try:
        # Read commissioning price from N193 first
        commissioning_price = sheet['N193'].value or 0
        if isinstance(commissioning_price, (int, float)):
            commissioning_price = float(commissioning_price)
        else:
            commissioning_price = 0
        
        # Read total delivery price from N182
        n182_value = sheet['N182'].value
        if n182_value and isinstance(n182_value, (int, float)):
            delivery_total = float(n182_value)
        else:
            # If N182 formula not evaluated, manually calculate the SUBTOTAL(9,N183:N197)
            # This includes ALL items in the range (including N193 commissioning)
            print(f"Warning: N182 formula not evaluated, manually calculating SUBTOTAL")
            delivery_total = 0
            for row in range(183, 198):  # N183 to N197 (SUBTOTAL range)
                cell_value = sheet[f'N{row}'].value
                if cell_value and isinstance(cell_value, (int, float)):
                    delivery_total += float(cell_value)
        
        # Calculate net delivery & installation (N182 - N193)
        net_delivery = delivery_total - commissioning_price
        return net_delivery
        
    except Exception as e:
        print(f"Warning: Could not calculate net delivery price: {str(e)}")
        return 0

def read_wall_cladding_from_canopy(sheet: Worksheet, base_row: int) -> Dict:
    """
    Read wall cladding data from canopy row in Excel.
    
    Args:
        sheet (Worksheet): The worksheet to read from
        base_row (int): Base row for this canopy (model row)
        
    Returns:
        Dict: Wall cladding data
    """
    try:
        cladding_indicator_row = base_row + 5  # Row for wall cladding indicator (row 19 for first canopy)
        cladding_data_row = base_row + 6  # Row for wall cladding data (row 20 for first canopy)
        
        # Check for "2M² (HFL)" indicator in column C
        cladding_indicator = sheet[f"C{cladding_indicator_row}"].value or ""
        has_cladding_indicator = "2M²" in str(cladding_indicator).upper() or "HFL" in str(cladding_indicator).upper()
        
        # Read wall cladding data from columns P, Q, S
        width = sheet[f"P{cladding_data_row}"].value or None
        height = sheet[f"Q{cladding_data_row}"].value or None
        position_str = sheet[f"S{cladding_data_row}"].value or None
        
        # Check if any wall cladding data exists (either indicator or actual data)
        if has_cladding_indicator or width or height or position_str:
            # Convert position string to list (handle "and" separator)
            if position_str and str(position_str).strip():
                # Split by "and" and clean up each position
                position_list = [pos.strip().lower() for pos in str(position_str).split(" and ")]
                # Filter out empty strings
                position_list = [pos for pos in position_list if pos]
            else:
                position_list = []
            
            return {
                'type': 'Custom',  # Indicate this is custom wall cladding
                'width': int(width) if width and str(width).replace('.', '').isdigit() else None,
                'height': int(height) if height and str(height).replace('.', '').isdigit() else None,
                'position': position_list
            }
        else:
            # No wall cladding data found
            return {
                'type': 'None',
                'width': None,
                'height': None,
                'position': None
            }
    except Exception as e:
        print(f"Warning: Could not read wall cladding data from row {base_row}: {str(e)}")
        return {
            'type': 'None',
            'width': None,
            'height': None,
            'position': None
        }

def safe_float_conversion(value) -> float:
    """
    Safely convert a value to float, handling various Excel data types.
    
    Args:
        value: Value from Excel cell (could be int, float, string, or None)
        
    Returns:
        float: Converted value, or 0.0 if conversion fails
    """
    if value is None:
        return 0.0
    
    try:
        return float(value)
    except (ValueError, TypeError):
        # Try to extract number from string
        if isinstance(value, str):
            # Remove common non-numeric characters and try again
            cleaned = value.strip().replace(',', '').replace('£', '').replace('$', '')
            try:
                return float(cleaned)
            except ValueError:
                pass
        return 0.0

def extract_tank_quantity(tank_value) -> int:
    """
    Extract tank quantity number from tank value strings like "1 TANK", "2 TANK", etc.
    
    Args:
        tank_value: Tank value from Excel cell (could be string, number, or None)
        
    Returns:
        int: Tank quantity number, or 0 if not found/invalid
    """
    if not tank_value:
        return 0
    
    # Convert to string and clean up
    tank_str = str(tank_value).strip().upper()
    
    # Handle empty or dash values
    if tank_str == "" or tank_str == "-":
        return 0
    
    # Extract number from strings like "1 TANK", "2 TANK", "3 TANKS", etc.
    try:
        # Split by space and look for the first part that's a number
        parts = tank_str.split()
        for part in parts:
            # Try to convert each part to int
            try:
                return int(part)
            except ValueError:
                continue
        
        # If no number found in parts, try to extract digits from the whole string
        import re
        numbers = re.findall(r'\d+', tank_str)
        if numbers:
            return int(numbers[0])  # Return the first number found
        
        return 0
    except (ValueError, AttributeError):
        return 0

def transform_recoair_model(model_str: str) -> str:
    """
    Transform RecoAir model names according to business rules.
    
    Examples:
    - RA1.0 STANDARD -> RAH1.0
    - RAH0.5 STANDARD -> RAH0.5 (already has H)
    - RAH0.5 VOID -> RAH0.5V
    - RA2.5 STANDARD (Prem Controls) -> RAH2.5 (no P suffix)
    - RA1.5 VOID (+10%) -> RAH1.5V
    
    Args:
        model_str (str): Original model string from Excel
        
    Returns:
        str: Transformed model string
    """
    if not model_str:
        return ""
    
    model = str(model_str).strip().upper()
    
    # Extract the base model number (e.g., "0.5", "1.0", "2.5")
    import re
    
    # Look for pattern like RA(H)?X.X where X.X is the model number
    base_match = re.search(r'RA(H)?(\d+\.\d+)', model)
    if not base_match:
        return model  # Return original if pattern doesn't match
    
    has_h = base_match.group(1) is not None  # Check if H is already present
    model_number = base_match.group(2)  # e.g., "1.0", "2.5"
    
    # Start with RAH + model number
    result = f"RAH{model_number}"
    
    # Add suffixes based on content (only VOID gets a suffix)
    if 'VOID' in model:
        result += 'V'
    # For STANDARD and PREM CONTROLS, no suffix needed
    
    return result

def get_recoair_specifications(model: str) -> dict:
    """
    Get technical specifications for RecoAir models based on the model name.
    
    Args:
        model (str): RecoAir model name (e.g., "RAH1.0", "RAH0.5V")
        
    Returns:
        dict: Technical specifications including p_drop, motor, weight
    """
    # RecoAir specifications table
    specifications = {
        # Standard models
        'RAH0.5': {'p_drop': 1050, 'motor': 2.2, 'weight': 436},
        'RAH0.8': {'p_drop': 1050, 'motor': 2.2, 'weight': 470},
        'RAH1.0': {'p_drop': 1050, 'motor': 4.7, 'weight': 572},
        'RAH1.5': {'p_drop': 1050, 'motor': 4.7, 'weight': 820},
        'RAH2.0': {'p_drop': 1050, 'motor': 5.25, 'weight': 974},
        'RAH2.5': {'p_drop': 1050, 'motor': 5.25, 'weight': 1170},
        'RAH3.0': {'p_drop': 1050, 'motor': 5.25, 'weight': 1210},
        'RAH3.5': {'p_drop': 1050, 'motor': 5.25, 'weight': 1395},
        'RAH4.0': {'p_drop': 1050, 'motor': 5.25, 'weight': 1500},
        
        # VOID models (V suffix)
        'RAH0.5V': {'p_drop': 1050, 'motor': 2.2, 'weight': 385},
        'RAH0.8V': {'p_drop': 1050, 'motor': 2.2, 'weight': 415},
        'RAH1.0V': {'p_drop': 1050, 'motor': 4.7, 'weight': 542},
        'RAH1.5V': {'p_drop': 1050, 'motor': 4.7, 'weight': 765},
        'RAH2.0V': {'p_drop': 1050, 'motor': 5.25, 'weight': 884},
        'RAH2.5V': {'p_drop': 1050, 'motor': 5.25, 'weight': 1093},
        'RAH3.0V': {'p_drop': 1050, 'motor': 5.25, 'weight': 1210},
        'RAH3.5V': {'p_drop': 1050, 'motor': 5.25, 'weight': 1395},
        'RAH4.0V': {'p_drop': 1050, 'motor': 5.25, 'weight': 1500},
    }
    
    # Return specifications for the model, or default values if not found
    return specifications.get(model, {'p_drop': 0, 'motor': 0, 'weight': 0})

def extract_recoair_volume(volume_str) -> float:
    """
    Extract volume number from RecoAir volume strings like "VERTICAL 1.2M3/S".
    
    Args:
        volume_str: Volume string from Excel cell (could be string or None)
        
    Returns:
        float: Volume number, or 0.0 if not found/invalid
    """
    if not volume_str:
        return 0.0
    
    # Convert to string and clean up
    volume_string = str(volume_str).strip()
    
    # Handle empty or dash values
    if volume_string == "" or volume_string == "-":
        return 0.0
    
    try:
        # Use regex to find decimal numbers in the string
        import re
        # Look for patterns like "1.2", "2.5", "10.0", etc.
        # But avoid matching single digits that are part of "M3/S"
        numbers = re.findall(r'\d+\.\d+|\d+(?!\d*[/])', volume_string)
        if numbers:
            # Return the first number found as float
            return float(numbers[0])
        
        return 0.0
    except (ValueError, AttributeError):
        return 0.0

def _read_mua_volume(sheet: Worksheet, base_row: int, model: str) -> str:
    """
    Read MUA volume from the correct location based on canopy model.
    
    Args:
        sheet: The worksheet to read from
        base_row: Base row for the canopy (14, 31, 48, etc.)
        model: Canopy model string
        
    Returns:
        str: MUA volume value or empty string
    """
    if not model:
        return ""
    
    # If canopy has 'F' (fresh air), read from column H at row base_row + 8 (H22, H39, H56, etc.)
    if 'F' in model.upper():
        mua_volume_row = base_row + 8  # H22, H39, H56, etc.
        return sheet[f'H{mua_volume_row}'].value or ""
    else:
        # For non-fresh air canopies, read from the old location (column K) for backward compatibility
        return sheet[f'K{base_row}'].value or ""

def read_recoair_data_from_sheet(sheet: Worksheet) -> Dict:
    """
    Read RecoAir unit data from a RECOAIR sheet.
    
    Args:
        sheet (Worksheet): The RECOAIR worksheet to read from
        
    Returns:
        List[Dict]: List of RecoAir units found in the sheet
    """
    recoair_units = []
    sheet_name = sheet.title
    
    try:
        # Get item reference from C12 (e.g., "1.01", "2.01")
        item_reference = sheet['C12'].value or ""
        
        # Get delivery and installation price (N36 - N46) with validation
        n36_valid, n36_value, n36_error = validate_cell_data(
            sheet_name, 'N36', sheet['N36'].value, 'number', 'Total Delivery and Installation (N36)'
        )
        if not n36_valid:
            add_validation_error(n36_error)
            n36_value = 0

        # Get commissioning price from N46 with validation
        n46_valid, n46_value, n46_error = validate_cell_data(
            sheet_name, 'N46', sheet['N46'].value, 'number', 'Commissioning Price (N46)'
        )
        if not n46_valid:
            add_validation_error(n46_error)
            n46_value = 0

        # Calculate delivery and installation price (N36 - N46)
        delivery_installation_price = n36_value - n46_value if n36_value > n46_value else 0
        
        # Get flat pack data from D40 and N40
        flat_pack_description = sheet['D40'].value or ""
        
        # Validate flat pack price
        flat_pack_valid, flat_pack_price, flat_pack_error = validate_cell_data(
            sheet_name, 'N40', sheet['N40'].value, 'number', 'Flat Pack Price'
        )
        if not flat_pack_valid:
            add_validation_error(flat_pack_error)
            flat_pack_price = 0
        
        # Check rows 14 to 28 for RecoAir unit selections
        for row in range(14, 29):  # 14 to 28 inclusive
            # Check if there's a value of 1 or more in column E (selection indicator)
            selection_value = sheet[f'E{row}'].value
            
            if selection_value and str(selection_value).strip() != "":
                # Validate selection quantity (use 'integer' for quantities)
                selection_valid, selection_num, selection_error = validate_cell_data(
                    sheet_name, f'E{row}', selection_value, 'integer', f'RecoAir Unit Quantity (Row {row})'
                )
                
                if not selection_valid:
                    add_validation_error(selection_error)
                    continue
                    
                if selection_num >= 1:
                    # This row has a selected RecoAir unit
                    # Collect data from this row
                    model = sheet[f'C{row}'].value or ""
                    extract_volume_str = sheet[f'D{row}'].value or ""
                    
                    # Validate dimensions (use 'integer' for dimensions to avoid .0 display)
                    width_valid, width, width_error = validate_cell_data(
                        sheet_name, f'F{row}', sheet[f'F{row}'].value, 'integer', f'RecoAir Unit Width (Row {row})'
                    )
                    if not width_valid:
                        add_validation_error(width_error)
                        width = 0
                    
                    length_valid, length, length_error = validate_cell_data(
                        sheet_name, f'G{row}', sheet[f'G{row}'].value, 'integer', f'RecoAir Unit Length (Row {row})'
                    )
                    if not length_valid:
                        add_validation_error(length_error)
                        length = 0
                    
                    height_valid, height, height_error = validate_cell_data(
                        sheet_name, f'H{row}', sheet[f'H{row}'].value, 'integer', f'RecoAir Unit Height (Row {row})'
                    )
                    if not height_valid:
                        add_validation_error(height_error)
                        height = 0
                    
                    location_raw = sheet[f'I{row}'].value or "INTERNAL"  # Default to INTERNAL
                    
                    # Read base price from N12 (fixed cell for all units)
                    price_valid, unit_price, price_error = validate_cell_data(
                        sheet_name, 'N12', sheet['N12'].value, 'number', 'RecoAir Unit Base Price (N12)'
                    )
                    if not price_valid:
                        add_validation_error(price_error)
                        unit_price = 0
                    
                    # Clean up location value - handle placeholder text
                    if location_raw:
                        location_str = str(location_raw).strip().upper()
                        # If location is a placeholder like "SELECT..." or empty, default to INTERNAL
                        if location_str in ["SELECT...", "SELECT", "", "-"] or "SELECT" in location_str:
                            location = "INTERNAL"
                        else:
                            location = location_str
                    else:
                        location = "INTERNAL"
                    
                    # Extract volume number from extract volume string
                    extract_volume = extract_recoair_volume(extract_volume_str)
                    
                    # Transform the model name according to business rules
                    original_model = str(model).strip() if model else ""
                    transformed_model = transform_recoair_model(original_model)
                    
                    # Get technical specifications for this model
                    specs = get_recoair_specifications(transformed_model)
                    
                    # Calculate final unit price (base price + delivery/install share + commissioning)
                    base_unit_price = unit_price  # Base price from N12
                    delivery_per_unit = delivery_installation_price / selection_num if selection_num > 0 else 0
                    commissioning_per_unit = n46_value / selection_num if selection_num > 0 else 0
                    # Total price is base price from N12 plus share of delivery and commissioning
                    final_unit_price = base_unit_price + delivery_per_unit + commissioning_per_unit
                    
                    # Create RecoAir unit data
                    recoair_unit = {
                        'item_reference': str(item_reference).strip() if item_reference else "",
                        'model': transformed_model,
                        'model_original': original_model,  # Keep original for reference
                        'extract_volume': extract_volume,
                        'extract_volume_raw': str(extract_volume_str).strip() if extract_volume_str else "",
                        'width': width,    # Already validated above
                        'length': length,  # Already validated above
                        'height': height,  # Already validated above
                        'location': location,
                        'unit_price': final_unit_price,  # Total price including all components
                        'base_unit_price': base_unit_price,  # Base price from N12
                        'delivery_installation_price': delivery_per_unit,  # Share of delivery/install price
                        'commissioning_price': commissioning_per_unit,  # Share of commissioning price
                        'quantity': selection_num,
                        'row': row,  # Keep track of which row this came from
                        
                        # Technical specifications
                        'p_drop': specs['p_drop'],  # Pressure drop (Pa)
                        'motor': specs['motor'],    # Motor power (kW/PH)
                        'weight': specs['weight']   # Weight (kg)
                    }
                    
                    recoair_units.append(recoair_unit)
        
        # Add delivery price to each unit (split equally if multiple units)
        if recoair_units and delivery_installation_price > 0:
            delivery_per_unit = delivery_installation_price / len(recoair_units)
            for unit in recoair_units:
                unit['delivery_installation_price'] = delivery_per_unit
        else:
            for unit in recoair_units:
                unit['delivery_installation_price'] = 0
        
        # Create result dictionary with units and flat pack data
        result = {
            'units': recoair_units,
            'flat_pack': {
                'item_reference': str(item_reference).strip() if item_reference else "",  # Add item reference to flat pack
                'description': flat_pack_description,
                'price': safe_float_conversion(flat_pack_price),
                'has_flat_pack': bool(flat_pack_description and str(flat_pack_description).strip())
            }
        }
        
        return result
        
    except Exception as e:
        print(f"Warning: Could not read RecoAir data from sheet: {str(e)}")
        return {
            'units': [],
            'flat_pack': {
                'item_reference': '',  # Add item reference to flat pack error case
                'description': '',
                'price': 0,
                'has_flat_pack': False
            }
        }

def read_vent_clg_data_from_sheet(sheet: Worksheet) -> Dict:
    """
    Read VENT CLG (Ventilated Ceiling) data from a VENT CLG sheet.
    
    Args:
        sheet (Worksheet): The VENT CLG worksheet to read from
        
    Returns:
        Dict: VENT CLG data including item number, coverage, unit price, delivery/install, and commissioning
    """
    vent_clg_data = {}
    sheet_name = sheet.title
    
    try:
        # Get item number from B12
        item_number = sheet['B12'].value or ""
        
        # Get square metre coverage from C37
        coverage_valid, coverage, coverage_error = validate_cell_data(
            sheet_name, 'C37', sheet['C37'].value, 'number', 'Square Metre Coverage (C37)'
        )
        if not coverage_valid:
            add_validation_error(coverage_error)
            coverage = 0
        
        # Get unit price from J13 as specified by user
        unit_price_valid, unit_price, unit_price_error = validate_cell_data(
            sheet_name, 'J13', sheet['J13'].value, 'number', 'Unit Price (J13)'
        )
        if not unit_price_valid:
            add_validation_error(unit_price_error)
            unit_price = 0
        
        # Get total cost from G10 as specified by user
        total_cost_valid, total_cost, total_cost_error = validate_cell_data(
            sheet_name, 'G10', sheet['G10'].value, 'number', 'Total Cost (G10)'
        )
        if not total_cost_valid:
            add_validation_error(total_cost_error)
            total_cost = 0
        
        # Get total selling price from J10 as specified by user
        total_price_valid, total_price, total_price_error = validate_cell_data(
            sheet_name, 'J10', sheet['J10'].value, 'number', 'Total Selling Price (J10)'
        )
        if not total_price_valid:
            add_validation_error(total_price_error)
            total_price = 0
        
        # Get delivery total from J43
        j43_valid, j43_value, j43_error = validate_cell_data(
            sheet_name, 'J43', sheet['J43'].value, 'number', 'Delivery Total (J43)'
        )
        if not j43_valid:
            add_validation_error(j43_error)
            j43_value = 0
        
        # Get commissioning price from J55
        j55_valid, j55_value, j55_error = validate_cell_data(
            sheet_name, 'J55', sheet['J55'].value, 'number', 'Commissioning (J55)'
        )
        if not j55_valid:
            add_validation_error(j55_error)
            j55_value = 0
        
        # Calculate delivery and installation price (J43 - J55)
        delivery_installation_price = j43_value - j55_value if j43_value >= j55_value else 0
        commissioning_price = j55_value
        
        vent_clg_data = {
            'item_number': str(item_number).strip() if item_number else "",
            'coverage_sqm': coverage,
            'unit_price': unit_price,
            'delivery_installation_price': delivery_installation_price,
            'commissioning_price': commissioning_price,
            'total_price': total_price,
            'total_cost': total_cost,  # Add total cost from G10
            'sheet_name': sheet_name
        }
        
        print(f"✓ Read VENT CLG data from {sheet_name}:")
        print(f"   Item Number: {vent_clg_data['item_number']}")
        print(f"   Coverage: {coverage} m²")
        print(f"   Unit Price (J13): £{unit_price:.2f}")
        print(f"   Delivery & Installation (J43-J55): £{delivery_installation_price:.2f}")
        print(f"   Commissioning (J55): £{commissioning_price:.2f}")
        print(f"   Total Cost (G10): £{total_cost:.2f}")
        print(f"   Total Selling Price (J10): £{total_price:.2f}")
        
        return vent_clg_data
        
    except Exception as e:
        print(f"Error reading VENT CLG data from {sheet_name}: {str(e)}")
        add_validation_error(f"Failed to read VENT CLG data from {sheet_name}: {str(e)}")
        return {
            'item_number': "",
            'coverage_sqm': 0,
            'unit_price': 0,              # Unit price from J13
            'delivery_installation_price': 0,  # J43 - J55
            'commissioning_price': 0,     # J55
            'total_price': 0,             # Total price from J10
            'total_cost': 0,              # Total cost from G10
            'sheet_name': sheet_name
        }

def write_cost_sheet_identifier(sheet: Worksheet, sheet_name: str, template_version: str = None):
    """
    Write cost sheet identifier to N2 of each sheet.
    Format: R19.2 (sheet type) COST SHEET for individual sheets
    Format: R19.2 COST SHEET for JOB TOTAL sheet
    
    Args:
        sheet (Worksheet): The worksheet to write to
        sheet_name (str): Name of the sheet to determine the identifier
        template_version (str, optional): Template version to use. If None, tries to detect from B1
    """
    try:
        # Use provided version or try to detect from B1
        version = template_version or "R19.2"  # Default version
        
        if not template_version:
            # Fallback: try to get version from B1 of this sheet
            title = sheet['B1'].value
            if title and ' - ' in title:
                parts = title.split(' ')[0].split('-')
                if len(parts) == 2:
                    detected_version = f"R{parts[1]}"
                    if detected_version in TEMPLATE_PATHS:
                        version = detected_version
        
        # Determine sheet type from sheet name
        if sheet_name == "JOB TOTAL":
            identifier = f"{version} COST SHEET"
        else:
            # Extract sheet type from sheet name
            sheet_type = ""
            if "CANOPY" in sheet_name:
                if "CANOPY (UV)" in sheet_name:
                    sheet_type = "CANOPY (UV)"
                else:
                    sheet_type = "CANOPY"
            elif "FIRE SUPP" in sheet_name:
                sheet_type = "FIRE SUPPRESSION"
            elif "EBOX" in sheet_name:
                sheet_type = "EBOX"
            elif "SDU" in sheet_name:
                sheet_type = "SDU"
            elif "RECOAIR" in sheet_name:
                sheet_type = "RECOAIR"
            elif "MARVEL" in sheet_name:
                sheet_type = "MARVEL"
            elif "VENT CLG" in sheet_name:
                sheet_type = "VENT CLG"
            else:
                sheet_type = "SYSTEM"  # Default fallback
            
            identifier = f"{version} ({sheet_type}) COST SHEET"
        
        # Write to N2
        sheet['N2'] = identifier
        
    except Exception as e:
        print(f"Warning: Could not write cost sheet identifier to N2 on {sheet_name}: {str(e)}")

def write_project_metadata(sheet: Worksheet, project_data: Dict, template_version: str = None):
    """
    Write project metadata to the specified cells in the sheet.
    
    Args:
        sheet (Worksheet): The worksheet to write to
        project_data (Dict): Project metadata
        template_version (str, optional): Template version to use for cost sheet identifier
    """
    # Define cell mappings
    CELL_MAPPINGS = {
        "project_number": "C3",    # Job No
        "company": "C5",           # Company (changed from customer)
        "estimator": "C7",         # Sales Manager / Estimator Initials
        "project_name": "G3",      # Project Name (changed from F3)
        "project_location": "G5",  # Project Location (changed from F5)
        "date": "G7",             # Date (changed from F7)
        "revision": "K7",         # Revision
    }
    
    for field, cell in CELL_MAPPINGS.items():
        value = project_data.get(field)
        
        try:
            # Special handling for revision - use the value from project_data (don't default to A for initial)
            if field == "revision":
                sheet[cell] = value or ""  # Use provided revision or leave blank for initial version
            elif value:
                # Special handling for estimator/sales manager initials (only for sheet display)
                if field == "estimator":
                    # Generate combined initials (Sales Contact + Estimator)
                    from utils.word import get_combined_initials
                    
                    # Use sales_contact from project_data directly instead of trying to match estimator
                    sales_contact_name = project_data.get('sales_contact', '')
                    estimator_name = value  # This is the estimator field value
                    
                    # Generate combined initials using the actual sales contact selection
                    value = get_combined_initials(sales_contact_name, estimator_name)
                # Title case for other fields except date
                elif field != "date":
                    value = str(value).title()
                # Date handling
                elif field == "date" and not value:
                    value = get_current_date()
                
                sheet[cell] = value
        except Exception as e:
            # Handle merged cells or other write errors
            print(f"Warning: Could not write {field} to cell {cell}: {str(e)}")
            try:
                # Try to unmerge the cell and write
                if hasattr(sheet, 'merged_cells'):
                    for merged_range in list(sheet.merged_cells.ranges):
                        if cell in merged_range:
                            sheet.unmerge_cells(str(merged_range))
                            break
                # Try writing again after unmerging
                if field == "revision":
                    sheet[cell] = value or ""  # Use provided revision or leave blank for initial version
                elif value:
                    if field == "estimator":
                        # Generate combined initials (Sales Contact + Estimator)
                        from utils.word import get_combined_initials
                        
                        # Use sales_contact from project_data directly
                        sales_contact_name = project_data.get('sales_contact', '')
                        estimator_name = value  # This is the estimator field value
                        
                        # Generate combined initials using the actual sales contact selection
                        value = get_combined_initials(sales_contact_name, estimator_name)
                    elif field != "date":
                        value = str(value).title()
                    elif field == "date" and not value:
                        value = get_current_date()
                    sheet[cell] = value
            except Exception as e2:
                print(f"Warning: Still could not write {field} to cell {cell} after unmerging: {str(e2)}")
                continue
    
    # Add cost sheet identifier to N2
    write_cost_sheet_identifier(sheet, sheet.title, template_version)

# DEPRECATED: This function was overwriting Excel template formulas with hard-coded values
# The Excel template has built-in formulas in N14, N15, N19, N31, N32, N36, etc. that calculate
# prices automatically using VLOOKUP formulas. We should NOT overwrite these formulas.
#
# def write_canopy_pricing_data(sheet: Worksheet, canopy: Dict, row_index: int):
#     """
#     [DEPRECATED] Write canopy pricing data to the sheet at the specified locations.
#     This function was overwriting Excel template pricing formulas and has been disabled.
#     """
#     pass

def write_area_delivery_install_pricing(sheet: Worksheet, area: Dict):
    """
    Write area-level delivery and installation pricing to N182-N193 range.
    
    Args:
        sheet (Worksheet): The worksheet to write to
        area (Dict): Area data containing delivery and installation pricing
    """
    try:
        # DO NOT write to N182 - it contains a SUBTOTAL formula that calculates everything
        # N182 = SUBTOTAL(9,N183:N197) - this is managed by Excel formulas
        
        # Commissioning price goes in N193
        commissioning_price = area.get('commissioning_price', 0)
        if commissioning_price:
            try:
                sheet['N193'] = commissioning_price
                print(f"✓ Wrote commissioning price {commissioning_price} to N193")
            except Exception as e:
                print(f"Warning: Could not write commissioning price to N193: {str(e)}")
        
        # Note: We read delivery & installation as N182-N193 directly in code
        # P182 is not needed since we calculate N182-N193 programmatically
        
        # Write delivery and installation items to feed into N182 subtotal (N183-N197 range)
        delivery_installation_price = area.get('delivery_installation_price', 0)
        if delivery_installation_price:
            try:
                # sheet['N183'] = delivery_installation_price
                print(f"✓ Wrote delivery/installation price {delivery_installation_price} to N183 (feeds into N182 subtotal)")
            except Exception as e:
                print(f"Warning: Could not write delivery/installation price to N183: {str(e)}")
        
        # WRITE PRESERVED AREA-LEVEL MANUAL INPUT FIELDS
        # Delivery number (C183)
        delivery_number = area.get('delivery_number', '')
        if delivery_number:
            try:
                sheet['C183'] = delivery_number
                print(f"✓ Wrote delivery number '{delivery_number}' to C183")
            except Exception as e:
                print(f"Warning: Could not write delivery number to C183: {str(e)}")
        
        # Access equipment entries
        access_equipment_1 = area.get('access_equipment_1', '')
        if access_equipment_1:
            try:
                sheet['D184'] = access_equipment_1
                print(f"✓ Wrote access equipment 1 '{access_equipment_1}' to D184")
            except Exception as e:
                print(f"Warning: Could not write access equipment 1 to D184: {str(e)}")
        
        access_equipment_2 = area.get('access_equipment_2', '')
        if access_equipment_2:
            try:
                sheet['D185'] = access_equipment_2
                print(f"✓ Wrote access equipment 2 '{access_equipment_2}' to D185")
            except Exception as e:
                print(f"Warning: Could not write access equipment 2 to D185: {str(e)}")
        
        # Testing and commissioning description
        testing_commissioning_description = area.get('testing_commissioning_description', '')
        if testing_commissioning_description:
            try:
                sheet['C193'] = testing_commissioning_description
                print(f"✓ Wrote T&C description '{testing_commissioning_description}' to C193")
            except Exception as e:
                print(f"Warning: Could not write T&C description to C193: {str(e)}")
        
    except Exception as e:
        print(f"Warning: Failed to write area delivery/install pricing: {str(e)}")

def write_canopy_data(sheet: Worksheet, canopy: Dict, row_index: int):
    """
    Write canopy specifications to the sheet at the specified row.
    
    Args:
        sheet (Worksheet): The worksheet to write to
        canopy (Dict): Canopy specification data
        row_index (int): Starting row for this canopy's data (this is the model/config row)
    """
    try:
        # Reference number starts 2 rows before configuration/model
        ref_row = row_index - 2  # If row_index is 14, ref_row will be 12
        ref_number = canopy.get("reference_number", "")
        if ref_number:
            try:
                sheet[f"B{ref_row}"] = safe_upper(ref_number)
            except Exception as e:
                print(f"Warning: Could not write reference number to B{ref_row}: {str(e)}")
        
        # Note: Do not write pricing data - let Excel template formulas calculate prices automatically
        
        # Configuration and Model on same row
        configuration = canopy.get("configuration", "")
        if configuration:
            try:
                sheet[f"C{row_index}"] = configuration.upper()
            except Exception as e:
                print(f"Warning: Could not write configuration to C{row_index}: {str(e)}")
        
        # Model in D14, D31, D48, etc.
        model = canopy.get("model", "")
        if model:
            try:
                sheet[f"D{row_index}"] = model.upper()
                
                # Add "1" to D18 for each canopy (4 rows below the model row)
                quantity_row = row_index + 4  # D18, D35, D52, etc.
                sheet[f"D{quantity_row}"] = 1
                
                # For CMWF/CMWI canopies, initialize C27 (base_row + 13) to 0
                if model.upper() in ['CMWF', 'CMWI']:
                    initial_value_row = row_index + 13  # C27, C44, C61, etc.
                    try:
                        sheet[f"C{initial_value_row}"] = 0
                    except Exception as e:
                        print(f"Warning: Could not initialize C{initial_value_row} to 0 for CMWF/CMWI canopy: {str(e)}")
                
                # If canopy has 'F' (fresh air), store MUA volume in column H starting from row 22
                if 'F' in model.upper():
                    mua_volume_row = row_index + 8  # H22, H39, H56, etc. (row_index 14 + 8 = 22)
                    mua_volume = canopy.get("mua_volume", "")
                    if mua_volume:
                        try:
                            # Convert to float if it's a numeric value, otherwise store as string
                            if isinstance(mua_volume, str) and mua_volume.strip():
                                try:
                                    mua_volume_float = float(mua_volume.strip())
                                    sheet[f"H{mua_volume_row}"] = mua_volume_float
                                except ValueError:
                                    sheet[f"H{mua_volume_row}"] = mua_volume.strip()
                            elif isinstance(mua_volume, (int, float)):
                                sheet[f"H{mua_volume_row}"] = mua_volume
                        except Exception as e:
                            print(f"Warning: Could not write MUA volume to H{mua_volume_row}: {str(e)}")
                        
            except Exception as e:
                print(f"Warning: Could not write model to D{row_index}: {str(e)}")
        
        # Write canopy dimensions in E14, F14, G14 (width, length, height)
        width = canopy.get("width", "")
        if width:
            try:
                sheet[f"E{row_index}"] = width
            except Exception as e:
                print(f"Warning: Could not write width to E{row_index}: {str(e)}")
        
        length = canopy.get("length", "")
        if length:
            try:
                sheet[f"F{row_index}"] = length
            except Exception as e:
                print(f"Warning: Could not write length to F{row_index}: {str(e)}")
        
        height = canopy.get("height", "")
        if height:
            try:
                sheet[f"G{row_index}"] = height
            except Exception as e:
                print(f"Warning: Could not write height to G{row_index}: {str(e)}")
        
        # Write number of sections in H14, H31, H48, etc.
        sections = canopy.get("sections", "")
        if sections:
            try:
                sheet[f"H{row_index}"] = sections
            except Exception as e:
                print(f"Warning: Could not write sections to H{row_index}: {str(e)}")
        
        # WRITE PRESERVED MANUAL INPUT FIELDS
        # Light inputs in D15 (base_row + 1)
        light_inputs = canopy.get("light_inputs", "")
        if light_inputs:
            try:
                sheet[f"D{row_index + 1}"] = light_inputs
            except Exception as e:
                print(f"Warning: Could not write light inputs to D{row_index + 1}: {str(e)}")
        
        # Special works entries (C16, C17, C18)
        special_works_1 = canopy.get("special_works_1", "")
        if special_works_1:
            try:
                sheet[f"C{row_index + 2}"] = special_works_1
            except Exception as e:
                print(f"Warning: Could not write special works 1 to C{row_index + 2}: {str(e)}")
        
        special_works_2 = canopy.get("special_works_2", "")
        if special_works_2:
            try:
                sheet[f"C{row_index + 3}"] = special_works_2
            except Exception as e:
                print(f"Warning: Could not write special works 2 to C{row_index + 3}: {str(e)}")
        
        special_works_3 = canopy.get("special_works_3", "")
        if special_works_3:
            try:
                sheet[f"C{row_index + 4}"] = special_works_3
            except Exception as e:
                print(f"Warning: Could not write special works 3 to C{row_index + 4}: {str(e)}")
        
        # Quantity override in D18 (if different from default 1)
        quantity_override = canopy.get("quantity_override", "")
        if quantity_override and str(quantity_override).strip() not in ['', '1']:
            try:
                sheet[f"D{row_index + 4}"] = quantity_override
            except Exception as e:
                print(f"Warning: Could not write quantity override to D{row_index + 4}: {str(e)}")
        
        # Options (only fire suppression at canopy level now)
        options_row = row_index + 4
        options = canopy.get("options", {})
        if options.get("fire_suppression"):
            try:
                sheet[f"B{options_row}"] = "FIRE SUPPRESSION SYSTEM"
            except Exception as e:
                print(f"Warning: Could not write fire suppression to B{options_row}: {str(e)}")
        
        # Wall cladding data (if present)
        wall_cladding = canopy.get("wall_cladding", {})
        if wall_cladding and wall_cladding.get('type') not in ['None', None, '']:
            cladding_indicator_row = row_index + 5  # Row for wall cladding indicator (row 19 for first canopy)
            cladding_data_row = row_index + 6  # Row for wall cladding data (row 20 for first canopy)
            try:
                # Write "2M² (HFL)" indicator in column C (C19, C36, C53, etc.)
                sheet[f"C{cladding_indicator_row}"] = "2M² (HFL)"
                
                # Write wall cladding width in column P
                if wall_cladding.get('width'):
                    sheet[f"P{cladding_data_row}"] = wall_cladding['width']
                
                # Write wall cladding height in column Q  
                if wall_cladding.get('height'):
                    sheet[f"Q{cladding_data_row}"] = wall_cladding['height']
                
                # Write wall cladding position in column S
                position = wall_cladding.get('position', [])
                if isinstance(position, list):
                    position_str = " and ".join(position) if position else ""
                else:
                    position_str = str(position) if position else ""
                
                if position_str:
                    sheet[f"S{cladding_data_row}"] = position_str
                    
            except Exception as e:
                print(f"Warning: Could not write wall cladding data to row {cladding_indicator_row}: {str(e)}")
    except Exception as e:
        raise Exception(f"Failed to write canopy data: {str(e)}")

def write_area_options(sheet: Worksheet, area: Dict):
    """
    Write area-level options (UV-C, SDU, RecoAir) to the sheet.
    
    Args:
        sheet (Worksheet): The worksheet to write to
        area (Dict): Area data containing options
    """
    try:
        # Write area options in a dedicated section (e.g., starting at row 6)
        options_start_row = 6
        area_options = area.get("options", {})
        
        if area_options.get("uvc"):
            try:
                sheet[f"B{options_start_row}"] = "UV-C SYSTEM"
            except Exception as e:
                print(f"Warning: Could not write UV-C option to B{options_start_row}: {str(e)}")
        if area_options.get("sdu"):
            try:
                sheet[f"B{options_start_row + 1}"] = "SDU"
            except Exception as e:
                print(f"Warning: Could not write SDU option to B{options_start_row + 1}: {str(e)}")
        if area_options.get("recoair"):
            try:
                sheet[f"B{options_start_row + 2}"] = "RECOAIR"
            except Exception as e:
                print(f"Warning: Could not write RecoAir option to B{options_start_row + 2}: {str(e)}")
    except Exception as e:
        print(f"Warning: Could not write area options: {str(e)}")
        pass

def write_fire_suppression_canopy_data(sheet: Worksheet, canopy: Dict, row_index: int):
    """
    Write canopy reference number to the fire suppression sheet at the specified row.
    
    Args:
        sheet (Worksheet): The fire suppression worksheet to write to
        canopy (Dict): Canopy specification data
        row_index (int): Starting row for this canopy's data (this is the model/config row)
    """
    try:
        # Reference number starts 2 rows before configuration/model (same pattern as canopy sheets)
        ref_row = row_index - 2  # If row_index is 14, ref_row will be 12
        sheet[f"B{ref_row}"] = safe_upper(canopy["reference_number"])
    except Exception as e:
        raise Exception(f"Failed to write fire suppression canopy data: {str(e)}")

def write_to_sheet(sheet: Worksheet, project_data: Dict, level_name: str, area_name: str, canopies: List[Dict]):
    """
    Write all data for an area to a sheet.
    
    Args:
        sheet (Worksheet): The worksheet to write to
        project_data (Dict): Project metadata
        level_name (str): Name of the level
        area_name (str): Name of the area
        canopies (List[Dict]): List of canopy specifications
    """
    try:
        # Write level and area information in B1
        sheet["B1"] = f"{level_name} - {area_name}"
        
        # Write project metadata
        write_project_metadata(sheet, project_data)
        
        # Write each canopy with proper spacing (17 rows)
        for idx, canopy in enumerate(canopies):
            row_start = CANOPY_START_ROW + (idx * CANOPY_ROW_SPACING)  # Starts at 14, then 31, 48, etc.
            write_canopy_data(sheet, canopy, row_start)
            
            # If this canopy has fire suppression and fire suppression sheet exists, write to it
            if canopy.get("options", {}).get("fire_suppression") and fs_sheet:
                fs_row_start = CANOPY_START_ROW + (fs_canopy_idx * CANOPY_ROW_SPACING)
                write_fire_suppression_canopy_data(fs_sheet, canopy, fs_row_start)
                fs_canopy_idx += 1  # Only increment for canopies with fire suppression
        
        # Add dropdowns
        add_dropdowns_to_sheet(wb, current_canopy_sheet)
        if fs_sheet:
            # Add fire suppression specific dropdowns
            add_fire_suppression_dropdowns(fs_sheet)
        
        sheet_count += 1
    except Exception as e:
        raise Exception(f"Failed to write sheet data: {str(e)}")

def add_dropdowns_to_sheet(wb: Workbook, sheet: Worksheet, start_row: int = 12):
    """
    Add data validation (dropdowns) to specific cells in the sheet.
    
    Args:
        wb (Workbook): The workbook containing the sheet
        sheet (Worksheet): The worksheet to add dropdowns to
        start_row (int): Starting row for dropdowns
    """
    try:
        # Define dropdown options (keeping them shorter to avoid Excel corruption)
        lighting_options = [
            'LIGHT SELECTION',
            'LED STRIP L6 Inc DALI',
            'LED STRIP L12 Inc DALI', 
            'LED STRIP L18 Inc DALI',
            'LED STRIP L6EM',
            'LED STRIP L12EM',
            'LED STRIP L18EM',
            'LM6',
            'LM12',
            'LM18',
            'Small LED Spots Inc DALI',
            'Large LED Spots Inc DALI',
            'HCL600 DALI',
            'HCL1200 DALI',
            'HCL1800 DALI',
            'EL215',
            'EL218'
        ]
        
        special_works_options = [
            'ROUND CORNERS',
            'CUT OUT',
            'CASTELLE LOCKING ',
            'HEADER DUCT S/S',
            'HEADER DUCT',
            'PAINT FINSH',  # Fixed typo from "PAINT FINSH"
            'UV ON DEMAND',
            'E/over for emergency strip light',
            'E/over for small emer. spot light',
            'E/over for large emer. spot light',
            'COLD MIST ON DEMAND',
            'CMW PIPEWORK HWS/CWS',  # Fixed spacing
            'CANOPY GROUND SUPPORT',
            '2nd EXTRACT PLENUM',  # Removed extra space
            'SUPPLY AIR PLENUM',
            'CAPTUREJET PLENUM',
            'COALESCER',
            
        ]
        
        cladding_options = [
            "Standard Stainless Steel",
            "Brushed Stainless Steel",
            "Painted Steel",
            "Galvanized Steel", 
            "Aluminum Composite",
            "No Cladding"
        ]
        
        # Wall cladding options for C19
        wall_cladding_options = [
            "",  # Empty option
            "2M² (HFL)"
        ]
        
        # Wall cladding position options for column S (comprehensive combinations)
        wall_cladding_position_options = [
            "",  # Empty option
            # Single positions
            "rear",
            "left", 
            "right",
            "front",
            # Two-position combinations
            "rear and left",
            "rear and right", 
            "rear and front",
            "left and right",
            "left and front",
            "right and front",
            # Three-position combinations
            "rear and left and right",
            "rear and left and front",
            "rear and right and front",
            "left and right and front",
            # All sides
            "all sides"
        ]
        
        # CMWF/CMWI panel options for wash canopies
        cmw_panel_type_options = [
            "",  # Empty option
            "CP1S",
            "CP2S", 
            "CP3S",
            "CP4S"
        ]
        
        cmw_panel_size_options = [
            "",  # Empty option
            "1000-S",
            "1500-S",
            "2000-S",
            "2500-S",
            "3000-S",
            "1000-D",
            "1500-D",
            "2000-D",
            "2500-D",
            "3000-D"
        ]
        
        # Access equipment options for E39 and E40
        access_equipment_options = [
            "",  # Empty option
            "SL10 GENIE",
            "EXTENSION FORKS",
            "2.5M COMBI LADDER",
            "1.5M PODIUM",
            "3M TOWER",
            "COMBI LADDER",
            "PECO LIFT",
            "3M YOUNGMAN BOARD",
            "GS1930 SCISSOR LIFT",
            "4-6 SHERASCOPIC",
            "7-9 SHERASCOPIC"
        ]
        
        # Create data validations with proper escaping
        def create_validation(options, validation_name=""):
            # For long option lists, we need a different approach since Excel has a 255-character formula limit
            formula = ",".join(options)
            if len(formula) <= 255:
                return DataValidation(type="list", formula1=f'"{formula}"', allow_blank=True)
            else:
                # For longer lists, write them to hidden cells and reference them
                # This allows for much longer option lists
                start_row = 300 + len(options)  # Start at row 300+ to avoid conflicts
                for i, option in enumerate(options):
                    try:
                        sheet[f'AA{start_row + i}'] = option  # Use column AA (hidden area)
                    except:
                        pass  # If we can't write, fall back to text validation
                
                # Create a validation that references the range
                try:
                    range_ref = f'$AA${start_row}:$AA${start_row + len(options) - 1}'
                    return DataValidation(type="list", formula1=range_ref, allow_blank=True)
                except:
                    # Fallback to allowing any text input
                    return DataValidation(type="textLength", operator="lessThan", formula1="100", allow_blank=True)
        
        lighting_dv = create_validation(lighting_options, "lighting")
        special_works_dv = create_validation(special_works_options, "special_works")
        cladding_dv = create_validation(cladding_options, "cladding")
        wall_cladding_dv = create_validation(wall_cladding_options, "wall_cladding")
        wall_cladding_position_dv = create_validation(wall_cladding_position_options, "wall_cladding_position")
        cmw_panel_type_dv = create_validation(cmw_panel_type_options, "cmw_panel_type")
        cmw_panel_size_dv = create_validation(cmw_panel_size_options, "cmw_panel_size")
        access_equipment_dv = create_validation(access_equipment_options, "access_equipment")
        
        # Add validations to sheet
        sheet.add_data_validation(lighting_dv)
        sheet.add_data_validation(special_works_dv)
        sheet.add_data_validation(cladding_dv)
        sheet.add_data_validation(wall_cladding_dv)
        sheet.add_data_validation(wall_cladding_position_dv)
        sheet.add_data_validation(cmw_panel_type_dv)
        sheet.add_data_validation(cmw_panel_size_dv)
        sheet.add_data_validation(access_equipment_dv)
        
        # Write wall cladding headers in row 19 (first canopy's cladding row - 1)
        try:
            sheet['P19'] = 'Width (mm)'
            sheet['Q19'] = 'Height (mm)'
            sheet['S19'] = 'Position'
        except Exception as e:
            print(f"Warning: Could not write wall cladding headers: {str(e)}")
        
        # Add wall cladding input fields for all 10 canopy rows (even if no canopy exists)
        for canopy_index in range(10):  # Support up to 10 canopies per sheet
            base_row = CANOPY_START_ROW + (canopy_index * CANOPY_ROW_SPACING)  # 14, 31, 48, 65, 82, 99, 116, 133, 150, 167
            cladding_row = base_row + 6  # 20, 37, 54, 71, 88, 105, 122, 139, 156, 173
            
            try:
                # Add placeholder values to ensure cells exist for dropdowns (only if cells are currently None)
                # Use direct cell assignment to ensure values are set
                width_cell = sheet[f'P{cladding_row}']
                height_cell = sheet[f'Q{cladding_row}']
                position_cell = sheet[f'S{cladding_row}']
                
                # Only set placeholder if cell is currently None (don't overwrite existing data)
                if width_cell.value is None:
                    width_cell.value = ""  # Width placeholder
                if height_cell.value is None:
                    height_cell.value = ""  # Height placeholder
                if position_cell.value is None:
                    position_cell.value = ""  # Position placeholder
                
                # Values are set to empty strings only if no data exists
            except Exception as e:
                print(f"Warning: Could not add wall cladding placeholders for row {cladding_row}: {str(e)}")
        
        # Apply dropdowns to multiple canopy sections (every 17 rows)
        for canopy_index in range(10):  # Support up to 10 canopies per sheet
            base_row = CANOPY_START_ROW + (canopy_index * CANOPY_ROW_SPACING)  # 14, 31, 48, 65, 82, 99, 116, 133, 150, 167
            
            try:
                # Lighting options - typically around row 15 (C15)
                lighting_row = base_row + 1
                lighting_dv.add(f"C{lighting_row}")
                
                # Special works options - C16, C17, C18 (base_row + 2, 3, 4)
                special_works_dv.add(f"C{base_row + 2}")  # C16
                special_works_dv.add(f"C{base_row + 3}")  # C17
                special_works_dv.add(f"C{base_row + 4}")  # C18
                
                # Wall cladding options - C19 (base_row + 5)
                wall_cladding_row = base_row + 5
                wall_cladding_dv.add(f"C{wall_cladding_row}")  # C19
                
                # Wall cladding position dropdown - S20 (base_row + 6)
                wall_cladding_position_row = base_row + 6
                wall_cladding_position_dv.add(f"S{wall_cladding_position_row}")  # S20, S37, S54, S71, S88, S105, S122, S139, S156, S173
                
                # Cladding options - typically around row 16 (D16) for cladding type
                cladding_row = base_row + 2
                cladding_dv.add(f"D{cladding_row}")
            except Exception as e:
                print(f"Warning: Could not add dropdown to canopy {canopy_index + 1}: {str(e)}")
                continue
        
        # Add some additional dropdowns for common fields
        # Configuration options for column C (model row)
        config_options = ["Wall", "Island", "Single", "Double", "Corner"]
        config_dv = create_validation(config_options, "config")
        sheet.add_data_validation(config_dv)
        
        # Model options for column D (model row)
        model_dv = create_validation(VALID_CANOPY_MODELS, "model")
        sheet.add_data_validation(model_dv)
        
        for canopy_index in range(10):
            try:
                base_row = CANOPY_START_ROW + (canopy_index * CANOPY_ROW_SPACING)
                config_dv.add(f"C{base_row}")  # Configuration in column C of the model row
                model_dv.add(f"D{base_row}")   # Model in column D of the model row (D14, D31, D48, etc.)
                
                # Add CMWF/CMWI panel options dropdowns for all canopies
                # These will be available for all canopies but are specifically for CMWF/CMWI models
                cmw_panel_type_row = base_row + 11  # C25, C42, C59, etc. (base_row + 11)
                cmw_panel_size_row = base_row + 12  # C26, C43, C60, etc. (base_row + 12)
                
                cmw_panel_type_dv.add(f"C{cmw_panel_type_row}")  # Panel type dropdown
                cmw_panel_size_dv.add(f"C{cmw_panel_size_row}")  # Panel size dropdown
                
            except Exception as e:
                print(f"Warning: Could not add config/model dropdown to canopy {canopy_index + 1}: {str(e)}")
                continue
        
        # Add access equipment dropdowns to specific cells D184 and D185
        try:
            access_equipment_dv.add("D184")  # Access equipment dropdown in D184
            access_equipment_dv.add("D185")  # Access equipment dropdown in D185
        except Exception as e:
            print(f"Warning: Could not add access equipment dropdowns to D184/D185: {str(e)}")
        
    except Exception as e:
        # Silently fail for dropdown addition to avoid breaking the main process
        print(f"Warning: Could not add dropdowns to sheet {sheet.title}: {str(e)}")
        pass

def set_sheet_tab_color(sheet: Worksheet, area_index: int):
    """
    Set the tab color for a worksheet based on the area index.
    
    Args:
        sheet (Worksheet): The worksheet to color
        area_index (int): The area index (0-based)
    """
    # Get color from list, cycling through colors if area index exceeds available colors
    color = TAB_COLORS[area_index % len(TAB_COLORS)]
    sheet.sheet_properties.tabColor = color

def write_ebox_metadata(sheet: Worksheet, project_data: Dict, template_version: str = None):
    """
    Write project metadata to EBOX sheet with specific cell mappings.
    
    Args:
        sheet (Worksheet): The EBOX worksheet to write to
        project_data (Dict): Project metadata
    """
    try:
        # EBOX-specific cell mappings
        ebox_cell_mappings = {
            "project_number": "D3",    # Job No
            "company": "D5",           # Company (changed from customer)
            "estimator": "D7",         # Sales Manager / Estimator Initials
            "project_name": "H3",      # Project Name
            "project_location": "H5",          # Project Location (was "location")
            "date": "H7",             # Date
            "revision": "O7",         # Revision
        }
        
        # Write EBOX-specific data
        try:
            sheet["C12"] = "UV-C"  # Model name
            sheet["D38"] = 1       # Quantity
            # Note: Delivery location written by general loop to E38
            
            # Add plant selection dropdowns to E39 and E40
            add_plant_selection_dropdowns_to_ebox(sheet)
        except Exception as e:
            print(f"Warning: Could not write EBOX-specific data: {str(e)}")
        
        for field, cell in ebox_cell_mappings.items():
            value = project_data.get(field)
            
            try:
                # Special handling for revision - use the value from project_data (don't default to A for initial)
                if field == "revision":
                    sheet[cell] = value or ""  # Use provided revision or leave blank for initial version
                elif value:
                    # Special handling for estimator/sales manager initials
                    if field == "estimator":
                        # Generate combined initials (Sales Contact + Estimator)
                        from utils.word import get_combined_initials
                        from config.business_data import SALES_CONTACTS
                        
                        # Get sales contact info based on estimator
                        sales_contact_name = ""
                        for contact_name, phone in SALES_CONTACTS.items():
                            if value and any(name.lower() in value.lower() for name in contact_name.split()):
                                sales_contact_name = contact_name
                                break
                        
                        # If no match found, use first sales contact
                        if not sales_contact_name:
                            sales_contact_name = list(SALES_CONTACTS.keys())[0]
                        
                        # Generate combined initials
                        value = get_combined_initials(sales_contact_name, value)
                    # Title case for other fields except date
                    elif field != "date":
                        value = str(value).title()
                    # Date handling
                    elif field == "date" and not value:
                        value = get_current_date()
                    
                    sheet[cell] = value
            except Exception as e:
                # Handle merged cells or other write errors
                print(f"Warning: Could not write {field} to EBOX cell {cell}: {str(e)}")
                try:
                    # Try to unmerge the cell and write
                    if hasattr(sheet, 'merged_cells'):
                        for merged_range in list(sheet.merged_cells.ranges):
                            if cell in merged_range:
                                sheet.unmerge_cells(str(merged_range))
                                break
                    # Try writing again after unmerging
                    if field == "revision":
                        sheet[cell] = value or ""  # Use provided revision or leave blank for initial version
                    elif value:
                        if field == "estimator":
                            # Generate combined initials (Sales Contact + Estimator)
                            from utils.word import get_combined_initials
                            from config.business_data import SALES_CONTACTS
                            
                            # Use sales_contact from project_data directly
                            sales_contact_name = project_data.get('sales_contact', '')
                            estimator_name = value  # This is the estimator field value
                            
                            # Generate combined initials using the actual sales contact selection
                            value = get_combined_initials(sales_contact_name, estimator_name)
                        elif field != "date":
                            value = str(value).title()
                        elif field == "date" and not value:
                            value = get_current_date()
                        sheet[cell] = value
                except Exception as e2:
                    print(f"Warning: Still could not write {field} to EBOX cell {cell} after unmerging: {str(e2)}")
                    continue
        
        # Add cost sheet identifier to N2
        write_cost_sheet_identifier(sheet, sheet.title, template_version)
        
    except Exception as e:
        print(f"Warning: Could not write EBOX metadata: {str(e)}")
        pass

def write_recoair_metadata(sheet: Worksheet, project_data: Dict, item_number: str = "1.01", template_version: str = None):
    """
    Write project metadata to RECOAIR sheet with specific cell mappings.
    
    Args:
        sheet (Worksheet): The RECOAIR worksheet to write to
        project_data (Dict): Project data dictionary
        item_number (str): Item number for this RecoAir sheet (e.g., "1.01", "2.01")
    """
    try:
        # RECOAIR-specific cell mappings (same as EBOX - D/H columns)
        recoair_cell_mappings = {
            "project_number": "D3",  # Job No
            "company": "D5",         # Company (changed from customer)
            "estimator": "D7",       # Sales Manager / Estimator Initials
            "project_name": "H3",    # Project Name
            "project_location": "H5",        # Project Location (was "location")
            "date": "H7",           # Date
            "revision": "O7",       # Revision
        }
        
        # Write RECOAIR-specific data
        try:
            sheet['C12'] = item_number  # Item number (1.01, 2.01, etc.)
            sheet['D37'] = 1  # Quantity
            # Note: Delivery location written by general loop to E37
            # N9 cell ready for RecoAir price (to be implemented)
            
            # Add plant selection dropdowns to E38 and E39
            add_plant_selection_dropdowns_to_recoair(sheet)
        except Exception as e:
            print(f"Warning: Could not write RECOAIR-specific data: {str(e)}")
        
        for field, cell in recoair_cell_mappings.items():
            try:
                value = project_data.get(field, "")
                
                # Handle special cases
                if field == "estimator":
                    # Generate combined initials (Sales Contact + Estimator) for RECOAIR sheets
                    from utils.word import get_combined_initials
                    
                    estimator_name = project_data.get("estimator", "")
                    
                    # Use sales_contact from project_data directly
                    sales_contact_name = project_data.get('sales_contact', '')
                    
                    # Generate combined initials using the actual sales contact selection
                    value = get_combined_initials(sales_contact_name, estimator_name)
                elif field == "revision":
                    value = project_data.get("revision", "")  # Use provided revision or leave blank for initial version
                elif field == "date":
                    # Keep date as is from project data
                    value = project_data.get("date", "")
                
                # Write the value to the cell
                sheet[cell] = value
                
            except Exception as e:
                print(f"Warning: Could not write {field} to RECOAIR cell {cell}: {str(e)}")
                
                # Try to handle merged cells
                try:
                    # Check if the cell is part of a merged range
                    for merged_range in sheet.merged_cells.ranges:
                        if cell in merged_range:
                            # Unmerge the range temporarily
                            sheet.unmerge_cells(str(merged_range))
                            # Write the value
                            sheet[cell] = value
                            # Re-merge the range
                            sheet.merge_cells(str(merged_range))
                            break
                    else:
                        # Cell is not merged, try writing again
                        sheet[cell] = value
                        
                except Exception as e2:
                    print(f"Warning: Still could not write {field} to RECOAIR cell {cell} after unmerging: {str(e2)}")
        
        # Add cost sheet identifier to N2
        write_cost_sheet_identifier(sheet, sheet.title, template_version)
        
    except Exception as e:
        print(f"Warning: Could not write RECOAIR metadata: {str(e)}")

def write_sdu_metadata(sheet: Worksheet, project_data: Dict, template_version: str = None, canopy_data: Dict = None):
    """
    Write project metadata to SDU sheet with specific cell mappings.
    
    Args:
        sheet (Worksheet): The SDU worksheet to write to
        project_data (Dict): Project data dictionary
        template_version (str, optional): Template version
        canopy_data (Dict, optional): Canopy data containing SDU item number
    """
    try:
        
        # Write SDU-specific data
        try:
            # Write SDU item number to B12 (if provided), otherwise default to "SDU"
            if canopy_data and canopy_data.get('sdu_item_number'):
                sheet['B12'] = canopy_data['sdu_item_number']
            else:
                sheet['B12'] = "SDU"  # Default if no item number provided
            
            # Write model name to C12
            sheet['C12'] = "SDU"  # Model name
            
            # Write quantity (1) to C97
            sheet['C97'] = 1
            
            # Note: Delivery location written by general loop to D97
        except Exception as e:
            print(f"Warning: Could not write SDU-specific data: {str(e)}")
        
        # Write project metadata to SDU-specific cells with merged cell handling
        def write_to_cell_safe(sheet, cell_ref, value):
            """Safely write to a cell, handling merged cells by unmerging first."""
            try:
                # Check if the cell is part of a merged range
                for merged_range in list(sheet.merged_cells.ranges):
                    if cell_ref in merged_range:
                        # Unmerge the range temporarily
                        sheet.unmerge_cells(str(merged_range))
                        # Write the value
                        sheet[cell_ref] = value
                        # Re-merge the range
                        sheet.merge_cells(str(merged_range))
                        return
                # If not merged, write directly
                sheet[cell_ref] = value
            except Exception as e:
                print(f"Warning: Could not write to {cell_ref}: {str(e)}")
        
        try:
            # Job No at C4
            write_to_cell_safe(sheet, 'C4', project_data.get('project_number', ''))
            
            # Company at C6 (changed from customer)
            write_to_cell_safe(sheet, 'C6', project_data.get('company', ''))
            
            # Sales Manager/Estimator Initials at C8
            estimator_name = project_data.get('estimator', '')
            if estimator_name:
                # Generate combined initials (Sales Contact + Estimator)
                from utils.word import get_combined_initials
                from config.business_data import SALES_CONTACTS
                
                # Get sales contact info based on estimator
                sales_contact_name = project_data.get('sales_contact', '')
                
                # Generate combined initials using the actual sales contact selection
                combined_initials = get_combined_initials(sales_contact_name, estimator_name)
                write_to_cell_safe(sheet, 'C8', combined_initials)
            
            # Project Name at F4 (corrected from G4)
            write_to_cell_safe(sheet, 'F4', project_data.get('project_name', ''))
            
            # Location at F6 (corrected from G6)
            write_to_cell_safe(sheet, 'F6', project_data.get('project_location', ''))
            
            # Date at F8 (corrected from G8)
            write_to_cell_safe(sheet, 'F8', project_data.get('date', ''))
            
            # Revision at K7
            write_to_cell_safe(sheet, 'K7', project_data.get('revision', ''))
            
        except Exception as e:
            print(f"Warning: Could not write SDU project metadata: {str(e)}")
        
        # Add cost sheet identifier to N2
        write_cost_sheet_identifier(sheet, sheet.title, template_version)
        
    except Exception as e:
        print(f"Warning: Could not write SDU metadata: {str(e)}")


def write_vent_clg_metadata(sheet: Worksheet, project_data: Dict, template_version: str = None):
    """
    Write VENT CLG-specific metadata to the VENT CLG sheet.
    
    Args:
        sheet (Worksheet): The VENT CLG worksheet
        project_data (Dict): Project data containing metadata
        template_version (str): Version of the template being used
    """
    try:
        # Use correct column mappings for VENT CLG sheet
        cell_mappings = {
            "project_number": "C3",     # Job No in column C
            "company": "C5",            # Customer in column C
            "estimator": "C7",          # Sales Manager / Estimator in column C
            "project_name": "F3",       # Project Name in column F
            "project_location": "F5",   # Project Location in column F
            "date": "F7",               # Date in column F
            "revision": "K7",           # Revision in column K
        }
        
        # Helper function to safely write to cells
        def write_to_cell_safe(sheet, cell_ref, value):
            try:
                sheet[cell_ref] = value
            except Exception as e:
                print(f"Warning: Could not write to cell {cell_ref} in write_vent_clg_metadata: {str(e)}")
        
        # Write project metadata using the mappings
        write_to_cell_safe(sheet, cell_mappings["project_number"], project_data.get("project_number", ""))
        write_to_cell_safe(sheet, cell_mappings["company"], project_data.get("company", ""))
        write_to_cell_safe(sheet, cell_mappings["estimator"], project_data.get("estimator", ""))
        write_to_cell_safe(sheet, cell_mappings["project_name"], project_data.get("project_name", ""))
        write_to_cell_safe(sheet, cell_mappings["project_location"], project_data.get("project_location", ""))
        write_to_cell_safe(sheet, cell_mappings["date"], project_data.get("date", ""))
        write_to_cell_safe(sheet, cell_mappings["revision"], project_data.get("revision", ""))
        
        # Write additional VENT CLG-specific metadata with default values
        # Do NOT overwrite D44 for vent sealing tabs - leave delivery location as is
        write_to_cell_safe(sheet, "D45", project_data.get("plant_selection", "PLANT SELECTION (weekly)"))    # Plant selection dropdown
        
        # Add cost sheet identifier to N2
        write_cost_sheet_identifier(sheet, sheet.title, template_version)
        
    except Exception as e:
        print(f"Error writing VENT CLG metadata: {str(e)}")


def add_vent_clg_dropdowns(sheet: Worksheet):
    """
    Add VENT CLG-specific dropdowns to the VENT CLG sheet.
    Note: Delivery location dropdown is handled by add_delivery_location_dropdown_to_sheet()
    
    Args:
        sheet (Worksheet): The VENT CLG worksheet
    """
    try:
        # Plant selection options for VENT CLG sheets
        plant_selection_options = [
            "",
            "SL10 GENIE",
            "EXTENSION FORKS",
            "2.5M COMBI LADDER",
            "1.5M PODIUM",
            "3M TOWER",
            "COMBI LADDER",
            "PECO LIFT",
            "3M YOUNGMAN BOARD",
            "GS1930 SCISSOR LIFT",
            "4-6 SHERASCOPIC",
            "7-9 SHERASCOPIC"
        ]
        
        def create_validation(options):
            """Create validation with proper formatting, handling Excel's 255-character formula limit"""
            formula = ",".join(options)
            if len(formula) > 255:  # Excel formula limit - use hidden cells approach
                # Write options to hidden cells
                start_row = 600  # Use row 600+ to avoid conflicts with other hidden data
                for i, option in enumerate(options):
                    try:
                        sheet[f'AD{start_row + i}'] = option  # Use column AD (hidden area)
                    except:
                        pass  # If we can't write, continue
                
                # Create range reference for validation
                end_row = start_row + len(options) - 1
                range_ref = f'$AD${start_row}:$AD${end_row}'
                return DataValidation(type="list", formula1=range_ref, allow_blank=True)
            else:
                return DataValidation(type="list", formula1=f'"{formula}"', allow_blank=True)
        
        # Create plant selection validation
        plant_selection_dv = create_validation(plant_selection_options)
        
        # Add validation to sheet
        sheet.add_data_validation(plant_selection_dv)
        
        # Apply dropdown to plant selection cell
        plant_selection_dv.add('D45')  # Plant selection dropdown
        
        # Set default value for plant selection
        sheet['D45'] = "PLANT SELECTION (weekly)"
        
        print(f"✅ Added VENT CLG plant selection dropdown to D45")
        
    except Exception as e:
        print(f"Warning: Could not add VENT CLG dropdowns: {str(e)}")


def add_fire_suppression_dropdowns(sheet: Worksheet):
    """
    Add specific dropdowns for fire suppression sheets.
    
    Args:
        sheet (Worksheet): The fire suppression worksheet to add dropdowns to
    """
    try:
        # Fire suppression specific options (shortened to avoid Excel corruption)
        system_types = [
            "1 TANK SYSTEM",
            "1 TANK TRAVEL HUB",
            "1 TANK DISTANCE",
            "NOBEL",
            "AMAREX",
            "OTHER",
            "2 TANK SYSTEM",
            "2 TANK TRAVEL HUB",
            "2 TANK DISTANCE",
            "3 TANK SYSTEM",
            "3 TANK TRAVEL HUB",
            "3 TANK DISTANCE",
            "4 TANK SYSTEM",
            "4 TANK TRAVEL HUB",
            "4 TANK DISTANCE",
            "5 TANK SYSTEM",
            "5 TANK TRAVEL HUB",
            "5 TANK DISTANCE",
            "6 TANK SYSTEM",
            "6 TANK TRAVEL HUB",
            "6 TANK DISTANCE"
            ]
        
        tank_sizes = [
            "1 TANK",
            "1 TANK DISTANCE",
            "2 TANK",
            "2 TANK DISTANCE",
            "3 TANK",
            "3 TANK DISTANCE",
            "4 TANK",
            "4 TANK DISTANCE",
            "5 TANK",
            "5 TANK DISTANCE",
            "6 TANK",
            "6 TANK DISTANCE"
        ]
        
        # Create data validations - use hidden cells for long lists to avoid Excel 255 char limit
        def create_validation(options):
            formula = ",".join(options)
            if len(formula) > 255:  # Excel formula limit - use hidden cells approach
                # Write options to hidden cells
                start_row = 500
                for i, option in enumerate(options):
                    cell_ref = f"AA{start_row + i}"
                    sheet[cell_ref] = option
                
                # Create range reference for validation
                end_row = start_row + len(options) - 1
                range_ref = f"$AA${start_row}:$AA${end_row}"
                return DataValidation(type="list", formula1=range_ref, allow_blank=True)
            else:
                return DataValidation(type="list", formula1=f'"{formula}"', allow_blank=True)
        
        system_dv = create_validation(system_types)
        tank_dv = create_validation(tank_sizes)
        
        # Add validations to sheet
        sheet.add_data_validation(system_dv)
        sheet.add_data_validation(tank_dv)
        
        # Apply to specific cells with error handling
        # Note: Only apply fire suppression dropdowns to fire suppression sheets, not canopy sheets
        try:
            # Fire suppression system type (C16, C33, C50) - ONLY on fire suppression sheets
            if "FIRE" in sheet.title.upper() or "SUPP" in sheet.title.upper():
                system_dv.add("C16")
                system_dv.add("C33") 
                system_dv.add("C50")
                
                # Tank installation options (C17, C34, C51) - ONLY on fire suppression sheets
                tank_dv.add("C17")
                tank_dv.add("C34")
                tank_dv.add("C51")
                
                # Add plant selection dropdown to D184 for fire suppression sheets
                add_plant_selection_dropdown_to_fire_supp(sheet)
            else:
                print(f"ℹ️  Skipping fire suppression dropdowns for non-fire suppression sheet: {sheet.title}")
        except Exception as e:
            print(f"Warning: Could not add fire suppression dropdown cells: {str(e)}")
        
    except Exception as e:
        print(f"Warning: Could not add fire suppression dropdowns to sheet {sheet.title}: {str(e)}")
        pass

def add_recoair_dropdowns(sheet: Worksheet):
    """
    Add specific dropdowns for RecoAir sheets.
    
    Args:
        sheet (Worksheet): The RecoAir worksheet to add dropdowns to
    """
    try:
        # RecoAir specific options
        internal_external_options = [
            "INTERNAL",
            "EXTERNAL"
        ]
        
        # Create data validation
        def create_validation(options):
            formula = ",".join(options)
            return DataValidation(type="list", formula1=f'"{formula}"', allow_blank=True)
        
        internal_external_dv = create_validation(internal_external_options)
        
        # Add validation to sheet
        sheet.add_data_validation(internal_external_dv)
        
        # Apply to range I14 to I28
        try:
            for row in range(14, 29):  # 14 to 28 inclusive
                internal_external_dv.add(f"I{row}")
        except Exception as e:
            print(f"Warning: Could not add RecoAir dropdown cells: {str(e)}")
        
    except Exception as e:
        print(f"Warning: Could not add RecoAir dropdowns to sheet {sheet.title}: {str(e)}")
        pass

def add_sdu_dropdowns(sheet: Worksheet):
    """
    Add specific dropdowns for SDU sheets.
    
    Args:
        sheet (Worksheet): The SDU worksheet to add dropdowns to
    """
    try:
        # Water connection types for E89 to E91 (CWS/HWS)
        water_types_options = [
            "",  # Empty option
            "CWS",
            "HWS"
        ]
        
        # Water connection sizes for D89 to D91
        water_sizes_options = [
            "",  # Empty option
            "15mm",
            "22mm", 
            "28mm"
        ]
        
        # MCB #-way options for D35
        mcb_way_options = [
            "",  # Empty option
            "MCB 4-WAY 125A",
            "MCB 6-WAY 125A",
            "MCB 8-WAY 125A",
            "MCB 16-WAY 125A",
            "MCB 18-WAY 125A",
            "MCB 24-WAY 125A",
            "MCB 4-WAY 160A",
            "MCB 6-WAY 160A",
            "MCB 8-WAY 160A",
            "MCB 16-WAY 160A",
            "MCB 18-WAY 160A",
            "MCB 4-WAY 250A",
            "MCB 6-WAY 250A",
            "MCB 8-WAY 250A",
            "MCB 16-WAY 250A",
            "MCB 18-WAY 250A",
            "MCB 24-WAY 250A"
        ]
        
        # # Amp (No MCB) options for D40 to D47
        amp_no_mcb_options = [
            "",  # Empty option
            "16 AMP 1-PH ISO/OUTLET(NO MCB)",
            "32 AMP 1-PH ISO/OUTLET(NO MCB)",
            "16 AMP 3-PH ISO/OUTLET(NO MCB)",
            "32 AMP 3-PH ISO/OUTLET(NO MCB)",
            "63 AMP 3-PH ISO/OUTLET(NO MCB)",
            "125 AMP 3-PH ISO/OUTLET(NO MCB)"
        ]
        
        # (MCB) options for D49 to D56
        amp_mcb_options = [
            "",  # Empty option
            "16 AMP 1-PH ISO/OUTLET(MCB)",
            "32 AMP 1-PH ISO/OUTLET(MCB)",
            "16 AMP 3-PH ISO/OUTLET(MCB + VIGI)",
            "32 AMP 3-PH ISO/OUTLET(MCB + VIGI)",
            "63 AMP 3-PH ISO/OUTLET(MCB)",
            "125 AMP 3-PH ISO/OUTLET(MCB)",
            "16 AMP 3-PH ISO/OUTLET(MCB + RCD)",
            "32 AMP 3-PH ISO/OUTLET(MCB + RCD)"
        ]
        
        # Create data validation function that handles long lists
        def create_validation(options):
            formula = ",".join(options)
            if len(formula) > 255:  # Excel formula limit
                # For longer lists, write them to hidden cells and reference them
                start_row = 500 + len(options)  # Start at row 500+ to avoid conflicts with delivery locations
                for i, option in enumerate(options):
                    try:
                        sheet[f'AC{start_row + i}'] = option  # Use column AC (hidden area)
                    except:
                        pass  # If we can't write, fall back
                
                # Create a validation that references the range
                try:
                    range_ref = f'$AC${start_row}:$AC${start_row + len(options) - 1}'
                    return DataValidation(type="list", formula1=range_ref, allow_blank=True)
                except:
                    # Fallback to allowing any text input
                    return DataValidation(type="textLength", operator="lessThan", formula1="100", allow_blank=True)
            else:
                return DataValidation(type="list", formula1=f'"{formula}"', allow_blank=True)
        
        # Create all validations
        water_types_dv = create_validation(water_types_options)
        water_sizes_dv = create_validation(water_sizes_options)
        mcb_way_dv = create_validation(mcb_way_options)
        amp_no_mcb_dv = create_validation(amp_no_mcb_options)
        amp_mcb_dv = create_validation(amp_mcb_options)
        
        # Add validations to sheet
        sheet.add_data_validation(water_types_dv)
        sheet.add_data_validation(water_sizes_dv)
        sheet.add_data_validation(mcb_way_dv)
        sheet.add_data_validation(amp_no_mcb_dv)
        sheet.add_data_validation(amp_mcb_dv)
        
        # Apply water types to cells E89 to E91 and set default to CWS
        try:
            for row in range(89, 92):  # E89 to E91 inclusive
                water_types_dv.add(f"E{row}")
                # Set default value to CWS
                sheet[f"E{row}"] = "CWS"
        except Exception as e:
            print(f"Warning: Could not add SDU water types dropdown cells: {str(e)}")
        
        # Apply water sizes to cells D89 to D91
        try:
            for row in range(89, 92):  # D89 to D91 inclusive
                water_sizes_dv.add(f"D{row}")
        except Exception as e:
            print(f"Warning: Could not add SDU water sizes dropdown cells: {str(e)}")
        
        # Apply MCB #-way dropdown to D35
        try:
            mcb_way_dv.add("D35")
            print(f"✅ Added MCB #-way dropdown to D35 on SDU sheet")
        except Exception as e:
            print(f"Warning: Could not add MCB #-way dropdown to D35: {str(e)}")
        
        # Apply # Amp (No MCB) dropdowns to D40 to D47
        try:
            for row in range(40, 48):  # D40 to D47 inclusive
                amp_no_mcb_dv.add(f"D{row}")
            print(f"✅ Added # Amp (No MCB) dropdowns to D40-D47 on SDU sheet")
        except Exception as e:
            print(f"Warning: Could not add # Amp (No MCB) dropdown cells: {str(e)}")
        
        # Apply (MCB) dropdowns to D49 to D56
        try:
            for row in range(49, 57):  # D49 to D56 inclusive
                amp_mcb_dv.add(f"D{row}")
            print(f"✅ Added (MCB) dropdowns to D49-D56 on SDU sheet")
        except Exception as e:
            print(f"Warning: Could not add (MCB) dropdown cells: {str(e)}")
        
    except Exception as e:
        print(f"Warning: Could not add SDU dropdowns to sheet {sheet.title}: {str(e)}")
        pass

def organize_sheets_by_area(wb: Workbook):
    """
    Organize sheets by area grouping: JOB TOTAL first, then if contract option enabled:
    CONTRACT, EXTRACT DUCT, SUPPLY DUCT, SPIRAL DUCT, then all sheets for each area together
    (CANOPY, CANOPY (UV), FIRE SUPP, EBOX, RECOAIR, SDU for area 1, then area 2, etc.)
    
    Args:
        wb (Workbook): The workbook to reorganize
    """
    try:
        # Categorize all sheets
        job_total_sheets = []
        contract_sheets = []
        area_sheets = {}  # Dictionary to group sheets by area
        misc_sheets = []
        
        for sheet_name in wb.sheetnames:
            if 'JOB TOTAL' in sheet_name:
                job_total_sheets.append(sheet_name)
            elif sheet_name in ['CONTRACT', 'EXTRACT DUCT', 'SUPPLY DUCT', 'SPIRAL DUCT']:
                contract_sheets.append(sheet_name)
            elif any(sys_type in sheet_name for sys_type in ['CANOPY', 'FIRE SUPP', 'EBOX', 'RECOAIR', 'SDU', 'MARVEL', 'VENT CLG']):
                # Extract area identifier for grouping
                if ' - ' in sheet_name and '(' in sheet_name and ')' in sheet_name:
                    parts = sheet_name.split(' - ', 1)
                    if len(parts) == 2:
                        area_identifier = parts[1]  # e.g., "LEVEL 1 (1)"
                        
                        if area_identifier not in area_sheets:
                            area_sheets[area_identifier] = []
                        area_sheets[area_identifier].append(sheet_name)
                else:
                    misc_sheets.append(sheet_name)
            else:
                misc_sheets.append(sheet_name)
        
        # Sort sheets within each area by system type priority
        def get_system_priority(sheet_name):
            """Return sort priority for different system types within an area"""
            if 'CANOPY (UV)' in sheet_name:
                return 0  # UV canopies first
            elif 'CANOPY - ' in sheet_name:
                return 1  # Regular canopies second
            elif 'FIRE SUPP' in sheet_name:
                return 2  # Fire suppression third
            elif 'EBOX' in sheet_name:
                return 3  # Edge boxes fourth
            elif 'RECOAIR' in sheet_name:
                return 4  # RecoAir fifth
            elif 'SDU' in sheet_name:
                return 5  # SDU sixth
            elif 'MARVEL' in sheet_name:
                return 6  # MARVEL seventh
            elif 'VENT CLG' in sheet_name:
                return 7  # VENT CLG eighth
            else:
                return 8  # Any other system types
        
        # Sort sheets within each area
        for area_id in area_sheets:
            area_sheets[area_id].sort(key=get_system_priority)
        
        # Sort areas by their identifier (LEVEL 1 (1), LEVEL 1 (2), LEVEL 2 (1), etc.)
        def get_area_sort_key(area_id):
            """Extract level and area number for sorting"""
            try:
                # Parse "LEVEL X (Y)" format
                if 'LEVEL' in area_id and '(' in area_id and ')' in area_id:
                    # Extract level number and area number
                    level_part = area_id.split('(')[0].strip()  # "LEVEL X"
                    area_part = area_id.split('(')[1].split(')')[0].strip()  # "Y"
                    
                    level_num = int(level_part.split()[-1])  # Extract X from "LEVEL X"
                    area_num = int(area_part)  # Extract Y from "Y"
                    
                    return (level_num, area_num)
                else:
                    return (999, 999)  # Put unrecognized formats at the end
            except:
                return (999, 999)  # Put unparseable formats at the end
        
        sorted_areas = sorted(area_sheets.keys(), key=get_area_sort_key)
        
        # Create final ordered list: JOB TOTAL → CONTRACT SHEETS → AREA 1 SHEETS → AREA 2 SHEETS → ... → MISC
        ordered_sheets = []
        ordered_sheets.extend(job_total_sheets)  # JOB TOTAL first
        
        # Add contract sheets after JOB TOTAL in specific order
        contract_names = ['CONTRACT', 'EXTRACT DUCT', 'SUPPLY DUCT', 'SPIRAL DUCT']  # Updated order
        for name in contract_names:
            # Find exact sheet name in current workbook
            if name in wb.sheetnames:
                ordered_sheets.append(name)
        
        # Add all sheets for each area in order
        for area_id in sorted_areas:
            ordered_sheets.extend(area_sheets[area_id])
        
        # Add miscellaneous sheets (excluding Lists at the end)
        ordered_sheets.extend([s for s in misc_sheets if s != 'Lists'])  # Other misc sheets
        ordered_sheets.extend([s for s in misc_sheets if s == 'Lists'])  # Lists last
        
        # Reorder the sheets in the workbook
        current_sheets = wb.sheetnames.copy()
        
        # Move sheets to the correct order
        for i, target_sheet in enumerate(ordered_sheets):
            if target_sheet in current_sheets:
                current_index = wb.sheetnames.index(target_sheet)
                if current_index != i:
                    # Move the sheet to the correct position
                    sheet_to_move = wb[target_sheet]
                    wb.move_sheet(sheet_to_move, offset=i - current_index)
        
    except Exception as e:
        print(f"Warning: Could not organize sheets by area: {str(e)}")
        pass

def write_company_data_to_hidden_sheet(wb: Workbook, project_data: Dict):
    """
    Write company and estimator data to a hidden sheet for later extraction.
    
    Args:
        wb (Workbook): The workbook to add the hidden sheet to
        project_data (Dict): Project data including company and estimator info
    """
    try:
        # Create or get the hidden sheet
        sheet_name = "ProjectData"
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            sheet = wb.create_sheet(sheet_name)
        
        # Hide the sheet
        sheet.sheet_state = 'hidden'
        
        # Write customer information
        sheet['A1'] = 'Customer'
        sheet['B1'] = project_data.get('customer', '')
        
        # Write company information
        sheet['A2'] = 'Company'
        sheet['B2'] = project_data.get('company', '')
        
        sheet['A3'] = 'Address'
        sheet['B3'] = project_data.get('address', '')
        
        # Write estimator information (full name, not initials)
        sheet['A4'] = 'Estimator_Full_Name'
        sheet['B4'] = project_data.get('estimator', '')
        
        # Get estimator rank from business data
        estimator_name = project_data.get('estimator', '')
        estimator_rank = 'Estimator'  # Default
        
        # Look up the rank from ESTIMATORS dictionary
        from config.business_data import ESTIMATORS
        for name, rank in ESTIMATORS.items():
            if name.lower() in estimator_name.lower():
                estimator_rank = rank
                break
        
        sheet['A5'] = 'Estimator_Rank'
        sheet['B5'] = estimator_rank
        
        # Write sales contact information
        sheet['A6'] = 'Sales_Contact'
        sheet['B6'] = project_data.get('sales_contact', '')
        
        # Write delivery location if available
        sheet['A7'] = 'Delivery_Location'
        sheet['B7'] = project_data.get('delivery_location', '')
        
        # Write revision information
        sheet['A8'] = 'Revision'
        sheet['B8'] = project_data.get('revision', '')
        
    except Exception as e:
        print(f"Warning: Could not write company data to hidden sheet: {str(e)}")
        pass

def add_delivery_location_dropdown_to_sheet(sheet: Worksheet, selected_delivery_location: str = ""):
    """
    Add delivery location dropdown to the appropriate cell based on sheet type.
    
    Args:
        sheet (Worksheet): The worksheet to add dropdown to
        selected_delivery_location (str): The pre-selected delivery location value
    """
    try:
        from config.business_data import DELIVERY_LOCATIONS
        from openpyxl.worksheet.datavalidation import DataValidation
        
        sheet_name = sheet.title.upper()
        
        # Determine cell based on sheet type
        if "SDU" in sheet_name:
            cell = "D97"
        elif "EBOX" in sheet_name or "EDGE BOX" in sheet_name:
            cell = "E38"
        elif "RECOAIR" in sheet_name:
            cell = "E37"
        elif "MARVEL" in sheet_name:
            cell = "D54"  # MARVEL sheets use D54 for delivery location
        elif "VENT CLG" in sheet_name:
            cell = "D44"  # VENT CLG sheets use D44 for delivery location
        elif "FIRE" in sheet_name and "SUPP" in sheet_name:
            cell = "D186"
        elif "CANOPY" in sheet_name:
            cell = "D183"
        else:
            # Default to D183 for other sheet types
            cell = "D183"
        
        # Create delivery location dropdown
        # Note: Need to handle the long list of delivery locations
        # Since Excel has a 255-character formula limit, we'll write to hidden cells and reference them
        start_row = 400  # Use row 400+ to avoid conflicts
        
        for i, location in enumerate(DELIVERY_LOCATIONS):
            try:
                sheet[f'AB{start_row + i}'] = location  # Use column AB (hidden area)
            except:
                pass  # If we can't write, continue
        
        # Create a validation that references the range
        try:
            range_ref = f'$AB${start_row}:$AB${start_row + len(DELIVERY_LOCATIONS) - 1}'
            delivery_dv = DataValidation(type="list", formula1=range_ref, allow_blank=True)
            
            # Add validation to sheet
            sheet.add_data_validation(delivery_dv)
            
            # Apply to the specific cell
            delivery_dv.add(cell)
            
            # Set the selected value if provided and not "Select..."
            if selected_delivery_location and selected_delivery_location != "Select...":
                sheet[cell] = selected_delivery_location
            elif not selected_delivery_location and "VENT CLG" in sheet_name:
                # Set default for VENT CLG sheets when no delivery location is provided
                sheet[cell] = "SELECT LOCATION..."
            
            print(f"📍 Added delivery location dropdown to {cell} on {sheet.title}")
            if selected_delivery_location and selected_delivery_location != "Select...":
                print(f"   Pre-selected: '{selected_delivery_location}'")
                
        except Exception as e:
            print(f"Warning: Could not create delivery location dropdown, writing value directly: {str(e)}")
            # Fallback to writing the value directly
            if selected_delivery_location and selected_delivery_location != "Select...":
                sheet[cell] = selected_delivery_location
    
    except Exception as e:
        print(f"Warning: Could not add delivery location dropdown to sheet {sheet.title}: {str(e)}")
        pass

def save_to_excel(project_data: Dict, template_path: str = None) -> str:
    """
    Generate a complete Excel workbook from project data.
    
    Args:
        project_data (Dict): Complete project specification data
        template_path (str, optional): Path to the template file to use
    
    Returns:
        str: Path to the saved Excel file
    """
    try:
        # Load template and detect version
        wb = load_template_workbook(template_path)
        template_version = detect_template_version(wb)
        print(f"🔍 Detected template version: {template_version}")
        
        # Get all sheets once and create lists of available sheets
        all_sheets = wb.sheetnames
        canopy_sheets = [sheet for sheet in all_sheets if 'CANOPY' in sheet]
        fire_supp_sheets = [sheet for sheet in all_sheets if 'FIRE SUPP' in sheet or 'FIRE SUPPRESSION' in sheet]
        edge_box_sheets = [sheet for sheet in all_sheets if 'EBOX' in sheet or 'EDGE BOX' in sheet]
        recoair_sheets = [sheet for sheet in all_sheets if 'RECOAIR' in sheet]
        sdu_sheets = [sheet for sheet in all_sheets if 'SDU' in sheet and 'CANOPY' not in sheet and 'FIRE' not in sheet]
        # MARVEL template sheets (for UV grease recovery option)
        marvel_sheets = [sheet for sheet in all_sheets if 'MARVEL' in sheet]
        # VENT CLG template sheets (for ventilated ceiling systems)
        vent_clg_sheets = [sheet for sheet in all_sheets if 'VENT CLG' in sheet or 'VENTILATED CEILING' in sheet]
        # Contract template sheets - handle exact matches and numbered variants
        contract_sheets = [sheet for sheet in all_sheets if sheet.strip() == 'CONTRACT' or sheet.startswith('CONTRACT')]
        spiral_duct_sheets = [sheet for sheet in all_sheets if sheet.strip() == 'SPIRAL DUCT' or sheet.startswith('SPIRAL DUCT')]
        supply_duct_sheets = [sheet for sheet in all_sheets if sheet.strip() == 'SUPPLY DUCT' or sheet.startswith('SUPPLY DUCT')]
        extract_duct_sheets = [sheet for sheet in all_sheets if sheet.strip() == 'EXTRACT DUCT' or sheet.startswith('EXTRACT DUCT')]
        
        # Hide the Lists sheet if it exists
        if 'Lists' in wb.sheetnames:
            wb['Lists'].sheet_state = 'hidden'
        
        # Add project metadata to JOB TOTAL sheet by default
        if 'JOB TOTAL' in wb.sheetnames:
            job_total_sheet = wb['JOB TOTAL']
            write_project_metadata(job_total_sheet, project_data, template_version)
            job_total_sheet.sheet_state = 'visible'
        
        # Write company and estimator data to hidden sheet
        write_company_data_to_hidden_sheet(wb, project_data)
        
        # Track created contract sheet names for organization
        created_contract_sheets = []
        
        # Create contract sheets if contract option is enabled
        if project_data.get('contract_option', False):
            print("🔨 Contract option enabled - creating contract sheets")
            
            # Create Contract sheet
            contract_actual_name = None
            if contract_sheets:
                contract_sheet_name = contract_sheets.pop(0)
                contract_sheet = wb[contract_sheet_name]
                print(f"Original contract sheet name: {contract_sheet_name}")
                
                # Try to rename the sheet, handle conflicts
                try:
                    # If a sheet named "CONTRACT" already exists, remove it first
                    if "CONTRACT" in wb.sheetnames and wb["CONTRACT"] != contract_sheet:
                        wb.remove(wb["CONTRACT"])
                        print("Removed existing CONTRACT sheet")
                    
                    contract_sheet.title = "CONTRACT"
                    contract_actual_name = "CONTRACT"
                    print(f"✓ Successfully renamed to: {contract_actual_name}")
                except Exception as e:
                    # If renaming fails, keep the original name
                    contract_actual_name = contract_sheet_name
                    print(f"Warning: Could not rename contract sheet to CONTRACT, keeping name: {contract_actual_name}. Error: {str(e)}")
                
                contract_sheet.sheet_state = 'visible'
                
                # Write basic project metadata to contract sheet
                write_project_metadata(contract_sheet, project_data, template_version)
                created_contract_sheets.append(contract_actual_name)
                print(f"✓ Created contract sheet with name: {contract_actual_name}")
                
                # Add dropdowns and fix column alignment
                try:
                    # Add delivery location dropdown (D57) using the same options as canopy sheets
                    from config.business_data import DELIVERY_LOCATIONS
                    # Write delivery locations to hidden cells and reference them
                    start_row = 400  # Use row 400+ to avoid conflicts
                    for i, location in enumerate(DELIVERY_LOCATIONS):
                        try:
                            contract_sheet[f'AB{start_row + i}'] = location  # Use column AB (hidden area)
                        except:
                            pass  # If we can't write, continue
                    
                    # Create a validation that references the range
                    range_ref = f'$AB${start_row}:$AB${start_row + len(DELIVERY_LOCATIONS) - 1}'
                    delivery_dv = DataValidation(type="list", formula1=range_ref, allow_blank=True)
                    contract_sheet.add_data_validation(delivery_dv)
                    delivery_dv.add('D57:D57')
                    
                    # Add plant selection dropdown (D58) using the same options as canopy sheets
                    plant_options = [
                        "",  # Empty option
                        "SL10 GENIE",
                        "EXTENSION FORKS",
                        "2.5M COMBI LADDER",
                        "1.5M PODIUM",
                        "3M TOWER",
                        "COMBI LADDER",
                        "PECO LIFT",
                        "3M YOUNGMAN BOARD",
                        "GS1930 SCISSOR LIFT",
                        "4-6 SHERASCOPIC",
                        "7-9 SHERASCOPIC"
                    ]
                    
                    # Create validation for plant selection
                    formula = ",".join(plant_options)
                    plant_dv = DataValidation(type="list", formula1=f'"{formula}"', allow_blank=True)
                    contract_sheet.add_data_validation(plant_dv)
                    plant_dv.add('D58:D58')
                    
                    # Contract sheet structure should remain as-is
                    # Do not move data between columns as it interferes with duct formulas
                    
                    print("✓ Added dropdowns and fixed column alignment in contract sheet")
                except Exception as e:
                    print(f"Warning: Could not add dropdowns or fix alignment in contract sheet: {str(e)}")
            else:
                print(f"Warning: No CONTRACT template sheet found")
            
            # Create Spiral Duct sheet
            spiral_duct_actual_name = None
            if spiral_duct_sheets:
                spiral_duct_sheet_name = spiral_duct_sheets.pop(0)
                spiral_duct_sheet = wb[spiral_duct_sheet_name]
                print(f"Original spiral duct sheet name: {spiral_duct_sheet_name}")
                
                try:
                    if "SPIRAL DUCT" in wb.sheetnames and wb["SPIRAL DUCT"] != spiral_duct_sheet:
                        wb.remove(wb["SPIRAL DUCT"])
                        print("Removed existing SPIRAL DUCT sheet")
                    
                    spiral_duct_sheet.title = "SPIRAL DUCT"
                    spiral_duct_actual_name = "SPIRAL DUCT"
                    print(f"✓ Successfully renamed to: {spiral_duct_actual_name}")
                except Exception as e:
                    spiral_duct_actual_name = spiral_duct_sheet_name
                    print(f"Warning: Could not rename spiral duct sheet, keeping name: {spiral_duct_actual_name}. Error: {str(e)}")
                
                spiral_duct_sheet.sheet_state = 'visible'
                created_contract_sheets.append(spiral_duct_actual_name)
                print(f"✓ Created spiral duct sheet with name: {spiral_duct_actual_name}")
            else:
                print(f"Warning: No SPIRAL DUCT template sheet found")
            
            # Create Supply Duct sheet
            supply_duct_actual_name = None
            if supply_duct_sheets:
                supply_duct_sheet_name = supply_duct_sheets.pop(0)
                supply_duct_sheet = wb[supply_duct_sheet_name]
                print(f"Original supply duct sheet name: {supply_duct_sheet_name}")
                
                try:
                    if "SUPPLY DUCT" in wb.sheetnames and wb["SUPPLY DUCT"] != supply_duct_sheet:
                        wb.remove(wb["SUPPLY DUCT"])
                        print("Removed existing SUPPLY DUCT sheet")
                    
                    supply_duct_sheet.title = "SUPPLY DUCT"
                    supply_duct_actual_name = "SUPPLY DUCT"
                    print(f"✓ Successfully renamed to: {supply_duct_actual_name}")
                except Exception as e:
                    supply_duct_actual_name = supply_duct_sheet_name
                    print(f"Warning: Could not rename supply duct sheet, keeping name: {supply_duct_actual_name}. Error: {str(e)}")
                
                supply_duct_sheet.sheet_state = 'visible'
                created_contract_sheets.append(supply_duct_actual_name)
                print(f"✓ Created supply duct sheet with name: {supply_duct_actual_name}")
            else:
                print(f"Warning: No SUPPLY DUCT template sheet found")
            
            # Create Extract Duct sheet
            extract_duct_actual_name = None
            if extract_duct_sheets:
                extract_duct_sheet_name = extract_duct_sheets.pop(0)
                extract_duct_sheet = wb[extract_duct_sheet_name]
                print(f"Original extract duct sheet name: {extract_duct_sheet_name}")
                
                try:
                    if "EXTRACT DUCT" in wb.sheetnames and wb["EXTRACT DUCT"] != extract_duct_sheet:
                        wb.remove(wb["EXTRACT DUCT"])
                        print("Removed existing EXTRACT DUCT sheet")
                    
                    extract_duct_sheet.title = "EXTRACT DUCT"
                    extract_duct_actual_name = "EXTRACT DUCT"
                    print(f"✓ Successfully renamed to: {extract_duct_actual_name}")
                except Exception as e:
                    extract_duct_actual_name = extract_duct_sheet_name
                    print(f"Warning: Could not rename extract duct sheet, keeping name: {extract_duct_actual_name}. Error: {str(e)}")
                
                extract_duct_sheet.sheet_state = 'visible'
                created_contract_sheets.append(extract_duct_actual_name)
                print(f"✓ Created extract duct sheet with name: {extract_duct_actual_name}")
            else:
                print(f"Warning: No EXTRACT DUCT template sheet found")
            
            # Add formulas and conditional values to CONTRACT sheet after all duct sheets are created
            if contract_actual_name and contract_actual_name in wb.sheetnames:
                contract_sheet = wb[contract_actual_name]
                print(f"Adding formulas to contract sheet: {contract_actual_name}")
                
                try:
                    # Add formulas to reference duct sheet totals using actual sheet names
                    if extract_duct_actual_name and extract_duct_actual_name in wb.sheetnames:
                        # Extract Duct System
                        contract_sheet['F16'] = f"='{extract_duct_actual_name}'!V5"  # Total price
                        contract_sheet['C16'] = f"=IF('{extract_duct_actual_name}'!V5>0,1,0)"  # Conditional value
                        # Keep extract system price in M12 (no change needed)
                        print(f"✓ Added extract duct formulas referencing: {extract_duct_actual_name}")
                    
                    if supply_duct_actual_name and supply_duct_actual_name in wb.sheetnames:
                        # Supply Duct System
                        contract_sheet['F30'] = f"='{supply_duct_actual_name}'!V5"  # Total price
                        contract_sheet['C30'] = f"=IF('{supply_duct_actual_name}'!V5>0,1,0)"  # Conditional value
                        # Keep supply system price in N12 (no change needed)
                        print(f"✓ Added supply duct formulas referencing: {supply_duct_actual_name}")
                    
                    if spiral_duct_actual_name and spiral_duct_actual_name in wb.sheetnames:
                        # Spiral Duct System
                        contract_sheet['F37'] = f"='{spiral_duct_actual_name}'!V5"  # Total price
                        contract_sheet['C37'] = f"=IF('{spiral_duct_actual_name}'!V5>0,1,0)"  # Conditional value
                        print(f"✓ Added spiral duct formulas referencing: {spiral_duct_actual_name}")
                    
                    print(f"✓ Successfully added all duct sheet references and conditional values to {contract_actual_name} sheet")
                except Exception as e:
                    print(f"Warning: Could not add duct references to {contract_actual_name} sheet: {str(e)}")
                    import traceback
                    print(f"Full error: {traceback.format_exc()}")
        
        # Counters for sheet numbering
        recoair_sheet_count = 0
        
        # Keep track of total areas for coloring
        area_count = 0
        
        # Create a mapping of level names to area numbers for proper sheet numbering
        level_area_numbers = {}
        
        # Process each level and area
        for level in project_data.get("levels", []):
            level_number = level["level_number"]
            level_name = level.get("level_name", f"Level {level_number}")
            
            # Initialize area counter for this level if not exists
            if level_name not in level_area_numbers:
                level_area_numbers[level_name] = 0
            
            for idx, area in enumerate(level["areas"], 1):
                # Increment area number for this level
                level_area_numbers[level_name] += 1
                area_number = level_area_numbers[level_name]
                area_name = area["name"]
                area_canopies = area.get("canopies", [])
                
                # Get tab color for this area
                tab_color = TAB_COLORS[area_count % len(TAB_COLORS)]
                
                # Check if area has fire suppression
                has_fire_suppression = any(canopy.get("options", {}).get("fire_suppression", False) for canopy in area_canopies)
                
                # Check if area has UV-C system (area-level option)
                has_uvc = area.get("options", {}).get("uvc", False)
                
                # Check if area has SDU system (area-level option) - DEPRECATED
                has_sdu = area.get("options", {}).get("sdu", False)
                
                # Check which canopies have SDU enabled (canopy-level option)
                sdu_canopies = [canopy for canopy in area_canopies if canopy.get('options', {}).get('sdu', False)]
                has_canopy_sdu = len(sdu_canopies) > 0
                
                # Check if area has RecoAir system (area-level option)
                has_recoair = area.get("options", {}).get("recoair", False)
                
                # Check if area has Marvel system (area-level option)
                has_marvel = area.get("options", {}).get("marvel", False)
                
                # Check if area has VENT CLG system (area-level option)
                has_vent_clg = area.get("options", {}).get("vent_clg", False)
                
                # Check if area has UV Extra Over option
                has_uv_extra_over = area.get("options", {}).get("uv_extra_over", False)
                
                # Check if area has UV canopies for UV Extra Over
                uv_canopies = [canopy for canopy in area_canopies if canopy.get('model', '').upper().startswith('UV')]
                non_uv_canopies = [canopy for canopy in area_canopies if not canopy.get('model', '').upper().startswith('UV')]
                has_uv_canopies = len(uv_canopies) > 0
                
                current_canopy_sheet = None
                fs_sheet = None
                ebox_sheet = None
                sdu_sheet = None
                recoair_sheet = None
                vent_clg_sheet = None
                uv_extra_over_sheet = None  # For UV Extra Over comparison
                
                # Process canopy sheet if canopies exist for this area
                if area_canopies:
                    if canopy_sheets:
                        sheet_name = canopy_sheets.pop(0)
                        current_canopy_sheet = wb[sheet_name]
                        
                        # Set title in B1
                        sheet_title_display = f"{level_name} - {area_name}"
                        current_canopy_sheet['B1'] = sheet_title_display
                        
                        # Determine sheet name based on UV Extra Over setting
                        if has_uv_extra_over and has_uv_canopies:
                            # This is the UV canopy sheet
                            canopy_sheet_tab_name = f"CANOPY (UV) - {level_name} ({area_number})"
                            if len(canopy_sheet_tab_name) > 31:  # Excel sheet name limit
                                canopy_sheet_tab_name = f"CANOPY (UV) - L{level_number} ({area_number})"
                        else:
                            # Normal canopy sheet naming
                            canopy_sheet_tab_name = f"CANOPY - {level_name} ({area_number})"
                            if len(canopy_sheet_tab_name) > 31:  # Excel sheet name limit
                                canopy_sheet_tab_name = f"CANOPY - L{level_number} ({area_number})"
                        
                        current_canopy_sheet.title = canopy_sheet_tab_name
                        current_canopy_sheet.sheet_state = 'visible'
                        current_canopy_sheet.sheet_properties.tabColor = tab_color
                        
                        # Write project metadata to canopy sheet (C/G columns)
                        write_project_metadata(current_canopy_sheet, project_data, template_version)
                        
                        # Write area-level delivery and installation pricing
                        write_area_delivery_install_pricing(current_canopy_sheet, area)
                        
                        # Create fire suppression sheet if needed
                        if has_fire_suppression:
                            if fire_supp_sheets:
                                fs_sheet_name = fire_supp_sheets.pop(0)
                                fs_sheet = wb[fs_sheet_name]
                                new_fs_name = f"FIRE SUPP - {level_name} ({area_number})"
                                fs_sheet.title = new_fs_name
                                fs_sheet.sheet_state = 'visible'
                                fs_sheet.sheet_properties.tabColor = tab_color
                                
                                # Write project metadata to fire suppression sheet
                                write_project_metadata(fs_sheet, project_data, template_version)
                                # Set fire suppression sheet title in B1
                                fs_sheet['B1'] = f"{level_name} - {area_name} - FIRE SUPPRESSION"
                        
                        # Create EBOX sheet if UV-C is selected for this area
                        if has_uvc:
                            if edge_box_sheets:
                                ebox_sheet_name = edge_box_sheets.pop(0)
                                ebox_sheet = wb[ebox_sheet_name]
                                new_ebox_name = f"EBOX - {level_name} ({area_number})"
                                ebox_sheet.title = new_ebox_name
                                ebox_sheet.sheet_state = 'visible'
                                ebox_sheet.sheet_properties.tabColor = tab_color
                                
                                # Write EBOX-specific metadata to EBOX sheet (D/H columns)
                                write_ebox_metadata(ebox_sheet, project_data, template_version)
                                # Set EBOX sheet title in C1
                                ebox_sheet['C1'] = f"{level_name} - {area_name} - UV-C SYSTEM"
                            else:
                                print(f"Warning: Not enough EBOX sheets in template for UV-C system in area {area_name}")
                        
                        # Create SDU sheets for each canopy that has SDU enabled
                        for canopy in sdu_canopies:
                            if sdu_sheets:
                                sdu_sheet_name = sdu_sheets.pop(0)
                                sdu_sheet = wb[sdu_sheet_name]
                                canopy_ref = canopy.get('reference_number', 'C???')
                                new_sdu_name = f"SDU - {level_name} ({area_number}) - {canopy_ref}"
                                # Ensure sheet name doesn't exceed Excel's 31 character limit
                                if len(new_sdu_name) > 31:
                                    new_sdu_name = f"SDU - L{level_number} ({area_number}) - {canopy_ref}"
                                    if len(new_sdu_name) > 31:
                                        new_sdu_name = f"SDU - L{level_number}({area_number}) - {canopy_ref}"
                                        if len(new_sdu_name) > 31:
                                            new_sdu_name = f"SDU - L{level_number}({area_number})-{canopy_ref}"
                                
                                sdu_sheet.title = new_sdu_name
                                sdu_sheet.sheet_state = 'visible'
                                sdu_sheet.sheet_properties.tabColor = tab_color
                                
                                # Write SDU-specific metadata to SDU sheet (C/G columns)
                                write_sdu_metadata(sdu_sheet, project_data, template_version, canopy)
                                # Set SDU sheet title in B1 - include canopy reference
                                sdu_sheet['B1'] = f"{level_name} - {area_name} - SDU SYSTEM - {canopy_ref}"
                                
                                # Add SDU specific dropdowns
                                add_sdu_dropdowns(sdu_sheet)
                            else:
                                print(f"Warning: Not enough SDU sheets in template for SDU system in canopy {canopy.get('reference_number', 'C???')} in area {area_name}")
                        
                        # Create RECOAIR sheet if RecoAir is selected for this area
                        if has_recoair:
                            if recoair_sheets:
                                recoair_sheet_name = recoair_sheets.pop(0)
                                recoair_sheet = wb[recoair_sheet_name]
                                new_recoair_name = f"RECOAIR - {level_name} ({area_number})"
                                recoair_sheet.title = new_recoair_name
                                recoair_sheet.sheet_state = 'visible'
                                recoair_sheet.sheet_properties.tabColor = tab_color
                                
                                # Generate item number for this RecoAir sheet
                                recoair_sheet_count += 1
                                item_number = f"{recoair_sheet_count}.01"
                                
                                # Write RECOAIR-specific metadata to RECOAIR sheet (D/H columns)
                                write_recoair_metadata(recoair_sheet, project_data, item_number, template_version)
                                # Set RECOAIR sheet title in C1
                                recoair_sheet['C1'] = f"{level_name} - {area_name} - RECOAIR SYSTEM"
                                
                                # Add RecoAir specific dropdowns
                                add_recoair_dropdowns(recoair_sheet)
                            else:
                                print(f"Warning: Not enough RECOAIR sheets in template for RecoAir system in area {area_name}")
                        
                        # Create VENT CLG sheet if VENT CLG is selected for this area
                        if has_vent_clg:
                            if vent_clg_sheets:
                                vent_clg_sheet_name = vent_clg_sheets.pop(0)
                                vent_clg_sheet = wb[vent_clg_sheet_name]
                                new_vent_clg_name = f"VENT CLG - {level_name} ({area_number})"
                                vent_clg_sheet.title = new_vent_clg_name
                                vent_clg_sheet.sheet_state = 'visible'
                                vent_clg_sheet.sheet_properties.tabColor = tab_color
                                
                                # Write VENT CLG-specific metadata to VENT CLG sheet (C/G columns)
                                write_vent_clg_metadata(vent_clg_sheet, project_data, template_version)
                                # Set VENT CLG sheet title in B1
                                vent_clg_sheet['B1'] = f"{level_name} - {area_name} - VENT CLG SYSTEM"
                                
                                # Add VENT CLG specific dropdowns
                                add_vent_clg_dropdowns(vent_clg_sheet)
                            else:
                                print(f"Warning: Not enough VENT CLG sheets in template for VENT CLG system in area {area_name}")
                        
                        # Create MARVEL sheet if Marvel is selected for this area
                        if has_marvel:
                            if marvel_sheets:
                                marvel_sheet_name = marvel_sheets.pop(0)
                                marvel_sheet = wb[marvel_sheet_name]
                                new_marvel_name = f"MARVEL - {level_name} ({area_number})"
                                marvel_sheet.title = new_marvel_name
                                marvel_sheet.sheet_state = 'visible'
                                marvel_sheet.sheet_properties.tabColor = tab_color
                                
                                # Write MARVEL-specific metadata to MARVEL sheet (F columns for project name/location/date)
                                write_marvel_metadata(marvel_sheet, project_data, template_version)
                                # Set MARVEL sheet title in B1 (like other sheets)
                                try:
                                    marvel_sheet['B1'] = f"{level_name} - {area_name} - MARVEL SYSTEM"
                                except Exception as e:
                                    print(f"Warning: Could not write title to B1 on MARVEL sheet: {str(e)}")
                            else:
                                print(f"Warning: Not enough MARVEL sheets in template for Marvel system in area {area_name}")
                        
                        # Write each canopy with proper spacing
                        fs_canopy_idx = 0  # Track fire suppression canopies separately
                        
                        if has_uv_extra_over and len(non_uv_canopies) > 0:
                            # UV Extra Over mode: Write all canopies to the main sheet (non-UV or all canopies if mixed)
                            for canopy_idx, canopy in enumerate(area_canopies):
                                row_start = CANOPY_START_ROW + (canopy_idx * CANOPY_ROW_SPACING)
                                write_canopy_data(current_canopy_sheet, canopy, row_start)
                                
                                # If this canopy has fire suppression and fire suppression sheet exists, write to it
                                if canopy.get("options", {}).get("fire_suppression") and fs_sheet:
                                    fs_row_start = CANOPY_START_ROW + (fs_canopy_idx * CANOPY_ROW_SPACING)
                                    write_fire_suppression_canopy_data(fs_sheet, canopy, fs_row_start)
                                    fs_canopy_idx += 1
                        else:
                            # Normal mode: Write all canopies
                            for canopy_idx, canopy in enumerate(area_canopies):
                                row_start = CANOPY_START_ROW + (canopy_idx * CANOPY_ROW_SPACING)
                                write_canopy_data(current_canopy_sheet, canopy, row_start)
                                
                                # If this canopy has fire suppression and fire suppression sheet exists, write to it
                                if canopy.get("options", {}).get("fire_suppression") and fs_sheet:
                                    fs_row_start = CANOPY_START_ROW + (fs_canopy_idx * CANOPY_ROW_SPACING)
                                    write_fire_suppression_canopy_data(fs_sheet, canopy, fs_row_start)
                                    fs_canopy_idx += 1  # Only increment for canopies with fire suppression
                        
                        # Add dropdowns
                        add_dropdowns_to_sheet(wb, current_canopy_sheet)
                        if fs_sheet:
                            # Add fire suppression specific dropdowns
                            add_fire_suppression_dropdowns(fs_sheet)
                        
                        # Create UV Extra Over sheet if enabled and there are non-UV canopies to convert
                        if has_uv_extra_over and len(non_uv_canopies) > 0:
                            if len(canopy_sheets) >= 1:  # Need another sheet for UV Extra Over
                                # Create UV Extra Over sheet with converted canopies
                                uv_extra_over_sheet_name = canopy_sheets.pop(0)
                                uv_extra_over_sheet = wb[uv_extra_over_sheet_name]
                                uv_extra_over_sheet.title = f"CANOPY (UV) - {level_name} ({area_number})"
                                if len(uv_extra_over_sheet.title) > 31:  # Excel sheet name limit
                                    uv_extra_over_sheet.title = f"CANOPY (UV) - L{level_number} ({area_number})"
                                
                                uv_extra_over_sheet.sheet_state = 'visible'
                                uv_extra_over_sheet.sheet_properties.tabColor = tab_color  # Use same color as non-UV sheet
                                
                                # Set title in B1 for UV Extra Over sheet
                                uv_extra_over_sheet['B1'] = f"{level_name} - {area_name} - UV EXTRA OVER"
                                write_project_metadata(uv_extra_over_sheet, project_data, template_version)
                                
                                # Convert eligible non-UV canopies to UV equivalents
                                uv_conversion_map = {
                                    'KVF': 'UVF',
                                    'KVI': 'UVI', 
                                    'KVX': 'UVX',
                                    'KVX-M': 'UVX-M'
                                }
                                
                                uv_converted_canopies = []
                                for non_uv_canopy in non_uv_canopies:
                                    canopy_model = non_uv_canopy.get('model', '').upper()
                                    
                                    # Check if this canopy type can be converted to UV equivalent
                                    if canopy_model in uv_conversion_map:
                                        # Create UV equivalent with same configuration
                                        uv_equivalent = non_uv_canopy.copy()  # Deep copy all properties
                                        uv_equivalent['model'] = uv_conversion_map[canopy_model]  # Convert to UV model
                                        
                                        # Keep all other properties (dimensions, wall cladding, options, etc.)
                                        uv_converted_canopies.append(uv_equivalent)
                                        
                                        print(f"   🔄 Converting {canopy_model} to {uv_conversion_map[canopy_model]} for UV Extra Over")
                                    else:
                                        print(f"   ⏭️  Skipping {canopy_model} - not eligible for UV conversion")
                                
                                # Write converted UV canopies to the UV Extra Over sheet
                                for canopy_idx, uv_canopy in enumerate(uv_converted_canopies):
                                    row_start = CANOPY_START_ROW + (canopy_idx * CANOPY_ROW_SPACING)
                                    write_canopy_data(uv_extra_over_sheet, uv_canopy, row_start)
                                
                                add_dropdowns_to_sheet(wb, uv_extra_over_sheet)
                                
                                print(f"✅ Created UV Extra Over sheet with {len(uv_converted_canopies)} converted UV canopies")
                                
                            else:
                                print(f"Warning: Not enough CANOPY sheets in template for UV Extra Over in area {area_name}")
                        
                        elif has_uv_extra_over and len(non_uv_canopies) == 0:
                            print(f"Info: No non-UV canopies found in area {area_name} for UV Extra Over conversion.")
                        
                    else:
                        raise Exception(f"Not enough CANOPY sheets in template for area {area_name}")
                
                # Handle case where UV-C, SDU, RecoAir, and/or Marvel are selected but no canopies exist (edge case)
                elif (has_uvc or has_sdu or has_canopy_sdu or has_recoair or has_marvel or has_vent_clg) and not area_canopies:
                    # Create EBOX sheet if UV-C is selected
                    if has_uvc:
                        if edge_box_sheets:
                            ebox_sheet_name = edge_box_sheets.pop(0)
                            ebox_sheet = wb[ebox_sheet_name]
                            new_ebox_name = f"EBOX - {level_name} ({area_number})"
                            ebox_sheet.title = new_ebox_name
                            ebox_sheet.sheet_state = 'visible'
                            ebox_sheet.sheet_properties.tabColor = tab_color
                            
                            # Write EBOX-specific metadata to EBOX sheet (D/H columns)
                            write_ebox_metadata(ebox_sheet, project_data, template_version)
                            # Set EBOX sheet title in C1
                            ebox_sheet['C1'] = f"{level_name} - {area_name} - UV-C SYSTEM"
                        else:
                            print(f"Warning: Not enough EBOX sheets in template for UV-C system in area {area_name}")
                    
                    # Create SDU sheet if area-level SDU is selected (edge case - no canopies exist)
                    # Note: Canopy-level SDU cannot exist without canopies
                    if has_sdu:
                        if sdu_sheets:
                            sdu_sheet_name = sdu_sheets.pop(0)
                            sdu_sheet = wb[sdu_sheet_name]
                            new_sdu_name = f"SDU - {level_name} ({area_number})"
                            sdu_sheet.title = new_sdu_name
                            sdu_sheet.sheet_state = 'visible'
                            sdu_sheet.sheet_properties.tabColor = tab_color
                            
                            # Write SDU-specific metadata to SDU sheet (C/G columns)
                            # For area-level SDU (old template), we don't have canopy data
                            write_sdu_metadata(sdu_sheet, project_data, template_version, None)
                            # Set SDU sheet title in B1
                            sdu_sheet['B1'] = f"{level_name} - {area_name} - SDU SYSTEM"
                            
                            # Add SDU specific dropdowns
                            add_sdu_dropdowns(sdu_sheet)
                        else:
                            print(f"Warning: Not enough SDU sheets in template for SDU system in area {area_name}")
                    
                    # Create RECOAIR sheet if RecoAir is selected
                    if has_recoair:
                        if recoair_sheets:
                            recoair_sheet_name = recoair_sheets.pop(0)
                            recoair_sheet = wb[recoair_sheet_name]
                            new_recoair_name = f"RECOAIR - {level_name} ({area_number})"
                            recoair_sheet.title = new_recoair_name
                            recoair_sheet.sheet_state = 'visible'
                            recoair_sheet.sheet_properties.tabColor = tab_color
                            
                            # Generate item number for this RecoAir sheet
                            recoair_sheet_count += 1
                            item_number = f"{recoair_sheet_count}.01"
                            
                            # Write RECOAIR-specific metadata to RECOAIR sheet (D/H columns)
                            write_recoair_metadata(recoair_sheet, project_data, item_number, template_version)
                            # Set RECOAIR sheet title in C1
                            recoair_sheet['C1'] = f"{level_name} - {area_name} - RECOAIR SYSTEM"
                            
                            # Add RecoAir specific dropdowns
                            add_recoair_dropdowns(recoair_sheet)
                        else:
                            print(f"Warning: Not enough RECOAIR sheets in template for RecoAir system in area {area_name}")
                    
                    # Create MARVEL sheet if Marvel is selected
                    if has_marvel:
                        if marvel_sheets:
                            marvel_sheet_name = marvel_sheets.pop(0)
                            marvel_sheet = wb[marvel_sheet_name]
                            new_marvel_name = f"MARVEL - {level_name} ({area_number})"
                            marvel_sheet.title = new_marvel_name
                            marvel_sheet.sheet_state = 'visible'
                            marvel_sheet.sheet_properties.tabColor = tab_color
                            
                            # Write MARVEL-specific metadata to MARVEL sheet (F columns for project name/location/date)
                            write_marvel_metadata(marvel_sheet, project_data, template_version)
                            # Set MARVEL sheet title in B1 (like other sheets)
                            try:
                                marvel_sheet['B1'] = f"{level_name} - {area_name} - MARVEL SYSTEM"
                            except Exception as e:
                                print(f"Warning: Could not write title to B1 on MARVEL sheet: {str(e)}")
                        else:
                            print(f"Warning: Not enough MARVEL sheets in template for Marvel system in area {area_name}")                    
                    # Create VENT CLG sheet if VENT CLG is selected
                    if has_vent_clg:
                        if vent_clg_sheets:
                            vent_clg_sheet_name = vent_clg_sheets.pop(0)
                            vent_clg_sheet = wb[vent_clg_sheet_name]
                            new_vent_clg_name = f"VENT CLG - {level_name} ({area_number})"
                            vent_clg_sheet.title = new_vent_clg_name
                            vent_clg_sheet.sheet_state = 'visible'
                            vent_clg_sheet.sheet_properties.tabColor = tab_color
                            
                            # Write VENT CLG-specific metadata to VENT CLG sheet (C/G columns)
                            write_vent_clg_metadata(vent_clg_sheet, project_data, template_version)
                            # Set VENT CLG sheet title in B1
                            vent_clg_sheet['B1'] = f"{level_name} - {area_name} - VENT CLG SYSTEM"
                            
                            # Add VENT CLG specific dropdowns
                            add_vent_clg_dropdowns(vent_clg_sheet)
                        else:
                            print(f"Warning: Not enough VENT CLG sheets in template for VENT CLG system in area {area_name}")

                
                area_count += 1
        
        # Write project metadata to any other visible sheets that might exist
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            # Check if sheet name starts with duct names (to handle cases like "SPIRAL DUCT1", "EXTRACT DUCT1", etc.)
            is_duct_sheet = (sheet_name.startswith('SPIRAL DUCT') or 
                           sheet_name.startswith('SUPPLY DUCT') or 
                           sheet_name.startswith('EXTRACT DUCT'))
            
            if (sheet.sheet_state == 'visible' and 
                not sheet_name.startswith(('CANOPY', 'FIRE SUPP', 'EBOX', 'RECOAIR', 'SDU', 'MARVEL', 'VENT CLG')) and 
                not is_duct_sheet and
                sheet_name not in ['Lists', 'JOB TOTAL']):
                # Write metadata to any other visible sheets (excluding EBOX, RECOAIR, SDU, MARVEL, and duct sheets which don't need metadata)
                try:
                    write_project_metadata(sheet, project_data, template_version)
                except Exception as e:
                    print(f"Warning: Could not write metadata to sheet {sheet_name}: {str(e)}")
        
        # Organize sheets by area for better navigation
        organize_sheets_by_area(wb)
        
        # Add delivery location dropdowns to all relevant sheets
        delivery_location = project_data.get('delivery_location', '')
        print(f"🚚 Adding delivery location dropdowns, pre-selected: '{delivery_location}'")
        
        sheets_updated = 0
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            if (sheet.sheet_state == 'visible' and 
                sheet_name not in ['JOB TOTAL', 'Lists', 'PRICING_SUMMARY', 'ProjectData'] and
                any(prefix in sheet_name for prefix in ['CANOPY', 'FIRE SUPP', 'EBOX', 'RECOAIR', 'SDU', 'MARVEL', 'VENT CLG'])):
                add_delivery_location_dropdown_to_sheet(sheet, delivery_location)
                sheets_updated += 1
        print(f"📝 Added delivery location dropdowns to {sheets_updated} sheets")
        
        # Delete only unused sheets for the specific systems we work with (CANOPY, FIRE SUPP, EBOX, SDU, RECOAIR, MARVEL, CONTRACT)
        # Exclude the actually created contract sheets from deletion
        created_contract_sheet_names = []
        if project_data.get('contract_option', False):
            created_contract_sheet_names = ['CONTRACT', 'SPIRAL DUCT', 'SUPPLY DUCT', 'EXTRACT DUCT']
        
        unused_sheets = canopy_sheets + fire_supp_sheets + edge_box_sheets + sdu_sheets + recoair_sheets + marvel_sheets + vent_clg_sheets + contract_sheets + spiral_duct_sheets + supply_duct_sheets + extract_duct_sheets
        
        # Filter out the created contract sheets from the deletion list
        unused_sheets = [sheet for sheet in unused_sheets if sheet not in created_contract_sheet_names]
        
        print(f"🗑️  Removing {len(unused_sheets)} unused system template sheets...")
        for sheet_name in unused_sheets:
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
                print(f"   Deleted: {sheet_name}")
        
        # Hide ALL template sheets except the ones we actually use
        # Only keep visible: used system sheets and essential management sheets
        allowed_visible_prefixes = (
            'CANOPY -', 'CANOPY (UV)', 'FIRE SUPP -', 'EBOX -', 'SDU -', 'RECOAIR -', 'MARVEL -', 'VENT CLG -',
            'JOB TOTAL', 'PRICING_SUMMARY', 'ProjectData', 'Lists',
            'CONTRACT', 'SPIRAL DUCT', 'SUPPLY DUCT', 'EXTRACT DUCT'
        )
        
        extra_hidden_count = 0
        for sheet in wb.worksheets:
            if sheet.sheet_state == 'visible':
                keep_visible = False
                
                # Check if it starts with an allowed prefix (used system sheets or management)
                for prefix in allowed_visible_prefixes:
                    if sheet.title.startswith(prefix) or sheet.title == prefix:
                        keep_visible = True
                        break
                
                if not keep_visible:
                    sheet.sheet_state = 'hidden'
                    extra_hidden_count += 1
                    print(f"   Hidden: {sheet.title}")
        
        if extra_hidden_count:
            print(f"🔒 Hidden {extra_hidden_count} unused template sheets (preserved for future use).")
        
        # Create pricing summary sheet for dynamic pricing aggregation
        print("Creating PRICING_SUMMARY sheet...")
        create_pricing_summary_sheet(wb)
        
        # Create UV Extra Over calculations sheet if there are any UV Extra Over configurations
        has_uv_extra_over = any('CANOPY (UV) - ' in sheet_name for sheet_name in wb.sheetnames)
        if has_uv_extra_over:
            print("Creating UV Extra Over calculations sheet...")
            create_uv_extra_over_calculations_sheet(wb)
        
        # Update JOB TOTAL sheet to reference pricing summary
        print("Updating JOB TOTAL sheet with dynamic pricing formulas...")
        update_job_total_sheet(wb)
        
        # Save the workbook
        project_number = project_data.get('project_number', 'unknown')
        date_str = project_data.get('date', '')
        
        # Format date for filename (remove slashes and make it filename-safe)
        if date_str:
            # Convert DD/MM/YYYY to DDMMYYYY or similar format
            formatted_date = date_str.replace('/', '').replace('-', '')
        else:
            # Use current date if no date provided
            formatted_date = get_current_date().replace('/', '')
        
        output_path = f"output/{project_number} Cost Sheet {formatted_date}.xlsx"
        os.makedirs("output", exist_ok=True)
        wb.save(output_path)
        
        return output_path
    
    except Exception as e:
        raise Exception(f"Failed to generate Excel file: {str(e)}")

def detect_template_version(wb: Workbook) -> str:
    """
    Detect which template version was used based on workbook structure.
    
    Args:
        wb (Workbook): The Excel workbook to analyze
        
    Returns:
        str: Template version identifier (e.g. "R19.2")
    """
    try:
        # Method 1: Check for version in CANOPY sheet B1
        sheet_names = wb.sheetnames
        
        # First try to find a CANOPY sheet
        for sheet_name in sheet_names:
            if 'CANOPY' in sheet_name:
                sheet = wb[sheet_name]
                title = sheet['B1'].value
                if title and ' - ' in title:
                    # Extract version from title (e.g. "F24-19.2 CANOPY COST SHEET" or "F24 - 19.1  CANOPY COST SHEET")
                    # Handle both formats: "F24-19.2" and "F24 - 19.1"
                    title_str = str(title)
                    if ' - ' in title_str:
                        # Format: "F24 - 19.1  CANOPY COST SHEET"
                        parts = title_str.split(' - ')
                        if len(parts) >= 2:
                            version_part = parts[1].split()[0]  # Get "19.1" from "19.1  CANOPY"
                            version = f"R{version_part}"
                            if version in TEMPLATE_PATHS:  # Only return known versions
                                return version
                    elif '-' in title_str:
                        # Format: "F24-19.2 CANOPY COST SHEET"
                        parts = title_str.split(' ')[0].split('-')  # Split "F24-19.2" into ["F24", "19.2"]
                        if len(parts) == 2:
                            version = f"R{parts[1]}"
                            if version in TEMPLATE_PATHS:  # Only return known versions
                                return version
        
        # Method 2: Check JOB TOTAL sheet if no CANOPY sheet found
        if 'JOB TOTAL' in sheet_names:
            job_total_sheet = wb['JOB TOTAL']
            title = job_total_sheet['B1'].value
            if title and ' - ' in title:
                parts = title.split(' ')[0].split('-')
                if len(parts) == 2:
                    version = f"R{parts[1]}"
                    if version in TEMPLATE_PATHS:  # Only return if it's a known version
                        return version
        
        # Default to R19.2 if no version found
        return "R19.2"
        
    except Exception as e:
        print(f"Warning: Could not detect template version: {str(e)}")
        # Default to R19.2 if detection fails
        return "R19.2"

def read_excel_project_data(excel_path: str) -> Dict:
    """
    Read project data back from a generated Excel file.
    
    Args:
        excel_path (str): Path to the Excel file to read
        
    Returns:
        Dict: Project data extracted from the Excel file
    """
    # Clear any previous validation errors
    clear_validation_errors()
    
    try:
        wb = load_workbook(excel_path, data_only=True)
        
        # Try to get data from JOB TOTAL sheet first, then any canopy sheet
        data_sheet = None
        if 'JOB TOTAL' in wb.sheetnames:
            data_sheet = wb['JOB TOTAL']
        else:
            # Find first CANOPY sheet
            for sheet_name in wb.sheetnames:
                if 'CANOPY' in sheet_name:
                    data_sheet = wb[sheet_name]
                    break
        
        if not data_sheet:
            raise Exception("No suitable sheet found to extract project data")
        
        # Extract project metadata using the same cell mappings
        project_data = {}
        
        # Detect which template version was used based on sheet structure/content
        template_used = detect_template_version(wb)
        project_data['template_used'] = template_used
        
        # Read basic project info
        project_data['project_number'] = data_sheet['C3'].value or ""
        project_data['company'] = data_sheet['C5'].value or ""
        project_data['estimator_initials'] = data_sheet['C7'].value or ""  # This is the initials version
        project_data['project_name'] = data_sheet['G3'].value or ""
        project_data['project_location'] = data_sheet['G5'].value or ""  # Project location from G5
        project_data['location'] = data_sheet['G5'].value or ""  # Keep for backward compatibility
        
        # Read and format date consistently
        date_value = data_sheet['G7'].value or ""
        if date_value:
            # If it's a datetime object from Excel, convert to string
            if hasattr(date_value, 'strftime'):
                project_data['date'] = date_value.strftime("%d/%m/%Y")
            else:
                # If it's already a string, ensure it's in the right format
                project_data['date'] = format_date_for_display(str(date_value))
        else:
            project_data['date'] = ""
            
        project_data['revision'] = data_sheet['K7'].value or ""  # Revision from K7, leave blank if not set
        
        # Read company and estimator data from hidden ProjectData sheet
        if 'ProjectData' in wb.sheetnames:
            hidden_sheet = wb['ProjectData']
            
            # Read customer information (new - row 1)
            project_data['customer'] = hidden_sheet['B1'].value or ""
            
            # Read company information (moved to row 2)
            project_data['company'] = hidden_sheet['B2'].value or ""
            project_data['address'] = hidden_sheet['B3'].value or ""
            
            # Read full estimator information (moved to row 4)
            project_data['estimator'] = hidden_sheet['B4'].value or project_data['estimator_initials']
            project_data['estimator_rank'] = hidden_sheet['B5'].value or "Estimator"
            
            # Read additional data (moved to rows 6-8)
            project_data['sales_contact'] = hidden_sheet['B6'].value or ""
            project_data['delivery_location'] = hidden_sheet['B7'].value or ""
            
            # Read revision from ProjectData sheet if not already set (moved to row 8)
            if not project_data.get('revision'):
                project_data['revision'] = hidden_sheet['B8'].value or ""
        else:
            # Fallback if no hidden sheet exists
            project_data['estimator'] = project_data['estimator_initials']
            project_data['estimator_rank'] = "Estimator"
            project_data['company'] = ""
            project_data['address'] = ""
            project_data['sales_contact'] = ""
            project_data['delivery_location'] = ""
        
        # Get level and area information from sheet titles
        levels_data = {}
        canopy_data = {}
        
        # First pass: Create areas from all sheet types (CANOPY, FIRE SUPP, EBOX, RECOAIR, SDU)
        for sheet_name in wb.sheetnames:
            if any(prefix in sheet_name for prefix in ['CANOPY - ', 'CANOPY (UV) - ', 'FIRE SUPP - ', 'EBOX - ', 'RECOAIR - ', 'SDU - ']):
                sheet = wb[sheet_name]
                
                # Determine which cell contains the title based on sheet type
                if 'EBOX - ' in sheet_name or 'RECOAIR - ' in sheet_name:
                    title_cell = sheet['C1'].value  # EBOX and RECOAIR sheets have title in C1
                else:
                    title_cell = sheet['B1'].value  # CANOPY, FIRE SUPP, and SDU sheets have title in B1
                
                if title_cell and ' - ' in title_cell:
                    # Handle different title formats
                    if 'UV-C SYSTEM' in title_cell or 'RECOAIR SYSTEM' in title_cell or 'SDU SYSTEM' in title_cell:
                        # For EBOX/RECOAIR/SDU: "Level 1 - Main Kitchen - UV-C SYSTEM" or "Level 1 - Main Kitchen - RECOAIR SYSTEM" or "Level 1 - Main Kitchen - SDU SYSTEM"
                        title_parts = title_cell.split(' - ')
                        if len(title_parts) >= 2:
                            level_name = title_parts[0]
                            area_name = title_parts[1]
                        else:
                            continue
                    else:
                        # For CANOPY/FIRE SUPP: "Level 1 - Main Kitchen" or "Level 1 - Main Kitchen - FIRE SUPPRESSION"
                        level_area = title_cell.split(' - ')
                        if len(level_area) >= 2:
                            level_name = level_area[0]
                            area_name = level_area[1]
                        else:
                            continue
                    
                    # Create level if it doesn't exist
                    if level_name not in levels_data:
                        levels_data[level_name] = []
                    
                    # Create area if it doesn't exist
                    if area_name not in [area['name'] for area in levels_data[level_name]]:
                        levels_data[level_name].append({
                            'name': area_name,
                            'canopies': []
                        })
        
        # Second pass: Read canopy data from CANOPY sheets (exclude UV Extra Over sheets)
        for sheet_name in wb.sheetnames:
            if 'CANOPY - ' in sheet_name and 'CANOPY (UV) - ' not in sheet_name:
                sheet = wb[sheet_name]
                title_cell = sheet['B1'].value
                
                if title_cell and ' - ' in title_cell:
                    level_area = title_cell.split(' - ')
                    if len(level_area) >= 2:
                        level_name = level_area[0]
                        area_name = level_area[1]
                        
                        # Read canopy specifications from the sheet
                        # This is a simplified read - you might want to enhance this
                        for canopy_idx in range(10):  # Support up to 10 canopies
                            base_row = CANOPY_START_ROW + (canopy_idx * CANOPY_ROW_SPACING)
                            ref_row = base_row - 2
                            
                            ref_number = sheet[f'B{ref_row}'].value
                            if ref_number:
                                # Get model to check for placeholder rows
                                model = sheet[f'D{base_row}'].value or ""
                                
                                # Skip placeholder rows
                                if (safe_upper(ref_number) == "ITEM" or 
                                    safe_upper(model) == "CANOPY TYPE" or
                                    safe_upper(str(ref_number).strip()) == "ITEM" or
                                    safe_upper(str(model).strip()) == "CANOPY TYPE"):
                                    continue
                                
                                canopy_info = {
                                    'reference_number': ref_number,
                                    'configuration': sheet[f'C{base_row}'].value or "",
                                    'model': model,
                                    
                                    # Additional specification data (convert dimensions to integers to avoid .0 display)
                                    'length': int(float(sheet[f'F{base_row}'].value)) if sheet[f'F{base_row}'].value and str(sheet[f'F{base_row}'].value).strip() not in ['', '-'] else "",
                                    'width': int(float(sheet[f'E{base_row}'].value)) if sheet[f'E{base_row}'].value and str(sheet[f'E{base_row}'].value).strip() not in ['', '-'] else "",
                                    'height': int(float(sheet[f'G{base_row}'].value)) if sheet[f'G{base_row}'].value and str(sheet[f'G{base_row}'].value).strip() not in ['', '-'] else "",
                                    'sections': int(float(sheet[f'H{base_row}'].value)) if sheet[f'H{base_row}'].value and str(sheet[f'H{base_row}'].value).strip() not in ['', '-'] else "",
                                    'lighting_type': sheet[f'C{base_row + 1}'].value or "",  # C15 (base_row + 1)
                                    
                                    # PRESERVE MANUAL INPUT CELLS - These are commonly edited by users
                                    'light_inputs': sheet[f'D{base_row + 1}'].value or "",  # D15 - Light inputs (commonly manually entered)
                                    'special_works_1': sheet[f'C{base_row + 2}'].value or "",  # C16 - Special works line 1
                                    'special_works_2': sheet[f'C{base_row + 3}'].value or "",  # C17 - Special works line 2  
                                    'special_works_3': sheet[f'C{base_row + 4}'].value or "",  # C18 - Special works line 3
                                    'quantity_override': sheet[f'D{base_row + 4}'].value or "",  # D18 - Quantity (sometimes manually changed)
                                    
                                    # Additional manual input preservation
                                    'extract_volume_manual': sheet[f'I{base_row}'].value or "",  # I14 - Extract volume (sometimes manually entered)
                                    'supply_static_manual': sheet[f'L{base_row}'].value or "",  # L14 - Supply static (sometimes manually entered)
                                    
                                    # Volume and static data (if available in your template)
                                    'extract_volume': sheet[f'I{base_row}'].value or "",
                                    'extract_static': sheet[f'F{base_row + 8}'].value or "",  # F22, F39, F56, etc.
                                    'mua_volume': _read_mua_volume(sheet, base_row, model),
                                    'supply_static': sheet[f'L{base_row}'].value or "",
                                    
                                    # Pricing data - calculate net canopy price (canopy total minus cladding)
                                    # Try to read from P12 formula result, or calculate manually if not available
                                    'canopy_price': _calculate_net_canopy_price(sheet, ref_row),  # P12, P29, P46, etc. (net amount after cladding subtraction)
                                    
                                    # Fire suppression data - will be populated from FIRE SUPP sheet
                                    'fire_suppression_tank_quantity': 0,  # Default to 0, will be updated from FIRE SUPP sheet
                                    'fire_suppression_price': 0,  # Default to 0, will be updated from FIRE SUPP sheet
                                    'fire_suppression_system_type': None,  # Default to None, will be updated from FIRE SUPP sheet
                                    
                                    # Read wall cladding data from Excel
                                    'wall_cladding': read_wall_cladding_from_canopy(sheet, base_row),
                                    
                                    # Read wall cladding price from Excel (N19, N36, N53, etc.)
                                    'cladding_price': sheet[f'N{ref_row + 7}'].value or 0,  # N19, N36, N53, etc. (ref_row + 7)
                                    
                                    # Initialize options (fire suppression will be set to True later if data exists)
                                    'options': {
                                        'fire_suppression': False  # Will be updated to True if fire suppression data is found
                                    }
                                }
                                
                                # Add CWS/HWS data for CMWF and CMWI canopies
                                if model.upper() in ['CMWF', 'CMWI']:
                                    # Calculate the row offset for CWS/HWS data (F25, F26, F27 relative to canopy)
                                    cws_row = base_row + 11  # F25 relative to base_row (14 + 11 = 25)
                                    hws_row = base_row + 12  # F26 relative to base_row
                                    storage_row = base_row + 13  # F27 relative to base_row
                                    
                                    canopy_info.update({
                                        'cws_capacity': sheet[f'F{cws_row}'].value or "",  # CWS @ 2 Bar
                                        'hws_requirement': sheet[f'F{hws_row}'].value or "",  # HWS @ 2 Bar  
                                        'hw_storage': sheet[f'F{storage_row}'].value or "",  # HWS Storage
                                        'has_wash_capabilities': True
                                    })
                                else:
                                    canopy_info.update({
                                        'cws_capacity': "",
                                        'hws_requirement': "",
                                        'hw_storage': "",
                                        'has_wash_capabilities': False
                                    })
                                
                                # Find the area and add canopy data
                                for area in levels_data[level_name]:
                                    if area['name'] == area_name:
                                        area['canopies'].append(canopy_info)
                                        break
        
        # Read fire suppression data from FIRE SUPP sheets
        for sheet_name in wb.sheetnames:
            if 'FIRE SUPP - ' in sheet_name:
                sheet = wb[sheet_name]
                title_cell = sheet['B1'].value
                
                if title_cell and ' - ' in title_cell:
                    # Extract level and area from title like "Level 1 - Main Kitchen - FIRE SUPPRESSION"
                    title_parts = title_cell.split(' - ')
                    if len(title_parts) >= 2:
                        level_name = title_parts[0]
                        area_name = title_parts[1]
                        
                        # Get fire suppression commissioning price from N193 and calculate net delivery price
                        fs_commissioning_price = sheet['N193'].value or 0
                        fs_delivery_price = _calculate_net_delivery_price(sheet)
                        
                        # Count how many fire suppression units are in this sheet
                        fs_units = []
                        for canopy_idx in range(10):  # Support up to 10 canopies
                            base_row = CANOPY_START_ROW + (canopy_idx * CANOPY_ROW_SPACING)
                            ref_row = base_row - 2
                            system_row = base_row + 2  # C16 relative to base_row (14 + 2 = 16)
                            tank_row = base_row + 3  # C17 relative to base_row (14 + 3 = 17)
                            
                            ref_number = sheet[f'B{ref_row}'].value
                            system_type = sheet[f'C{system_row}'].value  # Fire suppression system type from C16
                            tank_value = sheet[f'C{tank_row}'].value
                            base_fire_suppression_price = sheet[f'N{ref_row}'].value or 0  # Fire suppression base price at N12, N29, N46, etc.
                            
                            # Only count actual fire suppression units, not template entries
                            # Exclude entries with "ITEM" reference OR "TANK INSTALL"/"TANK INSTALLATION" tank values
                            if (ref_number and tank_value and 
                                safe_upper(str(ref_number)) != "ITEM" and 
                                safe_upper(str(tank_value)) not in ["TANK INSTALL", "TANK INSTALLATION"]):
                                
                                tank_quantity = extract_tank_quantity(tank_value)

                                fs_units.append({
                                    'ref_number': ref_number,
                                    'system_type': system_type,  # Add system type from C16
                                    'tank_quantity': tank_quantity,
                                    'base_price': base_fire_suppression_price
                                })
                        
                        # Calculate delivery price per unit (split equally among all units, or full amount if only one unit)
                        if len(fs_units) == 1:
                            delivery_per_unit = fs_delivery_price  # Single unit gets full delivery price
                        else:
                            delivery_per_unit = fs_delivery_price / len(fs_units) if fs_units else 0  # Multiple units split delivery
                        
                        # Update fire suppression data for each canopy
                        for fs_unit in fs_units:
                            # Calculate fire suppression price: base price + delivery share (no commissioning share)
                            total_fs_price = fs_unit['base_price'] + delivery_per_unit
                            
                            # Find the corresponding canopy and update its fire suppression data
                            for level_areas in levels_data.get(level_name, []):
                                if level_areas['name'] == area_name:
                                    for canopy in level_areas['canopies']:
                                        # Use flexible reference matching instead of exact match
                                        if references_match(canopy['reference_number'], fs_unit['ref_number']):
                                            canopy['fire_suppression_tank_quantity'] = fs_unit['tank_quantity']
                                            canopy['fire_suppression_price'] = total_fs_price
                                            canopy['fire_suppression_system_type'] = fs_unit['system_type']  # Add system type
                                            canopy['fire_suppression_reference_number'] = fs_unit['ref_number']  # Store the actual fire suppression reference
                                            
                                            # Set the fire suppression option flag for form compatibility
                                            if 'options' not in canopy:
                                                canopy['options'] = {}
                                            canopy['options']['fire_suppression'] = True
                                            
                                            print(f"✅ Matched fire suppression: Canopy '{canopy['reference_number']}' ↔ Fire Supp '{fs_unit['ref_number']}'")
                                            # print(f"   📊 Tank quantity: {fs_unit['tank_quantity']}, Price: £{total_fs_price}")
                                            # print(f"   🔗 Fire suppression reference stored: '{fs_unit['ref_number']}'")
                                            break
        
        # Read area-level pricing data (delivery & installation, commissioning) from non-UV sheets only
        for sheet_name in wb.sheetnames:
            if 'CANOPY - ' in sheet_name and 'CANOPY (UV) - ' not in sheet_name:
                sheet = wb[sheet_name]
                title_cell = sheet['B1'].value
                
                if title_cell and ' - ' in title_cell:
                    level_area = title_cell.split(' - ')
                    if len(level_area) >= 2:
                        level_name = level_area[0]
                        area_name = level_area[1]
                        
                        # Read area-level pricing
                        # Get total delivery price from N182 and commissioning from N193
                        n182_total_delivery = sheet['N182'].value or 0
                        commissioning_price = sheet['N193'].value or 0
                        
                        # Calculate net delivery & installation price (N182 - N193)
                        # This is what goes to Word document as "delivery_installation_price"
                        delivery_installation_price = n182_total_delivery - commissioning_price
                        
                        # PRESERVE AREA-LEVEL MANUAL INPUT CELLS
                        # Read delivery number (commonly on the left of delivery location)
                        delivery_number = sheet['C183'].value or ""  # C183 - Common location for delivery number
                        delivery_location_value = sheet['D183'].value or ""  # D183 - Delivery location
                        
                        # Read access equipment entries (commonly used fields)
                        access_equipment_1 = sheet['D184'].value or ""  # D184 - Access equipment 1
                        access_equipment_2 = sheet['D185'].value or ""  # D185 - Access equipment 2
                        
                        # Read testing and commissioning entries
                        testing_commissioning_description = sheet['C193'].value or ""  # C193 - T&C description
                        
                        # Find the area and add pricing data + manual inputs
                        area_found = False
                        if level_name in levels_data:
                            for area in levels_data[level_name]:
                                if area['name'] == area_name:
                                    area.update({
                                        'delivery_installation_price': delivery_installation_price,
                                        'commissioning_price': commissioning_price,
                                        # Preserve manual inputs
                                        'delivery_number': delivery_number,
                                        'delivery_location_value': delivery_location_value,
                                        'access_equipment_1': access_equipment_1,
                                        'access_equipment_2': access_equipment_2,
                                        'testing_commissioning_description': testing_commissioning_description
                                    })
                                    area_found = True
                                    break
                        
                        # If area wasn't found, create it (this shouldn't happen if first pass worked correctly)
                        if not area_found:
                            if level_name not in levels_data:
                                levels_data[level_name] = []
                            
                            levels_data[level_name].append({
                                'name': area_name,
                                'canopies': [],
                                'delivery_installation_price': delivery_installation_price,
                                'commissioning_price': commissioning_price
                            })
        
        # Read UV-C pricing from EBOX sheets
        for sheet_name in wb.sheetnames:
            if 'EBOX - ' in sheet_name:
                sheet = wb[sheet_name]
                title_cell = sheet['C1'].value  # EBOX sheets have title in C1
                
                if title_cell and ' - ' in title_cell and 'UV-C SYSTEM' in title_cell:
                    # Extract level and area from title like "Level 1 - Main Kitchen - UV-C SYSTEM"
                    title_parts = title_cell.split(' - ')
                    if len(title_parts) >= 2:
                        level_name = title_parts[0]
                        area_name = title_parts[1]
                        
                        # Read UV-C price from N9
                        uvc_price = sheet['N9'].value or 0
                        
                        # Find the area and add UV-C pricing data
                        area_found = False
                        if level_name in levels_data:
                            for area in levels_data[level_name]:
                                if area['name'] == area_name:
                                    area.update({
                                        'uvc_price': uvc_price
                                    })
                                    area_found = True
                                    break
                        
                        # If area wasn't found, create it (this shouldn't happen if first pass worked correctly)
                        if not area_found:
                            if level_name not in levels_data:
                                levels_data[level_name] = []
                            
                            levels_data[level_name].append({
                                'name': area_name,
                                'canopies': [],
                                'uvc_price': uvc_price
                            })
        
        # Read RecoAir data from RECOAIR sheets
        for sheet_name in wb.sheetnames:
            if 'RECOAIR - ' in sheet_name:
                sheet = wb[sheet_name]
                title_cell = sheet['C1'].value  # RECOAIR sheets have title in C1
                
                if title_cell and ' - ' in title_cell and 'RECOAIR SYSTEM' in title_cell:
                    # Extract level and area from title like "Level 1 - Main Kitchen - RECOAIR SYSTEM"
                    title_parts = title_cell.split(' - ')
                    if len(title_parts) >= 2:
                        level_name = title_parts[0]
                        area_name = title_parts[1]
                        
                        # Read RecoAir data from this sheet (now returns dict with units and flat pack)
                        recoair_data = read_recoair_data_from_sheet(sheet)
                        recoair_units = recoair_data['units']
                        flat_pack_data = recoair_data['flat_pack']
                        
                        # Read RecoAir commissioning price from N46
                        recoair_commissioning_price = sheet['N46'].value or 0
                        
                        # Calculate total RecoAir price (sum of all unit prices + delivery + commissioning + flat pack)
                        total_unit_price = sum(unit['unit_price'] for unit in recoair_units)
                        total_delivery_price = sum(unit['delivery_installation_price'] for unit in recoair_units)
                        flat_pack_price = flat_pack_data['price'] if flat_pack_data['has_flat_pack'] else 0
                        recoair_price = total_unit_price + total_delivery_price + recoair_commissioning_price + flat_pack_price
                        
                        # Find the area and add RecoAir data
                        area_found = False
                        if level_name in levels_data:
                            for area in levels_data[level_name]:
                                if area['name'] == area_name:
                                    area.update({
                                        'recoair_price': recoair_price,
                                        'recoair_commissioning_price': recoair_commissioning_price,  # Add commissioning price separately
                                        'recoair_units': recoair_units,  # Add detailed unit data
                                        'recoair_flat_pack': flat_pack_data  # Add flat pack data
                                    })
                                    area_found = True
                                    break
                        
                        # If area wasn't found, create it (this shouldn't happen if first pass worked correctly)
                        if not area_found:
                            if level_name not in levels_data:
                                levels_data[level_name] = []
                            
                            levels_data[level_name].append({
                                'name': area_name,
                                'canopies': [],
                                'recoair_price': recoair_price,
                                'recoair_commissioning_price': recoair_commissioning_price,
                                'recoair_units': recoair_units,
                                'recoair_flat_pack': flat_pack_data
                            })
        
        # Read MARVEL pricing from MARVEL sheets
        for sheet_name in wb.sheetnames:
            if 'MARVEL - ' in sheet_name:
                sheet = wb[sheet_name]
                title_cell = sheet['B1'].value  # MARVEL sheets have title in B1 (fixed to match creation)
                
                if title_cell and ' - ' in title_cell and 'MARVEL SYSTEM' in title_cell:
                    # Extract level and area from title like "Level 1 - Main Kitchen - MARVEL SYSTEM"
                    title_parts = title_cell.split(' - ')
                    if len(title_parts) >= 2:
                        level_name = title_parts[0]
                        area_name = title_parts[1]
                        
                        # Extract MARVEL pricing data using the specified cells
                        try:
                            # Factory components: J12+J23+J33+J44
                            j12_value = sheet['J12'].value or 0
                            j23_value = sheet['J23'].value or 0
                            j33_value = sheet['J33'].value or 0
                            j44_value = sheet['J44'].value or 0
                            factory_components = j12_value + j23_value + j33_value + j44_value
                            
                            # Commissioning: J59+J60
                            j59_value = sheet['J59'].value or 0
                            j60_value = sheet['J60'].value or 0
                            commissioning = j59_value + j60_value
                            
                            # Onsite installation: J53-(J59+J60)
                            j53_value = sheet['J53'].value or 0
                            onsite_installation = j53_value - commissioning
                            
                            # Total MARVEL price
                            marvel_price = factory_components + onsite_installation + commissioning
                            
                            marvel_data = {
                                'factory_components': factory_components,
                                'onsite_installation': onsite_installation,
                                'commissioning': commissioning,
                                'total_price': marvel_price
                            }
                            
                            print(f"✓ Read MARVEL pricing from {sheet_name}:")
                            print(f"   Factory Components: £{factory_components:.2f}")
                            print(f"   Onsite Installation: £{onsite_installation:.2f}")
                            print(f"   Commissioning: £{commissioning:.2f}")
                            print(f"   Total: £{marvel_price:.2f}")
                            
                        except Exception as e:
                            print(f"Warning: Could not read MARVEL pricing from {sheet_name}: {str(e)}")
                            marvel_data = {
                                'factory_components': 0,
                                'onsite_installation': 0,
                                'commissioning': 0,
                                'total_price': 0
                            }
                        
                        # Find the area and add MARVEL data
                        area_found = False
                        if level_name in levels_data:
                            for area in levels_data[level_name]:
                                if area['name'] == area_name:
                                    area.update({
                                        'marvel_price': marvel_data['total_price'],
                                        'marvel_pricing': marvel_data
                                    })
                                    area_found = True
                                    break
                        
                        # If area wasn't found, create it
                        if not area_found:
                            if level_name not in levels_data:
                                levels_data[level_name] = []
                            
                            levels_data[level_name].append({
                                'name': area_name,
                                'canopies': [],
                                'marvel_price': marvel_data['total_price'],
                                'marvel_pricing': marvel_data
                            })
        
        # Read VENT CLG pricing from VENT CLG sheets
        for sheet_name in wb.sheetnames:
            if 'VENT CLG - ' in sheet_name:
                sheet = wb[sheet_name]
                title_cell = sheet['B1'].value  # VENT CLG sheets have title in B1 (to match creation)
                
                if title_cell and ' - ' in title_cell and 'VENT CLG SYSTEM' in title_cell:
                    # Extract level and area from title like "Level 1 - Main Kitchen - VENT CLG SYSTEM"
                    title_parts = title_cell.split(' - ')
                    if len(title_parts) >= 2:
                        level_name = title_parts[0]
                        area_name = title_parts[1]
                        
                        # Read VENT CLG data from this sheet
                        vent_clg_data = read_vent_clg_data_from_sheet(sheet)
                        
                        # Find the area and add VENT CLG data
                        area_found = False
                        if level_name in levels_data:
                            for area in levels_data[level_name]:
                                if area['name'] == area_name:
                                    area.update({
                                        'vent_clg_price': vent_clg_data['total_price'],
                                        'vent_clg_cost': vent_clg_data['total_cost'],  # Add cost tracking
                                        'vent_clg_detailed_pricing': vent_clg_data
                                    })
                                    area_found = True
                                    break
                        
                        # If area wasn't found, create it (this shouldn't happen if first pass worked correctly)
                        if not area_found:
                            if level_name not in levels_data:
                                levels_data[level_name] = []
                            
                            levels_data[level_name].append({
                                'name': area_name,
                                'canopies': [],
                                'vent_clg_price': vent_clg_data['total_price'],
                                'vent_clg_detailed_pricing': vent_clg_data
                            })
        
        # Read area-level options from sheets
        # Initialize all areas with default options first
        for level_name, areas in levels_data.items():
            for area in areas:
                if 'options' not in area:
                    area['options'] = {'uvc': False, 'sdu': False, 'recoair': False, 'marvel': False, 'vent_clg': False}
        
        # Check CANOPY sheets for area options written in rows 6-8
        for sheet_name in wb.sheetnames:
            if 'CANOPY - ' in sheet_name or 'CANOPY (UV) - ' in sheet_name:
                sheet = wb[sheet_name]
                title_cell = sheet['B1'].value
                
                if title_cell and ' - ' in title_cell:
                    level_area = title_cell.split(' - ')
                    if len(level_area) >= 2:
                        level_name = level_area[0]
                        area_name = level_area[1]
                        
                        # Check for area options in rows 6-8 (where write_area_options writes them)
                        for row in range(6, 9):
                            cell_value = sheet[f'B{row}'].value
                            if cell_value:
                                cell_value_upper = str(cell_value).upper()
                                
                                # Find the area and update options
                                if level_name in levels_data:
                                    for area in levels_data[level_name]:
                                        if area['name'] == area_name:
                                            if 'UV-C' in cell_value_upper:
                                                area['options']['uvc'] = True
                                            elif 'SDU' in cell_value_upper:
                                                area['options']['sdu'] = True
                                            elif 'RECOAIR' in cell_value_upper:
                                                area['options']['recoair'] = True
                                            break
        
        # Check EBOX sheets for UV-C option (this will override CANOPY sheet if needed)
        for sheet_name in wb.sheetnames:
            if 'EBOX - ' in sheet_name:
                sheet = wb[sheet_name]
                title_cell = sheet['C1'].value  # EBOX sheets have title in C1
                
                if title_cell and ' - ' in title_cell and 'UV-C SYSTEM' in title_cell:
                    # Extract level and area from title like "Level 1 - Main Kitchen - UV-C SYSTEM"
                    title_parts = title_cell.split(' - ')
                    if len(title_parts) >= 2:
                        level_name = title_parts[0]
                        area_name = title_parts[1]
                        
                        # Find the area and set UV-C option to True
                        if level_name in levels_data:
                            for area in levels_data[level_name]:
                                if area['name'] == area_name:
                                    area['options']['uvc'] = True
                                    break
        
        # Check SDU sheets for SDU option - both area-level and canopy-level
        for sheet_name in wb.sheetnames:
            if 'SDU - ' in sheet_name:
                sheet = wb[sheet_name]
                title_cell = sheet['B1'].value  # SDU sheets have title in B1
                
                # Parse the sheet name to get canopy reference
                # Expected format: "SDU - Level Name (Area#) - CanopyRef"
                sheet_parts = sheet_name.split(' - ')
                canopy_ref = None
                if len(sheet_parts) >= 3:
                    canopy_ref = sheet_parts[-1].strip()
                
                if title_cell and ' - ' in title_cell and 'SDU SYSTEM' in title_cell:
                    # Extract level and area from title like "Level 1 - Main Kitchen - SDU SYSTEM - C001"
                    title_parts = title_cell.split(' - ')
                    if len(title_parts) >= 2:
                        level_name = title_parts[0]
                        area_name = title_parts[1]
                        
                        # Find the area and set SDU option to True
                        if level_name in levels_data:
                            for area in levels_data[level_name]:
                                if area['name'] == area_name:
                                    # Set area-level SDU option (for backward compatibility)
                                    area['options']['sdu'] = True
                                    
                                    # If we have a canopy reference, find and set the canopy-level SDU option
                                    if canopy_ref:
                                        for canopy in area.get('canopies', []):
                                            # Case-insensitive comparison for canopy references
                                            if canopy.get('reference_number', '').upper() == canopy_ref.upper():
                                                if 'options' not in canopy:
                                                    canopy['options'] = {}
                                                canopy['options']['sdu'] = True
                                                print(f"✅ Set SDU option for canopy {canopy_ref} in {level_name} - {area_name}")
                                                break
                                    break
        
        # Check RECOAIR sheets for RecoAir option
        for sheet_name in wb.sheetnames:
            if 'RECOAIR - ' in sheet_name:
                sheet = wb[sheet_name]
                title_cell = sheet['C1'].value  # RECOAIR sheets have title in C1
                
                if title_cell and ' - ' in title_cell and 'RECOAIR SYSTEM' in title_cell:
                    # Extract level and area from title like "Level 1 - Main Kitchen - RECOAIR SYSTEM"
                    title_parts = title_cell.split(' - ')
                    if len(title_parts) >= 2:
                        level_name = title_parts[0]
                        area_name = title_parts[1]
                        
                        # Find the area and set RecoAir option to True
                        if level_name in levels_data:
                            for area in levels_data[level_name]:
                                if area['name'] == area_name:
                                    area['options']['recoair'] = True
                                    break
        
        # Check MARVEL sheets for MARVEL option
        for sheet_name in wb.sheetnames:
            if 'MARVEL - ' in sheet_name:
                sheet = wb[sheet_name]
                title_cell = sheet['B1'].value  # MARVEL sheets have title in B1 (fixed to match creation)
                
                if title_cell and ' - ' in title_cell and 'MARVEL SYSTEM' in title_cell:
                    # Extract level and area from title like "Level 1 - Main Kitchen - MARVEL SYSTEM"
                    title_parts = title_cell.split(' - ')
                    if len(title_parts) >= 2:
                        level_name = title_parts[0]
                        area_name = title_parts[1]
                        
                        # Find the area and set MARVEL option to True
                        if level_name in levels_data:
                            for area in levels_data[level_name]:
                                if area['name'] == area_name:
                                    area['options']['marvel'] = True
                                    break
        
        # Check VENT CLG sheets for VENT CLG option
        for sheet_name in wb.sheetnames:
            if 'VENT CLG - ' in sheet_name:
                sheet = wb[sheet_name]
                title_cell = sheet['B1'].value  # VENT CLG sheets have title in B1
                
                if title_cell and ' - ' in title_cell and 'VENT CLG SYSTEM' in title_cell:
                    # Extract level and area from title like "Level 1 - Main Kitchen - VENT CLG SYSTEM"
                    title_parts = title_cell.split(' - ')
                    if len(title_parts) >= 2:
                        level_name = title_parts[0]
                        area_name = title_parts[1]
                        
                        # Find the area and set VENT CLG option to True
                        if level_name in levels_data:
                            for area in levels_data[level_name]:
                                if area['name'] == area_name:
                                    area['options']['vent_clg'] = True
                                    break
        
        # Check for contract sheets to set contract option (handle exact matches and numbered variants)
        contract_sheet_names = ['CONTRACT', 'SPIRAL DUCT', 'SUPPLY DUCT', 'EXTRACT DUCT']
        has_contract_sheets = False
        for sheet_name in wb.sheetnames:
            for contract_name in contract_sheet_names:
                if sheet_name == contract_name or (sheet_name.startswith(contract_name) and len(sheet_name) <= len(contract_name) + 2):
                    has_contract_sheets = True
                    break
            if has_contract_sheets:
                break
        project_data['contract_option'] = has_contract_sheets
        
        if has_contract_sheets:
            print(f"🔨 Detected contract sheets in Excel file, setting contract option to True")
        
        # Convert levels_data to the format expected by the system
        project_data['levels'] = []
        for level_idx, (level_name, areas) in enumerate(levels_data.items(), 1):
            project_data['levels'].append({
                'level_number': level_idx,  # Add level_number field required by save_to_excel
                'level_name': level_name,
                'areas': areas
            })
        
        # Extract UV Extra Over information from both CANOPY (UV) sheets and hidden calculations sheet
        uv_extra_over_data = {}
        
        # First, try to read from the hidden UV_EXTRA_OVER_CALC sheet (new dynamic approach)
        if 'UV_EXTRA_OVER_CALC' in wb.sheetnames:
            try:
                calc_sheet = wb['UV_EXTRA_OVER_CALC']
                print("Found UV_EXTRA_OVER_CALC sheet - reading dynamic UV Extra Over costs...")
                
                # Read data starting from row 2 (row 1 has headers)
                found_data = False
                for row in range(2, calc_sheet.max_row + 1):
                    area_id_cell = calc_sheet[f'A{row}']
                    uv_cost_cell = calc_sheet[f'F{row}']  # UV Extra Over Cost (Price)
                    
                    if area_id_cell.value and uv_cost_cell.value is not None:
                        area_identifier = str(area_id_cell.value).strip()
                        
                        # Skip empty rows and totals row
                        if not area_identifier or area_identifier.upper() == 'TOTALS':
                            continue
                            
                        try:
                            uv_cost = float(uv_cost_cell.value) if uv_cost_cell.value != 0 else 0
                            if uv_cost > 0:  # Only store non-zero costs
                                uv_extra_over_data[area_identifier] = uv_cost
                                print(f"   {area_identifier}: £{uv_cost:.2f}")
                                found_data = True
                        except (ValueError, TypeError):
                            print(f"   Warning: Invalid UV cost for {area_identifier}: {uv_cost_cell.value}")
                            uv_extra_over_data[area_identifier] = 0
                
                # If no data was found with data_only=True, try reading formulas and evaluating manually
                if not found_data:
                    print("   No calculated values found, trying formula-based reading...")
                    
                    # Re-open workbook without data_only to see formulas
                    wb_formulas = load_workbook(excel_path, data_only=False)
                    if 'UV_EXTRA_OVER_CALC' in wb_formulas.sheetnames:
                        calc_sheet_formulas = wb_formulas['UV_EXTRA_OVER_CALC']
                        
                        for row in range(2, calc_sheet_formulas.max_row + 1):
                            area_id_cell = calc_sheet_formulas[f'A{row}']
                            uv_cost_cell = calc_sheet_formulas[f'F{row}']
                            
                            if area_id_cell.value and uv_cost_cell.value:
                                area_identifier = str(area_id_cell.value).strip()
                                
                                if not area_identifier or area_identifier.upper() == 'TOTALS':
                                    continue
                                
                                # If it's a formula, try to evaluate it manually
                                if isinstance(uv_cost_cell.value, str) and uv_cost_cell.value.startswith('='):
                                    formula = uv_cost_cell.value
                                    print(f"   Found formula for {area_identifier}: {formula}")
                                    
                                    # Try to extract sheet references and calculate manually
                                    # This is a simple case for D{row}-E{row} formulas
                                    try:
                                        uv_price_cell = calc_sheet_formulas[f'D{row}']
                                        non_uv_price_cell = calc_sheet_formulas[f'E{row}']
                                        
                                        if (isinstance(uv_price_cell.value, str) and uv_price_cell.value.startswith('=') and
                                            isinstance(non_uv_price_cell.value, str) and non_uv_price_cell.value.startswith('=')):
                                            
                                            # Extract sheet names from formulas
                                            uv_formula = uv_price_cell.value
                                            non_uv_formula = non_uv_price_cell.value
                                            
                                            # Try to get the actual values from the referenced sheets
                                            # area_identifier is like "Level 1 (1)" - need to map to sheet names
                                            uv_sheet_name = f"CANOPY (UV) - {area_identifier}"
                                            non_uv_sheet_name = f"CANOPY - {area_identifier}"
                                            
                                            if uv_sheet_name in wb.sheetnames and non_uv_sheet_name in wb.sheetnames:
                                                uv_sheet = wb[uv_sheet_name]
                                                non_uv_sheet = wb[non_uv_sheet_name]
                                                
                                                print(f"   Checking sheets: {uv_sheet_name} and {non_uv_sheet_name}")
                                                
                                                uv_total = uv_sheet['N9'].value or 0
                                                non_uv_total = non_uv_sheet['N9'].value or 0
                                                
                                                print(f"   UV N9: {uv_total}, Non-UV N9: {non_uv_total}")
                                                
                                                # If N9 is empty, try other common pricing cells
                                                if not uv_total and not non_uv_total:
                                                    # Try K9 (cost cells)
                                                    uv_total = uv_sheet['K9'].value or 0
                                                    non_uv_total = non_uv_sheet['K9'].value or 0
                                                    print(f"   UV K9: {uv_total}, Non-UV K9: {non_uv_total}")
                                                
                                                uv_cost = float(uv_total) - float(non_uv_total) if uv_total and non_uv_total else 0
                                                
                                                if uv_cost > 0:
                                                    uv_extra_over_data[area_identifier] = uv_cost
                                                    print(f"   {area_identifier}: £{uv_cost:.2f} (calculated from sheets)")
                                                    found_data = True
                                                else:
                                                    print(f"   {area_identifier}: No positive UV Extra Over cost calculated (UV: {uv_total}, Non-UV: {non_uv_total})")
                                            else:
                                                print(f"   Sheets not found: {uv_sheet_name} or {non_uv_sheet_name}")
                                                print(f"   Available sheets: {[s for s in wb.sheetnames if 'CANOPY' in s]}")
                                            
                                    except Exception as e:
                                        print(f"   Warning: Could not calculate UV cost for {area_identifier}: {e}")
                                
                                elif isinstance(uv_cost_cell.value, (int, float)) and uv_cost_cell.value > 0:
                                    uv_cost = float(uv_cost_cell.value)
                                    uv_extra_over_data[area_identifier] = uv_cost
                                    print(f"   {area_identifier}: £{uv_cost:.2f}")
                                    found_data = True
                
                if not found_data:
                    print("   No UV Extra Over data found in UV_EXTRA_OVER_CALC sheet")
                            
            except Exception as e:
                print(f"Warning: Could not read from UV_EXTRA_OVER_CALC sheet: {str(e)}")
        
        # Fallback: Read UV Extra Over information from CANOPY (UV) sheets (old A1 cell approach)
        if not uv_extra_over_data:
            print("No UV_EXTRA_OVER_CALC sheet found - checking individual CANOPY (UV) sheets...")
            for sheet_name in wb.sheetnames:
                if 'CANOPY (UV) - ' in sheet_name:
                    sheet = wb[sheet_name]
                    level_area = sheet_name.replace('CANOPY (UV) - ', '').strip()
                    level_area = level_area.split(' - ')
                    
                    if len(level_area) >= 2:
                        level_name = level_area[0]
                        area_name = level_area[1]
                        
                        # Check if UV Extra Over cost information is stored in cell A1
                        uv_extra_over_cell = sheet['A1'].value
                        uv_extra_over_cost = 0
                        
                        if uv_extra_over_cell and isinstance(uv_extra_over_cell, str) and uv_extra_over_cell.startswith('UV_EXTRA_OVER_COST:'):
                            try:
                                uv_extra_over_cost = float(uv_extra_over_cell.split(':')[1])
                            except (ValueError, IndexError):
                                uv_extra_over_cost = 0
                        
                        area_identifier = f"{level_name} ({area_name})" if area_name.startswith('(') else f"{level_name} - {area_name}"
                        uv_extra_over_data[area_identifier] = uv_extra_over_cost
        
        # Apply UV Extra Over data to project areas
        print(f"Applying UV Extra Over data to {len(project_data['levels'])} levels...")
        
        for level_idx, level in enumerate(project_data['levels']):
            level_name = level.get('level_name', '')
            level_number = level.get('level_number', level_idx + 1)
            
            for area_idx, area in enumerate(level['areas']):
                area_name = area.get('name', '')
                area_number = area_idx + 1  # Area number starts from 1
                
                print(f"  Processing area: {level_name} - {area_name} (Level {level_number}, Area {area_number})")
                
                # Create the exact identifier used in UV_EXTRA_OVER_CALC sheet
                # Format: "Level X (Y)" where X is level number and Y is area number
                uv_calc_identifier = f"Level {level_number} ({area_number})"
                
                # Try different area identifier formats to match the hidden sheet data
                possible_identifiers = [
                    uv_calc_identifier,  # e.g., "Level 1 (1)" - MOST LIKELY MATCH
                    f"{level_name} ({area_number})",  # e.g., "LEVEL 1 (1)"
                    f"{level_name} ({area_name})",  # e.g., "LEVEL 1 (Main Kitchen)" 
                    f"{level_name} - {area_name}",  # e.g., "LEVEL 1 - Main Kitchen"
                    f"Level {level_number} ({area_name})",  # e.g., "Level 1 (Main Kitchen)"
                    area_name,  # Just the area name
                    f"({area_number})",  # Just area number with parentheses
                    f"({area_name})",  # Just area name with parentheses
                ]
                
                print(f"    Primary identifier: {uv_calc_identifier}")
                print(f"    Trying identifiers: {possible_identifiers}")
                print(f"    Available UV data keys: {list(uv_extra_over_data.keys())}")
                
                uv_cost_found = 0
                matched_identifier = None
                for identifier in possible_identifiers:
                    if identifier in uv_extra_over_data:
                        uv_cost_found = uv_extra_over_data[identifier]
                        matched_identifier = identifier
                        break
                
                # Set UV Extra Over information for this area
                if 'options' not in area:
                    area['options'] = {}
                
                if uv_cost_found > 0:
                    area['options']['uv_extra_over'] = True
                    area['uv_extra_over_cost'] = uv_cost_found
                    area['extra_over_price'] = uv_cost_found  # Add this field for template compatibility
                    print(f"    ✅ Applied UV Extra Over: £{uv_cost_found:.2f} (matched: {matched_identifier})")
                else:
                    area['options']['uv_extra_over'] = False
                    area['uv_extra_over_cost'] = 0
                    area['extra_over_price'] = 0  # Add this field for template compatibility
                    print(f"    ❌ No UV Extra Over match found")
        
        # Ensure all areas have uv_extra_over option set (default to False)
        for level in project_data['levels']:
            for area in level['areas']:
                if 'options' not in area:
                    area['options'] = {}
                if 'uv_extra_over' not in area['options']:
                    area['options']['uv_extra_over'] = False
                if 'uv_extra_over_cost' not in area:
                    area['uv_extra_over_cost'] = 0
                if 'extra_over_price' not in area:
                    area['extra_over_price'] = 0
                
                # Add extra_overs flag for easy Jinja template access
                area['extra_overs'] = area['options'].get('uv_extra_over', False)
                
                # Ensure extra_over_price matches uv_extra_over_cost
                if area['extra_overs']:
                    area['extra_over_price'] = area.get('uv_extra_over_cost', 0)
        
        # Check for validation errors and include them in the result
        validation_errors = collect_validation_errors()
        if validation_errors:
            # Create a detailed error message with all validation issues
            error_details = "\n\n".join(validation_errors)
            raise Exception(f"Failed to read Excel project data: Data validation errors found:\n\n{error_details}")
        
        return project_data
        
    except Exception as e:
        # Check if this is a validation error (already formatted)
        if "Data validation errors found:" in str(e):
            raise e
        else:
            # For other errors, check if we have validation errors to include
            validation_errors = collect_validation_errors()
            if validation_errors:
                error_details = "\n\n".join(validation_errors)
                raise Exception(f"Failed to read Excel project data: {str(e)}\n\nAdditional validation errors:\n\n{error_details}")
            else:
                raise Exception(f"Failed to read Excel project data: {str(e)}")

def collect_wall_cladding_data(project_data: Dict) -> List[Dict]:
    """
    Collect all wall cladding data from canopies across all levels and areas.
    
    Args:
        project_data (Dict): Project data containing levels and areas
        
    Returns:
        List[Dict]: List of wall cladding specifications with item numbers
    """
    cladding_data = []
    
    levels = project_data.get("levels", [])
    for level in levels:
        for area in level.get("areas", []):
            for canopy in area.get("canopies", []):
                wall_cladding = canopy.get("wall_cladding", {})
                
                # Check if this canopy has wall cladding
                if wall_cladding.get("type") != "None" and wall_cladding.get("type"):
                    # Handle position as list or string
                    position = wall_cladding.get("position", [])
                    if isinstance(position, list):
                        position_list = position if position else []
                    else:
                        position_list = [position] if position else []
                    
                    # Create proper description based on number of positions
                    if len(position_list) == 0:
                        description = "Cladding to walls"
                    elif len(position_list) == 1:
                        description = f"Cladding to {position_list[0]} walls"
                    elif len(position_list) == 2:
                        description = f"Cladding to {position_list[0]} and {position_list[1]} walls"
                    else:
                        # For 3 or more positions: "item1, item2 and item3 walls"
                        description = f"Cladding to {', '.join(position_list[:-1])} and {position_list[-1]} walls"
                    
                    # Join positions for other uses (use "and" format for consistency)
                    position_str = " and ".join(position_list) if position_list else ""
                    
                    cladding_info = {
                        'item_number': canopy.get("reference_number", ""),  # Use canopy reference number
                        'description': description,
                        'width': wall_cladding.get("width", 0),
                        'height': wall_cladding.get("height", 0),
                        'dimensions': f"{wall_cladding.get('width', 0)}X{wall_cladding.get('height', 0)}",
                        'position_description': position_str,
                        'canopy_ref': canopy.get("reference_number", ""),
                        'level_name': level.get("level_name", ""),
                        'area_name': area.get("name", "")
                    }
                    
                    cladding_data.append(cladding_info)
    
    return cladding_data

def write_wall_cladding_summary(sheet: Worksheet, cladding_data: List[Dict]):
    """
    Write wall cladding summary to the sheet starting at row 19.
    
    Args:
        sheet (Worksheet): The worksheet to write to
        cladding_data (List[Dict]): List of wall cladding specifications
    """
    if not cladding_data:
        return
    
    # Starting row for wall cladding section
    start_row = 19
    
    # Write "2M² (HFL)" in C19 to indicate cladding is present
    sheet[f"C{start_row}"] = "2M² (HFL)"
    
    # Write each cladding item
    for idx, cladding in enumerate(cladding_data):
        current_row = start_row + idx
        
        # Item number in column A (if needed)
        # sheet[f"A{current_row}"] = cladding['item_number']
        
        # Description in column B (or appropriate column based on template)
        # sheet[f"B{current_row}"] = cladding['description']
        
        # Dimensions in column P
        sheet[f"P{current_row}"] = cladding['dimensions']
        
        # Position description in column Q  
        sheet[f"Q{current_row}"] = cladding['position_description'] 

# Add this new function after the write_delivery_location_to_sheet function

def create_pricing_summary_sheet(wb: Workbook) -> None:
    """
    Create a hidden PRICING_SUMMARY sheet that aggregates totals from all sheets.
    Uses Excel formulas to reference N9 cells from individual sheets for dynamic updates.
    
    Args:
        wb (Workbook): The workbook to add the pricing summary sheet to
    """
    try:
        # Create or get the PRICING_SUMMARY sheet
        sheet_name = "PRICING_SUMMARY"
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])  # Remove existing sheet to recreate
        
        summary_sheet = wb.create_sheet(sheet_name)
        # Keep PRICING_SUMMARY visible for easier access
        
        # Set up headers
        summary_sheet['A1'] = 'Sheet Type'
        summary_sheet['B1'] = 'Sheet Name'
        summary_sheet['C1'] = 'Total Price (N9)'
        summary_sheet['D1'] = 'Total Cost (K9)'
        summary_sheet['E1'] = 'Price Formula Reference'
        summary_sheet['F1'] = 'Cost Formula Reference'
        
        # Get all visible sheets and categorize them
        canopy_sheets = []
        fire_supp_sheets = []
        ebox_sheets = []
        sdu_sheets = []
        recoair_sheets = []
        marvel_sheets = []
        vent_clg_sheets = []
        contract_sheets = []
        other_sheets = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            if sheet.sheet_state == 'visible':
                if 'CANOPY - ' in sheet_name or 'CANOPY (UV) - ' in sheet_name:
                    canopy_sheets.append(sheet_name)
                elif 'FIRE SUPP - ' in sheet_name:
                    fire_supp_sheets.append(sheet_name)
                elif 'EBOX - ' in sheet_name:
                    ebox_sheets.append(sheet_name)
                elif 'SDU - ' in sheet_name:
                    sdu_sheets.append(sheet_name)
                elif 'RECOAIR - ' in sheet_name:
                    recoair_sheets.append(sheet_name)
                elif 'MARVEL - ' in sheet_name:
                    marvel_sheets.append(sheet_name)
                elif 'VENT CLG - ' in sheet_name:
                    vent_clg_sheets.append(sheet_name)  # VENT CLG gets its own category
                elif sheet_name == 'CONTRACT' or sheet_name.startswith('CONTRACT'):
                    contract_sheets.append(sheet_name)
                elif sheet_name not in ['JOB TOTAL', 'Lists', 'PRICING_SUMMARY', 'ProjectData']:
                    other_sheets.append(sheet_name)
        
        # Write individual sheet references
        current_row = 2
        
        # CANOPY sheets - separate UV and non-UV sheets
        for sheet_name in canopy_sheets:
            # Check if this is a UV Extra Over sheet (exclude from job totals)
            if 'CANOPY (UV) - ' in sheet_name:
                summary_sheet[f'A{current_row}'] = 'UV_EXTRA_OVER'  # Special category for UV sheets
                summary_sheet[f'B{current_row}'] = sheet_name
                safe_sheet_name = f"'{sheet_name}'" if ' ' in sheet_name else sheet_name
                summary_sheet[f'C{current_row}'] = f"=IFERROR({safe_sheet_name}!N9,0)"  # Price formula
                summary_sheet[f'D{current_row}'] = f"=IFERROR({safe_sheet_name}!K9,0)"  # Cost formula
                summary_sheet[f'E{current_row}'] = f"{safe_sheet_name}!N9"  # Price reference
                summary_sheet[f'F{current_row}'] = f"{safe_sheet_name}!K9"  # Cost reference
                summary_sheet[f'G{current_row}'] = "EXCLUDED FROM JOB TOTAL"  # Note column
            else:
                # Regular canopy sheets (included in job totals)
                summary_sheet[f'A{current_row}'] = 'CANOPY'
                summary_sheet[f'B{current_row}'] = sheet_name
                safe_sheet_name = f"'{sheet_name}'" if ' ' in sheet_name else sheet_name
                summary_sheet[f'C{current_row}'] = f"=IFERROR({safe_sheet_name}!N9,0)"  # Price formula
                summary_sheet[f'D{current_row}'] = f"=IFERROR({safe_sheet_name}!K9,0)"  # Cost formula
                summary_sheet[f'E{current_row}'] = f"{safe_sheet_name}!N9"  # Price reference
                summary_sheet[f'F{current_row}'] = f"{safe_sheet_name}!K9"  # Cost reference
            current_row += 1
        
        # FIRE SUPP sheets
        for sheet_name in fire_supp_sheets:
            summary_sheet[f'A{current_row}'] = 'FIRE SUPP'
            summary_sheet[f'B{current_row}'] = sheet_name
            safe_sheet_name = f"'{sheet_name}'" if ' ' in sheet_name else sheet_name
            summary_sheet[f'C{current_row}'] = f"=IFERROR({safe_sheet_name}!N9,0)"  # Price
            summary_sheet[f'D{current_row}'] = f"=IFERROR({safe_sheet_name}!K9,0)"  # Cost
            summary_sheet[f'E{current_row}'] = f"{safe_sheet_name}!N9"  # Price reference
            summary_sheet[f'F{current_row}'] = f"{safe_sheet_name}!K9"  # Cost reference
            current_row += 1
        
        # EBOX sheets
        for sheet_name in ebox_sheets:
            summary_sheet[f'A{current_row}'] = 'EBOX'
            summary_sheet[f'B{current_row}'] = sheet_name
            safe_sheet_name = f"'{sheet_name}'" if ' ' in sheet_name else sheet_name
            summary_sheet[f'C{current_row}'] = f"=IFERROR({safe_sheet_name}!N9,0)"  # Price
            summary_sheet[f'D{current_row}'] = f"=IFERROR({safe_sheet_name}!K9,0)"  # Cost
            summary_sheet[f'E{current_row}'] = f"{safe_sheet_name}!N9"  # Price reference
            summary_sheet[f'F{current_row}'] = f"{safe_sheet_name}!K9"  # Cost reference
            current_row += 1
        
        # SDU sheets
        for sheet_name in sdu_sheets:
            summary_sheet[f'A{current_row}'] = 'SDU'
            summary_sheet[f'B{current_row}'] = sheet_name
            safe_sheet_name = f"'{sheet_name}'" if ' ' in sheet_name else sheet_name
            summary_sheet[f'C{current_row}'] = f"=IFERROR({safe_sheet_name}!J10,0)"  # Price
            summary_sheet[f'D{current_row}'] = f"=IFERROR({safe_sheet_name}!G10,0)"  # Cost
            summary_sheet[f'E{current_row}'] = f"{safe_sheet_name}!J10"  # Price reference
            summary_sheet[f'F{current_row}'] = f"{safe_sheet_name}!G10"  # Cost reference
            current_row += 1
        
        # RECOAIR sheets
        for sheet_name in recoair_sheets:
            summary_sheet[f'A{current_row}'] = 'RECOAIR'
            summary_sheet[f'B{current_row}'] = sheet_name
            safe_sheet_name = f"'{sheet_name}'" if ' ' in sheet_name else sheet_name
            summary_sheet[f'C{current_row}'] = f"=IFERROR({safe_sheet_name}!N9,0)"  # Price
            summary_sheet[f'D{current_row}'] = f"=IFERROR({safe_sheet_name}!K9,0)"  # Cost
            summary_sheet[f'E{current_row}'] = f"{safe_sheet_name}!N9"  # Price reference
            summary_sheet[f'F{current_row}'] = f"{safe_sheet_name}!K9"  # Cost reference
            current_row += 1
        
        # MARVEL sheets
        for sheet_name in marvel_sheets:
            summary_sheet[f'A{current_row}'] = 'MARVEL'
            summary_sheet[f'B{current_row}'] = sheet_name
            safe_sheet_name = f"'{sheet_name}'" if ' ' in sheet_name else sheet_name
            summary_sheet[f'C{current_row}'] = f"=IFERROR({safe_sheet_name}!J9,0)"  # Price - MARVEL uses J9
            summary_sheet[f'D{current_row}'] = f"=IFERROR({safe_sheet_name}!G9,0)"  # Cost - MARVEL uses G9
            summary_sheet[f'E{current_row}'] = f"{safe_sheet_name}!J9"  # Price reference
            summary_sheet[f'F{current_row}'] = f"{safe_sheet_name}!G9"  # Cost reference
            current_row += 1
        
        # VENT CLG sheets
        for sheet_name in vent_clg_sheets:
            summary_sheet[f'A{current_row}'] = 'VENT CLG'
            summary_sheet[f'B{current_row}'] = sheet_name
            safe_sheet_name = f"'{sheet_name}'" if ' ' in sheet_name else sheet_name
            summary_sheet[f'C{current_row}'] = f"=IFERROR({safe_sheet_name}!J10,0)"  # Price - VENT CLG uses J10
            summary_sheet[f'D{current_row}'] = f"=IFERROR({safe_sheet_name}!G10,0)"  # Cost - VENT CLG uses G10
            summary_sheet[f'E{current_row}'] = f"{safe_sheet_name}!J10"  # Price reference
            summary_sheet[f'F{current_row}'] = f"{safe_sheet_name}!G10"  # Cost reference
            current_row += 1
        
        # CONTRACT sheets
        for sheet_name in contract_sheets:
            summary_sheet[f'A{current_row}'] = 'CONTRACT'
            summary_sheet[f'B{current_row}'] = sheet_name
            safe_sheet_name = f"'{sheet_name}'" if ' ' in sheet_name else sheet_name
            summary_sheet[f'C{current_row}'] = f"=IFERROR({safe_sheet_name}!J9,0)"  # Price - CONTRACT uses J9
            summary_sheet[f'D{current_row}'] = f"=IFERROR({safe_sheet_name}!G9,0)"  # Cost - CONTRACT uses G9 (assumed based on pattern)
            summary_sheet[f'E{current_row}'] = f"{safe_sheet_name}!J9"  # Price reference
            summary_sheet[f'F{current_row}'] = f"{safe_sheet_name}!G9"  # Cost reference
            current_row += 1
        
        # OTHER sheets
        for sheet_name in other_sheets:
            summary_sheet[f'A{current_row}'] = 'OTHER'
            summary_sheet[f'B{current_row}'] = sheet_name
            safe_sheet_name = f"'{sheet_name}'" if ' ' in sheet_name else sheet_name
            summary_sheet[f'C{current_row}'] = f"=IFERROR({safe_sheet_name}!N9,0)"  # Price
            summary_sheet[f'D{current_row}'] = f"=IFERROR({safe_sheet_name}!K9,0)"  # Cost
            summary_sheet[f'E{current_row}'] = f"{safe_sheet_name}!N9"  # Price reference
            summary_sheet[f'F{current_row}'] = f"{safe_sheet_name}!K9"  # Cost reference
            current_row += 1
        
        # Add summary totals by type
        summary_row = current_row + 2
        summary_sheet[f'A{summary_row}'] = 'SUMMARY TOTALS'
        summary_sheet[f'B{summary_row + 1}'] = 'CANOPY TOTAL'
        summary_sheet[f'B{summary_row + 2}'] = 'FIRE SUPP TOTAL'
        summary_sheet[f'B{summary_row + 3}'] = 'EBOX TOTAL'
        summary_sheet[f'B{summary_row + 4}'] = 'SDU TOTAL'
        summary_sheet[f'B{summary_row + 5}'] = 'RECOAIR TOTAL'
        summary_sheet[f'B{summary_row + 6}'] = 'MARVEL TOTAL'
        summary_sheet[f'B{summary_row + 7}'] = 'VENT CLG TOTAL'
        summary_sheet[f'B{summary_row + 8}'] = 'CONTRACT TOTAL'
        summary_sheet[f'B{summary_row + 9}'] = 'OTHER TOTAL'
        summary_sheet[f'B{summary_row + 10}'] = 'UV EXTRA OVER TOTAL'
        summary_sheet[f'B{summary_row + 11}'] = 'PROJECT TOTAL'
        
        # Calculate totals using SUMIF formulas
        summary_sheet[f'C{summary_row + 1}'] = f'=SUMIF(A:A,"CANOPY",C:C)'  # Sum all CANOPY sheet prices
        summary_sheet[f'C{summary_row + 2}'] = f'=SUMIF(A:A,"FIRE SUPP",C:C)'  # Sum all FIRE SUPP sheet prices
        summary_sheet[f'C{summary_row + 3}'] = f'=SUMIF(A:A,"EBOX",C:C)'  # Sum all EBOX sheet prices
        summary_sheet[f'C{summary_row + 4}'] = f'=SUMIF(A:A,"SDU",C:C)'  # Sum all SDU sheet prices
        summary_sheet[f'C{summary_row + 5}'] = f'=SUMIF(A:A,"RECOAIR",C:C)'  # Sum all RECOAIR sheet prices
        summary_sheet[f'C{summary_row + 6}'] = f'=SUMIF(A:A,"MARVEL",C:C)'  # Sum all MARVEL sheet prices
        summary_sheet[f'C{summary_row + 7}'] = f'=SUMIF(A:A,"VENT CLG",C:C)'  # Sum all VENT CLG sheet prices
        summary_sheet[f'C{summary_row + 8}'] = f'=SUMIF(A:A,"CONTRACT",C:C)'  # Sum all CONTRACT sheet prices
        summary_sheet[f'C{summary_row + 9}'] = f'=SUMIF(A:A,"OTHER",C:C)'  # Sum all OTHER sheet prices
        summary_sheet[f'C{summary_row + 10}'] = f'=SUMIF(A:A,"UV_EXTRA_OVER",C:C)'  # Sum all UV Extra Over sheet prices (tracked but excluded)
        summary_sheet[f'C{summary_row + 11}'] = f'=C{summary_row + 1}+C{summary_row + 2}+C{summary_row + 3}+C{summary_row + 4}+C{summary_row + 5}+C{summary_row + 6}+C{summary_row + 7}+C{summary_row + 8}+C{summary_row + 9}'  # Project price total (excludes UV Extra Over)
        
        # Cost totals
        summary_sheet[f'D{summary_row + 1}'] = f'=SUMIF(A:A,"CANOPY",D:D)'  # Sum all CANOPY sheet costs
        summary_sheet[f'D{summary_row + 2}'] = f'=SUMIF(A:A,"FIRE SUPP",D:D)'  # Sum all FIRE SUPP sheet costs
        summary_sheet[f'D{summary_row + 3}'] = f'=SUMIF(A:A,"EBOX",D:D)'  # Sum all EBOX sheet costs
        summary_sheet[f'D{summary_row + 4}'] = f'=SUMIF(A:A,"SDU",D:D)'  # Sum all SDU sheet costs
        summary_sheet[f'D{summary_row + 5}'] = f'=SUMIF(A:A,"RECOAIR",D:D)'  # Sum all RECOAIR sheet costs
        summary_sheet[f'D{summary_row + 6}'] = f'=SUMIF(A:A,"MARVEL",D:D)'  # Sum all MARVEL sheet costs
        summary_sheet[f'D{summary_row + 7}'] = f'=SUMIF(A:A,"VENT CLG",D:D)'  # Sum all VENT CLG sheet costs
        summary_sheet[f'D{summary_row + 8}'] = f'=SUMIF(A:A,"CONTRACT",D:D)'  # Sum all CONTRACT sheet costs
        summary_sheet[f'D{summary_row + 9}'] = f'=SUMIF(A:A,"OTHER",D:D)'  # Sum all OTHER sheet costs
        summary_sheet[f'D{summary_row + 10}'] = f'=SUMIF(A:A,"UV_EXTRA_OVER",D:D)'  # Sum all UV Extra Over sheet costs (tracked but excluded)
        summary_sheet[f'D{summary_row + 11}'] = f'=D{summary_row + 1}+D{summary_row + 2}+D{summary_row + 3}+D{summary_row + 4}+D{summary_row + 5}+D{summary_row + 6}+D{summary_row + 7}+D{summary_row + 8}+D{summary_row + 9}'  # Project cost total (excludes UV Extra Over)
        
        # Store the summary row positions for JOB TOTAL to reference
        summary_sheet['H1'] = 'Reference Cells for JOB TOTAL'
        summary_sheet['H2'] = f'CANOPY_PRICE_TOTAL=C{summary_row + 1}'
        summary_sheet['H3'] = f'FIRE_SUPP_PRICE_TOTAL=C{summary_row + 2}'
        summary_sheet['H4'] = f'EBOX_PRICE_TOTAL=C{summary_row + 3}'
        summary_sheet['H5'] = f'SDU_PRICE_TOTAL=C{summary_row + 4}'
        summary_sheet['H6'] = f'RECOAIR_PRICE_TOTAL=C{summary_row + 5}'
        summary_sheet['H7'] = f'MARVEL_PRICE_TOTAL=C{summary_row + 6}'
        summary_sheet['H8'] = f'VENT_CLG_PRICE_TOTAL=C{summary_row + 7}'
        summary_sheet['H9'] = f'CONTRACT_PRICE_TOTAL=C{summary_row + 8}'
        summary_sheet['H10'] = f'OTHER_PRICE_TOTAL=C{summary_row + 9}'
        summary_sheet['H11'] = f'UV_EXTRA_OVER_PRICE_TOTAL=C{summary_row + 10}'
        summary_sheet['H12'] = f'PROJECT_PRICE_TOTAL=C{summary_row + 11}'
        summary_sheet['H13'] = f'CANOPY_COST_TOTAL=D{summary_row + 1}'
        summary_sheet['H14'] = f'FIRE_SUPP_COST_TOTAL=D{summary_row + 2}'
        summary_sheet['H15'] = f'EBOX_COST_TOTAL=D{summary_row + 3}'
        summary_sheet['H16'] = f'SDU_COST_TOTAL=D{summary_row + 4}'
        summary_sheet['H17'] = f'RECOAIR_COST_TOTAL=D{summary_row + 5}'
        summary_sheet['H18'] = f'MARVEL_COST_TOTAL=D{summary_row + 6}'
        summary_sheet['H19'] = f'VENT_CLG_COST_TOTAL=D{summary_row + 7}'
        summary_sheet['H20'] = f'CONTRACT_COST_TOTAL=D{summary_row + 8}'
        summary_sheet['H21'] = f'OTHER_COST_TOTAL=D{summary_row + 9}'
        summary_sheet['H22'] = f'UV_EXTRA_OVER_COST_TOTAL=D{summary_row + 10}'
        summary_sheet['H23'] = f'PROJECT_COST_TOTAL=D{summary_row + 11}'
        
        print(f"Created PRICING_SUMMARY sheet with {current_row - 2} individual sheet references")
        
    except Exception as e:
        print(f"Warning: Could not create PRICING_SUMMARY sheet: {str(e)}")

def update_job_total_sheet(wb: Workbook) -> None:
    """
    Update the JOB TOTAL sheet to reference the PRICING_SUMMARY sheet for dynamic pricing.
    Only references categories that actually exist in the project.
    
    Args:
        wb (Workbook): The workbook containing the JOB TOTAL sheet
    """
    try:
        if 'JOB TOTAL' not in wb.sheetnames:
            print("Warning: JOB TOTAL sheet not found")
            return
        
        if 'PRICING_SUMMARY' not in wb.sheetnames:
            print("Warning: PRICING_SUMMARY sheet not found")
            return
        
        job_total_sheet = wb['JOB TOTAL']
        pricing_summary = wb['PRICING_SUMMARY']
        
        # Clear D17 (other costs) - we don't want to include this
        job_total_sheet['D17'] = 0
        
        # Find the summary totals section
        summary_row = None
        for row in range(1, 100):  # Search first 100 rows
            cell_value = pricing_summary[f'A{row}'].value
            if cell_value and 'SUMMARY TOTALS' in str(cell_value):
                summary_row = row
                break
        
        if not summary_row:
            print("Warning: Could not find SUMMARY TOTALS section in PRICING_SUMMARY")
            return
        
        # Read what categories actually exist from PRICING_SUMMARY
        categories = {}
        for offset in range(1, 10):  # Check rows after SUMMARY TOTALS (increased to include UV EXTRA OVER and PROJECT)
            category_cell = pricing_summary[f'B{summary_row + offset}'].value
            if category_cell:
                category_name = str(category_cell).replace(' TOTAL', '')
                categories[category_name] = {
                    'price_cell': f"C{summary_row + offset}",
                    'cost_cell': f"D{summary_row + offset}"
                }
        
        # Map Job Total categories to PRICING_SUMMARY categories and clear all first
        # Only include active/used systems - exclude unused items like Aerolys, Pollustop, Reactaway
        job_total_mapping = {
            16: ('Canopy', 'CANOPY'),           # Row 16: Canopy
            17: ('Fire Suppression', 'FIRE SUPP'),  # Row 17: Fire Suppression
            18: ('SDU', 'SDU'),                 # Row 18: SDU
            19: ('Vent Clg', 'VENT CLG'),       # Row 19: Vent Clg -> VENT CLG (dedicated category)
            20: ('MARVEL', 'MARVEL'),           # Row 20: MARVEL
            21: ('Edge', 'EBOX'),               # Row 21: Edge -> EBOX
            26: ('Contract', 'CONTRACT'),       # Row 22: Contract -> CONTRACT (was Aerolys, now repurposed for Contract)
            24: ('Reco', 'RECOAIR'),            # Row 24: Reco -> RECOAIR
            # Row 23 (Pollustop), 25 (Reactaway) intentionally excluded - these are unused systems
        }
        
        # Clear all Job Total cells first
        for row_num in range(16, 26):
            job_total_sheet[f'S{row_num}'] = 0  # Cost
            job_total_sheet[f'T{row_num}'] = 0  # Price
        
        # Only populate rows for categories that actually exist
        for row_num, (display_name, pricing_category) in job_total_mapping.items():
            if pricing_category in categories:
                # Set price (column T)
                job_total_sheet[f'T{row_num}'] = f"=PRICING_SUMMARY!{categories[pricing_category]['price_cell']}"
                # Set cost (column S)
                job_total_sheet[f'S{row_num}'] = f"=PRICING_SUMMARY!{categories[pricing_category]['cost_cell']}"
                print(f"✓ Linked {display_name} (row {row_num}) to {pricing_category} category")
            else:
                # Category doesn't exist - leave as 0
                print(f"○ Skipped {display_name} (row {row_num}) - {pricing_category} category not found")
            # Add margin formula in column U with IFERROR wrapper (always set)
            job_total_sheet[f'U{row_num}'] = f"=IFERROR((T{row_num}-S{row_num})/T{row_num}, \"\")"
        
        print(f"Updated JOB TOTAL sheet with dynamic pricing formulas for {len([cat for cat in categories.keys() if cat in [mapping[1] for mapping in job_total_mapping.values()]])} categories and added margin formulas")
        
    except Exception as e:
        print(f"Warning: Could not update JOB TOTAL sheet: {str(e)}")

# Add this new function after the save_to_excel function

def create_revision_from_existing(excel_path: str, new_revision: str, new_date: str = None) -> str:
    """
    Create a new revision by copying an existing Excel file and updating only the revision and date.
    This preserves all existing data, formulas, and pricing.
    
    Args:
        excel_path (str): Path to the existing Excel file
        new_revision (str): New revision letter (e.g., "B", "C")
        new_date (str, optional): New date in DD/MM/YYYY format. If None, keeps existing date.
    
    Returns:
        str: Path to the new revision file
    """
    try:
        # Load the existing workbook (without data_only to preserve formulas)
        wb = load_workbook(excel_path)
        
        # Update revision in all sheets that have the revision field (K7 or O7)
        sheets_to_update = []
        revision_cells = {}  # Track which cell contains revision for each sheet
        date_cells = {}  # Track which cell contains date for each sheet
        
        # Define the patterns for different sheet types
        sheet_patterns = {
            'G7_date': ['JOB TOTAL', 'CANOPY', 'FIRE SUPP', 'CONTRACT', 'SPIRAL DUCT', 'SUPPLY DUCT', 'EXTRACT DUCT'],
            'H7_date': ['EBOX', 'EDGE BOX', 'RECOAIR'],
            'F7_date': ['MARVEL', 'SDU']
        }
        
        # Check all visible sheets for revision field
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            if sheet.sheet_state == 'visible':
                try:
                    # Determine date cell based on sheet type
                    date_cell = 'G7'  # Default
                    for pattern_key, patterns in sheet_patterns.items():
                        if any(pattern in sheet_name.upper() for pattern in patterns):
                            date_cell = pattern_key.split('_')[0]
                            break
                    
                    # Check K7 first (most common location)
                    has_revision = False
                    if sheet['K7'].value is not None:
                        sheets_to_update.append(sheet_name)
                        revision_cells[sheet_name] = 'K7'
                        date_cells[sheet_name] = date_cell
                        has_revision = True
                    # Then check O7 (some sheets use this)
                    elif sheet['O7'].value is not None:
                        sheets_to_update.append(sheet_name)
                        revision_cells[sheet_name] = 'O7'
                        date_cells[sheet_name] = date_cell
                        has_revision = True
                    
                    # If no revision field but has a date field, still include for date update
                    if not has_revision and sheet[date_cell].value is not None:
                        sheets_to_update.append(sheet_name)
                        date_cells[sheet_name] = date_cell
                except:
                    # Skip sheets that don't have these cells or can't access them
                    continue
        
        print(f"DEBUG: Found {len(sheets_to_update)} sheets to update: {sheets_to_update}")
        print(f"DEBUG: Revision cells: {revision_cells}")
        print(f"DEBUG: Date cells: {date_cells}")
        
        # Update revision in all identified sheets
        for sheet_name in sheets_to_update:
            sheet = wb[sheet_name]
            if sheet_name in revision_cells:  # Only update revision if the sheet has a revision field
                try:
                    revision_cell = revision_cells[sheet_name]
                    sheet[revision_cell] = new_revision
                    print(f"Updated revision to {new_revision} in cell {revision_cell} of sheet: {sheet_name}")
                except Exception as e:
                    print(f"Warning: Could not update revision in sheet {sheet_name}: {str(e)}")
        
        # Update date if provided
        if new_date:
            print(f"DEBUG: Updating date to '{new_date}' in {len(sheets_to_update)} sheets")
            for sheet_name in sheets_to_update:
                sheet = wb[sheet_name]
                try:
                    # Use the correct date cell for this sheet
                    date_cell = date_cells.get(sheet_name, 'G7')  # Default to G7 if not found
                    old_value = sheet[date_cell].value
                    sheet[date_cell] = new_date
                    print(f"Updated date from '{old_value}' to '{new_date}' in cell {date_cell} of sheet: {sheet_name}")
                except Exception as e:
                    print(f"Warning: Could not update date in sheet {sheet_name}: {str(e)}")
        else:
            print(f"DEBUG: No new date provided, keeping existing dates")
        
        # Update revision in ProjectData sheet if it exists
        if 'ProjectData' in wb.sheetnames:
            try:
                hidden_sheet = wb['ProjectData']
                hidden_sheet['B8'] = new_revision  # Update revision in ProjectData (moved from B7 to B8)
                if new_date:
                    # Add date to ProjectData if not already there
                    hidden_sheet['A9'] = 'Date'
                    hidden_sheet['B9'] = new_date
                print(f"Updated revision in ProjectData sheet to {new_revision}")
            except Exception as e:
                print(f"Warning: Could not update ProjectData sheet: {str(e)}")
        
        # Generate output filename
        # Extract project number from the original filename or from the data
        original_filename = os.path.basename(excel_path)
        
        # Try to extract project number from filename or from sheet data
        project_number = "unknown"
        try:
            # Try to get project number from JOB TOTAL sheet
            if 'JOB TOTAL' in wb.sheetnames:
                job_total_sheet = wb['JOB TOTAL']
                project_number = job_total_sheet['C3'].value or "unknown"
            elif sheets_to_update:
                # Get from first available sheet
                first_sheet = wb[sheets_to_update[0]]
                project_number = first_sheet['C3'].value or "unknown"
        except:
            # If we can't get project number from sheets, try to extract from filename
            if " Cost Sheet " in original_filename:
                project_number = original_filename.split(" Cost Sheet ")[0]
        
        # Format date for filename
        if new_date:
            formatted_date = new_date.replace('/', '')
        else:
            # Try to get existing date from sheet
            try:
                if 'JOB TOTAL' in wb.sheetnames:
                    existing_date = wb['JOB TOTAL']['G7'].value or ""
                elif sheets_to_update:
                    # Get date from first sheet using the correct cell
                    first_sheet_name = sheets_to_update[0]
                    date_cell = date_cells.get(first_sheet_name, 'G7')
                    existing_date = wb[first_sheet_name][date_cell].value or ""
                else:
                    existing_date = ""
                
                if existing_date:
                    formatted_date = str(existing_date).replace('/', '')
                else:
                    formatted_date = get_current_date().replace('/', '')
            except:
                formatted_date = get_current_date().replace('/', '')
        
        # Create output filename: "Project Number Cost Sheet Date"
        output_filename = f"{project_number} Cost Sheet {formatted_date}.xlsx"
        output_path = f"output/{output_filename}"
        
        # Ensure output directory exists
        os.makedirs("output", exist_ok=True)
        
        # Save the updated workbook
        wb.save(output_path)
        wb.close()
        
        print(f"Created revision {new_revision} at: {output_path}")
        return output_path
        
    except Exception as e:
        raise Exception(f"Failed to create revision from existing file: {str(e)}")

def extract_sdu_electrical_services(sheet: Worksheet) -> Dict:
    """
    Extract electrical and gas services data from SDU sheet.
    Reads from the electrical services section (B35-B68) and gas services section (C71-C82).
    Also reads the SDU item number from B12.
    
    Args:
        sheet (Worksheet): The SDU worksheet to read from
        
    Returns:
        Dict: Electrical and gas services data with mapped values, plus SDU item number
    """
    sheet_name = sheet.title
    
    try:
        # Get SDU item number from B12
        sdu_item_number = sheet['B12'].value or ""
        electrical_services = {
            'distribution_board': 0,
            'single_phase_switched_spur': 0,
            'three_phase_socket_outlet': 0,
            'switched_socket_outlet': 0,
            'emergency_knock_off': 0,
            'ring_main_inc_2no_sso': 0
        }
        
        # Check distribution board value at C35 with validation (use 'integer' for quantities)
        distribution_valid, distribution_value, distribution_error = validate_cell_data(
            sheet_name, 'C35', sheet['C35'].value, 'integer', 'Distribution Board Quantity'
        )
        if not distribution_valid:
            add_validation_error(distribution_error)
            electrical_services['distribution_board'] = 0
        else:
            electrical_services['distribution_board'] = distribution_value if distribution_value > 0 else 0
        
        # Count single phase and three phase socket outlets based on dropdown selections
        single_phase_count = 0
        three_phase_count = 0
        
        # Check D40-D47 for ISO/OUTLET (NO MCB) options and their quantities in C40-C47
        for row in range(40, 48):  # D40 to D47 and C40 to C47
            dropdown_value = sheet[f'D{row}'].value
            quantity_value = sheet[f'C{row}'].value
            
            if dropdown_value and quantity_value:
                dropdown_str = str(dropdown_value).strip()
                try:
                    quantity = int(float(str(quantity_value).strip()))
                    if quantity > 0:
                        if '1-PH' in dropdown_str:
                            single_phase_count += quantity
                        elif '3-PH' in dropdown_str:
                            three_phase_count += quantity
                except (ValueError, TypeError):
                    continue
        
        # Check D49-D56 for ISO/OUTLET (MCB) options and their quantities in C49-C56
        for row in range(49, 57):  # D49 to D56 and C49 to C56
            dropdown_value = sheet[f'D{row}'].value
            quantity_value = sheet[f'C{row}'].value
            
            if dropdown_value and quantity_value:
                dropdown_str = str(dropdown_value).strip()
                try:
                    quantity = int(float(str(quantity_value).strip()))
                    if quantity > 0:
                        if '1-PH' in dropdown_str:
                            single_phase_count += quantity
                        elif '3-PH' in dropdown_str:
                            three_phase_count += quantity
                except (ValueError, TypeError):
                    continue
        
        # Store the totals
        electrical_services['single_phase_switched_spur'] = single_phase_count
        electrical_services['three_phase_socket_outlet'] = three_phase_count
        
        # Switched socket outlet value at C65
        switched_socket_value = sheet['C65'].value
        if switched_socket_value and str(switched_socket_value).strip() not in ['', '0', '-']:
            try:
                electrical_services['switched_socket_outlet'] = int(float(str(switched_socket_value).strip()))
            except (ValueError, TypeError):
                electrical_services['switched_socket_outlet'] = 0
        
        # Emergency knock-off value (assuming it's around the electrical services section)
        # You may need to specify the exact cell for this
        emergency_value = sheet['C57'].value  # Adjust this cell reference as needed
        if emergency_value and str(emergency_value).strip() not in ['', '0', '-']:
            try:
                electrical_services['emergency_knock_off'] = int(float(str(emergency_value).strip()))
            except (ValueError, TypeError):
                electrical_services['emergency_knock_off'] = 0
        
        # Ring main inc. 2no SSO value at C68
        ring_main_value = sheet['C68'].value
        if ring_main_value and str(ring_main_value).strip() not in ['', '0', '-']:
            try:
                electrical_services['ring_main_inc_2no_sso'] = int(float(str(ring_main_value).strip()))
            except (ValueError, TypeError):
                electrical_services['ring_main_inc_2no_sso'] = 0
        
        # Gas services extraction
        gas_services = {
            'gas_manifold': 0,
            'gas_connection_15mm': 0,
            'gas_connection_20mm': 0,
            'gas_connection_25mm': 0,
            'gas_connection_32mm': 0,
            'gas_solenoid_valve': 0
        }
        
        # Gas manifold value from C71-C74 (take first non-zero value)
        for row in range(71, 75):  # C71 to C74
            cell_value = sheet[f'C{row}'].value
            if cell_value and str(cell_value).strip() not in ['', '0', '-']:
                try:
                    gas_services['gas_manifold'] = int(float(str(cell_value).strip()))
                    break  # Take the first non-zero value found
                except (ValueError, TypeError):
                    continue
        
        # Gas connections - specific cell locations from C75 to C78
        gas_connection_cells = {
            'gas_connection_15mm': 'C75',   # 15MM gas connection
            'gas_connection_20mm': 'C76',   # 20MM gas connection  
            'gas_connection_25mm': 'C77',   # 25MM gas connection
            'gas_connection_32mm': 'C78'    # 32MM gas connection
        }
        
        for service_name, cell_ref in gas_connection_cells.items():
            try:
                cell_value = sheet[cell_ref].value
                if cell_value and str(cell_value).strip() not in ['', '0', '-']:
                    gas_services[service_name] = int(float(str(cell_value).strip()))
            except (ValueError, TypeError, KeyError):
                gas_services[service_name] = 0
        
        # Gas solenoid valve from C79-C82 (take first non-zero value)
        for row in range(79, 83):  # C79 to C82
            cell_value = sheet[f'C{row}'].value
            if cell_value and str(cell_value).strip() not in ['', '0', '-']:
                try:
                    gas_services['gas_solenoid_valve'] = int(float(str(cell_value).strip()))
                    break  # Take the first non-zero value found
                except (ValueError, TypeError):
                    continue
        
        # Water services extraction
        water_services = {
            'cws_manifold_22mm': 0,
            'cws_manifold_15mm': 0,
            'hws_manifold': 0,
            'water_connection_15mm': 0,
            'water_connection_22mm': 0,
            'water_connection_28mm': 0
        }
        
        # Extract manifold values
        # 22mm CWS manifold at C86
        cws_22mm_manifold = sheet['C86'].value
        if cws_22mm_manifold and str(cws_22mm_manifold).strip() not in ['', '0', '-']:
            try:
                water_services['cws_manifold_22mm'] = int(float(str(cws_22mm_manifold).strip()))
            except (ValueError, TypeError):
                water_services['cws_manifold_22mm'] = 0
        
        # 15mm CWS manifold at C87
        cws_15mm_manifold = sheet['C87'].value
        if cws_15mm_manifold and str(cws_15mm_manifold).strip() not in ['', '0', '-']:
            try:
                water_services['cws_manifold_15mm'] = int(float(str(cws_15mm_manifold).strip()))
            except (ValueError, TypeError):
                water_services['cws_manifold_15mm'] = 0
        
        # HWS manifold at C88
        hws_manifold = sheet['C88'].value
        if hws_manifold and str(hws_manifold).strip() not in ['', '0', '-']:
            try:
                water_services['hws_manifold'] = int(float(str(hws_manifold).strip()))
            except (ValueError, TypeError):
                water_services['hws_manifold'] = 0
        
        # Extract water connection values from fixed cells
        # C89: 15mm connection
        connection_15mm_value = sheet['C89'].value
        if connection_15mm_value and str(connection_15mm_value).strip() not in ['', '0', '-']:
            try:
                water_services['water_connection_15mm'] = int(float(str(connection_15mm_value).strip()))
            except (ValueError, TypeError):
                water_services['water_connection_15mm'] = 0
        else:
            water_services['water_connection_15mm'] = 0
        
        # C90: 22mm connection
        connection_22mm_value = sheet['C90'].value
        if connection_22mm_value and str(connection_22mm_value).strip() not in ['', '0', '-']:
            try:
                water_services['water_connection_22mm'] = int(float(str(connection_22mm_value).strip()))
            except (ValueError, TypeError):
                water_services['water_connection_22mm'] = 0
        else:
            water_services['water_connection_22mm'] = 0
        
        # C91: 28mm connection
        connection_28mm_value = sheet['C91'].value
        if connection_28mm_value and str(connection_28mm_value).strip() not in ['', '0', '-']:
            try:
                water_services['water_connection_28mm'] = int(float(str(connection_28mm_value).strip()))
            except (ValueError, TypeError):
                water_services['water_connection_28mm'] = 0
        else:
            water_services['water_connection_28mm'] = 0
        
        # Extract pricing information
        pricing = {
            'carcass_only_price': 0,
            'electrical_mechanical_price': 0,
            'live_site_test_price': 0,
            'delivery_price': 0,
            'final_carcass_price': 0,
            'final_electrical_price': 0,
            'has_live_test': False
        }
        
        # Carcass only price at C105
        carcass_price = sheet['C105'].value
        if carcass_price and str(carcass_price).strip() not in ['', '0', '-']:
            try:
                pricing['carcass_only_price'] = float(str(carcass_price).strip())
            except (ValueError, TypeError):
                pricing['carcass_only_price'] = 0
        
        # Electrical & mechanical services price at C106
        electrical_price = sheet['C106'].value
        if electrical_price and str(electrical_price).strip() not in ['', '0', '-']:
            try:
                pricing['electrical_mechanical_price'] = float(str(electrical_price).strip())
            except (ValueError, TypeError):
                pricing['electrical_mechanical_price'] = 0
        
        # Check for live site test at C102 and cost at J102
        live_test_quantity = sheet['C102'].value
        if live_test_quantity and str(live_test_quantity).strip() not in ['', '0', '-']:
            try:
                test_qty = float(str(live_test_quantity).strip())
                if test_qty > 0:
                    pricing['has_live_test'] = True
                    # Get cost from J102
                    live_test_cost = sheet['J102'].value
                    if live_test_cost and str(live_test_cost).strip() not in ['', '0', '-']:
                        try:
                            pricing['live_site_test_price'] = float(str(live_test_cost).strip())
                        except (ValueError, TypeError):
                            pricing['live_site_test_price'] = 0
            except (ValueError, TypeError):
                pass
        
        # Delivery price at C107 (to be divided by 2)
        delivery_price = sheet['C107'].value
        if delivery_price and str(delivery_price).strip() not in ['', '0', '-']:
            try:
                pricing['delivery_price'] = float(str(delivery_price).strip())
            except (ValueError, TypeError):
                pricing['delivery_price'] = 0
        
        # Calculate final prices
        # Delivery price divided by 2, half added to each
        half_delivery = pricing['delivery_price'] / 2
        pricing['final_carcass_price'] = pricing['carcass_only_price'] + half_delivery
        pricing['final_electrical_price'] = pricing['electrical_mechanical_price'] + half_delivery
        
        # Combine electrical, gas, water services, pricing, and SDU item number
        result = {
            'electrical_services': electrical_services,
            'gas_services': gas_services,
            'water_services': water_services,
            'pricing': pricing,
            'sdu_item_number': sdu_item_number
        }
        
        return result
        
    except Exception as e:
        print(f"Warning: Could not extract electrical, gas, and water services data from SDU sheet: {str(e)}")
        return {
            'electrical_services': {
                'distribution_board': 0,
                'single_phase_switched_spur': 0,
                'three_phase_socket_outlet': 0,
                'switched_socket_outlet': 0,
                'emergency_knock_off': 0,
                'ring_main_inc_2no_sso': 0
            },
            'gas_services': {
                'gas_manifold': 0,
                'gas_connection_15mm': 0,
                'gas_connection_20mm': 0,
                'gas_connection_25mm': 0,
                'gas_connection_32mm': 0,
                'gas_solenoid_valve': 0
            },
            'water_services': {
                'cws_manifold_22mm': 0,
                'cws_manifold_15mm': 0,
                'hws_manifold': 0,
                'water_connection_15mm': 0,
                'water_connection_22mm': 0,
                'water_connection_28mm': 0
            },
            'pricing': {
                'carcass_only_price': 0,
                'electrical_mechanical_price': 0,
                'live_site_test_price': 0,
                'delivery_price': 0,
                'final_carcass_price': 0,
                'final_electrical_price': 0,
                'has_live_test': False
            },
            'sdu_item_number': ''
        }

def validate_cell_data(sheet_name: str, cell_ref: str, value, expected_type: str, context: str = "") -> tuple:
    """
    Validate cell data and return validation result with detailed error information.
    
    Args:
        sheet_name (str): Name of the Excel sheet
        cell_ref (str): Cell reference (e.g., 'C35', 'N29')
        value: The value from the Excel cell
        expected_type (str): Expected data type ('number', 'integer', 'text', 'boolean')
        context (str): Additional context about what this value is used for
        
    Returns:
        tuple: (is_valid: bool, converted_value, error_message: str)
    """
    if value is None or str(value).strip() == "":
        return True, 0 if expected_type in ['number', 'integer'] else "", ""
    
    try:
        if expected_type == 'number':
            # Try to convert to float first
            converted = float(str(value).strip())
            return True, converted, ""
        elif expected_type == 'integer':
            # Try to convert to integer (for dimensions, quantities, etc.)
            converted = int(float(str(value).strip()))  # Convert via float first to handle "1815.0" -> 1815
            return True, converted, ""
        elif expected_type == 'text':
            return True, str(value).strip(), ""
        elif expected_type == 'boolean':
            if isinstance(value, bool):
                return True, value, ""
            # Try to interpret as boolean
            str_val = str(value).strip().lower()
            if str_val in ['true', '1', 'yes', 'y']:
                return True, True, ""
            elif str_val in ['false', '0', 'no', 'n']:
                return True, False, ""
            else:
                return False, False, f"Cannot convert '{value}' to boolean"
        else:
            return True, value, ""
            
    except (ValueError, TypeError) as e:
        error_msg = f"**Data Type Error in {sheet_name}**\n"
        error_msg += f"   **Location:** Cell {cell_ref}\n"
        error_msg += f"   **Found Value:** '{value}' (type: {type(value).__name__})\n"
        error_msg += f"   **Expected:** {expected_type}\n"
        if context:
            error_msg += f"   **Used For:** {context}\n"
        error_msg += f"   **Fix:** Please enter a valid {expected_type} value in cell {cell_ref}"
        
        return False, None, error_msg

def collect_validation_errors() -> list:
    """
    Global list to collect validation errors during Excel reading.
    """
    if not hasattr(collect_validation_errors, 'errors'):
        collect_validation_errors.errors = []
    return collect_validation_errors.errors

def clear_validation_errors():
    """
    Clear the global validation errors list.
    """
    collect_validation_errors.errors = []

def add_validation_error(error_message: str):
    """
    Add a validation error to the global list.
    """
    errors = collect_validation_errors()
    errors.append(error_message)

def calculate_uv_extra_over_cost(wb: Workbook, level_name: str, area_number: int, uv_sheet_name: str, non_uv_sheet_name: str) -> float:
    """
    Calculate the UV Extra Over cost by comparing UV canopies vs non-UV equivalent canopies.
    
    Args:
        wb (Workbook): The workbook containing the UV comparison sheets
        level_name (str): Level name for sheet identification
        area_number (int): Area number for sheet identification
        uv_sheet_name (str): Name of the sheet with UV canopies
        non_uv_sheet_name (str): Name of the sheet with non-UV equivalent canopies
        
    Returns:
        float: The UV Extra Over cost (UV cost - non-UV cost)
    """
    try:
        if uv_sheet_name not in wb.sheetnames or non_uv_sheet_name not in wb.sheetnames:
            return 0.0
        
        uv_sheet = wb[uv_sheet_name]
        non_uv_sheet = wb[non_uv_sheet_name]
        
        # Get the total cost from both sheets (assuming it's in a standard location like N9 or similar)
        # You may need to adjust this based on your Excel template structure
        uv_total_cell = uv_sheet['N9']  # Adjust cell reference as needed
        non_uv_total_cell = non_uv_sheet['N9']  # Adjust cell reference as needed
        
        uv_total = uv_total_cell.value or 0
        non_uv_total = non_uv_total_cell.value or 0
        
        # Calculate the difference (UV cost - non-UV cost)
        uv_extra_over_cost = float(uv_total) - float(non_uv_total)
        
        return max(0, uv_extra_over_cost)  # Ensure non-negative cost
        
    except Exception as e:
        print(f"Warning: Could not calculate UV Extra Over cost for {level_name} ({area_number}): {str(e)}")
        return 0.0

def create_uv_extra_over_calculations_sheet(wb: Workbook) -> None:
    """
    Create a hidden sheet to track UV Extra Over calculations with dynamic formulas.
    This sheet calculates the difference between UV canopy costs and non-UV equivalent costs per area.
    
    Args:
        wb (Workbook): The workbook to add the calculations sheet to
    """
    try:
        # Create or get the hidden calculations sheet
        sheet_name = "UV_EXTRA_OVER_CALC"
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            # Clear existing data
            sheet.delete_rows(1, sheet.max_row)
        else:
            sheet = wb.create_sheet(sheet_name)
        
        # Hide the sheet
        sheet.sheet_state = 'hidden'
        
        # Set up headers
        sheet['A1'] = 'Area Identifier'
        sheet['B1'] = 'UV Sheet Name'
        sheet['C1'] = 'Non-UV Sheet Name'
        sheet['D1'] = 'UV Total Price'
        sheet['E1'] = 'Non-UV Total Price'
        sheet['F1'] = 'UV Extra Over Cost'
        sheet['G1'] = 'UV Total Cost'
        sheet['H1'] = 'Non-UV Total Cost'
        sheet['I1'] = 'UV Extra Over Cost (Cost)'
        
        # Style headers
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            cell = sheet[f'{col}1']
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Find all UV Extra Over pairs and add formulas
        row = 2
        for sheet_name in wb.sheetnames:
            if 'CANOPY (UV) - ' in sheet_name:
                # Find corresponding non-UV sheet
                uv_sheet_title = sheet_name
                # Look for the corresponding non-UV sheet (same area, but without UV)
                area_part = sheet_name.replace('CANOPY (UV) - ', '')  # e.g., "LEVEL 1 (1)"
                non_uv_sheet_name = f"CANOPY - {area_part}"
                
                if non_uv_sheet_name in wb.sheetnames:
                    # Add calculation row
                    sheet[f'A{row}'] = area_part  # Area identifier
                    sheet[f'B{row}'] = uv_sheet_title  # UV sheet name
                    sheet[f'C{row}'] = non_uv_sheet_name  # Non-UV sheet name
                    
                    # Dynamic formulas that reference the actual sheet totals
                    uv_safe_name = f"'{uv_sheet_title}'" if ' ' in uv_sheet_title else uv_sheet_title
                    non_uv_safe_name = f"'{non_uv_sheet_name}'" if ' ' in non_uv_sheet_name else non_uv_sheet_name
                    
                    # Price formulas (N9 contains the total price)
                    sheet[f'D{row}'] = f"=IFERROR({uv_safe_name}!N9,0)"  # UV Total Price
                    sheet[f'E{row}'] = f"=IFERROR({non_uv_safe_name}!N9,0)"  # Non-UV Total Price
                    sheet[f'F{row}'] = f"=D{row}-E{row}"  # UV Extra Over Cost (Price)
                    
                    # Cost formulas (K9 contains the total cost)
                    sheet[f'G{row}'] = f"=IFERROR({uv_safe_name}!K9,0)"  # UV Total Cost
                    sheet[f'H{row}'] = f"=IFERROR({non_uv_safe_name}!K9,0)"  # Non-UV Total Cost
                    sheet[f'I{row}'] = f"=G{row}-H{row}"  # UV Extra Over Cost (Cost)
                    
                    row += 1
        
        # Add summary totals at the bottom
        if row > 2:  # Only if we have data
            summary_row = row + 1
            sheet[f'A{summary_row}'] = 'TOTALS'
            sheet[f'A{summary_row}'].font = Font(bold=True)
            
            # Sum all UV Extra Over costs
            sheet[f'F{summary_row}'] = f"=SUM(F2:F{row-1})"  # Total UV Extra Over (Price)
            sheet[f'I{summary_row}'] = f"=SUM(I2:I{row-1})"  # Total UV Extra Over (Cost)
            
            # Style totals row
            for col in ['A', 'F', 'I']:
                cell = sheet[f'{col}{summary_row}']
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        
        print(f"Created UV Extra Over calculations sheet with {row-2} area calculations")
        
    except Exception as e:
        print(f"Warning: Could not create UV Extra Over calculations sheet: {str(e)}")
        pass

def add_plant_selection_dropdowns_to_ebox(sheet: Worksheet):
    """
    Add plant selection dropdowns to EBOX sheet at E39 and E40.
    
    Args:
        sheet (Worksheet): The EBOX worksheet to add dropdowns to
    """
    try:
        # Plant selection options
        plant_options = [
            "",  # Empty option
            "SL10 GENIE",
            "EXTENSION FORKS",
            "2.5M COMBI LADDER",
            "1.5M PODIUM",
            "3M TOWER",
            "COMBI LADDER",
            "PECO LIFT",
            "3M YOUNGMAN BOARD",
            "GS1930 SCISSOR LIFT",
            "4-6 SHERASCOPIC",
            "7-9 SHERASCOPIC"
        ]
        
        # Create validation
        from openpyxl.worksheet.datavalidation import DataValidation
        formula = ",".join(plant_options)
        plant_dv = DataValidation(type="list", formula1=f'"{formula}"', allow_blank=True)
        
        # Add validation to sheet
        sheet.add_data_validation(plant_dv)
        
        # Apply to E39 and E40
        plant_dv.add("E39")
        plant_dv.add("E40")
        
        print(f"✅ Added plant selection dropdowns to E39 and E40 on EBOX sheet")
        
    except Exception as e:
        print(f"Warning: Could not add plant selection dropdowns to EBOX sheet: {str(e)}")

def add_plant_selection_dropdowns_to_recoair(sheet: Worksheet):
    """
    Add plant selection dropdowns to RecoAir sheet at E38 and E39.
    
    Args:
        sheet (Worksheet): The RecoAir worksheet to add dropdowns to
    """
    try:
        # Plant selection options
        plant_options = [
            "",  # Empty option
            "SL10 GENIE",
            "EXTENSION FORKS",
            "2.5M COMBI LADDER",
            "1.5M PODIUM",
            "3M TOWER",
            "COMBI LADDER",
            "PECO LIFT",
            "3M YOUNGMAN BOARD",
            "GS1930 SCISSOR LIFT",
            "4-6 SHERASCOPIC",
            "7-9 SHERASCOPIC"
        ]
        
        # Create validation
        from openpyxl.worksheet.datavalidation import DataValidation
        formula = ",".join(plant_options)
        plant_dv = DataValidation(type="list", formula1=f'"{formula}"', allow_blank=True)
        
        # Add validation to sheet
        sheet.add_data_validation(plant_dv)
        
        # Apply to E38 and E39
        plant_dv.add("E38")
        plant_dv.add("E39")
        
        print(f"✅ Added plant selection dropdowns to E38 and E39 on RecoAir sheet")
        
    except Exception as e:
        print(f"Warning: Could not add plant selection dropdowns to RecoAir sheet: {str(e)}")

def add_plant_selection_dropdown_to_fire_supp(sheet: Worksheet):
    """
    Add plant selection dropdown to Fire Suppression sheet at D184.
    
    Args:
        sheet (Worksheet): The Fire Suppression worksheet to add dropdown to
    """
    try:
        # Plant selection options
        plant_options = [
            "",  # Empty option
            "SL10 GENIE",
            "EXTENSION FORKS",
            "2.5M COMBI LADDER",
            "1.5M PODIUM",
            "3M TOWER",
            "COMBI LADDER",
            "PECO LIFT",
            "3M YOUNGMAN BOARD",
            "GS1930 SCISSOR LIFT",
            "4-6 SHERASCOPIC",
            "7-9 SHERASCOPIC"
        ]
        
        # Create validation
        from openpyxl.worksheet.datavalidation import DataValidation
        formula = ",".join(plant_options)
        plant_dv = DataValidation(type="list", formula1=f'"{formula}"', allow_blank=True)
        
        # Add validation to sheet
        sheet.add_data_validation(plant_dv)
        
        # Apply to D184
        plant_dv.add("D184")
        
        print(f"✅ Added plant selection dropdown to D184 on Fire Suppression sheet")
        
    except Exception as e:
        print(f"Warning: Could not add plant selection dropdown to Fire Suppression sheet: {str(e)}")

# ---------------------- MARVEL METADATA ----------------------

def write_marvel_metadata(sheet: Worksheet, project_data: Dict, template_version: str = None):
    """
    Write project metadata to MARVEL sheet with specific cell mappings.
    
    Args:
        sheet (Worksheet): The MARVEL worksheet to write to
        project_data (Dict): Project metadata
    """
    try:
        # MARVEL-specific cell mappings (F columns for project name/location/date)
        marvel_cell_mappings = {
            "project_number": "C3",  # Job No
            "company": "C5",         # Company (changed from customer)
            "estimator": "C7",       # Sales Manager / Estimator Initials
            "project_name": "F3",    # Project Name (stays in F3 for MARVEL)
            "project_location": "F5", # Project Location (stays in F5 for MARVEL)
            "date": "F7",            # Date (stays in F7 for MARVEL)
            "revision": "K7",        # Revision
        }
        
        # Write project metadata using MARVEL-specific mappings
        write_metadata_with_mappings(sheet, project_data, marvel_cell_mappings, template_version)
        
    except Exception as e:
        print(f"Warning: Could not write MARVEL metadata: {str(e)}")

def normalize_reference_number(ref_num: str) -> str:
    """
    Normalize a reference number for matching by removing variations that don't affect identity.
    
    This function now preserves uppercase letters that are part of the reference (like A, B in 30.13A, 30.13B)
    while only normalizing case differences.
    
    Args:
        ref_num (str): The reference number to normalize
        
    Returns:
        str: Normalized reference number
    """
    if not ref_num:
        return ""
    
    # Convert to string and strip whitespace
    ref_str = str(ref_num).strip()
    
    # Convert to uppercase for case-insensitive matching
    # This preserves important distinctions like 30.13A vs 30.13B
    return ref_str.upper()

def references_match(canopy_ref: str, fire_supp_ref: str) -> bool:
    """
    Check if two reference numbers should be considered a match.
    
    Exact matches are required for references like "30.13A" vs "30.13B".
    Flexible matching is allowed for minor variants like "1.01" vs "1.01a".
    
    Args:
        canopy_ref (str): Reference number from canopy sheet
        fire_supp_ref (str): Reference number from fire suppression sheet
        
    Returns:
        bool: True if the references should be considered a match
    """
    if not canopy_ref or not fire_supp_ref:
        return False
    
    # Normalize both references (now just uppercases them)
    normalized_canopy = normalize_reference_number(canopy_ref)
    normalized_fire_supp = normalize_reference_number(fire_supp_ref)
    
    # Check if normalized versions match exactly
    if normalized_canopy == normalized_fire_supp:
        # print(f"🔗 Reference match (exact): '{canopy_ref}' ↔ '{fire_supp_ref}'")
        return True
    
    # For backward compatibility, allow flexible matching only for lowercase suffix variants
    # e.g., "1.01" should match "1.01a" but "30.13A" should NOT match "30.13B"
    import re
    
    # Extract base numbers without any letters
    canopy_base = re.match(r'^([0-9\.\-]+)', normalized_canopy)
    fire_supp_base = re.match(r'^([0-9\.\-]+)', normalized_fire_supp)
    
    if canopy_base and fire_supp_base:
        canopy_base_str = canopy_base.group(1)
        fire_supp_base_str = fire_supp_base.group(1)
        
        # If the base numbers match and one has a lowercase letter suffix in the original
        if canopy_base_str == fire_supp_base_str:
            # Check if the difference is just a lowercase letter suffix
            canopy_suffix = normalized_canopy[len(canopy_base_str):]
            fire_supp_suffix = normalized_fire_supp[len(fire_supp_base_str):]
            
            # Allow match if one has no suffix and the other has a single lowercase letter
            # (when uppercased, we check if it's a single letter)
            if (not canopy_suffix and len(fire_supp_suffix) == 1 and fire_supp_suffix.isalpha()) or \
               (not fire_supp_suffix and len(canopy_suffix) == 1 and canopy_suffix.isalpha()):
                # Check original refs to see if suffix was lowercase
                orig_canopy = str(canopy_ref).strip()
                orig_fire_supp = str(fire_supp_ref).strip()
                
                # Only allow if the suffix was originally lowercase
                if (not canopy_suffix and orig_fire_supp.endswith(('a','b','c','d','e','f'))) or \
                   (not fire_supp_suffix and orig_canopy.endswith(('a','b','c','d','e','f'))):
                    # print(f"🔗 Reference match (lowercase suffix variant): '{canopy_ref}' ↔ '{fire_supp_ref}'")
                    return True
    
    return False

def write_metadata_with_mappings(sheet: Worksheet, project_data: Dict, cell_mappings: Dict, template_version: str = None):
    """
    Write project metadata to a sheet using custom cell mappings.
    
    Args:
        sheet (Worksheet): The worksheet to write to
        project_data (Dict): Project metadata
        cell_mappings (Dict): Dictionary mapping field names to cell references
        template_version (str, optional): Template version for cost sheet identifier
    """
    try:
        for field, cell in cell_mappings.items():
            value = project_data.get(field)
            try:
                if field == "revision":
                    # Use provided revision or leave blank for initial version (no revision)
                    sheet[cell] = value or ""
                elif value:
                    if field == "estimator":
                        from utils.word import get_combined_initials
                        value = get_combined_initials(project_data.get('sales_contact',''), value)
                    elif field != "date":
                        value = str(value).title()
                    elif field == "date" and not value:
                        value = get_current_date()
                    sheet[cell] = value
            except Exception as e:
                # Handle merged cells
                print(f"Warning: Could not write {field} to cell {cell}: {str(e)}")
                try:
                    # Try to unmerge the cell and write
                    if hasattr(sheet, 'merged_cells'):
                        for merged_range in list(sheet.merged_cells.ranges):
                            if cell in merged_range:
                                sheet.unmerge_cells(str(merged_range))
                                break
                    # Try writing again after unmerging
                    if field == "revision":
                        # Use provided revision or leave blank for initial version (no revision)
                        sheet[cell] = value or ""
                    elif value:
                        if field == "estimator":
                            from utils.word import get_combined_initials
                            value = get_combined_initials(project_data.get('sales_contact',''), value)
                        elif field != "date":
                            value = str(value).title()
                        elif field == "date" and not value:
                            value = get_current_date()
                        sheet[cell] = value
                except Exception as e2:
                    print(f"Warning: Still could not write {field} to cell {cell} after unmerging: {str(e2)}")
                    continue
        
        # Cost sheet identifier in N2 (same as others)
        write_cost_sheet_identifier(sheet, sheet.title, template_version)
        
    except Exception as e:
        print(f"Warning: Could not write metadata with mappings: {str(e)}")

def update_revision_with_edits(excel_path: str, edited_data: Dict, new_revision: str, new_date: str = None) -> str:
    """
    Update an existing Excel file with edited data while preserving all other content.
    This function specifically updates canopy configurations while keeping lights, formulas, etc.
    
    Args:
        excel_path (str): Path to the existing Excel file
        edited_data (Dict): Edited project data with canopy changes
        new_revision (str): New revision letter
        new_date (str, optional): New date in DD/MM/YYYY format
    
    Returns:
        str: Path to the updated Excel file
    """
    import tempfile
    from openpyxl import load_workbook
    
    # Create a temporary file for the output
    output_fd, output_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(output_fd)
    
    try:
        # Load the existing workbook
        wb = load_workbook(excel_path, data_only=False)  # Keep formulas
        
        # Update revision and date in all sheets
        sheets_to_update = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Update revision based on sheet patterns
            if 'JOB TOTAL' in sheet_name or 'CANOPY' in sheet_name:
                sheet['K7'] = new_revision
                if new_date:
                    sheet['G7'] = new_date
            elif any(x in sheet_name for x in ['FIRE SUPP', 'CONTRACT', 'SPIRAL DUCT', 'SUPPLY DUCT', 'EXTRACT DUCT']):
                sheet['K7'] = new_revision
                if new_date:
                    sheet['G7'] = new_date
            elif any(x in sheet_name for x in ['EBOX', 'EDGE BOX', 'RECOAIR']):
                sheet['K7'] = new_revision
                if new_date:
                    sheet['H7'] = new_date
            elif any(x in sheet_name for x in ['MARVEL', 'SDU']):
                sheet['K7'] = new_revision
                if new_date:
                    sheet['F7'] = new_date
            
            # Check if revision was in O7 instead
            if sheet['O7'].value and str(sheet['O7'].value).strip():
                sheet['O7'] = new_revision
        
        # Update canopy data in CANOPY sheets
        for sheet_name in wb.sheetnames:
            if 'CANOPY' in sheet_name:
                sheet = wb[sheet_name]
                
                # Extract level and area from sheet name
                level_num = None
                area_num = None
                for level in edited_data.get('levels', []):
                    level_name = level.get('level_name', '')
                    if level_name in sheet_name:
                        level_num = level.get('level_number')
                        # Find area number
                        for area_idx, area in enumerate(level.get('areas', [])):
                            if f"AREA {area_idx + 1}" in sheet_name:
                                area_num = area_idx + 1
                                
                                # Update canopies for this area
                                for canopy_idx, canopy in enumerate(area.get('canopies', [])):
                                    base_row = 14 + (canopy_idx * 17)  # Each canopy block is 17 rows
                                    
                                    # Update model (B14, B31, etc.)
                                    if 'model' in canopy:
                                        sheet[f'B{base_row}'] = canopy['model']
                                    
                                    # Update configuration (C14, C31, etc.)
                                    if 'configuration' in canopy:
                                        sheet[f'C{base_row}'] = canopy['configuration']
                                    
                                    # Update dimensions
                                    if 'length' in canopy:
                                        sheet[f'E{base_row}'] = canopy['length']
                                    if 'width' in canopy:
                                        sheet[f'G{base_row}'] = canopy['width']
                                    if 'height' in canopy:
                                        sheet[f'I{base_row}'] = canopy['height']
                                    if 'sections' in canopy:
                                        sheet[f'K{base_row}'] = canopy['sections']
                                    
                                    # Update wall cladding if edited
                                    if 'wall_cladding' in canopy:
                                        wall_cladding = canopy['wall_cladding']
                                        cladding_row = base_row + 6  # Row 20 for first canopy
                                        
                                        # Only update if wall cladding is enabled
                                        if wall_cladding.get('type') != 'None':
                                            # Update width (P20, P37, etc.)
                                            if 'width' in wall_cladding:
                                                sheet[f'P{cladding_row}'] = wall_cladding['width']
                                            
                                            # Update height (Q20, Q37, etc.)
                                            if 'height' in wall_cladding:
                                                sheet[f'Q{cladding_row}'] = wall_cladding['height']
                                            
                                            # Update position (S20, S37, etc.)
                                            if 'position' in wall_cladding:
                                                position_list = wall_cladding['position']
                                                if isinstance(position_list, list):
                                                    position_str = " and ".join(position_list)
                                                    sheet[f'S{cladding_row}'] = position_str
                                        else:
                                            # Clear wall cladding data if disabled
                                            sheet[f'P{cladding_row}'] = None
                                            sheet[f'Q{cladding_row}'] = None
                                            sheet[f'S{cladding_row}'] = None
                                break
                        break
        
        # Save the updated workbook
        wb.save(output_path)
        wb.close()
        
        return output_path
        
    except Exception as e:
        # Clean up on error
        if os.path.exists(output_path):
            os.remove(output_path)
        raise e

