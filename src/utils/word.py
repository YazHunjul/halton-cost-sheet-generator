"""
Word document generation utilities for Halton quotation system.
Handles creation of quotation documents from Excel data using Jinja templating.
"""
from typing import Dict, List, Tuple, Union
import os
import zipfile
from datetime import datetime
from docxtpl import DocxTemplate
from config.business_data import SALES_CONTACTS, ESTIMATORS
from config.constants import is_feature_enabled
import streamlit as st
from openpyxl import load_workbook # Added for contract data extraction
# Import template storage utilities
from utils.template_storage import download_template_to_local

# Template paths - will be downloaded from Supabase Storage when needed
WORD_TEMPLATE_PATH = "templates/word/Halton Quote Feb 2024.docx"
RECOAIR_TEMPLATE_PATH = "templates/word/Halton RECO Quotation Jan 2025 (2).docx"
AHU_TEMPLATE_PATH = "templates/word/Halton AHU quote JAN2020.docx"


def ensure_template_available(template_key: str, local_path: str) -> str:
    """
    Ensure template is available locally by downloading from Supabase Storage if needed.

    Args:
        template_key: Template key (e.g., 'canopy_quotation')
        local_path: Local filepath where template should be

    Returns:
        Local filepath of template

    Raises:
        Exception if template cannot be retrieved
    """
    import os

    # Check if template exists locally
    if os.path.exists(local_path):
        return local_path

    # Template doesn't exist locally, download from Supabase Storage
    print(f"Template not found locally at {local_path}, downloading from Supabase Storage...")
    success, downloaded_path = download_template_to_local(template_key)

    if not success:
        raise Exception(f"Failed to download template from storage: {downloaded_path}")

    print(f"✓ Template downloaded successfully to {downloaded_path}")
    return downloaded_path



def get_fire_suppression_system_description(system_type: str) -> str:
    """
    Determine the fire suppression system description based on the system type from C16.
    
    Args:
        system_type (str): The system type value from C16 (NOBEL, AMAREX, or other)
        
    Returns:
        str: The appropriate system description
    """
    if not system_type:
        return 'Ansul R102 system. Supplied, installed & commissioned.'
    
    system_type_upper = str(system_type).upper().strip()
    
    if 'NOBEL' in system_type_upper:
        return 'NOBEL System. Supplied, installed & commissioned.'
    elif 'AMAREX' in system_type_upper:
        return 'AMAREX System. Supplied, installed & commissioned.'
    else:
        # For "FIRE SUPPRESSION", "1 TANK SYSTEM", or any other value, default to Ansul R102
        return 'Ansul R102 system. Supplied, installed & commissioned.'

def get_sales_contact_info(estimator_name: str, project_data: Dict = None) -> Dict[str, str]:
    """
    Get sales contact information based on project data or estimator name.
    
    Args:
        estimator_name (str): Name of the estimator (kept for backward compatibility)
        project_data (Dict, optional): Project data containing sales_contact
        
    Returns:
        Dict: Contact information including name and phone
    """
    # First, try to use sales_contact from project_data if available
    if project_data and project_data.get('sales_contact'):
        sales_contact_name = project_data['sales_contact']
        if sales_contact_name in SALES_CONTACTS:
            return {
                'name': sales_contact_name,
                'phone': SALES_CONTACTS[sales_contact_name]
            }
    
    # Fallback: try to match estimator to sales contact (old logic, likely won't match)
    for contact_name, phone in SALES_CONTACTS.items():
        if estimator_name and any(name.lower() in estimator_name.lower() for name in contact_name.split()):
            return {
                'name': contact_name,
                'phone': phone
            }
    
    # Default to first sales contact if no match
    first_contact = list(SALES_CONTACTS.items())[0]
    return {
        'name': first_contact[0],
        'phone': first_contact[1]
    }

def format_halton_reference(project_number: str, date: str, revision: str = None) -> str:
    """
    Format the Halton reference number.
    
    Args:
        project_number (str): Project number
        date (str): Project date
        revision (str, optional): Revision letter
        
    Returns:
        str: Formatted Halton reference
    """
    try:
        if isinstance(date, str) and '/' in date:
            # Extract year from date (assume format DD/MM/YYYY)
            year = date.split('/')[-1][-2:]  # Get last 2 digits of year
        else:
            year = str(datetime.now().year)[-2:]
        
        # Format as project_number/month/year/revision
        month = datetime.now().strftime("%m")
        reference = f"{project_number}/{month}/{year}"
        
        # Add revision if present
        if revision and revision.strip():
            reference = f"{reference}/{revision}"
            
        return reference
    except:
        return f"{project_number}/XX/XX"

def format_date_for_display(date_str: str) -> str:
    """
    Format date for display in the document.
    
    Args:
        date_str (str): Date string from Excel
        
    Returns:
        str: Formatted date string
    """
    try:
        if isinstance(date_str, str) and '/' in date_str:
            # Convert DD/MM/YYYY to DD Month YYYY
            parts = date_str.split('/')
            if len(parts) == 3:
                day, month, year = parts
                month_names = [
                    'January', 'February', 'March', 'April', 'May', 'June',
                    'July', 'August', 'September', 'October', 'November', 'December'
                ]
                month_name = month_names[int(month) - 1] if 1 <= int(month) <= 12 else month
                return f"{day} {month_name} {year}"
        return date_str or datetime.now().strftime('%d %B %Y')
    except:
        return datetime.now().strftime('%d %B %Y')

def transform_lighting_type(lighting_type: str) -> str:
    """
    Transform lighting type for Word document display.
    
    Args:
        lighting_type (str): Raw lighting type from Excel
        
    Returns:
        str: Transformed lighting type
    """
    # Handle None, empty string, or whitespace-only strings
    if not lighting_type or str(lighting_type).strip() == "":
        return "-"
    
    lighting_str = str(lighting_type).upper().strip()
    
    # Handle "LIGHT SELECTION" as empty value
    if lighting_str == "LIGHT SELECTION":
        return "-"
    
    # Check for LED strip variations (any string containing "LED STRIP")
    if "LED STRIP" in lighting_str:
        # Always return just "LED STRIP" regardless of variations (L6, L12, L18, DALI, EM, etc.)
        return "LED STRIP"
    
    # Check for spots variations
    elif "SPOTS" in lighting_str or "SPOT" in lighting_str:
        # Always return just "LED SPOTS" regardless of variations (Small, Large, DALI, etc.)
        return "LED SPOTS"
    
    # Check for HCL variations
    elif lighting_str.startswith("HCL"):
        # Always return "HCL DALI" regardless of the number (600, 1200, 1800, etc.)
        return "HCL DALI"
    
    # Check for EL variations
    elif lighting_str.startswith("EL"):
        # Always return "EL" regardless of the number (215, 218, etc.)
        return "EL"
    
    # For any other value or unrecognized lighting type, return "-"
    return "-"

def handle_empty_value(value) -> str:
    """
    Convert empty values to '-' for Word document display.
    
    Args:
        value: Any value that might be empty
        
    Returns:
        str: The value or '-' if empty
    """
    # Handle None values
    if value is None:
        return "-"
    
    # Convert to string and check for empty or whitespace-only strings
    str_value = str(value).strip()
    if str_value == "" or str_value.lower() == "none":
        return "-"
    
    return str_value

def format_extract_static(value) -> str:
    """
    Format extract static value by removing 'Pa' and returning clean number.
    
    Args:
        value: Extract static value that might contain 'Pa'
        
    Returns:
        str: Formatted extract static value without 'Pa'
    """
    if not value:
        return "-"
    
    # Convert to string and clean up
    str_value = str(value).strip()
    
    # Handle empty or dash values
    if str_value == "" or str_value == "-":
        return "-"
    
    # Remove 'Pa' (case insensitive) and any extra whitespace
    cleaned_value = str_value.replace("Pa", "").replace("pa", "").replace("PA", "").strip()
    
    # If nothing left after removing Pa, return dash
    if not cleaned_value:
        return "-"
    
    return cleaned_value

def format_mua_volume(value) -> str:
    """
    Format MUA volume rounded to 2 decimal places.
    
    Args:
        value: MUA volume value to format
        
    Returns:
        str: MUA volume rounded to 2 decimal places or "-" if empty
    """
    if not value:
        return "-"
    
    # Convert to string and clean up
    str_value = str(value).strip()
    
    # Handle empty or dash values
    if str_value == "" or str_value == "-":
        return "-"
    
    # Try to convert to float and round to 2 decimal places
    try:
        float_value = float(str_value)
        # Round to 2 decimal places and format
        rounded_value = round(float_value, 2)
        # Format to avoid unnecessary trailing zeros (e.g., 1.50 becomes 1.5, 1.00 becomes 1)
        if rounded_value == int(rounded_value):
            return str(int(rounded_value))
        else:
            return f"{rounded_value:.2f}".rstrip('0').rstrip('.')
    except (ValueError, TypeError):
        # If conversion fails, return the original value
        return str_value

def get_combined_initials(sales_contact_name: str, estimator_name: str) -> str:
    """
    Generate combined initials from sales contact and estimator names.
    Format: Sales Contact Initials / Estimator Initials
    
    Args:
        sales_contact_name (str): Full name of sales contact
        estimator_name (str): Full name of estimator
        
    Returns:
        str: Combined initials (e.g., "YH/JS" for "Yazan Hunjul" + "Joe Salloum")
    """
    from utils.excel import get_initials
    
    sales_initials = get_initials(sales_contact_name) if sales_contact_name else ""
    estimator_initials = get_initials(estimator_name) if estimator_name else ""
    
    # Combine with '/' separator
    if sales_initials and estimator_initials:
        return f"{sales_initials}/{estimator_initials}"
    elif sales_initials:
        return sales_initials
    elif estimator_initials:
        return estimator_initials
    else:
        return ""

def get_customer_first_name(customer_name: str) -> str:
    """
    Extract the first name from a customer name.
    
    Args:
        customer_name (str): Full customer name
        
    Returns:
        str: First name of the customer
    """
    if not customer_name:
        return ""
    
    # Split by spaces and take the first part
    name_parts = customer_name.strip().split()
    if name_parts:
        return name_parts[0]
    else:
        return ""

def generate_reference_variable(project_number: str, sales_contact_name: str, estimator_name: str, revision: str = None) -> str:
    """
    Generate reference variable in format: projectnumber/salesinitials/estimatorintials/revision
    
    Args:
        project_number (str): Project number
        sales_contact_name (str): Full name of sales contact
        estimator_name (str): Full name of estimator
        revision (str, optional): Revision letter
        
    Returns:
        str: Reference variable (e.g., "P12345/YH/JS" or "P12345/YH/JS/A")
    """
    from utils.excel import get_initials
    
    sales_initials = get_initials(sales_contact_name) if sales_contact_name else ""
    estimator_initials = get_initials(estimator_name) if estimator_name else ""
    
    # Format: projectnumber/salesinitials/estimatorintials/revision
    reference = f"{project_number}/{sales_initials}/{estimator_initials}"
    
    # Add revision if present
    if revision and revision.strip():
        reference = f"{reference}/{revision}"
        
    return reference

def generate_quote_title(revision: str) -> str:
    """
    Generate quote title based on revision.
    
    Args:
        revision (str): Project revision (A, B, C, etc.) or empty for initial version
        
    Returns:
        str: Quote title ("QUOTATION" for no revision or blank, "QUOTATION - REVISION A" for A and beyond)
    """
    # If no revision or empty/blank revision, just return "QUOTATION"
    if not revision or str(revision).strip() == "":
        return "QUOTATION"
    
    revision = str(revision).strip().upper()
    
    # For blank/empty revision (initial version), just return "QUOTATION"
    if revision == "":
        return "QUOTATION"
    
    # For any actual revision letter (A, B, C, etc.), include it in the title
    return f"QUOTATION - REVISION {revision}"

def normalize_area_object(area: Dict) -> Dict:
    """
    Ensure area object has all required attributes for template compatibility.
    
    Args:
        area (Dict): Area object to normalize
        
    Returns:
        Dict: Normalized area object with all required attributes
    """
    normalized = area.copy()
    
    # Ensure options exists and is a dictionary
    if 'options' not in normalized:
        normalized['options'] = {}
    elif normalized['options'] is None:
        normalized['options'] = {}
    
    # Ensure all expected option keys exist
    if not isinstance(normalized['options'], dict):
        normalized['options'] = {}
    
    option_defaults = {'uvc': False, 'sdu': False, 'recoair': False, 'vent_clg': False}
    for key, default_value in option_defaults.items():
        if key not in normalized['options']:
            normalized['options'][key] = default_value
    
    return normalized

def prepare_template_context(project_data: Dict, excel_file_path: str = None) -> Dict:
    """
    Prepare the template context for Word document generation.
    
    Args:
        project_data (Dict): Project data
        excel_file_path (str, optional): Path to Excel file to extract detailed data
        
    Returns:
        Dict: Template context with all required data
    """
    import time
    start_time = time.time()
    print(f"🚀 Starting template context preparation...")
    
    # Initialize lists for tracking items across all areas
    all_canopies = []
    wall_cladding_items = []
    fire_suppression_items = []
    enhanced_levels = []
    
    # Initialize contract data
    contract_data = {
        'has_extract_system': False,
        'has_supply_system': False,
        'extract_system_price': 0,
        'supply_system_price': 0,
        'extract_system_total': 0,  # New: total including all costs
        'supply_system_total': 0,   # New: total including all costs
        'contract_total_price': 0,
        'delivery_location': '',  # D57 in Contract
        'plant_selection': '',    # D58 in Contract
        'extract_installation_price': 0,
        'supply_installation_price': 0,
        'extract_delivery_price': 0,
        'supply_delivery_price': 0,
        'extract_commissioning_price': 0,
        'supply_commissioning_price': 0,
        'total_installation_price': 0,  # C72
        'total_commissioning_price': 0, # J66
        'total_delivery_base': 0,       # J57
    }
    
    init_time = time.time()
    print(f"   ✅ Initialization complete: {init_time - start_time:.3f}s")
    
    # Get sales contact information
    sales_contact = get_sales_contact_info(
        project_data.get('estimator', ''),  # Pass estimator name for backward compatibility
        project_data  # Pass full project data to use sales_contact if available
    )
    
    # Get estimator information
    estimator = project_data.get('estimator', '')
    estimator_rank = project_data.get('estimator_rank', 'Estimator')
    
    contact_time = time.time()
    print(f"   ✅ Sales contact info retrieved: {contact_time - init_time:.3f}s")
    
    # Load Excel workbook once and reuse it for all operations
    cached_wb = None
    wb_load_time = 0
    contract_start = time.time()  # Initialize contract_start for all paths
    
    if excel_file_path and os.path.exists(excel_file_path):
        print(f"🔍 Starting contract sheet processing...")
        print(f"   📖 Loading Excel workbook: {excel_file_path}")
        wb_load_start = time.time()
        
        try:
            cached_wb = load_workbook(excel_file_path, data_only=True)  # Use data_only=True to read values
            wb_load_time = time.time() - wb_load_start
            print(f"   ✅ Workbook loaded: {wb_load_time:.3f}s")
            
            # Look for CONTRACT sheet (exact match or numbered variant)
            sheet_search_start = time.time()
            contract_sheet = None
            print(f"   🔍 Searching for CONTRACT sheet in {len(cached_wb.sheetnames)} sheets...")
            
            for sheet_name in cached_wb.sheetnames:
                if (sheet_name == 'CONTRACT' or 
                    (sheet_name.startswith('CONTRACT') and 
                     (sheet_name.replace('CONTRACT', '').strip() == '' or 
                      sheet_name.replace('CONTRACT', '').strip().isdigit()))):
                    contract_sheet = cached_wb[sheet_name]
                    print(f"   ✅ Found contract sheet: {sheet_name}")
                    break
            
            sheet_search_time = time.time() - sheet_search_start
            print(f"   ✅ Sheet search complete: {sheet_search_time:.3f}s")
            
            # Process contract data if a contract sheet was found
            if contract_sheet:
                print(f"   🔨 Processing contract data...")
                contract_process_start = time.time()
                
                # Get installation, delivery, and commissioning prices first
                print(f"      📊 Reading pricing data from cells...")
                c72_value = float(contract_sheet['C72'].value or 0)
                c57_value = float(contract_sheet['J57'].value or 0)
                total_commissioning = float(contract_sheet['J66'].value or 0)
                total_delivery_base = float(contract_sheet['J57'].value or 0)  # Changed from J56 to J57
                
                # Calculate installation price as C72 - C57
                total_installation = c72_value - c57_value
                
                # Delivery price should be C57
                total_delivery = c57_value
                
                contract_data['total_installation_price'] = total_installation
                contract_data['total_commissioning_price'] = total_commissioning
                contract_data['total_delivery_base'] = total_delivery_base
                
                print(f"      💰 Processing extract/supply system pricing...")
                # Check M12 for extract system price
                extract_total = contract_sheet['M12'].value
                if extract_total and isinstance(extract_total, (int, float)) and extract_total > 0:
                    contract_data['has_extract_system'] = True
                    # Store the base price from M12
                    contract_data['extract_system_price'] = float(extract_total)
                
                # Check N12 for supply system price
                supply_total = contract_sheet['N12'].value
                if supply_total and isinstance(supply_total, (int, float)) and supply_total > 0:
                    contract_data['has_supply_system'] = True
                    # Store the base price from N12
                    contract_data['supply_system_price'] = float(supply_total)
                
                print(f"      🧮 Calculating cost splits...")
                # Split costs between extract and supply if both exist
                if contract_data['has_extract_system'] and contract_data['has_supply_system']:
                    # Equal distribution of delivery, installation, and commissioning (divided by 2)
                    
                    # Split installation equally (divide by 2)
                    contract_data['extract_installation_price'] = total_installation / 2
                    contract_data['supply_installation_price'] = total_installation / 2
                    
                    # Split delivery equally (divide by 2)
                    contract_data['extract_delivery_price'] = total_delivery / 2
                    contract_data['supply_delivery_price'] = total_delivery / 2
                    
                    # Split commissioning equally (divide by 2)
                    contract_data['extract_commissioning_price'] = total_commissioning / 2
                    contract_data['supply_commissioning_price'] = total_commissioning / 2
                    
                    # Calculate total including all costs
                    contract_data['extract_system_total'] = (
                        contract_data['extract_system_price'] +
                        contract_data['extract_installation_price'] +
                        contract_data['extract_delivery_price'] +
                        contract_data['extract_commissioning_price']
                    )
                    contract_data['supply_system_total'] = (
                        contract_data['supply_system_price'] +
                        contract_data['supply_installation_price'] +
                        contract_data['supply_delivery_price'] +
                        contract_data['supply_commissioning_price']
                    )
                elif contract_data['has_extract_system']:
                    # All costs go to extract system
                    contract_data['extract_installation_price'] = total_installation
                    contract_data['extract_delivery_price'] = total_delivery
                    contract_data['extract_commissioning_price'] = total_commissioning
                    # Calculate total including all costs
                    contract_data['extract_system_total'] = (
                        contract_data['extract_system_price'] +
                        contract_data['extract_installation_price'] +
                        contract_data['extract_delivery_price'] +
                        contract_data['extract_commissioning_price']
                    )
                elif contract_data['has_supply_system']:
                    # All costs go to supply system
                    contract_data['supply_installation_price'] = total_installation
                    contract_data['supply_delivery_price'] = total_delivery
                    contract_data['supply_commissioning_price'] = total_commissioning
                    # Calculate total including all costs
                    contract_data['supply_system_total'] = (
                        contract_data['supply_system_price'] +
                        contract_data['supply_installation_price'] +
                        contract_data['supply_delivery_price'] +
                        contract_data['supply_commissioning_price']
                    )
                
                print(f"      📋 Reading additional contract data...")
                # Get contract total from J9
                contract_total = contract_sheet['J9'].value
                if contract_total and isinstance(contract_total, (int, float)) and contract_total > 0:
                    contract_data['contract_total_price'] = float(contract_total)
                    print(f"Found contract total price: {contract_data['contract_total_price']}")
                
                # Get delivery location and plant selection
                delivery_location = contract_sheet['D57'].value
                plant_selection = contract_sheet['D58'].value
                contract_data['delivery_location'] = str(delivery_location) if delivery_location else ''
                contract_data['plant_selection'] = str(plant_selection) if plant_selection else ''
                
                contract_process_time = time.time() - contract_process_start
                print(f"   ✅ Contract processing complete: {contract_process_time:.3f}s")
            else:
                print(f"   ℹ️  No CONTRACT sheet found")
                
        except Exception as e:
            print(f"   ❌ Contract processing error: {str(e)}")
            cached_wb = None
        
        contract_time = time.time()
        total_contract_time = contract_time - contract_start
        print(f"🔍 Contract processing complete: {total_contract_time:.3f}s")
    else:
        print(f"   ℹ️  No Excel file provided for contract processing")
        contract_time = time.time()
        total_contract_time = contract_time - contract_start
    
    # Check if fire suppression sheets exist by looking for any areas with fire suppression data
    fire_supp_start = time.time()
    print(f"🔥 Processing fire suppression data...")
    has_fire_suppression_sheets = False
    
    levels_start = time.time()
    print(f"🏢 Processing levels and areas data...")
    
    for level in project_data.get('levels', []):
        level_name = level.get('level_name', '')
        enhanced_areas = []
        
        for area in level.get('areas', []):
            # Normalize area object to ensure it has proper structure
            area = normalize_area_object(area)
            area_name = area.get('name', '')
            level_area_combined = f"{level_name} - {area_name}"
            
            # Check if any canopy in this area actually has fire suppression (tank quantity > 0 or price > 0)
            area_has_fire_suppression = any(
                (canopy.get('fire_suppression_tank_quantity', 0) > 0 or 
                 canopy.get('fire_suppression_price', 0) > 0)
                for canopy in area.get('canopies', [])
            )
            
            if area_has_fire_suppression:
                has_fire_suppression_sheets = True
            
            # Process canopies for this area and create transformed versions
            transformed_canopies = []
            
            # Add to all canopies with enhanced info
            for canopy in area.get('canopies', []):
                # Get model for business logic
                model = canopy.get('model', '').upper()
                
                # Apply business rules for volume and static data
                # CXW models are extract-only, so they never have supply static or MUA volume
                if model == "CXW":
                    mua_volume = "-"
                    supply_static = "-"
                # If canopy doesn't have 'F' in its name, set MUA volume and supply static to '-'
                elif 'F' not in model:
                    mua_volume = "-"
                    supply_static = "-"
                else:
                    mua_volume = format_mua_volume(canopy.get('mua_volume', ''))
                    
                    # Handle supply static for F models (excluding CXW)
                    raw_supply_static = canopy.get('supply_static', '')
                    
                    # Check if we should use the existing value or apply default
                    should_use_default = False
                    
                    if not raw_supply_static:
                        should_use_default = True
                    elif str(raw_supply_static).strip() == "":
                        should_use_default = True
                    else:
                        # Check if it's a very small value that should be treated as empty
                        try:
                            numeric_value = float(str(raw_supply_static).strip())
                            # Treat values less than 1 as empty (likely Excel calculation artifacts)
                            if numeric_value < 1:
                                should_use_default = True
                        except (ValueError, TypeError):
                            # If it's not a number, use the existing value
                            pass
                    
                    if should_use_default:
                        # Apply default value for F models
                        supply_static = "45"
                    else:
                        # Use existing value from Excel
                        supply_static = format_extract_static(raw_supply_static)
                
                # Extract static logic based on model type
                if model in ['CMWF', 'CMWI']:
                    # CMWF/CMWI models always show '-'
                    extract_static = "-"
                elif model == "CXW":
                    # CXW models always show '45'
                    extract_static = "45"
                else:
                    # KV, UV, and all other models get value from Excel
                    raw_extract_static = canopy.get('extract_static', '')
                    extract_static = format_extract_static(raw_extract_static)
                
                # Check for wall cladding on this canopy
                wall_cladding = canopy.get('wall_cladding', {})
                
                # Initialize position_str for use later (even if no cladding)
                position = wall_cladding.get('position', [])
                if isinstance(position, list):
                    position_str = ', '.join(position) if position else 'No positions'
                else:
                    position_str = str(position) if position else 'No positions'
                
                if wall_cladding.get('type') != 'None' and wall_cladding.get('type'):
                    cladding_item = {
                        'area': level_area_combined,
                        'canopy_ref': canopy.get('reference_number', 'N/A'),
                        'item_number': canopy.get('reference_number', 'N/A'),
                        'type': wall_cladding.get('type', 'Unknown'),
                        'position': position_str,
                        'height': wall_cladding.get('height', 'N/A'),
                        'length': wall_cladding.get('length', 'N/A'),
                        'width': wall_cladding.get('width', wall_cladding.get('length', 'N/A')),  # Use width or fallback to length
                        'price': wall_cladding.get('price', 0),
                        'description': f"Cladding below Item {canopy.get('reference_number', 'N/A')}, supplied and installed"
                    }
                    wall_cladding_items.append(cladding_item)
                
                # Check for fire suppression on this canopy
                fire_suppression_tank_quantity = canopy.get('fire_suppression_tank_quantity', 0)
                fire_suppression_price = canopy.get('fire_suppression_price', 0)
                # Display fire suppression if either tank quantity > 0 OR price > 0
                if fire_suppression_tank_quantity > 0 or fire_suppression_price > 0:
                    # Generate fire suppression system description
                    system_type = canopy.get('fire_suppression_system_type', '')
                    fs_system_desc = get_fire_suppression_system_description(system_type)
                    
                    fire_suppression_item = {
                        'area': level_area_combined,
                        'canopy_ref': canopy.get('reference_number', 'N/A'),
                        'item_number': canopy.get('reference_number', 'N/A'),
                        'tank_quantity': fire_suppression_tank_quantity,
                        'system_description': fs_system_desc,
                        'price': canopy.get('fire_suppression_price', 0),
                        'level_area_combined': level_area_combined
                    }
                    fire_suppression_items.append(fire_suppression_item)
                else:
                    fs_system_desc = ""
                
                # Sections formatting - handle both numeric and text values
                raw_sections = canopy.get('sections', '')
                if raw_sections:
                    try:
                        # Try to convert to int if it's a number
                        sections_num = int(float(raw_sections))
                        display_sections = str(sections_num)
                    except (ValueError, TypeError):
                        # If it's not a number, use as string
                        display_sections = str(raw_sections)
                else:
                    display_sections = ""
                
                # Determine if this canopy has cladding
                has_cladding = (wall_cladding.get('type') not in ['None', None, ''] and 
                               wall_cladding.get('type') and 
                               (canopy.get('cladding_price', 0) > 0 or wall_cladding.get('price', 0) > 0))
                
                transformed_canopy = {
                    'reference_number': handle_empty_value(canopy.get('reference_number', '')),
                    'model': handle_empty_value(canopy.get('model', '')),
                    'configuration': handle_empty_value(canopy.get('configuration', '')),
                    'length': handle_empty_value(canopy.get('length', '')),
                    'width': handle_empty_value(canopy.get('width', '')),
                    'height': handle_empty_value(canopy.get('height', '')),
                    'sections': display_sections,  # Use the processed sections value
                    'sections_raw': handle_empty_value(raw_sections),  # Keep original value for reference
                    'lighting_type': transform_lighting_type(canopy.get('lighting_type', '')),
                    'extract_volume': format_mua_volume(canopy.get('extract_volume', '')),
                    'extract_static': extract_static,
                    'mua_volume': mua_volume, 
                    'supply_static': supply_static,
                    
                    # CWS/HWS data for wash canopies
                    'cws_capacity': handle_empty_value(canopy.get('cws_capacity', '')),
                    'hws_requirement': handle_empty_value(canopy.get('hws_requirement', '')),
                    'hw_storage': handle_empty_value(canopy.get('hw_storage', '')),
                    'has_wash_capabilities': canopy.get('has_wash_capabilities', False),
                    
                    # Fire suppression data
                    'fire_suppression_tank_quantity': canopy.get('fire_suppression_tank_quantity', 0),
                    'fire_suppression_system_type': handle_empty_value(canopy.get('fire_suppression_system_type', '')),
                    'fire_suppression_system_description': fs_system_desc,
                    
                    # Pricing data
                    'canopy_price': canopy.get('canopy_price', 0),
                    'fire_suppression_price': canopy.get('fire_suppression_price', 0),
                    'cladding_price': canopy.get('cladding_price', 0) or wall_cladding.get('price', 0),
                    
                    # Wall cladding data and flags
                    'has_cladding': has_cladding,  # Template needs this to show cladding items
                    'wall_cladding_type': wall_cladding.get('type', 'None'),
                    'wall_cladding_position': position_str if wall_cladding.get('type') != 'None' else 'None',
                    'wall_cladding_price': wall_cladding.get('price', 0)
                }
                
                transformed_canopies.append(transformed_canopy)
                all_canopies.append(transformed_canopy)
            
            # Calculate area-level totals for template
            area_canopy_total = sum(canopy.get('canopy_price', 0) for canopy in transformed_canopies)
            area_fire_suppression_total = sum(canopy.get('fire_suppression_price', 0) for canopy in transformed_canopies)
            area_cladding_total = sum(canopy.get('cladding_price', 0) for canopy in transformed_canopies)
            
            # Get area-level pricing first
            area_delivery_installation = area.get('delivery_installation_price', 0)
            area_commissioning = area.get('commissioning_price', 0)
            
            # Calculate area subtotals and totals
            # CANOPY SCHEDULE subtotal should ONLY include canopy prices + delivery + commissioning
            # Fire suppression and cladding are separate schedules with their own subtotals
            area_canopy_schedule_subtotal = area_canopy_total + area_delivery_installation + area_commissioning
            area_uvc_price = area.get('uvc_price', 0)
            # Calculate total SDU price from all canopies in this area
            area_sdu_price = sum(canopy.get('sdu_price', 0) for canopy in area.get('canopies', []))
            area_recoair_price = area.get('recoair_price', 0)
            area_vent_clg_price = area.get('vent_clg_price', 0)
            area_marvel_price = area.get('marvel_price', 0)
            area_uv_extra_over_cost = area.get('uv_extra_over_cost', 0)
            
            # Area total includes: canopy schedule + fire suppression + cladding + other systems
            # Note: area_delivery_installation and area_commissioning are already included in area_canopy_schedule_subtotal
            # Note: RecoAir pricing should NOT be included in area totals - it has its own separate pricing schedule
            # Note: UV Extra Over cost should NOT be included in area totals - it's a comparison/information only
            area_total = (area_canopy_schedule_subtotal + area_fire_suppression_total + area_cladding_total + 
                         area_uvc_price + area_sdu_price + area_vent_clg_price + area_marvel_price)  # Removed uv_extra_over_cost

            # Enhanced area with transformed canopy data
            enhanced_area = {
                'name': area_name,
                'area_name': area_name,  # Add area_name key for consistency
                'level_area_name': level_area_combined,
                'level_area_combined': level_area_combined,  # Add this for template compatibility
                'canopies': transformed_canopies,  # Use transformed data instead of raw data
                
                # Area-level options
                'options': area.get('options', {}),
                'has_canopies': len(transformed_canopies) > 0,
                
                # Area-level pricing data
                'delivery_installation_price': area_delivery_installation,
                'commissioning_price': area_commissioning,
                
                # Area-level option pricing
                'uvc_price': area_uvc_price,
                'sdu_price': area_sdu_price,
                'recoair_price': area_recoair_price,
                'vent_clg_price': area_vent_clg_price,
                'vent_clg_detailed_pricing': area.get('vent_clg_detailed_pricing', {}),
                
                # RecoAir units data (detailed unit specifications)
                'recoair_units': area.get('recoair_units', []),
                
                # UV Extra Over data
                'has_uv_extra_over': area.get('options', {}).get('uv_extra_over', False),
                'uv_extra_over_cost': area_uv_extra_over_cost,
                
                # Calculated totals for template
                'canopy_total': area_canopy_total,
                'fire_suppression_total': area_fire_suppression_total,
                'cladding_total': area_cladding_total,
                'canopy_schedule_subtotal': area_canopy_schedule_subtotal,
                'area_total': area_total,
                
                # SDU pricing data (if available)
                'sdu_pricing': area.get('sdu_pricing', {}),
                # Check if any canopy in this area has SDU
                'has_sdu': any(canopy.get('options', {}).get('sdu', False) for canopy in area.get('canopies', [])),
                
                # VENT CLG data (if available)
                'has_vent_clg': area.get('options', {}).get('vent_clg', False)
            }
            enhanced_areas.append(enhanced_area)
        
        # Enhanced level with combined names in areas
        enhanced_level = {
            'level_name': level_name,
            'areas': enhanced_areas
        }
        enhanced_levels.append(enhanced_level)
    
    # Generate combined initials (Sales Contact / Estimator)
    combined_initials = get_combined_initials(sales_contact['name'], estimator)
    
    # Generate reference variable (projectnumber/salesinitials/estimatorintials/revision)
    reference_variable = generate_reference_variable(
        project_data.get('project_number', ''), 
        sales_contact['name'], 
        estimator,
        project_data.get('revision', '')
    )
    
    # Generate quote title based on revision
    quote_title = generate_quote_title(project_data.get('revision', ''))
    
    # Extract customer first name
    customer_first_name = get_customer_first_name(project_data.get('customer', ''))
    
    levels_time = time.time()
    total_levels_time = levels_time - levels_start
    print(f"🏢 Levels processing complete: {total_levels_time:.3f}s")
    
    # Collect RecoAir pricing data (areas and job totals)
    recoair_start = time.time()
    print(f"🌀 Collecting RecoAir pricing data...")
    recoair_pricing_data = collect_recoair_pricing_schedule_data(project_data)
    recoair_time = time.time() - recoair_start
    print(f"🌀 RecoAir data collection complete: {recoair_time:.3f}s")
    
    # Collect SDU data for areas with SDU systems
    sdu_start = time.time()
    print(f"📡 Collecting SDU data...")
    sdu_data = collect_sdu_data(project_data, excel_file_path, cached_wb)  # Pass cached workbook
    sdu_time = time.time() - sdu_start
    print(f"📡 SDU data collection complete: {sdu_time:.3f}s")
    
    # Analyze project for global flags
    analysis_start = time.time()
    print(f"🔬 Analyzing project areas...")
    has_canopies, has_recoair, is_recoair_only, has_uv, has_marvel, has_vent_clg, has_pollustop, has_aerolys, has_xeu = analyze_project_areas(project_data)
    analysis_time = time.time() - analysis_start
    print(f"🔬 Project analysis complete: {analysis_time:.3f}s")
    
    # Calculate pricing totals once
    pricing_start = time.time()
    print(f"💰 Calculating pricing totals...")
    pricing_totals = calculate_pricing_totals(project_data, excel_file_path, cached_wb)  # Pass cached workbook
    pricing_time = time.time() - pricing_start
    print(f"💰 Pricing calculations complete: {pricing_time:.3f}s")
    
    # Prepare the context
    context_start = time.time()
    print(f"📋 Preparing final template context...")
    context = {
        # Basic project information
        'client_name': project_data.get('customer', ''),  # Don't use handle_empty_value for customer name
        'customer_first_name': customer_first_name,  # Don't use handle_empty_value for customer first name
        'company': handle_empty_value(project_data.get('company', '')),
        'address': handle_empty_value(project_data.get('address', '')),
        'project_name': handle_empty_value(project_data.get('project_name', '')),
        'location': handle_empty_value(project_data.get('project_location') or project_data.get('location', '')),
        'project_number': handle_empty_value(project_data.get('project_number', '')),
        'estimator': estimator,  # Full name
        'estimator_rank': estimator_rank,  # Lead Estimator, Estimator, etc.
        'estimator_initials': combined_initials,  # Combined Sales Contact / Estimator initials
        'reference_variable': reference_variable,  # Project reference (projectnumber/salesinitials/estimatorintials)
        'quote_title': quote_title,  # Quote title based on revision (QUOTATION or QUOTATION - REVISION X)
        'revision': project_data.get('revision', ''),  # Project revision - keep blank for initial version
        
        # Formatted data
        'date': format_date_for_display(project_data.get('date', '')),
        'halton_ref': format_halton_reference(project_data.get('project_number', ''), project_data.get('date', ''), project_data.get('revision', '')),
        
        # Sales contact information
        'sales_contact_name': sales_contact['name'],
        'sales_contact_phone': sales_contact['phone'],
        
        # Additional business data
        'delivery_location': handle_empty_value(project_data.get('delivery_location', '')),
        
        # Project structure (enhanced with combined names)
        'levels': enhanced_levels,
        'canopies': all_canopies,
        'total_canopies': len(all_canopies),
        
        # Derived information
        'dear_line': f"{project_data.get('customer', '')}," if project_data.get('customer') else "Sir/Madam,",
        'subject_line': f"{project_data.get('project_name', '')}, {project_data.get('project_location') or project_data.get('location', '')}",
        
        # Estimator with rank for signatures
        'estimator_with_rank': f"{estimator}\n{estimator_rank}" if estimator and estimator_rank else estimator,
        
        # Current date for any additional formatting needs
        'current_date': datetime.now().strftime('%d %B %Y'),
        'current_year': datetime.now().year,
        
        # Wall cladding data
        'wall_cladding_items': wall_cladding_items,
        'cladding_items': wall_cladding_items,  # Alternative name for template compatibility
        
        # Fire suppression data
        'fire_suppression_items': fire_suppression_items,
        'fs_items': fire_suppression_items,  # Alternative name for template compatibility
        
        # Scope of works data
        'scope_of_works': generate_scope_of_works(project_data),
        
        # Pricing data
        'pricing_totals': pricing_totals,
        'recoair_pricing_schedules': recoair_pricing_data['areas'],  # RecoAir area-by-area pricing schedules
        'recoair_job_totals': recoair_pricing_data['job_totals'],  # RecoAir job totals
        'format_currency': format_currency,  # Make currency formatter available in templates
        'format_current': format_currency,  # Alias for format_currency (for template compatibility)
        'currency_format': format_currency,  # Additional alias for template compatibility
        
        # Contract system data
        'has_contract_system': contract_data['has_extract_system'] or contract_data['has_supply_system'],
        'has_extract_system': contract_data['has_extract_system'],
        'has_supply_system': contract_data['has_supply_system'],
        'extract_system_price': contract_data['extract_system_price'],  # System price excluding costs
        'supply_system_price': contract_data['supply_system_price'],    # System price excluding costs
        'extract_system_total': contract_data['extract_system_total'],  # Total including all costs
        'supply_system_total': contract_data['supply_system_total'],    # Total including all costs
        'suply_system_total': contract_data['supply_system_total'],     # Alternative spelling for template compatibility
        'contract_total_price': contract_data['contract_total_price'],
        
        # New contract data fields
        'contract_delivery_location': contract_data['delivery_location'],
        'contract_plant_selection': contract_data['plant_selection'],
        'extract_installation_price': contract_data['extract_installation_price'],
        'supply_installation_price': contract_data['supply_installation_price'],
        'extract_delivery_price': contract_data['extract_delivery_price'],
        'supply_delivery_price': contract_data['supply_delivery_price'],
        'extract_commissioning_price': contract_data['extract_commissioning_price'],
        'extract_comissioning_price': contract_data['extract_commissioning_price'],  # Alternative spelling for template compatibility
        'supply_commissioning_price': contract_data['supply_commissioning_price'],
        'total_installation_price': contract_data['total_installation_price'],
        'total_commissioning_price': contract_data['total_commissioning_price'],
        'total_delivery_base': contract_data['total_delivery_base'],
        
        # Individual pricing totals for template compatibility
        'total_canopy_price': pricing_totals.get('total_canopy_price', 0),
        'total_fire_suppression_price': pricing_totals.get('total_fire_suppression_price', 0),
        'total_cladding_price': pricing_totals.get('total_cladding_price', 0),
        'total_delivery_installation': pricing_totals.get('total_delivery_installation', 0),
        'total_commissioning': pricing_totals.get('total_commissioning', 0),
        'total_uvc_price': pricing_totals.get('total_uvc_price', 0),
        'total_sdu_price': pricing_totals.get('total_sdu_price', 0),
        'total_recoair_price': pricing_totals.get('total_recoair_price', 0),
        'total_vent_clg_price': pricing_totals.get('total_vent_clg_price', 0),
        'total_marvel_price': pricing_totals.get('total_marvel_price', 0),
        'project_total': pricing_totals.get('project_total', 0),  # Includes contract systems
        
        # Base systems total plus contract total (includes all pricing except contract systems M12/N12)
        'base_systems_total': (
            pricing_totals.get('total_canopy_price', 0) +
            pricing_totals.get('total_fire_suppression_price', 0) +
            pricing_totals.get('total_cladding_price', 0) +
            pricing_totals.get('total_delivery_installation', 0) +
            pricing_totals.get('total_commissioning', 0) +
            pricing_totals.get('total_uvc_price', 0) +
            pricing_totals.get('total_sdu_price', 0) +
            pricing_totals.get('total_vent_clg_price', 0) +
            pricing_totals.get('total_marvel_price', 0) +
            # UV Extra Over cost excluded from project total - it's comparison/information only
            pricing_totals.get('contract_total_price', 0)  # Include contract total from J9
        ),
        
        # Job total from Excel T28 (most accurate - uses Excel's own calculations)
        'job_total_t28': pricing_totals.get('job_total_t28', 0),
        
        # Main quote total excluding RecoAir (T28 - T24)
        'job_total_excluding_recoair': pricing_totals.get('job_total_t28', 0) - pricing_totals.get('recoair_price_t24', 0),
        
        # RecoAir-specific data (for RecoAir templates)
        'recoair_areas': [area for level in enhanced_levels for area in level.get('areas', []) if area.get('options', {}).get('recoair', False)],
        'total_recoair_units': sum(len(area.get('recoair_units', [])) for level in enhanced_levels for area in level.get('areas', [])),
        
        # SDU-specific data - ensure each SDU area has the services data directly accessible
        'sdu_areas': [
            {
                **sdu_item,
                # Make services data directly accessible on the sdu object for template compatibility
                'electrical_services': sdu_item.get('electrical_services', {}),
                'gas_services': sdu_item.get('gas_services', {}),
                'water_services': sdu_item.get('water_services', {}),
                'pricing': sdu_item.get('pricing', {})
            }
            for sdu_item in sdu_data
        ],
        'has_sdu': len(sdu_data) > 0,
        'total_sdu_areas': len(sdu_data),
        
        # Global project flags
        'has_canopies': has_canopies,
        'has_recoair': has_recoair,
        'is_recoair_only': is_recoair_only,
        'has_uv': has_uv,
        'has_marvel': has_marvel,
        'has_vent_clg': has_vent_clg,
        'has_pollustop': has_pollustop,
        'has_aerolys': has_aerolys,
        'has_xeu': has_xeu,
        
        # Feature flags for conditional display of systems
        'show_kitchen_extract_system': is_feature_enabled('kitchen_extract_system'),
        'show_kitchen_makeup_air_system': is_feature_enabled('kitchen_makeup_air_system'),
        'show_marvel_system': is_feature_enabled('marvel_system'),
        'show_cyclocell_cassette_ceiling': is_feature_enabled('cyclocell_cassette_ceiling'),
        'show_reactaway_unit': is_feature_enabled('reactaway_unit'),
        'show_dishwasher_extract': is_feature_enabled('dishwasher_extract'),
        'show_gas_interlocking': is_feature_enabled('gas_interlocking'),
        'show_pollustop_unit': is_feature_enabled('pollustop_unit'),
        'vent_ceiling': is_feature_enabled('cyclocell_cassette_ceiling'),
        
        # Add gas, water, electrical, and pricing data with default values (from first SDU if available)
        'electrical_services': sdu_data[0].get('electrical_services', {
            'distribution_board': 0,
            'single_phase_switched_spur': 0,
            'three_phase_socket_outlet': 0,
            'switched_socket_outlet': 0,
            'emergency_knock_off': 0,
            'ring_main_inc_2no_sso': 0
        }) if sdu_data else {
            'distribution_board': 0,
            'single_phase_switched_spur': 0,
            'three_phase_socket_outlet': 0,
            'switched_socket_outlet': 0,
            'emergency_knock_off': 0,
            'ring_main_inc_2no_sso': 0
        },
        'gas_services': sdu_data[0].get('gas_services', {
            'gas_manifold': 0,
            'gas_connection_15mm': 0,
            'gas_connection_20mm': 0,
            'gas_connection_25mm': 0,
            'gas_connection_32mm': 0,
            'gas_solenoid_valve': 0
        }) if sdu_data else {
            'gas_manifold': 0,
            'gas_connection_15mm': 0,
            'gas_connection_20mm': 0,
            'gas_connection_25mm': 0,
            'gas_connection_32mm': 0,
            'gas_solenoid_valve': 0
        },  # Match Excel structure for SDU gas services
        'water_services': sdu_data[0].get('water_services', {
            'cws_manifold_22mm': 0,
            'cws_manifold_15mm': 0,
            'hws_manifold': 0,
            'water_connection_15mm': 0,
            'water_connection_22mm': 0,
            'water_connection_28mm': 0
        }) if sdu_data else {
            'cws_manifold_22mm': 0,
            'cws_manifold_15mm': 0,
            'hws_manifold': 0,
            'water_connection_15mm': 0,
            'water_connection_22mm': 0,
            'water_connection_28mm': 0
        },  # Match Excel structure for SDU water services
        'pricing': sdu_data[0].get('pricing', {
            'final_carcass_price': 0,
            'final_electrical_price': 0,
            'live_site_test_price': 0,
            'has_live_test': False,
            'total_price': 0
        }) if sdu_data else {
            'final_carcass_price': 0,
            'final_electrical_price': 0,
            'live_site_test_price': 0,
            'has_live_test': False,
            'total_price': 0
        },  # Match Excel structure for SDU pricing
        
        # Fallback variables for template compatibility
        'level': enhanced_levels[0] if enhanced_levels else {'level_name': '', 'areas': []},  # First level as fallback
        'area': enhanced_levels[0].get('areas', [{}])[0] if enhanced_levels and enhanced_levels[0].get('areas') else {
            'name': '', 
            'level_area_combined': '',
            'canopies': [],
            'recoair_units': [], 
            'options': {},
            'delivery_installation_price': 0,
            'commissioning_price': 0,
            'uvc_price': 0,
            'sdu_price': 0,
            'recoair_price': 0,
            'has_uv_extra_over': False,
            'uv_extra_over_cost': 0,
            'canopy_total': 0,
            'fire_suppression_total': 0,
            'cladding_total': 0,
            'canopy_schedule_subtotal': 0,
            'area_total': 0,
            'has_canopies': False,
            'has_sdu': False,
            'sdu_pricing': {}
        },
        
        # RecoAir unit fallback variables
        'model': '',  # Fallback for RecoAir model
        'extract_volume': 0,  # Fallback for extract volume
        'width': 0,  # Fallback for width
        'length': 0,  # Fallback for length
        'height': 0,  # Fallback for height
        'recoair_location': 'INTERNAL',  # Fallback for RecoAir unit location
        'unit_price': 0,  # Fallback for unit price
        'quantity': 0,  # Fallback for quantity
        'delivery_installation_price': 0,  # Fallback for delivery price
        'total_uv_extra_over_cost': pricing_totals.get('total_uv_extra_over_cost', 0),  # Use calculated total UV Extra Over costs
        'has_any_uv_extra_over': pricing_totals.get('has_any_uv_extra_over', False),  # Use calculated UV Extra Over flag
        'extra_overs': (enhanced_levels[0].get('areas', [{}])[0] if enhanced_levels and enhanced_levels[0].get('areas') else {}).get('options', {}).get('uv_extra_over', False),  # Easy flag for templates
    }
    
    context_time = time.time()
    total_context_time = context_time - context_start
    print(f"📋 Context preparation complete: {total_context_time:.3f}s")
    
    total_time = time.time() - start_time
    print(f"🚀 Template context preparation TOTAL: {total_time:.3f}s")
    print(f"   📊 Breakdown:")
    print(f"      - Initialization: {init_time - start_time:.3f}s")
    print(f"      - Contact info: {contact_time - init_time:.3f}s")
    print(f"      - Contract processing: {total_contract_time:.3f}s")
    print(f"      - Levels processing: {total_levels_time:.3f}s")
    print(f"      - RecoAir data: {recoair_time:.3f}s")
    print(f"      - SDU data: {sdu_time:.3f}s")
    print(f"      - Project analysis: {analysis_time:.3f}s")
    print(f"      - Pricing calculations: {pricing_time:.3f}s")
    print(f"      - Context preparation: {total_context_time:.3f}s")
    
    return context

def analyze_project_areas(project_data: Dict) -> Tuple[bool, bool, bool, bool, bool, bool]:
    """
    Analyze project areas to determine what types of systems are present.
    
    Args:
        project_data (Dict): Project data with levels and areas
        
    Returns:
        Tuple[bool, bool, bool, bool, bool, bool, bool, bool, bool]: (has_canopies, has_recoair, is_recoair_only, has_uv, has_marvel, has_vent_clg, has_pollustop, has_aerolys, has_xeu)
    """
    has_canopies = False
    has_recoair = False
    has_uv = False
    has_marvel = False
    has_vent_clg = False
    has_pollustop = False
    has_aerolys = False
    has_xeu = False
    
    for level in project_data.get('levels', []):
        for area in level.get('areas', []):
            # Check if area has canopies (check length, not just existence)
            canopies = area.get('canopies', [])
            if len(canopies) > 0:
                has_canopies = True
            
            # Check for UV canopy models (UVI, UVF, etc.) across all canopies in the project
            for canopy in canopies:
                model = canopy.get('model', '').upper().strip()
                if model.startswith('UV'):  # UV models like UVI, UVF, etc.
                    has_uv = True
            
            # Check if area has RecoAir option
            if area.get('options', {}).get('recoair', False):
                has_recoair = True
            
            # Check if area has MARVEL option
            if area.get('options', {}).get('marvel', False):
                has_marvel = True
            
            # Check if area has VENT CLG option
            if area.get('options', {}).get('vent_clg', False):
                has_vent_clg = True

            # Check if area has Pollustop option
            pollustop_value = area.get('options', {}).get('pollustop', False)
            if pollustop_value:
                print(f"🟢 Pollustop detected in area: {area.get('name', 'Unknown')}")
                has_pollustop = True

            # Check if area has Aerolys option
            aerolys_value = area.get('options', {}).get('aerolys', False)
            if aerolys_value:
                print(f"🟢 Aerolys detected in area: {area.get('name', 'Unknown')}")
                has_aerolys = True

            # Check if area has XEU option (XEU creates both Pollustop AND Aerolys sheets)
            xeu_value = area.get('options', {}).get('xeu', False)
            if xeu_value:
                print(f"🟢 XEU detected in area: {area.get('name', 'Unknown')} - this creates both Pollustop AND Aerolys sheets")
                has_xeu = True
                has_pollustop = True  # XEU creates Pollustop sheet
                has_aerolys = True    # XEU creates Aerolys sheet
    
    # Determine if project is RecoAir-only (has RecoAir but no other systems)
    is_recoair_only = has_recoair and not (has_canopies or has_vent_clg or has_marvel or has_uv or has_pollustop or has_aerolys or has_xeu)

    return has_canopies, has_recoair, is_recoair_only, has_uv, has_marvel, has_vent_clg, has_pollustop, has_aerolys, has_xeu

def generate_single_document(project_data: Dict, template_path: str, output_filename: str, excel_file_path: str = None, template_key: str = None) -> str:
    """
    Generate a single Word document from project data using specified template.

    Args:
        project_data (Dict): Project data extracted from Excel
        template_path (str): Path to the Word template (will download from Supabase if needed)
        output_filename (str): Name for the output file
        template_key (str): Template key for Supabase Storage (e.g., 'canopy_quotation')

    Returns:
        str: Path to the generated Word document
    """
    import time
    start_time = time.time()
    print(f"📄 Starting Word document generation...")
    print(f"   📁 Template: {template_path}")
    print(f"   📁 Output: {output_filename}")
    
    try:
        # Ensure template is available (download from Supabase Storage if needed)
        template_check_start = time.time()
        if template_key:
            print(f"   📥 Ensuring template is available (downloading from Supabase if needed)...")
            template_path = ensure_template_available(template_key, template_path)
        else:
            # Fallback to old behavior if no template_key provided
            if not os.path.exists(template_path):
                raise Exception(f"Template file not found at {template_path}")
        template_check_time = time.time() - template_check_start
        print(f"   ✅ Template ready: {template_check_time:.3f}s")

        # Load the template
        template_load_start = time.time()
        print(f"   📖 Loading Word template...")
        doc = DocxTemplate(template_path)
        template_load_time = time.time() - template_load_start
        print(f"   ✅ Template loaded: {template_load_time:.3f}s")
        
        # Prepare the context for template rendering
        context_prep_start = time.time()
        print(f"   🔧 Preparing template context...")
        context = prepare_template_context(project_data, excel_file_path)
        context_prep_time = time.time() - context_prep_start
        print(f"   ✅ Context prepared: {context_prep_time:.3f}s")
        
        # Render the template with the context
        render_start = time.time()
        print(f"   🎨 Rendering template with context...")
        doc.render(context)
        render_time = time.time() - render_start
        print(f"   ✅ Template rendered: {render_time:.3f}s")
        
        # Save the document
        save_start = time.time()
        print(f"   💾 Saving document...")
        output_path = f"output/{output_filename}"
        os.makedirs("output", exist_ok=True)
        doc.save(output_path)
        save_time = time.time() - save_start
        print(f"   ✅ Document saved: {save_time:.3f}s")
        
        total_time = time.time() - start_time
        print(f"📄 Word document generation COMPLETE: {total_time:.3f}s")
        print(f"   📊 Breakdown:")
        print(f"      - Template check: {template_check_time:.3f}s")
        print(f"      - Template load: {template_load_time:.3f}s")
        print(f"      - Context prep: {context_prep_time:.3f}s")
        print(f"      - Template render: {render_time:.3f}s")
        print(f"      - Document save: {save_time:.3f}s")
        
        return output_path
        
    except Exception as e:
        error_time = time.time() - start_time
        print(f"   ❌ Error after {error_time:.3f}s: {str(e)}")
        raise Exception(f"Failed to generate Word document: {str(e)}")

def format_date_for_filename(date_str: str) -> str:
    """
    Format date for filename (remove slashes and make it filename-safe).
    
    Args:
        date_str (str): Date string from project data
        
    Returns:
        str: Formatted date string for filename
    """
    if date_str:
        # Convert DD/MM/YYYY to DDMMYYYY or similar format
        return date_str.replace('/', '').replace('-', '')
    else:
        # Use current date if no date provided
        return datetime.now().strftime("%d%m%Y")

def generate_quotation_document(project_data: Dict, excel_file_path: str = None) -> Union[str, str]:
    """
    Generate Word quotation document(s) from project data using Jinja templating.
    Returns either a single document path or a zip file path containing multiple documents.
    
    Args:
        project_data (Dict): Project data extracted from Excel
        
    Returns:
        str: Path to the generated Word document or zip file
    """
    try:
        # Analyze project to determine what documents to generate
        has_canopies, has_recoair, is_recoair_only, has_uv, has_marvel, has_vent_clg, has_pollustop, has_aerolys, has_xeu = analyze_project_areas(project_data)
        
        project_number = project_data.get('project_number', 'unknown')
        date_str = format_date_for_filename(project_data.get('date', ''))
        
        # Case 1: RecoAir-only project - generate only RecoAir quotation
        # Format: "Project Number RecoAir Quotation Date Rev X"
        revision = project_data.get('revision', '')
        if is_recoair_only:
            if revision and revision.strip():
                output_filename = f"{project_number} RecoAir Quotation {date_str} Rev {revision}.docx"
            else:
                output_filename = f"{project_number} RecoAir Quotation {date_str}.docx"
            return generate_single_document(project_data, RECOAIR_TEMPLATE_PATH, output_filename, excel_file_path, template_key="recoair_quotation")

        # Case 2: Mixed project or canopy-only project
        documents_to_generate = []

        # Generate main quotation if there are canopies OR ventilated ceilings (or other non-RecoAir systems)
        # Format: "Project Number Quotation Date Rev X"
        if has_canopies or has_vent_clg or has_marvel or has_uv or has_pollustop or has_aerolys or has_xeu:
            if revision and revision.strip():
                main_filename = f"{project_number} Quotation {date_str} Rev {revision}.docx"
            else:
                main_filename = f"{project_number} Quotation {date_str}.docx"
            documents_to_generate.append((WORD_TEMPLATE_PATH, main_filename, "Main Quotation", "canopy_quotation"))

        # Generate RecoAir quotation if there are RecoAir areas
        # Format: "Project Number RecoAir Quotation Date Rev X"
        if has_recoair:
            if revision and revision.strip():
                recoair_filename = f"{project_number} RecoAir Quotation {date_str} Rev {revision}.docx"
            else:
                recoair_filename = f"{project_number} RecoAir Quotation {date_str}.docx"
            documents_to_generate.append((RECOAIR_TEMPLATE_PATH, recoair_filename, "RecoAir Quotation", "recoair_quotation"))

        # Generate AHU quotation if there are Pollustop or Aerolys areas (including XEU which creates both)
        # Format: "Project Number AHU Quotation Date Rev X"
        print(f"🔍 AHU Detection: has_pollustop={has_pollustop}, has_aerolys={has_aerolys}, has_xeu={has_xeu}")
        if has_pollustop or has_aerolys:
            if revision and revision.strip():
                ahu_filename = f"{project_number} AHU Quotation {date_str} Rev {revision}.docx"
            else:
                ahu_filename = f"{project_number} AHU Quotation {date_str}.docx"
            print(f"✅ Adding AHU quotation to generation list: {ahu_filename}")
            documents_to_generate.append((AHU_TEMPLATE_PATH, ahu_filename, "AHU Quotation", "ahu_quotation"))
        else:
            print(f"❌ No AHU quotation needed - no Pollustop, Aerolys, or XEU areas detected")

        # If only one document to generate, return it directly
        if len(documents_to_generate) == 1:
            template_path, filename, _, template_key = documents_to_generate[0]
            return generate_single_document(project_data, template_path, filename, excel_file_path, template_key=template_key)

        # If multiple documents, generate all and create zip file
        if len(documents_to_generate) > 1:
            generated_files = []

            # Generate each document
            for template_path, filename, description, template_key in documents_to_generate:
                try:
                    file_path = generate_single_document(project_data, template_path, filename, excel_file_path, template_key=template_key)
                    generated_files.append((file_path, filename))
                except Exception as e:
                    print(f"Warning: Failed to generate {description}: {str(e)}")
                    continue
            
            # Create zip file if we have multiple documents
            if len(generated_files) > 1:
                if revision and revision.strip():
                    zip_filename = f"{project_number} Quotations {date_str} Rev {revision}.zip"
                else:
                    zip_filename = f"{project_number} Quotations {date_str}.zip"
                zip_path = f"output/{zip_filename}"
                
                with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    # Add Word documents only
                    for file_path, filename in generated_files:
                        zipf.write(file_path, filename)
                
                return zip_path
            
            # If only one document was successfully generated, return it
            elif len(generated_files) == 1:
                return generated_files[0][0]
        
        # Fallback: generate main quotation only
        if revision and revision.strip():
            main_filename = f"{project_number} Quotation {date_str} Rev {revision}.docx"
        else:
            main_filename = f"{project_number} Quotation {date_str}.docx"
        return generate_single_document(project_data, WORD_TEMPLATE_PATH, main_filename, excel_file_path, template_key="canopy_quotation")
        
    except Exception as e:
        raise Exception(f"Failed to generate Word document(s): {str(e)}")

def collect_recoair_pricing_schedule_data(project_data: Dict) -> Dict:
    """
    Collect RecoAir pricing schedule data per area for Word document display.
    
    Args:
        project_data (Dict): Project data with levels and areas
        
    Returns:
        Dict: Dictionary containing area pricing schedules and job totals
    """
    pricing_schedules = []
    
    # Job totals tracking
    job_totals = {
        'total_units_price': 0,
        'total_delivery_price': 0,
        'total_commissioning_price': 0,
        'total_flat_pack_price': 0,
        'job_total': 0,
        'total_areas': 0,
        'total_units': 0
    }
    
    for level in project_data.get('levels', []):
        level_name = level.get('level_name', '')
        
        for area in level.get('areas', []):
            area_name = area.get('name', '')
            level_area_combined = f"{level_name} - {area_name}"
            
            # Only process areas that have RecoAir systems
            if not area.get('options', {}).get('recoair', False):
                continue
            
            # Get RecoAir units for this area
            recoair_units = area.get('recoair_units', [])
            
            if not recoair_units:
                continue
            
            # Collect RecoAir unit items for this area
            recoair_items = []
            recoair_units_full = []  # Full unit specifications
            area_units_total = 0
            total_delivery_price = 0
            
            for unit in recoair_units:
                reference_number = unit.get('item_reference', '')
                model = unit.get('model', '')
                base_unit_price = unit.get('base_unit_price', 0) or 0  # Use base price from N12
                delivery_price = unit.get('delivery_installation_price', 0) or 0
                
                if reference_number and model:  # Only include units with reference and model
                    # Basic pricing item (for backward compatibility)
                    recoair_item = {
                        'reference_number': reference_number,
                        'model': model,
                        'price': base_unit_price,  # Use base price from N12
                        'delivery_price': delivery_price
                    }
                    recoair_items.append(recoair_item)
                    
                    # Full unit specification (includes all technical details)
                    recoair_unit_full = {
                        'reference_number': reference_number,
                        'model': model,
                        'price': base_unit_price,  # Use base price from N12
                        'delivery_price': delivery_price,
                        # Technical specifications
                        'length': unit.get('length', 0),
                        'width': unit.get('width', 0),
                        'height': unit.get('height', 0),
                        'extract_volume': unit.get('extract_volume', 0),
                        'p_drop': unit.get('p_drop', 0),
                        'motor': unit.get('motor', 0),
                        'weight': unit.get('weight', 0),
                        'location': unit.get('location', 'INTERNAL'),
                        # Additional data
                        'quantity': unit.get('quantity', 1),
                        'model_original': unit.get('model_original', model),
                        'extract_volume_raw': unit.get('extract_volume_raw', ''),
                        'base_unit_price': base_unit_price,  # Keep base price for reference
                        'n29_addition': unit.get('n29_addition', 0)
                    }
                    recoair_units_full.append(recoair_unit_full)
                    
                    area_units_total += base_unit_price  # Use base price from N12
                    total_delivery_price += delivery_price
            
            # Get commissioning price from N46 in Excel (recoair_commissioning_price)
            # Do NOT fallback to general area commissioning price - use only RecoAir-specific commissioning
            commissioning_price = area.get('recoair_commissioning_price', 0)
            print(f"RecoAir commissioning price for {level_area_combined}: {commissioning_price}")
            
            # Get flat pack data if available
            flat_pack_data = area.get('recoair_flat_pack', {})
            flat_pack_price = flat_pack_data.get('price', 0) if flat_pack_data.get('has_flat_pack', False) else 0
            
            # Calculate different subtotals for different purposes
            recoair_subtotal = area_units_total + total_delivery_price + commissioning_price  # RecoAir units + delivery + commissioning (excluding flat pack)
            area_total_with_flat_pack = recoair_subtotal + flat_pack_price  # Everything including flat pack
            
            # Only create pricing schedule if there are RecoAir units in this area
            if recoair_items:
                area_pricing = {
                    'level_name': level_name,
                    'area_name': area_name,
                    'level_area_combined': level_area_combined,
                    'recoair_items': recoair_items,  # Basic pricing items (for backward compatibility)
                    'units': recoair_units_full,  # Full unit specifications with technical details
                    'units_total': area_units_total,
                    'delivery_installation_price': total_delivery_price,
                    'commissioning_price': commissioning_price,
                    'flat_pack_price': flat_pack_price,
                    'flat_pack_description': flat_pack_data.get('description', ''),
                    'flat_pack_item_reference': flat_pack_data.get('item_reference', ''),
                    'has_flat_pack': flat_pack_data.get('has_flat_pack', False),
                    'area_subtotal': recoair_subtotal,  # RecoAir units + delivery + commissioning (excluding flat pack)
                    'area_total_with_flat_pack': area_total_with_flat_pack,  # Everything including flat pack
                    'unit_count': len(recoair_items)
                }
                pricing_schedules.append(area_pricing)
                
                # Add to job totals (without rounding)
                job_totals['total_units_price'] += area_units_total
                job_totals['total_delivery_price'] += total_delivery_price
                job_totals['total_commissioning_price'] += commissioning_price
                job_totals['total_flat_pack_price'] += flat_pack_price
                job_totals['total_areas'] += 1
                job_totals['total_units'] += len(recoair_items)
    
    # Calculate overall job total (excluding flat pack - flat pack is shown separately as "Additional Items")
    job_totals['job_total'] = (
        job_totals['total_units_price'] + 
        job_totals['total_delivery_price'] + 
        job_totals['total_commissioning_price']
        # Note: flat_pack_price excluded from main total to match Excel
    )
    
    return {
        'areas': pricing_schedules,
        'job_totals': job_totals
    }

def calculate_pricing_totals(project_data: Dict, excel_file_path: str = None, cached_wb=None) -> Dict:
    """
    Calculate comprehensive pricing totals for the project.
    
    Args:
        project_data (Dict): Project data
        excel_file_path (str, optional): Path to Excel file for contract data
        cached_wb (optional): Pre-loaded Excel workbook to avoid reloading
        
    Returns:
        Dict: Comprehensive pricing totals
    """
    import time
    start_time = time.time()
    print(f"💰 Starting pricing totals calculation...")
    
    totals = {
        'total_canopy_price': 0,
        'total_fire_suppression_price': 0,
        'total_cladding_price': 0,
        'total_delivery_installation': 0,
        'total_commissioning': 0,
        'total_uvc_price': 0,
        'total_sdu_price': 0,
        'total_recoair_price': 0,
        'total_vent_clg_price': 0,
        'total_uv_extra_over_cost': 0,
        'has_any_uv_extra_over': False,
        'project_total': 0,
        'contract_total_price': 0,
        'job_total_t28': 0,  # Job total from Excel T28
        'areas': []  # Store area-level data for templates
    }
    
    init_time = time.time()
    print(f"   ✅ Totals initialization: {init_time - start_time:.3f}s")
    
    base_project_total = 0
    contract_systems_total = 0
    
    # Process each level and area
    area_process_start = time.time()
    print(f"   🏢 Processing areas for pricing...")
    
    for level in project_data.get('levels', []):
        level_name = level.get('level_name', '')
        
        for area in level.get('areas', []):
            area_name = area.get('name', '')
            
            # Canopy pricing
            area_canopy_total = sum(canopy.get('canopy_price', 0) for canopy in area.get('canopies', []))
            totals['total_canopy_price'] += area_canopy_total
            
            # Fire suppression pricing
            area_fire_supp_total = sum(canopy.get('fire_suppression_price', 0) for canopy in area.get('canopies', []))
            totals['total_fire_suppression_price'] += area_fire_supp_total
            
            # Wall cladding pricing - use same logic as transformed canopies
            area_cladding_total = 0
            for canopy in area.get('canopies', []):
                # Use same logic as in transformed canopies for consistency
                cladding_price = canopy.get('cladding_price', 0)
                if not cladding_price:
                    wall_cladding = canopy.get('wall_cladding', {})
                    cladding_price = wall_cladding.get('price', 0)
                area_cladding_total += cladding_price
            totals['total_cladding_price'] += area_cladding_total
            
            # Area-level pricing
            delivery_installation = area.get('delivery_installation_price', 0)
            commissioning = area.get('commissioning_price', 0)
            uvc_price = area.get('uvc_price', 0)
            # Calculate total SDU price from all canopies in this area
            sdu_price = sum(canopy.get('sdu_price', 0) for canopy in area.get('canopies', []))
            recoair_price = area.get('recoair_price', 0)
            vent_clg_price = area.get('vent_clg_price', 0)
            marvel_price = area.get('marvel_price', 0)
            
            totals['total_delivery_installation'] += delivery_installation
            totals['total_commissioning'] += commissioning
            totals['total_uvc_price'] += uvc_price
            totals['total_sdu_price'] += sdu_price
            totals['total_recoair_price'] += recoair_price
            totals['total_vent_clg_price'] += vent_clg_price
            totals['total_marvel_price'] = totals.get('total_marvel_price', 0) + marvel_price
            
            # Check for UV Extra Over
            has_uv_extra_over = area.get('options', {}).get('uv_extra_over', False)
            uv_extra_over_cost = area.get('uv_extra_over_cost', 0)
            
            if has_uv_extra_over and uv_extra_over_cost > 0:
                totals['has_any_uv_extra_over'] = True
                totals['total_uv_extra_over_cost'] += uv_extra_over_cost
            
            # Calculate area subtotals and totals for template access
            # CANOPY SCHEDULE subtotal should ONLY include canopy prices + delivery + commissioning
            # Fire suppression and cladding are separate schedules with their own subtotals
            area_canopy_schedule_subtotal = area_canopy_total + delivery_installation + commissioning
            # Area total includes: canopy schedule + fire suppression + cladding + other systems
            # Note: delivery_installation and commissioning are already included in area_canopy_schedule_subtotal
            # Note: RecoAir pricing should NOT be included in area totals - it has its own separate pricing schedule
            # Note: UV Extra Over cost should NOT be included in area totals - it's a comparison/information only
            area_total = (area_canopy_schedule_subtotal + area_fire_supp_total + area_cladding_total + 
                         uvc_price + sdu_price + vent_clg_price + marvel_price)  # Removed uv_extra_over_cost

            # Process canopies to add has_cladding flag for template compatibility
            processed_canopies = []
            for canopy in area.get('canopies', []):
                # Create a copy of the canopy data
                processed_canopy = dict(canopy)
                
                # Apply same has_cladding logic as in template context preparation
                wall_cladding = canopy.get('wall_cladding', {})
                cladding_price = canopy.get('cladding_price', 0)
                has_cladding = (wall_cladding.get('type') not in ['None', None, ''] and 
                               wall_cladding.get('type') and 
                               (cladding_price > 0 or wall_cladding.get('price', 0) > 0))
                
                processed_canopy['has_cladding'] = has_cladding
                
                # Ensure fire suppression system description is included
                if processed_canopy.get('fire_suppression_system_type'):
                    fs_type = processed_canopy.get('fire_suppression_system_type', '')
                    processed_canopy['fire_suppression_system_description'] = get_fire_suppression_system_description(fs_type)
                
                processed_canopies.append(processed_canopy)
            
            # Store area data for template access (moved outside canopy loop)
            area_data = {
                'level_area_combined': f"{level_name} - {area_name}",
                'name': area_name,
                'area_name': area_name,  # Add area_name key for consistency
                'level_name': level_name,
                'has_canopies': len(area.get('canopies', [])) > 0,
                'has_uv_extra_over': has_uv_extra_over,
                'uv_extra_over_cost': uv_extra_over_cost,
                'options': area.get('options', {}),
                'delivery_installation_price': delivery_installation,
                'commissioning_price': commissioning,
                'uvc_price': uvc_price,
                'sdu_price': sdu_price,
                'recoair_price': recoair_price,
                'vent_clg_price': vent_clg_price,
                'marvel_price': marvel_price,
                'vent_clg_detailed_pricing': area.get('vent_clg_detailed_pricing', {}),
                'recoair_units': area.get('recoair_units', []),
                'canopy_total': area_canopy_total,
                'fire_suppression_total': area_fire_supp_total,
                'cladding_total': area_cladding_total,
                'canopy_schedule_subtotal': area_canopy_schedule_subtotal,
                'area_total': area_total,
                'area_subtotal': area_total,  # Alternative name for template compatibility
                'canopies': processed_canopies,  # Use processed canopies with has_cladding flag
                # Check if any canopy in this area has SDU
                'has_sdu': any(canopy.get('options', {}).get('sdu', False) for canopy in area.get('canopies', [])),
                'sdu_pricing': area.get('sdu_pricing', {}),
                
                # Additional template compatibility variables
                'has_marvel': area.get('options', {}).get('marvel', False),
                'marvel_pricing': area.get('marvel_pricing', {}),
                'has_vent_clg': area.get('options', {}).get('vent_clg', False),
                'has_pollustop': area.get('options', {}).get('pollustop', False),
                'has_aerolys': area.get('options', {}).get('aerolys', False),
                'has_xeu': area.get('options', {}).get('xeu', False),
                'uve_price': uvc_price,  # Alternative spelling for template compatibility
                'sdu_subtotal': sdu_price,  # SDU subtotal for template compatibility
                'sdu': {  # SDU data object with pricing structure that matches template expectations
                    'pricing': {
                        'final_carcass_price': 0,
                        'final_electrical_price': 0,
                        'live_site_test_price': 0,
                        'has_live_test': False,
                        'total_price': 0
                    }
                },
            }
            totals['areas'].append(area_data)
    
    area_process_time = time.time() - area_process_start
    print(f"   ✅ Area processing complete: {area_process_time:.3f}s")
    
    # Calculate base project total
    base_calc_start = time.time()
    print(f"   🧮 Calculating base project total...")
    
    # Base project total calculation (excluding UV Extra Over)
    # UV Extra Over is excluded as it's for comparison/information only
    base_project_total = (
        totals['total_canopy_price'] +
        totals['total_fire_suppression_price'] +
        totals['total_cladding_price'] +
        totals['total_delivery_installation'] +
        totals['total_commissioning'] +
        totals['total_uvc_price'] +
        totals['total_sdu_price'] +
        totals['total_vent_clg_price'] +
        totals.get('total_marvel_price', 0)
        # UV Extra Over cost excluded from project total
    )
    
    base_calc_time = time.time() - base_calc_start
    print(f"   ✅ Base calculation complete: {base_calc_time:.3f}s")
    
    # Collect SDU data to merge with area data
    sdu_merge_start = time.time()
    print(f"   📡 Collecting SDU data for merging...")
    sdu_data_list = collect_sdu_data(project_data, excel_file_path, cached_wb)
    
    # Create a lookup dictionary for SDU data by canopy reference
    sdu_lookup = {sdu['canopy_reference']: sdu for sdu in sdu_data_list}
    
    # Update area data with actual SDU pricing and recalculate totals
    for area_data in totals['areas']:
        level_area_combined = area_data['level_area_combined']
        
        # Aggregate SDU pricing from all canopies in this area
        area_sdu_subtotal = 0
        area_has_detailed_sdu = False
        
        # Get all SDU data for canopies in this area
        for level in project_data.get('levels', []):
            if level.get('level_name') == area_data['level_name']:
                for area in level.get('areas', []):
                    if area.get('name') == area_data['area_name']:
                        for canopy in area.get('canopies', []):
                            canopy_ref = canopy.get('reference_number', '')
                            if canopy_ref in sdu_lookup:
                                sdu_info = sdu_lookup[canopy_ref]
                                sdu_detailed_pricing = sdu_info.get('pricing', {})
                                canopy_sdu_total = sdu_detailed_pricing.get('total_price', 0)
                                if canopy_sdu_total > 0:
                                    area_sdu_subtotal += canopy_sdu_total
                                    area_has_detailed_sdu = True
        
        # Update area SDU pricing if we have detailed data
        if area_has_detailed_sdu and area_sdu_subtotal > 0:
            area_data['sdu_subtotal'] = area_sdu_subtotal  # Update subtotal
            
            # Recalculate area total using actual SDU pricing instead of basic sdu_price
            old_sdu_price = area_data['sdu_price']  # Basic SDU price used in original calculation
            area_data['sdu_price'] = area_sdu_subtotal   # Update to use detailed pricing
            
            # Update area total by replacing the old SDU price with the new detailed pricing
            area_data['area_total'] = area_data['area_total'] - old_sdu_price + area_sdu_subtotal
            area_data['area_subtotal'] = area_data['area_total']  # Keep alternative name in sync
            
            # Update global totals as well
            totals['total_sdu_price'] = totals['total_sdu_price'] - old_sdu_price + area_sdu_subtotal
            
            print(f"         ✅ Updated {level_area_combined}: SDU subtotal ${area_sdu_subtotal}, Area total ${area_data['area_total']}")
        else:
            # Keep basic SDU pricing - no detailed pricing available
            area_data['sdu_subtotal'] = area_data['sdu_price']  # Use basic price as subtotal
            if area_data['sdu_price'] > 0:
                print(f"         ℹ️  Kept basic SDU pricing for {level_area_combined}: ${area_data['sdu_price']}")
    
    sdu_merge_time = time.time() - sdu_merge_start
    print(f"   ✅ SDU data merging complete: {sdu_merge_time:.3f}s")
    
    # Get contract total from Excel if available
    contract_start = time.time()
    print(f"   📋 Reading contract totals from Excel...")
    
    if excel_file_path and os.path.exists(excel_file_path):
        try:
            # Use cached workbook if available, otherwise load fresh
            if cached_wb:
                wb = cached_wb
                print(f"      ✅ Using cached workbook")
                wb_load_time = 0
            else:
                wb_load_start = time.time()
                wb = load_workbook(excel_file_path, data_only=True)
                wb_load_time = time.time() - wb_load_start
                print(f"      📖 Excel loaded: {wb_load_time:.3f}s")
            
            # Look for CONTRACT sheet (exact match or numbered variant like CONTRACT1)
            contract_sheet = None
            for sheet_name in wb.sheetnames:
                if sheet_name == 'CONTRACT' or sheet_name.startswith('CONTRACT') and len(sheet_name) <= 10:  # Handle CONTRACT1, CONTRACT2, etc.
                    contract_sheet = wb[sheet_name]
                    break
            
            if contract_sheet:
                print(f"      📊 Reading contract totals...")
                # Get contract total from J9
                contract_total = contract_sheet['J9'].value
                if contract_total and float(contract_total) > 0:
                    totals['contract_total_price'] = float(contract_total)
                    print(f"      ✅ Contract total found: {totals['contract_total_price']}")
            else:
                print(f"      ℹ️  No CONTRACT sheet found")
            
            # Read job total from JOB TOTAL sheet T28 and RecoAir price from T24
            if 'JOB TOTAL' in wb.sheetnames:
                job_total_sheet = wb['JOB TOTAL']
                t28_value = job_total_sheet['T28'].value
                if t28_value and isinstance(t28_value, (int, float)) and t28_value > 0:
                    totals['job_total_t28'] = float(t28_value)
                    print(f"      ✅ Job total T28 found: {totals['job_total_t28']}")
                else:
                    print(f"      ℹ️  No valid T28 value in JOB TOTAL sheet (value: {t28_value})")
                
                # Read RecoAir price from T24
                t24_value = job_total_sheet['T24'].value
                if t24_value and isinstance(t24_value, (int, float)):
                    totals['recoair_price_t24'] = float(t24_value)
                    print(f"      ✅ RecoAir price T24 found: {totals['recoair_price_t24']}")
                else:
                    print(f"      ℹ️  No valid T24 value in JOB TOTAL sheet (value: {t24_value})")
                    totals['recoair_price_t24'] = 0
            else:
                print(f"      ℹ️  No JOB TOTAL sheet found")
        except Exception as e:
            print(f"      ❌ Contract total read error: {str(e)}")
    else:
        print(f"      ℹ️  No Excel file for contract totals")
    
    contract_time = time.time() - contract_start
    print(f"   ✅ Contract processing complete: {contract_time:.3f}s")
    
    # Read contract systems totals for project total calculation
    contract_systems_start = time.time()
    print(f"   🏗️  Reading contract systems totals...")
    
    if excel_file_path and os.path.exists(excel_file_path):
        try:
            # Use cached workbook if available, otherwise load fresh
            if cached_wb:
                wb = cached_wb
                print(f"      ✅ Using cached workbook")
            else:
                wb = load_workbook(excel_file_path, data_only=True)
            
            # Look for CONTRACT sheet
            contract_sheet = None
            for sheet_name in wb.sheetnames:
                if sheet_name == 'CONTRACT' or sheet_name.startswith('CONTRACT') and len(sheet_name) <= 10:
                    contract_sheet = wb[sheet_name]
                    break
            
            if contract_sheet:
                # Get extract system total from M12
                extract_total = contract_sheet['M12'].value
                if extract_total and isinstance(extract_total, (int, float)) and extract_total > 0:
                    contract_systems_total += float(extract_total)
                
                # Get supply system total from N12
                supply_total = contract_sheet['N12'].value
                if supply_total and isinstance(supply_total, (int, float)) and supply_total > 0:
                    contract_systems_total += float(supply_total)
                    
                print(f"      ✅ Contract systems total: {contract_systems_total}")
        except Exception as e:
            print(f"      ❌ Contract systems read error: {str(e)}")
    
    contract_systems_time = time.time() - contract_systems_start
    print(f"   ✅ Contract systems processing complete: {contract_systems_time:.3f}s")
    
    # Recalculate base project total with updated SDU pricing
    final_calc_start = time.time()
    # Recalculate base total (excluding UV Extra Over)
    # UV Extra Over is excluded as it's for comparison/information only
    updated_base_total = (
        totals['total_canopy_price'] +
        totals['total_fire_suppression_price'] +
        totals['total_cladding_price'] +
        totals['total_delivery_installation'] +
        totals['total_commissioning'] +
        totals['total_uvc_price'] +
        totals['total_sdu_price'] +  # This now includes updated detailed SDU pricing
        totals['total_vent_clg_price'] +
        totals.get('total_marvel_price', 0)
        # UV Extra Over cost excluded from project total
    )
    
    # Add contract systems total and contract total price to project total
    totals['project_total'] = updated_base_total + contract_systems_total + totals.get('contract_total_price', 0)
    final_calc_time = time.time() - final_calc_start
    print(f"   ✅ Final total calculation (with updated SDU pricing): {final_calc_time:.3f}s")
    
    total_time = time.time() - start_time
    print(f"💰 Pricing totals calculation COMPLETE: {total_time:.3f}s")
    print(f"   📊 Breakdown:")
    print(f"      - Initialization: {init_time - start_time:.3f}s")
    print(f"      - Area processing: {area_process_time:.3f}s")
    print(f"      - Base calculation: {base_calc_time:.3f}s")
    print(f"      - SDU data merging: {sdu_merge_time:.3f}s")
    print(f"      - Contract totals: {contract_time:.3f}s")
    print(f"      - Contract systems: {contract_systems_time:.3f}s")
    print(f"      - Final calculation: {final_calc_time:.3f}s")
    
    return totals

def format_currency(amount) -> str:
    """
    Format currency amount for display with ceiling rounding.
    All amounts are rounded UP to the nearest whole pound.
    
    Args:
        amount: Currency amount to format
        
    Returns:
        str: Formatted currency string (always ends in .00)
    """
    if not amount:
        return "£0.00"
    
    try:
        import math
        # Round UP to the nearest whole number (ceiling)
        # Only round if this is a final total (amount ends in .99 or similar)
        float_amount = float(amount)
        decimal_part = float_amount - int(float_amount)
        
        if decimal_part > 0:
            # This is a final total that needs rounding
            rounded_amount = math.ceil(float_amount)
        else:
            # This is a component price that shouldn't be rounded
            rounded_amount = float_amount
            
        return f"£{rounded_amount:,.2f}"
    except (ValueError, TypeError):
        return "£0.00"

def generate_scope_of_works(project_data: Dict) -> List[Dict]:
    """
    Generate comprehensive scope of works including all equipment types.
    
    Args:
        project_data (Dict): Project data with all equipment information
        
    Returns:
        List[Dict]: List of scope items with counts and descriptions
    """
    from config.constants import is_feature_enabled
    
    scope_items = []
    
    # 1. Count canopies by model and lighting type
    model_lighting_counts = {}
    areas_with_cladding = set()
    areas_with_sdu = set()
    areas_with_vent_clg = set()
    areas_with_marvel = set()
    areas_with_recoair = set()
    areas_with_reactaway = set()
    
    # Track fire suppression by type
    fire_suppression_by_type = {
        'NOBEL': set(),
        'AMAREX': set(),
        'ANSUL': set()
    }
    
    for level in project_data.get('levels', []):
        level_name = level.get('level_name', '')
        
        for area in level.get('areas', []):
            area_name = area.get('name', '')
            area_identifier = f"{level_name} - {area_name}"
            
            # Check for ventilated ceiling
            if area.get('options', {}).get('vent_clg', False):
                areas_with_vent_clg.add(area_identifier)
            
            # Check for MARVEL system
            if area.get('options', {}).get('marvel', False):
                areas_with_marvel.add(area_identifier)
            
            # Check for RecoAir system
            if area.get('options', {}).get('recoair', False):
                areas_with_recoair.add(area_identifier)
            
            # Check for Reactaway unit (if feature enabled)
            if is_feature_enabled('reactaway_unit') and area.get('options', {}).get('reactaway', False):
                areas_with_reactaway.add(area_identifier)
            
            area_has_cladding = False
            area_has_sdu = False
            
            for canopy in area.get('canopies', []):
                model = canopy.get('model', '').upper().strip()
                lighting_type = canopy.get('lighting_type', '').upper().strip()
                
                # Check for wall cladding
                wall_cladding = canopy.get('wall_cladding', {})
                if wall_cladding.get('type') not in ['None', None, ''] and wall_cladding.get('type'):
                    area_has_cladding = True
                
                # Check for SDU
                if canopy.get('options', {}).get('sdu', False):
                    area_has_sdu = True
                
                # Check for fire suppression and track by type
                fire_supp_qty = canopy.get('fire_suppression_tank_quantity', 0)
                fire_supp_price = canopy.get('fire_suppression_price', 0)
                if fire_supp_qty > 0 or fire_supp_price > 0:
                    # Determine fire suppression type
                    fs_type = canopy.get('fire_suppression_system_type', '').upper().strip()
                    if 'NOBEL' in fs_type:
                        fire_suppression_by_type['NOBEL'].add(area_identifier)
                    elif 'AMAREX' in fs_type:
                        fire_suppression_by_type['AMAREX'].add(area_identifier)
                    else:
                        # Default to Ansul R102 for any other type
                        fire_suppression_by_type['ANSUL'].add(area_identifier)
                
                if model:
                    # Normalize lighting type
                    if lighting_type and lighting_type not in ['-', 'NONE', 'LIGHT SELECTION', '']:
                        if 'LED STRIP' in lighting_type:
                            lighting_normalized = 'LED STRIP'
                        elif 'SPOT' in lighting_type:
                            lighting_normalized = 'LED SPOTS'
                        elif lighting_type.startswith('HCL'):
                            lighting_normalized = 'HCL DALI'
                        elif lighting_type.startswith('EL'):
                            lighting_normalized = 'EL'
                        else:
                            lighting_normalized = lighting_type
                    else:
                        lighting_normalized = None
                    
                    # Create key combining model and lighting
                    key = (model, lighting_normalized)
                    model_lighting_counts[key] = model_lighting_counts.get(key, 0) + 1
            
            # Track areas with additional equipment
            if area_has_cladding:
                areas_with_cladding.add(area_identifier)
            if area_has_sdu:
                areas_with_sdu.add(area_identifier)
    
    # 2. Generate canopy scope items
    for (model, lighting), count in sorted(model_lighting_counts.items()):
        count_str = f"{count}no"
        
        # Create concise descriptions based on model type
        if model.startswith('CMW'):
            if 'F' in model:
                base_desc = f"Water Wash Extract/Supply Canopies"
            else:
                base_desc = f"Water Wash Extract Canopies"
        elif model.startswith('UV'):
            if 'F' in model:
                base_desc = f"UV-C Extract/Supply Canopies"
            else:
                base_desc = f"UV-C Extract Canopies"
        elif model.startswith('CXW'):
            base_desc = f"Condense Extract Canopies"
        elif model.startswith('KV'):
            if 'F' in model:
                base_desc = f"Extract/Supply Canopies"
            else:
                base_desc = f"Extract Canopies"
        else:
            # Generic for other models
            if 'F' in model:
                base_desc = f"Extract/Supply Canopies ({model})"
            else:
                base_desc = f"Extract Canopies ({model})"
        
        # Add lighting if present
        if lighting:
            description = f"{count_str} {base_desc} with {lighting}"
        else:
            description = f"{count_str} {base_desc}"
        
        scope_items.append({
            'model': model,
            'lighting': lighting,
            'count': count,
            'count_str': count_str,
            'description': description
        })
    
    # 3. Add ventilated ceiling if present
    if areas_with_vent_clg:
        count = len(areas_with_vent_clg)
        count_str = f"{count}no" if count > 1 else "1no"
        description = f"{count_str} Ventilated Ceiling System{'s' if count > 1 else ''}"
        scope_items.append({
            'model': 'VENT_CLG',
            'lighting': None,
            'count': count,
            'count_str': count_str,
            'description': description
        })
    
    # 4. Add SDU systems if present
    if areas_with_sdu:
        count = len(areas_with_sdu)
        count_str = f"{count}no" if count > 1 else "1no"
        description = f"{count_str} Service Distribution Unit{'s' if count > 1 else ''} (SDU)"
        scope_items.append({
            'model': 'SDU',
            'lighting': None,
            'count': count,
            'count_str': count_str,
            'description': description
        })
    
    # 5. Add MARVEL system if present
    if areas_with_marvel:
        count = len(areas_with_marvel)
        count_str = f"{count}no" if count > 1 else "1no"
        description = f"{count_str} M.A.R.V.E.L. Demand Control System{'s' if count > 1 else ''}"
        scope_items.append({
            'model': 'MARVEL',
            'lighting': None,
            'count': count,
            'count_str': count_str,
            'description': description
        })
    
    # 6. Add fire suppression systems by type
    # NOBEL systems
    if fire_suppression_by_type['NOBEL']:
        count = len(fire_suppression_by_type['NOBEL'])
        count_str = f"{count}no" if count > 1 else "1no"
        description = f"{count_str} NOBEL Fire Suppression System{'s' if count > 1 else ''}"
        scope_items.append({
            'model': 'FIRE_NOBEL',
            'lighting': None,
            'count': count,
            'count_str': count_str,
            'description': description
        })
    
    # AMAREX systems
    if fire_suppression_by_type['AMAREX']:
        count = len(fire_suppression_by_type['AMAREX'])
        count_str = f"{count}no" if count > 1 else "1no"
        description = f"{count_str} AMAREX Fire Suppression System{'s' if count > 1 else ''}"
        scope_items.append({
            'model': 'FIRE_AMAREX',
            'lighting': None,
            'count': count,
            'count_str': count_str,
            'description': description
        })
    
    # Ansul R102 systems (default)
    if fire_suppression_by_type['ANSUL']:
        count = len(fire_suppression_by_type['ANSUL'])
        count_str = f"{count}no" if count > 1 else "1no"
        description = f"{count_str} Ansul R102 Fire Suppression System{'s' if count > 1 else ''}"
        scope_items.append({
            'model': 'FIRE_ANSUL',
            'lighting': None,
            'count': count,
            'count_str': count_str,
            'description': description
        })
    
    # 7. Add wall cladding if present
    if areas_with_cladding:
        count = len(areas_with_cladding)
        count_str = f"{count}no" if count > 1 else "1no"
        description = f"{count_str} Stainless Steel Wall Cladding Area{'s' if count > 1 else ''}"
        scope_items.append({
            'model': 'CLADDING',
            'lighting': None,
            'count': count,
            'count_str': count_str,
            'description': description,
            'areas': list(areas_with_cladding)
        })
    
    # 8. Add RecoAir systems if present
    if areas_with_recoair:
        count = len(areas_with_recoair)
        count_str = f"{count}no" if count > 1 else "1no"
        description = f"{count_str} RecoAir Air Handling Unit{'s' if count > 1 else ''}"
        scope_items.append({
            'model': 'RECOAIR',
            'lighting': None,
            'count': count,
            'count_str': count_str,
            'description': description
        })
    
    # 9. Add Reactaway units if present (and feature enabled)
    if areas_with_reactaway:
        count = len(areas_with_reactaway)
        count_str = f"{count}no" if count > 1 else "1no"
        description = f"{count_str} Reactaway UV-C Filtration Unit{'s' if count > 1 else ''}"
        scope_items.append({
            'model': 'REACTAWAY',
            'lighting': None,
            'count': count,
            'count_str': count_str,
            'description': description
        })
    
    # 10. Add extract/supply systems if defined
    if project_data.get('has_extract_system', False):
        scope_items.append({
            'model': 'EXTRACT_SYS',
            'lighting': None,
            'count': 1,
            'count_str': '1no',
            'description': 'Extract System with Ductwork and Controls'
        })
    
    if project_data.get('has_supply_system', False):
        scope_items.append({
            'model': 'SUPPLY_SYS',
            'lighting': None,
            'count': 1,
            'count_str': '1no',
            'description': 'Supply Air System with Ductwork and Controls'
        })
    
    return scope_items

def collect_sdu_data(project_data: Dict, excel_file_path: str = None, cached_wb=None) -> List[Dict]:
    """
    Collect SDU (Supply Diffusion Unit) data for canopies with SDU systems.
    
    Args:
        project_data (Dict): Project data
        excel_file_path (str, optional): Path to Excel file to extract detailed SDU data
        cached_wb (optional): Pre-loaded Excel workbook to avoid reloading
        
    Returns:
        List[Dict]: SDU data for each canopy that has SDU systems
    """
    import time
    start_time = time.time()
    print(f"📡 Starting SDU data collection...")
    
    sdu_areas = []
    
    # Process each level, area, and canopy to find SDU systems
    canopy_scan_start = time.time()
    print(f"   🔍 Scanning canopies for SDU systems...")
    
    for level in project_data.get('levels', []):
        level_name = level.get('level_name', '')
        
        for area_index, area in enumerate(level.get('areas', [])):
            area_name = area.get('name', '')
            level_area_combined = f"{level_name} - {area_name}"
            area_number = area_index + 1  # Areas are numbered starting from 1
            
            # Process each canopy in this area
            for canopy in area.get('canopies', []):
                # Check if this canopy has SDU systems
                if not canopy.get('options', {}).get('sdu', False):
                    continue
                
                canopy_ref = canopy.get('reference_number', '')
                print(f"      📡 Found SDU canopy: {canopy_ref} in {level_area_combined}")
                
                # Initialize SDU canopy data
                sdu_canopy = {
                    'level_name': level_name,
                    'area_name': area_name,
                    'area_number': area_number,  # Keep for compatibility
                    'canopy_reference': canopy_ref,
                    'level_area_combined': level_area_combined,
                    'has_sdu': True,
                    'sdu_price': canopy.get('sdu_price', 0),
                    'sdu_item_number': canopy.get('sdu_item_number', ''),  # Get SDU item number from canopy data
                    'sdu_length': 'XXXX',  # Default SDU length - to be updated from Excel if available
                    'potrack': 'xxxxxxxxxxxxx',  # Default potrack - to be updated from Excel if available
                    'salamander_support': 'xxxxxxxxxxxxxx',  # Default - to be updated from Excel if available
                    'pricing': {},  # Detailed pricing from Excel
                    'electrical_services': {  # Initialize electrical services with defaults
                        'distribution_board': 0,
                        'single_phase_switched_spur': 0,
                        'three_phase_socket_outlet': 0,
                        'switched_socket_outlet': 0,
                        'emergency_knock_off': 0,
                        'ring_main_inc_2no_sso': 0
                    },
                    'gas_services': {  # Initialize gas services
                        'gas_manifold': 0,
                        'gas_connection_15mm': 0,
                        'gas_connection_20mm': 0,
                        'gas_connection_25mm': 0,
                        'gas_connection_32mm': 0,
                        'gas_solenoid_valve': 0
                    },
                    'water_services': {  # Initialize water services
                        'cws_manifold_22mm': 0,
                        'cws_manifold_15mm': 0,
                        'hws_manifold': 0,
                        'water_connection_15mm': 0,
                        'water_connection_22mm': 0,
                        'water_connection_28mm': 0
                    }
                }
                    
                sdu_areas.append(sdu_canopy)
    
    canopy_scan_time = time.time() - canopy_scan_start
    print(f"   ✅ Canopy scan complete: {canopy_scan_time:.3f}s - Found {len(sdu_areas)} SDU canopies")
    
    # Extract detailed data from Excel if available
    if excel_file_path and os.path.exists(excel_file_path) and len(sdu_areas) > 0:
        excel_start = time.time()
        print(f"   📖 Loading Excel for detailed SDU data...")
        
        try:
            # Use cached workbook if available, otherwise load fresh
            if cached_wb:
                wb = cached_wb
                print(f"      ✅ Using cached workbook")
                wb_load_time = 0
            else:
                wb_load_start = time.time()
                from openpyxl import load_workbook
                wb = load_workbook(excel_file_path, data_only=True)
                wb_load_time = time.time() - wb_load_start
                print(f"      ✅ Excel loaded: {wb_load_time:.3f}s")
            
            # Process each SDU area
            for sdu_area in sdu_areas:
                area_process_start = time.time()
                print(f"      🔧 Processing {sdu_area['level_area_combined']}...")
                
                level_name = sdu_area['level_name']
                area_number = sdu_area['area_number']
                
                # Look for SDU sheet for this canopy using the correct naming pattern
                canopy_ref = sdu_area['canopy_reference']
                
                # Try multiple naming patterns to find the SDU sheet
                possible_sheet_names = [
                    f"SDU - {level_name} ({area_number}) - {canopy_ref}",
                    f"SDU - L{area_number} ({area_number}) - {canopy_ref}",
                    f"SDU - {canopy_ref}",
                    # Also try with lowercase canopy reference
                    f"SDU - {level_name} ({area_number}) - {canopy_ref.lower()}",
                    f"SDU - L{area_number} ({area_number}) - {canopy_ref.lower()}",
                    f"SDU - {canopy_ref.lower()}",
                ]
                
                sdu_sheet_name = None
                for possible_name in possible_sheet_names:
                    if possible_name in wb.sheetnames:
                        sdu_sheet_name = possible_name
                        break
                
                # If not found by exact match, search for sheets containing the canopy reference
                if not sdu_sheet_name:
                    for sheet_name in wb.sheetnames:
                        # Case-insensitive search for SDU sheets with canopy reference
                        if "SDU" in sheet_name.upper() and canopy_ref.upper() in sheet_name.upper():
                            sdu_sheet_name = sheet_name
                            break
                
                if sdu_sheet_name:
                    print(f"         📊 Reading SDU sheet: {sdu_sheet_name}")
                    sdu_sheet = wb[sdu_sheet_name]
                    
                    # Extract electrical, gas, and water services data
                    from utils.excel import extract_sdu_electrical_services
                    services_data = extract_sdu_electrical_services(sdu_sheet)
                    sdu_area['electrical_services'] = services_data.get('electrical_services', {})
                    
                    # Update SDU item number from Excel if available
                    excel_sdu_item_number = services_data.get('sdu_item_number', '')
                    if excel_sdu_item_number:
                        sdu_area['sdu_item_number'] = excel_sdu_item_number
                    sdu_area['gas_services'] = services_data.get('gas_services', {
                        'gas_manifold': 0,
                        'gas_connection_15mm': 0,
                        'gas_connection_20mm': 0,
                        'gas_connection_25mm': 0,
                        'gas_connection_32mm': 0,
                        'gas_solenoid_valve': 0
                    })
                    sdu_area['water_services'] = services_data.get('water_services', {
                        'cws_manifold_22mm': 0,
                        'cws_manifold_15mm': 0,
                        'hws_manifold': 0,
                        'water_connection_15mm': 0,
                        'water_connection_22mm': 0,
                        'water_connection_28mm': 0
                    })
                    
                    # Extract pricing data using the services data structure
                    # The extract_sdu_electrical_services function already extracted pricing data
                    extracted_pricing = services_data.get('pricing', {})
                    
                    # Update pricing data with the extracted values
                    sdu_area['pricing'] = {
                        'final_carcass_price': extracted_pricing.get('final_carcass_price', 0),
                        'final_electrical_price': extracted_pricing.get('final_electrical_price', 0),
                        'live_site_test_price': extracted_pricing.get('live_site_test_price', 0),
                        'has_live_test': extracted_pricing.get('has_live_test', False),
                        'total_price': (
                            extracted_pricing.get('final_carcass_price', 0) +
                            extracted_pricing.get('final_electrical_price', 0) +
                            (extracted_pricing.get('live_site_test_price', 0) if extracted_pricing.get('has_live_test', False) else 0)
                        )
                    }
                    
                    print(f"         ✅ Pricing extracted: Total ${sdu_area['pricing']['total_price']}")
                else:
                    print(f"         ⚠️  No SDU sheet found: {sdu_sheet_name}")
                
                area_process_time = time.time() - area_process_start
                print(f"      ✅ {sdu_area['level_area_combined']} processed: {area_process_time:.3f}s")
                
        except Exception as e:
            excel_error_time = time.time() - excel_start
            print(f"      ❌ Excel processing error after {excel_error_time:.3f}s: {str(e)}")
        
        excel_time = time.time() - excel_start
        print(f"   ✅ Excel processing complete: {excel_time:.3f}s")
    else:
        if len(sdu_areas) == 0:
            print(f"   ℹ️  No SDU canopies found - skipping Excel processing")
        else:
            print(f"   ℹ️  No Excel file provided - using basic SDU data only")
    
    total_time = time.time() - start_time
    print(f"📡 SDU data collection COMPLETE: {total_time:.3f}s")
    print(f"   📊 Breakdown:")
    print(f"      - Canopy scanning: {canopy_scan_time:.3f}s")
    if excel_file_path and len(sdu_areas) > 0:
        excel_time = time.time() - (excel_start if 'excel_start' in locals() else start_time)
        print(f"      - Excel processing: {excel_time:.3f}s")
    print(f"   ✅ Collected data for {len(sdu_areas)} SDU canopies")
    
    return sdu_areas 