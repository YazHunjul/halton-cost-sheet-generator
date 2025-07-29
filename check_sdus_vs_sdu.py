#!/usr/bin/env python3
"""Check if SDUS is a typo or a specific term."""

import sys
import os
sys.path.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'src'))

from openpyxl import load_workbook

def check_sdus_in_excel(excel_path):
    """Check for SDUS vs SDU in Excel."""
    print(f"\n📊 Checking for SDUS vs SDU in: {excel_path}")
    print("=" * 80)
    
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    
    # Search all sheets for "SDUS" text
    sdus_found = False
    sdu_found = False
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # Check common cells and a sample of the sheet
        for row in range(1, min(sheet.max_row + 1, 200)):  # Check first 200 rows
            for col in range(1, min(sheet.max_column + 1, 20)):  # Check first 20 columns
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    if "SDUS" in cell_value.upper():
                        print(f"\n🔍 Found 'SDUS' in sheet '{sheet_name}' at cell {chr(64+col)}{row}: {cell_value}")
                        sdus_found = True
                    elif "SDU" in cell_value.upper() and "SDUS" not in cell_value.upper():
                        if sdu_found == False:  # Only print first few SDU occurrences
                            print(f"\n✅ Found 'SDU' (not SDUS) in sheet '{sheet_name}' at cell {chr(64+col)}{row}: {cell_value}")
                            sdu_found = True
    
    if not sdus_found:
        print("\n❌ No 'SDUS' found in the Excel file")
        print("✅ Only 'SDU' (without the extra 'S') was found")
    else:
        print("\n⚠️  'SDUS' was found - this might be a typo for 'SDU'")

if __name__ == "__main__":
    excel_path = "/Users/yazan/Downloads/36108 Cost Sheet 28072025.xlsx"
    
    if os.path.exists(excel_path):
        check_sdus_in_excel(excel_path)
    else:
        print(f"❌ Excel file not found: {excel_path}")