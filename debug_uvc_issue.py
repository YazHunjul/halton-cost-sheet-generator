#!/usr/bin/env python3
"""
Debug script to help identify why UV-C table is not showing when there's an EBOX sheet.
"""

import sys
import os
sys.path.append('src')

from utils.excel import read_excel_project_data
from utils.word import prepare_template_context, calculate_pricing_totals
import json

def debug_uvc_issue(excel_file_path: str):
    """
    Debug UV-C table display issue by examining the data at each step.
    
    Args:
        excel_file_path (str): Path to the Excel file with EBOX sheet
    """
    print("🔍 Debugging UV-C Table Display Issue")
    print("=" * 50)
    
    try:
        # Step 1: Read project data from Excel
        print("\n📊 Step 1: Reading project data from Excel...")
        project_data = read_excel_project_data(excel_file_path)
        
        # Step 2: Check for EBOX sheets in the Excel file
        print("\n📋 Step 2: Checking for EBOX sheets...")
        from openpyxl import load_workbook
        wb = load_workbook(excel_file_path, data_only=True)
        
        ebox_sheets = [sheet for sheet in wb.sheetnames if 'EBOX' in sheet]
        print(f"Found EBOX sheets: {ebox_sheets}")
        
        for sheet_name in ebox_sheets:
            sheet = wb[sheet_name]
            title_cell = sheet['C1'].value
            uvc_price_cell = sheet['N9'].value
            print(f"  - {sheet_name}:")
            print(f"    Title (C1): {title_cell}")
            print(f"    UV-C Price (N9): {uvc_price_cell}")
        
        # Step 3: Check project data structure
        print("\n🏗️ Step 3: Checking project data structure...")
        for level in project_data.get('levels', []):
            level_name = level.get('level_name', '')
            print(f"\nLevel: {level_name}")
            
            for area in level.get('areas', []):
                area_name = area.get('name', '')
                print(f"  Area: {area_name}")
                print(f"    Options: {area.get('options', {})}")
                print(f"    UV-C Price: {area.get('uvc_price', 0)}")
                print(f"    Has UV-C option: {area.get('options', {}).get('uvc', False)}")
        
        # Step 4: Check template context preparation
        print("\n🎨 Step 4: Checking template context preparation...")
        context = prepare_template_context(project_data, excel_file_path)
        
        print("\nTemplate context levels:")
        for level in context.get('levels', []):
            level_name = level.get('level_name', '')
            print(f"\nLevel: {level_name}")
            
            for area in level.get('areas', []):
                area_name = area.get('name', '')
                print(f"  Area: {area_name}")
                print(f"    Options: {area.get('options', {})}")
                print(f"    UV-C Price: {area.get('uvc_price', 0)}")
                print(f"    Has UV-C option: {area.get('options', {}).get('uvc', False)}")
        
        # Step 5: Check pricing totals
        print("\n💰 Step 5: Checking pricing totals...")
        pricing_totals = calculate_pricing_totals(project_data, excel_file_path)
        
        print(f"Total UV-C Price: {pricing_totals.get('total_uvc_price', 0)}")
        
        print("\nArea pricing totals:")
        for area in pricing_totals.get('areas', []):
            area_name = area.get('level_area_combined', '')
            print(f"  {area_name}:")
            print(f"    UV-C Price: {area.get('uvc_price', 0)}")
        
        # Step 6: Template condition check
        print("\n✅ Step 6: Template condition check...")
        print("\nFor UV-C table to show, one of these conditions must be true:")
        
        for level in context.get('levels', []):
            for area in level.get('areas', []):
                area_name = f"{level.get('level_name', '')} - {area.get('name', '')}"
                uvc_option = area.get('options', {}).get('uvc', False)
                uvc_price = area.get('uvc_price', 0)
                
                print(f"\n  {area_name}:")
                print(f"    {{% if area.options.uvc %}} = {uvc_option}")
                print(f"    {{% if area.uvc_price > 0 %}} = {uvc_price > 0} (price: {uvc_price})")
                
                if uvc_option or uvc_price > 0:
                    print(f"    ✅ UV-C table SHOULD show for this area")
                else:
                    print(f"    ❌ UV-C table will NOT show for this area")
                    print(f"    💡 Possible issues:")
                    print(f"       - UV-C price in N9 is 0 or empty")
                    print(f"       - EBOX sheet title doesn't contain 'UV-C SYSTEM'")
                    print(f"       - Area name mismatch between EBOX sheet and project data")
        
        # Step 7: Recommendations
        print("\n🔧 Step 7: Recommendations...")
        print("\nTo fix UV-C table not showing:")
        print("1. Check that EBOX sheet title (C1) contains 'UV-C SYSTEM'")
        print("2. Check that UV-C price in N9 is not 0 or empty")
        print("3. Check that area name in EBOX sheet matches project data exactly")
        print("4. Verify template uses: {{% if area.options.uvc or area.uvc_price > 0 %}}")
        
    except Exception as e:
        print(f"❌ Error during debugging: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python debug_uvc_issue.py <excel_file_path>")
        print("Example: python debug_uvc_issue.py 'path/to/your/costsheet.xlsx'")
        sys.exit(1)
    
    excel_file_path = sys.argv[1]
    
    if not os.path.exists(excel_file_path):
        print(f"❌ File not found: {excel_file_path}")
        sys.exit(1)
    
    debug_uvc_issue(excel_file_path) 