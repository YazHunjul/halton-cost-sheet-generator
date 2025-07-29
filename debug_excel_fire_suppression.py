#!/usr/bin/env python3
"""Debug script to analyze fire suppression data from Excel file"""

import openpyxl
from pathlib import Path
import json

def analyze_fire_suppression(excel_path):
    """Analyze fire suppression data from Excel file"""
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    print(f"\n=== Analyzing Excel File: {excel_path} ===\n")
    print(f"Available sheets: {wb.sheetnames}\n")
    
    # Look for ALL FIRE SUPP sheets
    fire_supp_sheets = []
    for sheet_name in wb.sheetnames:
        if 'FIRE SUPP' in sheet_name.upper():
            fire_supp_sheets.append(sheet_name)
    
    if not fire_supp_sheets:
        print("❌ No FIRE SUPP sheets found!")
        return
    
    print(f"✅ Found {len(fire_supp_sheets)} FIRE SUPP sheets: {fire_supp_sheets}")
    
    all_fs_data = []
    
    # Process each FIRE SUPP sheet
    for fire_supp_sheet in fire_supp_sheets:
        print(f"\n=== Processing: {fire_supp_sheet} ===")
        sheet = wb[fire_supp_sheet]
        
        # Read fire suppression data
        print("\n=== Fire Suppression Units ===")
        
        # Check multiple possible rows for fire suppression data
        base_rows = [14, 31, 48, 65, 82, 99]  # Standard canopy rows
        
        fs_units = []
        for base_row in base_rows:
            ref_row = base_row - 2  # B12, B29, B46, etc.
            system_row = base_row + 2  # C16, C33, C50, etc.
            tank_row = base_row + 3  # C17, C34, C51, etc.
            price_row = base_row - 2  # N12, N29, N46, etc.
            
            # Read values
            ref_number = sheet[f'B{ref_row}'].value
            system_type = sheet[f'C{system_row}'].value
            tank_value = sheet[f'C{tank_row}'].value
            base_price = sheet[f'N{price_row}'].value
            
            if ref_number:
                print(f"\nRow {ref_row}:")
                print(f"  Reference: {ref_number}")
                print(f"  System Type (C{system_row}): {system_type}")
                print(f"  Tank Value (C{tank_row}): {tank_value}")
                print(f"  Base Price (N{price_row}): {base_price}")
                
                # Parse tank quantity
                tank_quantity = 0
                if tank_value:
                    tank_str = str(tank_value).upper().strip()
                    if tank_str and tank_str != "-":
                        # Extract number from strings like "1 TANK", "2 TANK", etc.
                        import re
                        numbers = re.findall(r'\d+', tank_str)
                        if numbers:
                            tank_quantity = int(numbers[0])
                
                print(f"  Parsed Tank Quantity: {tank_quantity}")
                
                if tank_quantity > 0 or base_price:
                    fs_units.append({
                        'ref_number': ref_number,
                        'system_type': system_type,
                        'tank_quantity': tank_quantity,
                        'base_price': base_price or 0
                    })
    
        # Check delivery price
        delivery_price = sheet['N182'].value or 0
        print(f"\n=== Delivery Price (N182): {delivery_price} ===")
        
        print(f"\n=== Summary for {fire_supp_sheet} ===")
        print(f"Fire suppression units found: {len(fs_units)}")
        for unit in fs_units:
            print(f"  - {unit['ref_number']}: {unit['tank_quantity']} tanks, {unit['system_type']} system, £{unit['base_price']}")
        
        all_fs_data.append({
            'sheet_name': fire_supp_sheet,
            'delivery_price': delivery_price,
            'fs_units': fs_units
        })
    
    # Now let's check the canopy sheets to see which ones should have fire suppression
    print("\n=== Checking Canopy Sheets ===")
    
    # Check CANOPY sheets
    canopy_sheets = [s for s in wb.sheetnames if s.startswith('CANOPY')]
    
    for sheet_name in canopy_sheets:
        canopy_sheet = wb[sheet_name]
        print(f"\n{sheet_name}:")
        
        # Get area name from C8
        area_name = canopy_sheet['C8'].value or "Unknown Area"
        print(f"  Area: {area_name}")
        
        # Check canopy references
        base_rows = [14, 31, 48, 65, 82, 99]
        for base_row in base_rows:
            ref_row = base_row - 2  # B12, B29, B46, etc.
            ref_number = canopy_sheet[f'B{ref_row}'].value
            if ref_number and ref_number != 'ITEM':
                print(f"  Canopy Reference B{ref_row}: {ref_number}")
                
                # Check fire suppression option
                for row in range(base_row + 20, base_row + 26):  # Look around row 34-39 relative to base
                    if canopy_sheet[f'I{row}'].value == 'Fire suppression':
                        fs_option_value = canopy_sheet[f'K{row}'].value
                        print(f"    Fire Suppression Option K{row}: {fs_option_value}")
                        break
    
    print(f"\n=== Overall Summary ===")
    print(f"Total FIRE SUPP sheets: {len(fire_supp_sheets)}")
    for fs_data in all_fs_data:
        print(f"\n{fs_data['sheet_name']}:")
        print(f"  Units: {len(fs_data['fs_units'])}")
        for unit in fs_data['fs_units']:
            print(f"    - {unit['ref_number']}: {unit['tank_quantity']} tanks, {unit['system_type']} system")
    
    # Save to JSON for further analysis
    output_data = {
        'file': str(excel_path),
        'fire_supp_sheets': fire_supp_sheets,
        'all_data': all_fs_data
    }
    
    output_path = Path('debug_fire_suppression_data.json')
    with open(output_path, 'w') as f:
        json.dump(output_data, f, indent=2)
    
    print(f"\n✅ Debug data saved to: {output_path}")
    
    wb.close()

if __name__ == "__main__":
    excel_file = "/Users/yazan/Downloads/36108 Cost Sheet 28072025 (1).xlsx"
    analyze_fire_suppression(excel_file)