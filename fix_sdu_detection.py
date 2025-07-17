#!/usr/bin/env python3
"""
Fix SDU detection by updating canopy options based on SDU sheets in the Excel file.
"""

import os
import sys
from openpyxl import load_workbook

# Add the src directory to Python path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src'))

from utils.excel import read_excel_project_data

def fix_sdu_detection_in_project_data(project_data, excel_file_path):
    """
    Fix SDU detection by checking SDU sheets and updating canopy options accordingly.
    """
    print("\n🔧 Fixing SDU detection in project data...")
    
    # Load the workbook to check for SDU sheets
    wb = load_workbook(excel_file_path, data_only=True)
    sdu_sheets = [sheet for sheet in wb.sheetnames if sheet.startswith('SDU - ')]
    
    print(f"Found {len(sdu_sheets)} SDU sheets:")
    for sheet in sdu_sheets:
        print(f"  - {sheet}")
    
    # Parse SDU sheet names to identify which canopies have SDU
    sdu_canopy_refs = []
    for sheet_name in sdu_sheets:
        # Expected format: "SDU - Level Name (Area#) - CanopyRef"
        parts = sheet_name.split(' - ')
        if len(parts) >= 3:
            canopy_ref = parts[-1].strip()
            sdu_canopy_refs.append(canopy_ref)
            print(f"  → Canopy {canopy_ref} has SDU")
    
    # Update canopy options in project data
    updates_made = 0
    for level in project_data.get('levels', []):
        for area in level.get('areas', []):
            for canopy in area.get('canopies', []):
                canopy_ref = canopy.get('reference_number', '')
                
                if canopy_ref in sdu_canopy_refs:
                    # Ensure options dict exists
                    if 'options' not in canopy:
                        canopy['options'] = {}
                    
                    # Set SDU option to True
                    if not canopy['options'].get('sdu', False):
                        canopy['options']['sdu'] = True
                        updates_made += 1
                        print(f"✅ Updated canopy {canopy_ref}: SDU option set to True")
    
    print(f"\n✅ Fixed {updates_made} canopies with SDU option")
    return project_data

def test_fixed_sdu_collection(excel_file_path):
    """Test SDU collection after fixing the detection."""
    from utils.word import collect_sdu_data, prepare_template_context
    
    print(f"\n{'='*80}")
    print("Testing SDU Collection After Fix")
    print(f"{'='*80}\n")
    
    # Read project data
    print("Reading project data...")
    project_data = read_excel_project_data(excel_file_path)
    
    # Fix SDU detection
    project_data = fix_sdu_detection_in_project_data(project_data, excel_file_path)
    
    # Test SDU collection
    print("\n📡 Testing SDU collection after fix...")
    sdu_data = collect_sdu_data(project_data, excel_file_path)
    print(f"✅ collect_sdu_data returned {len(sdu_data)} SDU items")
    
    if sdu_data:
        for idx, sdu in enumerate(sdu_data):
            print(f"\nSDU {idx + 1}:")
            print(f"  - Canopy Reference: {sdu.get('canopy_reference')}")
            print(f"  - Level/Area: {sdu.get('level_area_combined')}")
            print(f"  - Electrical Services: {sum(sdu.get('electrical_services', {}).values())} items")
            print(f"  - Gas Services: {sum(sdu.get('gas_services', {}).values())} items")
            print(f"  - Water Services: {sum(sdu.get('water_services', {}).values())} items")
    
    # Test template context
    print("\n📋 Testing template context...")
    context = prepare_template_context(project_data, excel_file_path)
    has_sdu = context.get('has_sdu', False)
    sdu_areas = context.get('sdu_areas', [])
    
    print(f"✅ Template context:")
    print(f"  - has_sdu: {has_sdu}")
    print(f"  - sdu_areas length: {len(sdu_areas)}")
    
    if has_sdu:
        print("\n✅ SUCCESS: SDU data is now available for Word template!")
    else:
        print("\n❌ FAILED: SDU data still not available")
    
    return project_data, context

def main():
    """Main function."""
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    else:
        # Look for recent Excel files
        output_dir = "output"
        if os.path.exists(output_dir):
            excel_files = [f for f in os.listdir(output_dir) if f.endswith('.xlsx')]
            if excel_files:
                excel_files.sort(key=lambda x: os.path.getmtime(os.path.join(output_dir, x)), reverse=True)
                excel_file = os.path.join(output_dir, excel_files[0])
                print(f"Using most recent Excel file: {excel_file}")
            else:
                print("No Excel files found.")
                return
        else:
            print("Output directory not found.")
            return
    
    # Run the fix and test
    project_data, context = test_fixed_sdu_collection(excel_file)
    
    # Save the fixed project data if needed
    if context.get('has_sdu', False):
        print("\n💡 To use this fix in your application:")
        print("   Add the fix_sdu_detection_in_project_data function after reading Excel data")

if __name__ == "__main__":
    main()