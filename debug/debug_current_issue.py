#!/usr/bin/env python3
"""
Debug script to check the current state of Excel file with RecoAir units selected.
"""

import sys
import os
sys.path.insert(0, 'src')

from utils.excel import read_excel_project_data
from openpyxl import load_workbook

def debug_current_issue(excel_path: str):
    """Debug the current RecoAir issue."""
    print(f"🔍 Debugging current RecoAir issue in: {excel_path}")
    print("=" * 60)
    
    try:
        # First, check the raw Excel sheets
        print("📋 Step 1: Checking raw Excel RECOAIR sheets...")
        wb = load_workbook(excel_path, data_only=True)
        
        recoair_sheets = [name for name in wb.sheetnames if 'RECOAIR - ' in name]
        print(f"Found {len(recoair_sheets)} RECOAIR sheets:")
        
        for sheet_name in recoair_sheets:
            sheet = wb[sheet_name]
            title = sheet['C1'].value
            item_ref = sheet['C12'].value
            
            print(f"\n📄 {sheet_name}:")
            print(f"  Title (C1): {title}")
            print(f"  Item Ref (C12): {item_ref}")
            
            # Check for selected units
            units_found = 0
            for row in range(14, 29):
                selection = sheet[f'E{row}'].value
                if selection and str(selection).strip() != "":
                    try:
                        qty = float(str(selection).strip())
                        if qty >= 1:
                            model = sheet[f'C{row}'].value or ""
                            price = sheet[f'N{row}'].value or 0
                            print(f"    ✅ Row {row}: {model} (Qty: {qty}) - £{price:,.2f}")
                            units_found += 1
                    except (ValueError, TypeError):
                        continue
            
            if units_found == 0:
                print(f"    ❌ No units selected")
            else:
                print(f"    ✅ {units_found} units selected")
        
        # Now check how the data is being read
        print(f"\n📋 Step 2: Checking how data is being read...")
        project_data = read_excel_project_data(excel_path)
        
        levels = project_data.get('levels', [])
        print(f"Total levels found: {len(levels)}")
        
        for i, level in enumerate(levels, 1):
            level_name = level.get('level_name', f'Level {i}')
            areas = level.get('areas', [])
            
            print(f"\n🏢 Level {i}: {level_name}")
            print(f"  Areas: {len(areas)}")
            
            for j, area in enumerate(areas, 1):
                area_name = area.get('name', f'Area {j}')
                options = area.get('options', {})
                recoair_units = area.get('recoair_units', [])
                recoair_price = area.get('recoair_price', 0)
                
                print(f"    📍 Area {j}: {area_name}")
                print(f"      Options: {options}")
                print(f"      RecoAir units: {len(recoair_units)}")
                print(f"      RecoAir price: £{recoair_price:,.2f}")
                
                if recoair_units:
                    for k, unit in enumerate(recoair_units, 1):
                        model = unit.get('model', 'N/A')
                        ref = unit.get('item_reference', 'N/A')
                        price = unit.get('unit_price', 0)
                        print(f"        Unit {k}: {model} (Ref: {ref}) - £{price:,.2f}")
        
        # Test Word generation logic
        print(f"\n📋 Step 3: Testing Word generation logic...")
        from utils.word import collect_recoair_pricing_schedule_data
        
        recoair_data = collect_recoair_pricing_schedule_data(project_data)
        areas_found = len(recoair_data['areas'])
        
        print(f"RecoAir pricing areas found: {areas_found}")
        for i, area in enumerate(recoair_data['areas'], 1):
            level_name = area['level_name']
            area_name = area['area_name']
            unit_count = area['unit_count']
            
            print(f"  Area {i}: {level_name} - {area_name} ({unit_count} units)")
        
        print(f"\nJob totals:")
        job_totals = recoair_data['job_totals']
        print(f"  Total areas: {job_totals['total_areas']}")
        print(f"  Total units: {job_totals['total_units']}")
        print(f"  Job total: £{job_totals['job_total']:,.2f}")
        
        print("\n" + "=" * 60)
        print("✅ Debug complete!")
        
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python debug_current_issue.py <excel_file_path>")
        print("Please provide the path to your Excel file with RecoAir units selected")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    if not os.path.exists(excel_path):
        print(f"❌ File not found: {excel_path}")
        sys.exit(1)
    
    debug_current_issue(excel_path) 