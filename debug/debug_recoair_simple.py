#!/usr/bin/env python3
"""Debug RecoAir pricing without importing word module."""

import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from openpyxl import load_workbook
from utils.excel import read_excel_project_data

excel_path = '/Users/yazan/Downloads/36068 Cost Sheet 25032025 (1).xlsx'

print("=" * 80)
print("SIMPLE RECOAIR PRICING DEBUG")
print("=" * 80)

# 1. Check Excel totals directly
wb = load_workbook(excel_path, data_only=True)

print("\n1. EXCEL JOB TOTAL:")
print("-" * 40)
if 'JOB TOTAL' in wb.sheetnames:
    job_total_sheet = wb['JOB TOTAL']
    t28_value = job_total_sheet['T28'].value
    print(f"  T28 (Excel total): £{t28_value:,.2f}" if t28_value else "  T28: None")

# 2. Manual calculation from RecoAir sheets
print("\n2. MANUAL CALCULATION FROM RECOAIR SHEETS:")
print("-" * 40)

manual_total_excluding_flat_pack = 0
manual_total_including_flat_pack = 0
total_flat_pack = 0

for sheet_name in wb.sheetnames:
    if 'RECOAIR' in sheet_name:
        sheet = wb[sheet_name]
        print(f"\n  {sheet_name}:")
        
        # Base unit price (N12)
        n12 = sheet['N12'].value or 0
        print(f"    N12 (Base unit price): £{n12:,.2f}")
        
        # Flat pack (N40)
        n40 = sheet['N40'].value or 0
        print(f"    N40 (Flat pack): £{n40:,.2f}")
        total_flat_pack += n40
        
        # Area total without flat pack (just N12)
        area_total_no_flat_pack = n12
        manual_total_excluding_flat_pack += area_total_no_flat_pack
        
        # Area total with flat pack
        area_total_with_flat_pack = n12 + n40
        manual_total_including_flat_pack += area_total_with_flat_pack
        
        print(f"    Area total (excluding flat pack): £{area_total_no_flat_pack:,.2f}")
        print(f"    Area total (including flat pack): £{area_total_with_flat_pack:,.2f}")

print(f"\n  MANUAL TOTALS:")
print(f"    Total (excluding flat pack): £{manual_total_excluding_flat_pack:,.2f}")
print(f"    Total (including flat pack): £{manual_total_including_flat_pack:,.2f}")
print(f"    Total flat pack: £{total_flat_pack:,.2f}")

# 3. Read project data and check what we get
print("\n3. PROJECT DATA FROM UTILS:")
print("-" * 40)

project_data = read_excel_project_data(excel_path)

total_recoair_from_utils = 0
total_flat_pack_from_utils = 0

for level in project_data.get('levels', []):
    for area in level.get('areas', []):
        recoair_price = area.get('recoair_price', 0)
        flat_pack_data = area.get('recoair_flat_pack', {})
        flat_pack_price = flat_pack_data.get('price', 0) if flat_pack_data.get('has_flat_pack', False) else 0
        
        if recoair_price > 0 or flat_pack_price > 0:
            print(f"  Area: {area.get('name')}")
            print(f"    RecoAir Price: £{recoair_price:,.2f}")
            if flat_pack_price > 0:
                print(f"    Flat Pack Price: £{flat_pack_price:,.2f}")
            
            total_recoair_from_utils += recoair_price
            total_flat_pack_from_utils += flat_pack_price

print(f"\n  UTILS TOTALS:")
print(f"    RecoAir total: £{total_recoair_from_utils:,.2f}")
print(f"    Flat pack total: £{total_flat_pack_from_utils:,.2f}")
print(f"    Combined total: £{total_recoair_from_utils + total_flat_pack_from_utils:,.2f}")

# 4. Final comparison
print("\n4. COMPARISON:")
print("-" * 40)
excel_total = 148355.41
quote_total = 151087.00  # From the screenshot

print(f"  Excel T28: £{excel_total:,.2f}")
print(f"  Quote shows: £{quote_total:,.2f}")
print(f"  Manual (excl flat pack): £{manual_total_excluding_flat_pack:,.2f}")
print(f"  Manual (incl flat pack): £{manual_total_including_flat_pack:,.2f}")
print(f"  Utils (excl flat pack): £{total_recoair_from_utils:,.2f}")
print(f"  Utils (incl flat pack): £{total_recoair_from_utils + total_flat_pack_from_utils:,.2f}")

print(f"\n  Differences:")
print(f"    Quote vs Excel: £{quote_total - excel_total:,.2f}")
print(f"    Manual vs Excel: £{manual_total_excluding_flat_pack - excel_total:,.2f}")
print(f"    Utils vs Excel: £{total_recoair_from_utils - excel_total:,.2f}")

print("\n" + "=" * 80)