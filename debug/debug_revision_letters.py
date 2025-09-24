#!/usr/bin/env python3
"""
Debug script to test revision letter placement in different Excel sheet types.
"""

import os
import sys
sys.path.insert(0, 'src')

from openpyxl import load_workbook
from utils.excel import save_to_excel

def test_revision_letters():
    """Test revision letter placement in different sheet types."""
    
    # Test project data with revision
    project_data = {
        'project_number': 'TEST001',
        'company': 'Test Company',
        'customer': 'Test Customer',
        'estimator': 'Test Estimator',
        'project_name': 'Test Project',
        'project_location': 'Test Location',
        'date': '15/01/2025',
        'revision': 'B',  # Test with revision B
        'sales_contact': 'Test Sales',
        'delivery_location': 'Site',
        'contract_option': True,  # Enable contract sheets
        'levels': [
            {
                'level_number': 1,
                'level_name': 'Level 1',
                'areas': [
                    {
                        'name': 'Kitchen',
                        'canopies': [
                            {
                                'reference': '1.01',
                                'model': 'Test Model',
                                'length': 2000,
                                'width': 1000,
                                'height': 600,
                                'configuration': 'Standard',
                                'options': {}
                            }
                        ],
                        'options': {
                            'recoair': True,
                            'uv_grease_recovery': True,  # This should create MARVEL sheet
                            'vent_clg': True
                        }
                    }
                ]
            }
        ]
    }
    
    print("🔍 Testing revision letter placement in different sheet types...")
    print(f"   Test revision: {project_data['revision']}")
    
    # Generate Excel file
    try:
        output_path = save_to_excel(project_data)
        print(f"✅ Excel file generated: {output_path}")
        
        # Load the generated file and check revision placement
        wb = load_workbook(output_path)
        
        print("\n📊 Checking revision letter placement:")
        
        # Check each sheet type
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Skip hidden sheets
            if sheet.sheet_state != 'visible':
                continue
                
            print(f"\n📋 Sheet: {sheet_name}")
            
            # Check both K7 and O7 for revision
            k7_value = sheet['K7'].value
            o7_value = sheet['O7'].value
            
            print(f"   K7: {k7_value}")
            print(f"   O7: {o7_value}")
            
            # Determine expected revision cell based on sheet type
            expected_cell = None
            if 'CANOPY' in sheet_name:
                expected_cell = 'O7'  # User says should be O7
            elif 'RECOAIR' in sheet_name:
                expected_cell = 'O7'  # User says should be O7
            elif 'MARVEL' in sheet_name:
                expected_cell = 'K7'  # User says should be K7
            elif 'VENT CLG' in sheet_name:
                expected_cell = 'K7'  # User says should be K7
            elif sheet_name in ['CONTRACT', 'SPIRAL DUCT', 'SUPPLY DUCT', 'EXTRACT DUCT']:
                expected_cell = 'K7'  # User says should be K7
            elif sheet_name == 'JOB TOTAL':
                expected_cell = 'K7'  # Standard location
            
            if expected_cell:
                expected_value = sheet[expected_cell].value
                if expected_value == project_data['revision']:
                    print(f"   ✅ Revision correctly placed in {expected_cell}: {expected_value}")
                else:
                    print(f"   ❌ Revision missing in {expected_cell}. Expected: {project_data['revision']}, Got: {expected_value}")
            else:
                print(f"   ℹ️  Unknown sheet type, skipping revision check")
        
        return output_path
        
    except Exception as e:
        print(f"❌ Error generating Excel file: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    test_revision_letters() 