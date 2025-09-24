#!/usr/bin/env python3

"""
Debug the UV_EXTRA_OVER_CALC sheet values and formulas
"""

import sys
import os
sys.path.append('src')

from openpyxl import load_workbook

def debug_uv_sheet_values():
    print("🔍 Debugging UV_EXTRA_OVER_CALC sheet values...")
    
    # Find the most recent Excel file
    output_dir = "output"
    if not os.path.exists(output_dir):
        print("❌ Output directory not found")
        return
    
    excel_files = [f for f in os.listdir(output_dir) if f.endswith('.xlsx')]
    if not excel_files:
        print("❌ No Excel files found in output directory")
        return
    
    latest_file = max(excel_files, key=lambda x: os.path.getctime(os.path.join(output_dir, x)))
    excel_path = os.path.join(output_dir, latest_file)
    
    print(f"📂 Examining file: {latest_file}")
    
    try:
        # Load workbook with formulas (data_only=False)
        wb_formulas = load_workbook(excel_path, data_only=False)
        # Load workbook with values (data_only=True) 
        wb_values = load_workbook(excel_path, data_only=True)
        
        print(f"\n📋 All sheets in workbook:")
        for sheet_name in wb_formulas.sheetnames:
            print(f"  - {sheet_name}")
        
        if 'UV_EXTRA_OVER_CALC' in wb_formulas.sheetnames:
            print(f"\n🔍 UV_EXTRA_OVER_CALC sheet analysis:")
            sheet_formulas = wb_formulas['UV_EXTRA_OVER_CALC']
            sheet_values = wb_values['UV_EXTRA_OVER_CALC']
            
            print(f"  Max row: {sheet_formulas.max_row}")
            print(f"  Max col: {sheet_formulas.max_column}")
            
            # Check all cells in first few rows
            for row in range(1, min(sheet_formulas.max_row + 1, 6)):
                print(f"\n  Row {row}:")
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                    cell_formula = sheet_formulas[f'{col}{row}']
                    cell_value = sheet_values[f'{col}{row}']
                    
                    formula_val = cell_formula.value
                    calculated_val = cell_value.value
                    
                    if formula_val or calculated_val:
                        print(f"    {col}{row}: Formula='{formula_val}' | Value='{calculated_val}'")
        
        # Check UV and Non-UV sheet values
        uv_sheets = [s for s in wb_formulas.sheetnames if 'CANOPY (UV)' in s]
        canopy_sheets = [s for s in wb_formulas.sheetnames if 'CANOPY - ' in s and 'CANOPY (UV)' not in s]
        
        print(f"\n🔍 Sheet analysis:")
        print(f"  UV sheets: {uv_sheets}")
        print(f"  Non-UV canopy sheets: {canopy_sheets}")
        
        # Check specific pricing cells in UV and non-UV sheets
        for uv_sheet_name in uv_sheets:
            if uv_sheet_name in wb_values.sheetnames:
                uv_sheet = wb_values[uv_sheet_name]
                print(f"\n📊 {uv_sheet_name}:")
                
                # Check pricing cells
                for cell_ref in ['N9', 'K9', 'N197', 'K197']:
                    try:
                        cell_val = uv_sheet[cell_ref].value
                        print(f"    {cell_ref}: {cell_val}")
                    except:
                        print(f"    {cell_ref}: Error reading")
        
        for canopy_sheet_name in canopy_sheets:
            if canopy_sheet_name in wb_values.sheetnames:
                canopy_sheet = wb_values[canopy_sheet_name]
                print(f"\n📊 {canopy_sheet_name}:")
                
                # Check pricing cells  
                for cell_ref in ['N9', 'K9', 'N197', 'K197']:
                    try:
                        cell_val = canopy_sheet[cell_ref].value
                        print(f"    {cell_ref}: {cell_val}")
                    except:
                        print(f"    {cell_ref}: Error reading")
                        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    debug_uv_sheet_values() 