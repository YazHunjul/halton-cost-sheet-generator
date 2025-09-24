#!/usr/bin/env python3
"""Debug RecoAir unit selection detection."""

import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from openpyxl import load_workbook

excel_path = '/Users/yazan/Downloads/36068 Cost Sheet 25032025 (1).xlsx'

print("=" * 80)
print("DEBUG RECOAIR UNIT SELECTION DETECTION")
print("=" * 80)

# Load the workbook
wb = load_workbook(excel_path, data_only=True)

for sheet_name in wb.sheetnames:
    if 'RECOAIR' in sheet_name:
        sheet = wb[sheet_name]
        print(f"\n{sheet_name}:")
        print("-" * 40)
        
        print("Checking rows 14-28 for unit selections:")
        
        for row in range(14, 29):
            # Check selection column E
            e_cell = f'E{row}'
            c_cell = f'C{row}'
            
            e_value = sheet[e_cell].value
            c_value = sheet[c_cell].value
            
            # Show all rows with any content
            if e_value is not None or (c_value and str(c_value).strip()):
                print(f"  Row {row}:")
                print(f"    E{row} (selection): {repr(e_value)} (type: {type(e_value)})")
                print(f"    C{row} (model): {repr(c_value)}")
                
                # Check if this would be detected as selected
                if e_value and str(e_value).strip() != "":
                    try:
                        selection_num = int(float(str(e_value)))
                        if selection_num >= 1:
                            print(f"    ✅ SHOULD BE DETECTED (selection = {selection_num})")
                        else:
                            print(f"    ❌ Would not be detected (selection < 1)")
                    except:
                        print(f"    ❌ Invalid selection value")
                else:
                    print(f"    ❌ No selection value")
        
        # Check if there are any obvious patterns we're missing
        print(f"\nLooking for non-empty cells in column E:")
        for row in range(1, 50):  # Check more rows
            e_value = sheet[f'E{row}'].value
            if e_value is not None and str(e_value).strip():
                print(f"  E{row}: {repr(e_value)}")

print("\n" + "=" * 80)