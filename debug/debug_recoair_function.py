#!/usr/bin/env python3
"""Debug the actual RecoAir reading function step by step."""

import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from openpyxl import load_workbook
from utils.excel import validate_cell_data, clear_validation_errors

excel_path = '/Users/yazan/Downloads/36068 Cost Sheet 25032025 (1).xlsx'

def debug_read_recoair_data_from_sheet(sheet):
    """
    Debug version of read_recoair_data_from_sheet with extensive logging.
    """
    recoair_units = []
    sheet_name = sheet.title
    
    print(f"\n{'='*60}")
    print(f"DEBUGGING {sheet_name}")
    print(f"{'='*60}")
    
    try:
        # Get item reference from C12
        item_reference = sheet['C12'].value or ""
        print(f"1. Item reference (C12): {repr(item_reference)}")
        
        # Get delivery and installation price (N36 - N46) with validation
        n36_valid, n36_value, n36_error = validate_cell_data(
            sheet_name, 'N36', sheet['N36'].value, 'number', 'Total Delivery and Installation (N36)'
        )
        if not n36_valid:
            print(f"   ERROR: N36 validation failed: {n36_error}")
            n36_value = 0
        else:
            print(f"2. N36 (Total D&I): {n36_value}")

        # Get commissioning price from N46 with validation
        n46_valid, n46_value, n46_error = validate_cell_data(
            sheet_name, 'N46', sheet['N46'].value, 'number', 'Commissioning Price (N46)'
        )
        if not n46_valid:
            print(f"   ERROR: N46 validation failed: {n46_error}")
            n46_value = 0
        else:
            print(f"3. N46 (Commissioning): {n46_value}")

        # Calculate delivery and installation price (N36 - N46)
        delivery_installation_price = n36_value - n46_value if n36_value > n46_value else 0
        print(f"4. Calculated delivery price: {delivery_installation_price}")
        
        # Get flat pack data from D40 and N40
        flat_pack_description = sheet['D40'].value or ""
        print(f"5. Flat pack description (D40): {repr(flat_pack_description)}")
        
        # Validate flat pack price
        flat_pack_valid, flat_pack_price, flat_pack_error = validate_cell_data(
            sheet_name, 'N40', sheet['N40'].value, 'number', 'Flat Pack Price'
        )
        if not flat_pack_valid:
            print(f"   ERROR: N40 validation failed: {flat_pack_error}")
            flat_pack_price = 0
        else:
            print(f"6. Flat pack price (N40): {flat_pack_price}")
        
        print(f"\n7. Checking rows 14-28 for unit selections:")
        
        # Check rows 14 to 28 for RecoAir unit selections
        units_found = 0
        for row in range(14, 29):
            # Check if there's a value of 1 or more in column E (selection indicator)
            selection_value = sheet[f'E{row}'].value
            
            if selection_value and str(selection_value).strip() != "":
                print(f"   Row {row}: E{row} = {repr(selection_value)}")
                
                # Validate selection quantity (use 'integer' for quantities)
                selection_valid, selection_num, selection_error = validate_cell_data(
                    sheet_name, f'E{row}', selection_value, 'integer', f'RecoAir Unit Quantity (Row {row})'
                )
                
                if not selection_valid:
                    print(f"     ERROR: Selection validation failed: {selection_error}")
                    continue
                    
                print(f"     Selection validation passed: {selection_num}")
                    
                if selection_num >= 1:
                    print(f"     ✅ Unit selected (quantity >= 1)")
                    
                    # This row has a selected RecoAir unit
                    # Collect data from this row
                    model = sheet[f'C{row}'].value or ""
                    extract_volume_str = sheet[f'D{row}'].value or ""
                    
                    print(f"     Model (C{row}): {repr(model)}")
                    print(f"     Extract volume (D{row}): {repr(extract_volume_str)}")
                    
                    # Validate dimensions
                    width_valid, width, width_error = validate_cell_data(
                        sheet_name, f'F{row}', sheet[f'F{row}'].value, 'integer', f'RecoAir Unit Width (Row {row})'
                    )
                    if not width_valid:
                        print(f"     ERROR: Width validation failed: {width_error}")
                        width = 0
                    
                    length_valid, length, length_error = validate_cell_data(
                        sheet_name, f'G{row}', sheet[f'G{row}'].value, 'integer', f'RecoAir Unit Length (Row {row})'
                    )
                    if not length_valid:
                        print(f"     ERROR: Length validation failed: {length_error}")
                        length = 0
                    
                    height_valid, height, height_error = validate_cell_data(
                        sheet_name, f'H{row}', sheet[f'H{row}'].value, 'integer', f'RecoAir Unit Height (Row {row})'
                    )
                    if not height_valid:
                        print(f"     ERROR: Height validation failed: {height_error}")
                        height = 0
                    
                    location_raw = sheet[f'I{row}'].value or "INTERNAL"  # Default to INTERNAL
                    
                    print(f"     Dimensions: {width} x {length} x {height}")
                    print(f"     Location (I{row}): {repr(location_raw)}")
                    
                    # Read base price from N12 (fixed cell for all units)
                    price_valid, unit_price, price_error = validate_cell_data(
                        sheet_name, 'N12', sheet['N12'].value, 'number', 'RecoAir Unit Base Price (N12)'
                    )
                    if not price_valid:
                        print(f"     ERROR: Price validation failed: {price_error}")
                        unit_price = 0
                    else:
                        print(f"     Unit price (N12): {unit_price}")
                    
                    # If we get here, we should have a valid unit
                    print(f"     ✅ Unit data collected successfully")
                    units_found += 1
                else:
                    print(f"     ❌ Selection quantity < 1: {selection_num}")
        
        print(f"\n8. Summary:")
        print(f"   Units found: {units_found}")
        print(f"   Expected units: 1 per sheet")
        
        if units_found == 0:
            print(f"   ❌ NO UNITS FOUND - This is the problem!")
        else:
            print(f"   ✅ Units found as expected")
        
    except Exception as e:
        print(f"EXCEPTION in debug_read_recoair_data_from_sheet: {str(e)}")
        import traceback
        traceback.print_exc()

# Load the workbook and test the function
wb = load_workbook(excel_path, data_only=True)

for sheet_name in wb.sheetnames:
    if 'RECOAIR' in sheet_name:
        sheet = wb[sheet_name]
        clear_validation_errors()
        debug_read_recoair_data_from_sheet(sheet)