#!/usr/bin/env python3
"""Test the actual read_recoair_data_from_sheet function directly."""

import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from openpyxl import load_workbook
from utils.excel import read_recoair_data_from_sheet, clear_validation_errors, collect_validation_errors

excel_path = '/Users/yazan/Downloads/36068 Cost Sheet 25032025 (1).xlsx'

print("=" * 80)
print("TESTING ACTUAL read_recoair_data_from_sheet FUNCTION")
print("=" * 80)

# Load the workbook
wb = load_workbook(excel_path, data_only=True)

for sheet_name in wb.sheetnames:
    if 'RECOAIR' in sheet_name:
        sheet = wb[sheet_name]
        print(f"\nTesting {sheet_name}:")
        print("-" * 40)
        
        # Clear validation errors before calling
        clear_validation_errors()
        
        try:
            # Call the actual function
            result = read_recoair_data_from_sheet(sheet)
            
            print(f"  Function returned: {type(result)}")
            if isinstance(result, dict):
                units = result.get('units', [])
                flat_pack = result.get('flat_pack', {})
                print(f"  Units found: {len(units)}")
                print(f"  Flat pack data: {flat_pack}")
                
                if units:
                    for i, unit in enumerate(units):
                        print(f"    Unit {i+1}: {unit.get('model', 'No model')} at row {unit.get('row', 'No row')}")
                else:
                    print("    No units found")
            else:
                print(f"  Unexpected return type: {result}")
            
            # Check for validation errors
            errors = collect_validation_errors()
            if errors:
                print(f"  Validation errors ({len(errors)}):")
                for error in errors:
                    print(f"    {error}")
            else:
                print("  No validation errors")
                
        except Exception as e:
            print(f"  EXCEPTION: {e}")
            import traceback
            traceback.print_exc()

print("\n" + "=" * 80)