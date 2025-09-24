#!/usr/bin/env python3
"""
Check calculated values of Excel formulas.
"""

import sys
sys.path.append('src')
from openpyxl import load_workbook

def check_formula_values(excel_path: str):
    """Check calculated values of Excel formulas."""
    print(f"🔍 CHECKING FORMULA VALUES: {excel_path}")
    print("=" * 50)
    
    # Load with data_only=True to get calculated values
    wb_data = load_workbook(excel_path, data_only=True)
    # Load without data_only to get formulas
    wb_formula = load_workbook(excel_path, data_only=False)
    
    for sheet_name in wb_data.sheetnames:
        if 'CANOPY' in sheet_name:
            sheet_data = wb_data[sheet_name]
            sheet_formula = wb_formula[sheet_name]
            
            print(f"\n📋 {sheet_name}:")
            
            for row in [19, 20, 21]:
                formula = sheet_formula[f'N{row}'].value
                calculated = sheet_data[f'N{row}'].value
                
                print(f"  N{row}:")
                print(f"    Formula: {repr(formula)}")
                print(f"    Calculated: {repr(calculated)} (type: {type(calculated)})")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python check_formula_values.py <excel_file>")
        sys.exit(1)
    
    check_formula_values(sys.argv[1]) 