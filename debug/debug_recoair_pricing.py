#!/usr/bin/env python3
"""Debug RecoAir pricing discrepancy between Excel and Word document."""

import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from openpyxl import load_workbook
from utils.excel import read_excel_project_data
import json

excel_path = '/Users/yazan/Downloads/36068 Cost Sheet 25032025 (1).xlsx'

print("=" * 80)
print("DEBUGGING RECOAIR PRICING DISCREPANCY")
print("=" * 80)

# Load the workbook
wb = load_workbook(excel_path, data_only=True)

print("\n1. ANALYZING EXCEL SHEETS:")
print("-" * 40)

# List all sheets
for sheet_name in wb.sheetnames:
    if 'RECOAIR' in sheet_name:
        print(f"  Found RecoAir sheet: {sheet_name}")

print("\n2. CHECKING JOB TOTAL SHEET:")
print("-" * 40)

if 'JOB TOTAL' in wb.sheetnames:
    job_total_sheet = wb['JOB TOTAL']
    
    # Check various total cells
    print("  Checking common total cells:")
    cells_to_check = ['C24', 'T24', 'T28', 'C28', 'N24', 'N28']
    for cell in cells_to_check:
        value = job_total_sheet[cell].value
        if value and value != 0:
            print(f"    {cell}: {value:,.2f}" if isinstance(value, (int, float)) else f"    {cell}: {value}")

print("\n3. ANALYZING RECOAIR SHEETS:")
print("-" * 40)

total_from_sheets = 0
area_totals = {}

for sheet_name in wb.sheetnames:
    if 'RECOAIR' in sheet_name:
        sheet = wb[sheet_name]
        print(f"\n  Sheet: {sheet_name}")
        
        # Get area name from B1
        area_title = sheet['B1'].value
        print(f"    Title (B1): {area_title}")
        
        # Check key pricing cells
        # Unit prices
        print("    Unit prices:")
        for row in range(14, 29):  # Check rows 14-28 for units
            ref_cell = f'B{row}'
            model_cell = f'D{row}'
            price_cell = f'N{row}'
            
            ref_val = sheet[ref_cell].value
            model_val = sheet[model_cell].value
            price_val = sheet[price_cell].value
            
            if ref_val and price_val:
                print(f"      Row {row}: Ref={ref_val}, Model={model_val}, Price={price_val:,.2f}" if isinstance(price_val, (int, float)) else f"      Row {row}: Ref={ref_val}, Model={model_val}, Price={price_val}")
        
        # Check subtotals and totals
        print("    Key totals:")
        # N12 - Base unit price
        n12 = sheet['N12'].value
        if n12:
            print(f"      N12 (Base price): {n12:,.2f}" if isinstance(n12, (int, float)) else f"      N12: {n12}")
        
        # N40 - Flat pack price
        n40 = sheet['N40'].value
        if n40:
            print(f"      N40 (Flat pack): {n40:,.2f}" if isinstance(n40, (int, float)) else f"      N40: {n40}")
        
        # N182 - Delivery subtotal
        n182 = sheet['N182'].value
        if n182:
            print(f"      N182 (Delivery subtotal): {n182:,.2f}" if isinstance(n182, (int, float)) else f"      N182: {n182}")
        
        # N193 - Commissioning
        n193 = sheet['N193'].value
        if n193:
            print(f"      N193 (Commissioning): {n193:,.2f}" if isinstance(n193, (int, float)) else f"      N193: {n193}")
        
        # Calculate area total
        area_total = 0
        if n12:
            area_total += n12 if isinstance(n12, (int, float)) else 0
        if n40:
            area_total += n40 if isinstance(n40, (int, float)) else 0
        if n182:
            area_total += n182 if isinstance(n182, (int, float)) else 0
        
        area_totals[sheet_name] = area_total
        total_from_sheets += area_total
        print(f"    AREA TOTAL: £{area_total:,.2f}")

print("\n4. TOTAL FROM ALL RECOAIR SHEETS:")
print("-" * 40)
print(f"  Calculated total: £{total_from_sheets:,.2f}")

print("\n5. READING PROJECT DATA THROUGH UTILS:")
print("-" * 40)

# Read using the utility function
project_data = read_excel_project_data(excel_path)

# Calculate totals from project data
total_from_project_data = 0
for level in project_data.get('levels', []):
    for area in level.get('areas', []):
        recoair_price = area.get('recoair_price', 0)
        if recoair_price > 0:
            print(f"  Area: {area.get('name')} - RecoAir price: £{recoair_price:,.2f}")
            total_from_project_data += recoair_price

print(f"\n  Total from project data: £{total_from_project_data:,.2f}")

print("\n6. CHECKING FOR PRICING DISCREPANCIES:")
print("-" * 40)

# Expected total from Excel (as mentioned by user)
expected_total = 148355.00
calculated_total = total_from_sheets

print(f"  Expected total (from Excel): £{expected_total:,.2f}")
print(f"  Calculated from sheets: £{calculated_total:,.2f}")
print(f"  From project data: £{total_from_project_data:,.2f}")
print(f"  Shown in quote: £151,087.00")

difference = 151087.00 - expected_total
print(f"\n  Difference (Quote - Excel): £{difference:,.2f}")

# Look for the source of the difference
print("\n7. ANALYZING POTENTIAL SOURCES OF DISCREPANCY:")
print("-" * 40)

# Check if the difference might be in flat pack pricing
total_flat_pack = 0
for sheet_name in wb.sheetnames:
    if 'RECOAIR' in sheet_name:
        sheet = wb[sheet_name]
        n40 = sheet['N40'].value
        if n40 and isinstance(n40, (int, float)):
            total_flat_pack += n40
            print(f"  Flat pack from {sheet_name}: £{n40:,.2f}")

print(f"  Total flat pack: £{total_flat_pack:,.2f}")

# Check delivery and commissioning separately
total_delivery = 0
total_commissioning = 0
for sheet_name in wb.sheetnames:
    if 'RECOAIR' in sheet_name:
        sheet = wb[sheet_name]
        
        # Get delivery (N182 - N193)
        n182 = sheet['N182'].value
        n193 = sheet['N193'].value
        
        if n182 and n193 and isinstance(n182, (int, float)) and isinstance(n193, (int, float)):
            delivery = n182 - n193
            total_delivery += delivery
            total_commissioning += n193
            print(f"  {sheet_name}: Delivery={delivery:,.2f}, Commissioning={n193:,.2f}")

print(f"\n  Total delivery: £{total_delivery:,.2f}")
print(f"  Total commissioning: £{total_commissioning:,.2f}")

print("\n" + "=" * 80)