#!/usr/bin/env python3
"""Debug script to trace complete SDU data flow from Excel to Word template."""

import sys
import os
import json
sys.path.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'src'))

from openpyxl import load_workbook
from utils.excel import read_excel_project_data, extract_sdu_electrical_services
from utils.word import collect_sdu_data

def debug_sdu_flow(excel_path):
    """Debug the complete SDU data flow."""
    print(f"\n🔍 Debugging SDU data flow for: {excel_path}")
    print("=" * 80)
    
    # Step 1: Read project data
    print("\n📊 Step 1: Reading project data from Excel...")
    project_data = read_excel_project_data(excel_path)
    
    # Check if SDU option is detected on canopies
    print("\n✅ Checking canopies for SDU option:")
    sdu_canopies = []
    for level in project_data.get('levels', []):
        for area in level.get('areas', []):
            for canopy in area.get('canopies', []):
                if canopy.get('options', {}).get('sdu', False):
                    sdu_canopies.append({
                        'reference': canopy.get('reference_number'),
                        'area': area.get('name'),
                        'level': level.get('level_name'),
                        'sdu_item_number': canopy.get('sdu_item_number', 'NOT SET'),
                        'sdu_price': canopy.get('sdu_price', 0)
                    })
                    print(f"  - Canopy {canopy.get('reference_number')} has SDU enabled")
                    print(f"    sdu_item_number in canopy data: {canopy.get('sdu_item_number', 'NOT SET')}")
    
    # Step 2: Collect SDU data using Word utility
    print(f"\n📡 Step 2: Collecting SDU data using collect_sdu_data()...")
    sdu_data = collect_sdu_data(project_data, excel_path)
    
    print(f"\n📋 SDU data collected: {len(sdu_data)} items")
    for i, sdu in enumerate(sdu_data):
        print(f"\n  SDU #{i+1}:")
        print(f"    level_area_combined: {sdu.get('level_area_combined')}")
        print(f"    canopy_reference: {sdu.get('canopy_reference')}")
        print(f"    sdu_item_number: {sdu.get('sdu_item_number', 'NOT SET')}")
        print(f"    sdu_price: {sdu.get('sdu_price', 0)}")
        print(f"    has electrical_services: {'electrical_services' in sdu}")
        print(f"    has gas_services: {'gas_services' in sdu}")
        print(f"    has water_services: {'water_services' in sdu}")
        print(f"    has pricing: {'pricing' in sdu}")
        if 'pricing' in sdu:
            pricing = sdu['pricing']
            print(f"      - final_carcass_price: {pricing.get('final_carcass_price', 0)}")
            print(f"      - final_electrical_price: {pricing.get('final_electrical_price', 0)}")
            print(f"      - total_price: {pricing.get('total_price', 0)}")
    
    # Step 3: Check Excel sheets directly
    print(f"\n📖 Step 3: Checking Excel SDU sheets directly...")
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    
    sdu_sheets = [s for s in wb.sheetnames if 'SDU' in s]
    print(f"Found {len(sdu_sheets)} SDU sheets: {sdu_sheets}")
    
    for sheet_name in sdu_sheets:
        sheet = wb[sheet_name]
        print(f"\n  Sheet: {sheet_name}")
        print(f"    B1 (Title): {sheet['B1'].value}")
        print(f"    B12 (Item Number): {sheet['B12'].value}")
        print(f"    C12 (Model): {sheet['C12'].value}")
        
        # Extract using the same function
        print(f"\n    Extracting services data...")
        services_data = extract_sdu_electrical_services(sheet)
        print(f"    sdu_item_number from extraction: {services_data.get('sdu_item_number', 'NOT FOUND')}")
        print(f"    electrical_services: {json.dumps(services_data.get('electrical_services', {}), indent=6)}")
        print(f"    pricing total: {services_data.get('pricing', {}).get('total_price', 0)}")
    
    # Step 4: Simulate Word template data structure
    print(f"\n📝 Step 4: Simulating Word template data structure...")
    # This is what gets passed to the Word template
    template_data = {
        'sdu_areas': sdu_data,
        'has_sdu': len(sdu_data) > 0,
        'total_sdu_areas': len(sdu_data)
    }
    
    print(f"\nTemplate would receive:")
    print(f"  has_sdu: {template_data['has_sdu']}")
    print(f"  total_sdu_areas: {template_data['total_sdu_areas']}")
    print(f"  sdu_areas items: {len(template_data['sdu_areas'])}")
    
    if template_data['sdu_areas']:
        print(f"\n  First SDU area structure:")
        first_sdu = template_data['sdu_areas'][0]
        for key in sorted(first_sdu.keys()):
            value = first_sdu[key]
            if isinstance(value, dict):
                print(f"    {key}: <dict with {len(value)} keys>")
            else:
                print(f"    {key}: {value}")

if __name__ == "__main__":
    excel_path = "/Users/yazan/Downloads/36108 Cost Sheet 28072025.xlsx"
    
    if os.path.exists(excel_path):
        debug_sdu_flow(excel_path)
    else:
        print(f"❌ Excel file not found: {excel_path}")