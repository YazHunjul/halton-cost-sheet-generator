#!/usr/bin/env python3
"""Debug script to investigate SDUS detection issue."""

import sys
import os
sys.path.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'src'))

from openpyxl import load_workbook
from utils.excel import read_excel_project_data

def analyze_excel_for_sdus(excel_path):
    """Analyze Excel file for SDUS sheets and data."""
    print(f"\n📊 Analyzing Excel file: {excel_path}")
    print("=" * 80)
    
    # Load workbook
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    
    # List all sheet names
    print("\n📋 All sheet names:")
    for i, sheet_name in enumerate(wb.sheetnames):
        print(f"  {i+1}. {sheet_name}")
    
    # Look for SDU-related sheets
    print("\n🔍 SDU-related sheets:")
    sdu_sheets = []
    for sheet_name in wb.sheetnames:
        if 'SDU' in sheet_name.upper():
            sdu_sheets.append(sheet_name)
            print(f"  - {sheet_name}")
            
            # Check if it's "SDUS" specifically
            if 'SDUS' in sheet_name.upper():
                print(f"    ⚠️  This is an SDUS sheet (with 'S')")
    
    if not sdu_sheets:
        print("  ❌ No SDU-related sheets found")
    
    # Read project data to see what's detected
    print("\n📖 Reading project data...")
    try:
        project_data = read_excel_project_data(excel_path)
        
        # Check for SDU options in canopies
        print("\n🔍 Checking canopies for SDU options:")
        sdu_canopy_count = 0
        for level in project_data.get('levels', []):
            for area in level.get('areas', []):
                for canopy in area.get('canopies', []):
                    if canopy.get('options', {}).get('sdu', False):
                        sdu_canopy_count += 1
                        print(f"  - Canopy {canopy.get('reference_number')} in {area.get('name')} has SDU enabled")
        
        if sdu_canopy_count == 0:
            print("  ❌ No canopies with SDU option found")
        else:
            print(f"  ✅ Found {sdu_canopy_count} canopies with SDU option")
            
        # List all canopy references
        print("\n📋 All canopy references found:")
        for level in project_data.get('levels', []):
            for area in level.get('areas', []):
                for canopy in area.get('canopies', []):
                    ref = canopy.get('reference_number', 'No ref')
                    print(f"  - {ref} in {area.get('name')} (Level {level.get('level_number')})")
            
    except Exception as e:
        print(f"  ❌ Error reading project data: {str(e)}")
    
    # Check sheet name patterns
    print("\n🔍 Checking sheet naming patterns:")
    for sheet_name in sdu_sheets:
        sheet = wb[sheet_name]
        print(f"\n  Sheet: {sheet_name}")
        
        # Check B1 cell (title)
        title = sheet['B1'].value
        print(f"    B1 (Title): {title}")
        
        # Check B12 (SDU item number)
        item_number = sheet['B12'].value
        print(f"    B12 (Item): {item_number}")
        
        # Check C12 (Model name)
        model = sheet['C12'].value
        print(f"    C12 (Model): {model}")

if __name__ == "__main__":
    excel_path = "/Users/yazan/Downloads/36108 Cost Sheet 28072025.xlsx"
    
    if os.path.exists(excel_path):
        analyze_excel_for_sdus(excel_path)
    else:
        print(f"❌ Excel file not found: {excel_path}")