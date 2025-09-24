#!/usr/bin/env python3
"""
Debug script to check UV_EXTRA_OVER_CALC sheet directly.
"""

import sys
import os
sys.path.append('src')

from openpyxl import load_workbook

def debug_uv_sheet():
    """Debug the UV_EXTRA_OVER_CALC sheet."""
    
    # Find the most recent Excel file
    excel_files = []
    if os.path.exists('output'):
        for file in os.listdir('output'):
            if file.endswith('.xlsx'):
                excel_files.append(os.path.join('output', file))
    
    if not excel_files:
        print("❌ No Excel files found")
        return
    
    latest_file = max(excel_files, key=os.path.getmtime)
    print(f"📁 Checking: {latest_file}")
    
    try:
        wb = load_workbook(latest_file, data_only=True)
        print(f"📋 All sheets: {wb.sheetnames}")
        
        if 'UV_EXTRA_OVER_CALC' in wb.sheetnames:
            print("✅ UV_EXTRA_OVER_CALC sheet found!")
            
            calc_sheet = wb['UV_EXTRA_OVER_CALC']
            print(f"📊 Sheet state: {calc_sheet.sheet_state}")
            print(f"📏 Max row: {calc_sheet.max_row}")
            
            # Read all data from the sheet
            print("\n📋 Sheet contents:")
            for row in range(1, calc_sheet.max_row + 1):
                row_data = []
                for col in range(1, 10):  # A through I
                    cell = calc_sheet.cell(row=row, column=col)
                    row_data.append(str(cell.value) if cell.value is not None else "")
                print(f"  Row {row}: {row_data}")
                
                if row > 10:  # Limit output
                    break
        else:
            print("❌ UV_EXTRA_OVER_CALC sheet not found")
            
        # Also check for CANOPY (UV) sheets
        uv_sheets = [s for s in wb.sheetnames if 'CANOPY (UV)' in s]
        print(f"\n🔍 UV Canopy sheets found: {uv_sheets}")
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    debug_uv_sheet() 