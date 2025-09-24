#!/usr/bin/env python3
"""Detailed debugging of RecoAir price calculation."""

import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from openpyxl import load_workbook
from utils.excel import read_recoair_data_from_sheet
import json

excel_path = '/Users/yazan/Downloads/36068 Cost Sheet 25032025 (1).xlsx'

print("=" * 80)
print("DETAILED RECOAIR PRICING DEBUG")
print("=" * 80)

# Load the workbook
wb = load_workbook(excel_path, data_only=True)

print("\n1. RAW EXCEL DATA FROM EACH SHEET:")
print("-" * 50)

sheet_data = {}
for sheet_name in wb.sheetnames:
    if 'RECOAIR' in sheet_name:
        sheet = wb[sheet_name]
        print(f"\nSheet: {sheet_name}")
        
        # Key cells
        n12 = sheet['N12'].value or 0  # Base unit price
        n36 = sheet['N36'].value or 0  # Total delivery and installation
        n40 = sheet['N40'].value or 0  # Flat pack price
        n46 = sheet['N46'].value or 0  # Commissioning price
        c12 = sheet['C12'].value or ""  # Item reference
        
        # Calculate delivery (N36 - N46)
        delivery = n36 - n46 if n36 > n46 else 0
        
        print(f"  C12 (Item Reference): {c12}")
        print(f"  N12 (Base Unit Price): £{n12:,.2f}")
        print(f"  N36 (Total Delivery & Installation): £{n36:,.2f}")
        print(f"  N46 (Commissioning): £{n46:,.2f}")
        print(f"  N40 (Flat Pack): £{n40:,.2f}")
        print(f"  Calculated Delivery (N36-N46): £{delivery:,.2f}")
        
        # Check for selected units (rows 14-28)
        selected_units = []
        for row in range(14, 29):
            selection = sheet[f'E{row}'].value
            if selection and str(selection).strip() and selection >= 1:
                model = sheet[f'C{row}'].value or ""
                selected_units.append(f"Row {row}: {model}")
        
        print(f"  Selected Units: {len(selected_units)} units")
        for unit in selected_units:
            print(f"    {unit}")
        
        # Manual total calculation
        manual_total = n12 + delivery + n46  # Base + Delivery + Commissioning (no flat pack)
        manual_total_with_flat = manual_total + n40
        print(f"  Manual Total (excl flat pack): £{manual_total:,.2f}")
        print(f"  Manual Total (incl flat pack): £{manual_total_with_flat:,.2f}")
        
        sheet_data[sheet_name] = {
            'base_price': n12,
            'delivery': delivery,
            'commissioning': n46,
            'flat_pack': n40,
            'manual_total_excl_flat': manual_total,
            'manual_total_incl_flat': manual_total_with_flat,
            'selected_units_count': len(selected_units)
        }

print("\n2. USING RECOAIR READING FUNCTION:")
print("-" * 50)

function_totals = {}
for sheet_name in wb.sheetnames:
    if 'RECOAIR' in sheet_name:
        sheet = wb[sheet_name]
        print(f"\nProcessing {sheet_name} with function:")
        
        try:
            # Use the actual reading function
            recoair_data = read_recoair_data_from_sheet(sheet)
            
            if recoair_data and 'units' in recoair_data:
                units = recoair_data['units']
                flat_pack_data = recoair_data['flat_pack']
                # Commissioning is already included in each unit's commissioning_price
                commissioning = sum(unit.get('commissioning_price', 0) for unit in units)
                
                print(f"  Found {len(units)} units")
                
                total_unit_price = sum(unit.get('unit_price', 0) for unit in units)
                total_delivery = sum(unit.get('delivery_installation_price', 0) for unit in units)
                flat_pack_price = flat_pack_data.get('price', 0) if flat_pack_data.get('has_flat_pack') else 0
                
                print(f"  Total Unit Price: £{total_unit_price:,.2f}")
                print(f"  Total Delivery: £{total_delivery:,.2f}")
                print(f"  Commissioning: £{commissioning:,.2f}")
                print(f"  Flat Pack: £{flat_pack_price:,.2f}")
                
                function_total_excl = total_unit_price + total_delivery + commissioning
                function_total_incl = function_total_excl + flat_pack_price
                
                print(f"  Function Total (excl flat): £{function_total_excl:,.2f}")
                print(f"  Function Total (incl flat): £{function_total_incl:,.2f}")
                
                function_totals[sheet_name] = {
                    'total_excl_flat': function_total_excl,
                    'total_incl_flat': function_total_incl
                }
                
            else:
                print(f"  No RecoAir units found")
                
        except Exception as e:
            print(f"  Error processing sheet: {e}")

print("\n3. COMPARISON:")
print("-" * 50)

manual_total_excl = sum(data['manual_total_excl_flat'] for data in sheet_data.values())
manual_total_incl = sum(data['manual_total_incl_flat'] for data in sheet_data.values())

function_total_excl = sum(data['total_excl_flat'] for data in function_totals.values())
function_total_incl = sum(data['total_incl_flat'] for data in function_totals.values())

print(f"Manual calculation totals:")
print(f"  Excluding flat pack: £{manual_total_excl:,.2f}")
print(f"  Including flat pack: £{manual_total_incl:,.2f}")

print(f"\nFunction calculation totals:")
print(f"  Excluding flat pack: £{function_total_excl:,.2f}")  
print(f"  Including flat pack: £{function_total_incl:,.2f}")

print(f"\nTarget totals:")
print(f"  Excel total: £148,355.00")
print(f"  Quote total: £151,087.00")

print(f"\nDifferences:")
print(f"  Manual vs Excel (excl flat): £{manual_total_excl - 148355:,.2f}")
print(f"  Manual vs Quote (incl flat): £{manual_total_incl - 151087:,.2f}")
print(f"  Function vs Excel (excl flat): £{function_total_excl - 148355:,.2f}")
print(f"  Function vs Quote (incl flat): £{function_total_incl - 151087:,.2f}")

print("\n" + "=" * 80)