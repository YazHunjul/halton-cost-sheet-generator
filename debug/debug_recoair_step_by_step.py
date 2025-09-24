#!/usr/bin/env python3
"""Debug RecoAir function step by step to find the exact failure point."""

import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from openpyxl import load_workbook
from utils.excel import (
    validate_cell_data, clear_validation_errors, 
    transform_recoair_model, get_recoair_specifications,
    extract_recoair_volume, safe_float_conversion
)

excel_path = '/Users/yazan/Downloads/36068 Cost Sheet 25032025 (1).xlsx'

def debug_read_recoair_step_by_step(sheet):
    """
    Replicate the read_recoair_data_from_sheet function step by step with extensive logging.
    """
    recoair_units = []
    sheet_name = sheet.title
    
    print(f"\n{'='*60}")
    print(f"STEP-BY-STEP DEBUG: {sheet_name}")
    print(f"{'='*60}")
    
    try:
        print("STEP 1: Get basic data from sheet")
        
        # Get item reference from C12
        item_reference = sheet['C12'].value or ""
        print(f"  Item reference (C12): {repr(item_reference)}")
        
        # Get delivery and installation price (N36 - N46) with validation
        n36_valid, n36_value, n36_error = validate_cell_data(
            sheet_name, 'N36', sheet['N36'].value, 'number', 'Total Delivery and Installation (N36)'
        )
        if not n36_valid:
            print(f"  ERROR: N36 validation failed: {n36_error}")
            n36_value = 0
        
        # Get commissioning price from N46 with validation  
        n46_valid, n46_value, n46_error = validate_cell_data(
            sheet_name, 'N46', sheet['N46'].value, 'number', 'Commissioning Price (N46)'
        )
        if not n46_valid:
            print(f"  ERROR: N46 validation failed: {n46_error}")
            n46_value = 0
            
        # Calculate delivery and installation price (N36 - N46)
        delivery_installation_price = n36_value - n46_value if n36_value > n46_value else 0
        print(f"  Delivery price (N36-N46): {delivery_installation_price}")
        
        print("STEP 2: Get flat pack data")
        
        # Get flat pack data from D40 and N40
        flat_pack_description = sheet['D40'].value or ""
        print(f"  Flat pack description (D40): {repr(flat_pack_description)}")
        
        # Validate flat pack price
        flat_pack_valid, flat_pack_price, flat_pack_error = validate_cell_data(
            sheet_name, 'N40', sheet['N40'].value, 'number', 'Flat Pack Price'
        )
        if not flat_pack_valid:
            print(f"  ERROR: N40 validation failed: {flat_pack_error}")
            flat_pack_price = 0
        else:
            print(f"  Flat pack price (N40): {flat_pack_price}")
        
        print("STEP 3: Check for selected units (rows 14-28)")
        
        # Check rows 14 to 28 for RecoAir unit selections
        for row in range(14, 29):
            # Check if there's a value of 1 or more in column E (selection indicator)
            selection_value = sheet[f'E{row}'].value
            
            if selection_value and str(selection_value).strip() != "":
                print(f"  Row {row}: Found selection value {repr(selection_value)}")
                
                # Validate selection quantity
                selection_valid, selection_num, selection_error = validate_cell_data(
                    sheet_name, f'E{row}', selection_value, 'integer', f'RecoAir Unit Quantity (Row {row})'
                )
                
                if not selection_valid:
                    print(f"    ERROR: Selection validation failed: {selection_error}")
                    continue
                    
                if selection_num >= 1:
                    print(f"    ✅ Unit selected (quantity={selection_num})")
                    
                    print("STEP 4: Collect unit data")
                    
                    # This row has a selected RecoAir unit - collect data from this row
                    model = sheet[f'C{row}'].value or ""
                    extract_volume_str = sheet[f'D{row}'].value or ""
                    
                    print(f"    Model (C{row}): {repr(model)}")
                    print(f"    Extract volume (D{row}): {repr(extract_volume_str)}")
                    
                    # Validate dimensions
                    width_valid, width, width_error = validate_cell_data(
                        sheet_name, f'F{row}', sheet[f'F{row}'].value, 'integer', f'RecoAir Unit Width (Row {row})'
                    )
                    if not width_valid:
                        print(f"    ERROR: Width validation failed: {width_error}")
                        continue
                    
                    length_valid, length, length_error = validate_cell_data(
                        sheet_name, f'G{row}', sheet[f'G{row}'].value, 'integer', f'RecoAir Unit Length (Row {row})'
                    )
                    if not length_valid:
                        print(f"    ERROR: Length validation failed: {length_error}")
                        continue
                    
                    height_valid, height, height_error = validate_cell_data(
                        sheet_name, f'H{row}', sheet[f'H{row}'].value, 'integer', f'RecoAir Unit Height (Row {row})'
                    )
                    if not height_valid:
                        print(f"    ERROR: Height validation failed: {height_error}")
                        continue
                    
                    print(f"    Dimensions: {width} x {length} x {height}")
                    
                    # Location processing
                    location_raw = sheet[f'I{row}'].value or "INTERNAL"
                    location_str = str(location_raw).strip().upper() if location_raw else "INTERNAL"
                    
                    if location_str in ["EXTERNAL", "EXT"]:
                        location = "EXTERNAL"
                    elif location_str in ["INTERNAL", "INT"]:
                        location = "INTERNAL"
                    else:
                        location = location_str
                    
                    print(f"    Location: {repr(location)}")
                    
                    # Read base price from N12 (fixed cell for all units)
                    price_valid, unit_price, price_error = validate_cell_data(
                        sheet_name, 'N12', sheet['N12'].value, 'number', 'RecoAir Unit Base Price (N12)'
                    )
                    if not price_valid:
                        print(f"    ERROR: Price validation failed: {price_error}")
                        continue
                    else:
                        print(f"    Unit price (N12): {unit_price}")
                    
                    print("STEP 5: Process model transformation and specifications")
                    
                    # Extract volume number from extract volume string
                    extract_volume = extract_recoair_volume(extract_volume_str)
                    print(f"    Extract volume: {extract_volume}")
                    
                    # Transform the model name according to business rules
                    original_model = str(model).strip() if model else ""
                    transformed_model = transform_recoair_model(original_model)
                    print(f"    Original model: {repr(original_model)}")
                    print(f"    Transformed model: {repr(transformed_model)}")
                    
                    # Get technical specifications for this model
                    specs = get_recoair_specifications(transformed_model)
                    print(f"    Specifications: {specs}")
                    
                    print("STEP 6: Calculate prices")
                    
                    # Calculate prices
                    base_unit_price = unit_price
                    delivery_per_unit = delivery_installation_price / selection_num if selection_num > 0 else 0
                    commissioning_per_unit = n46_value / selection_num if selection_num > 0 else 0
                    final_unit_price = base_unit_price + delivery_per_unit + commissioning_per_unit
                    
                    print(f"    Base unit price: {base_unit_price}")
                    print(f"    Delivery per unit: {delivery_per_unit}")
                    print(f"    Commissioning per unit: {commissioning_per_unit}")
                    print(f"    Final unit price: {final_unit_price}")
                    
                    print("STEP 7: Create unit dictionary")
                    
                    # Create the unit dictionary (this is where the exception might occur)
                    recoair_unit = {
                        'item_reference': str(item_reference).strip() if item_reference else "",
                        'model': original_model,
                        'transformed_model': transformed_model,
                        'extract_volume': extract_volume,
                        'width': width,
                        'length': length,
                        'height': height,
                        'location': location,
                        'unit_price': final_unit_price,
                        'base_unit_price': base_unit_price,
                        'delivery_installation_price': delivery_per_unit,
                        'commissioning_price': commissioning_per_unit,
                        'quantity': selection_num,
                        'row': row,
                        
                        # Technical specifications - THIS IS WHERE THE EXCEPTION MIGHT HAPPEN
                        'p_drop': specs['p_drop'],  # Pressure drop (Pa)
                        'motor': specs['motor'],    # Motor power (kW/PH)
                        'weight': specs['weight']   # Weight (kg)
                    }
                    
                    print(f"    ✅ Unit dictionary created successfully")
                    recoair_units.append(recoair_unit)
                    print(f"    ✅ Unit added to list. Total units: {len(recoair_units)}")
        
        print("STEP 8: Finalize result")
        
        # Final result creation
        result = {
            'units': recoair_units,
            'flat_pack': {
                'item_reference': str(item_reference).strip() if item_reference else "",
                'description': flat_pack_description,
                'price': safe_float_conversion(flat_pack_price),
                'has_flat_pack': bool(flat_pack_description and str(flat_pack_description).strip())
            }
        }
        
        print(f"  Final result: {len(result['units'])} units found")
        return result
        
    except Exception as e:
        print(f"EXCEPTION: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            'units': [],
            'flat_pack': {
                'item_reference': '',
                'description': '',
                'price': 0,
                'has_flat_pack': False
            }
        }

# Load the workbook and test the function
wb = load_workbook(excel_path, data_only=True)

for sheet_name in wb.sheetnames:
    if 'RECOAIR' in sheet_name:
        sheet = wb[sheet_name]
        clear_validation_errors()
        result = debug_read_recoair_step_by_step(sheet)
        
        if result['units']:
            print(f"\n✅ SUCCESS: Found {len(result['units'])} units in {sheet_name}")
        else:
            print(f"\n❌ FAILED: No units found in {sheet_name}")