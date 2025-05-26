#!/usr/bin/env python3
"""
Check formula dependencies for cladding pricing.
"""

import sys
sys.path.append('src')
from openpyxl import load_workbook

def check_formula_dependencies(excel_path: str):
    """Check formula dependencies."""
    print(f"🔍 CHECKING FORMULA DEPENDENCIES: {excel_path}")
    print("=" * 50)
    
    wb = load_workbook(excel_path, data_only=True)
    
    for sheet_name in wb.sheetnames:
        if 'CANOPY' in sheet_name:
            sheet = wb[sheet_name]
            
            print(f"\n📋 {sheet_name}:")
            
            # Check B9 (used in VLOOKUP)
            b9_value = sheet['B9'].value
            print(f"  B9 (VLOOKUP key): {repr(b9_value)}")
            
            # Check M column values (multiplier in formula)
            print(f"\n  M column values:")
            for row in range(15, 25):
                m_val = sheet[f'M{row}'].value
                if m_val is not None:
                    print(f"    M{row}: {repr(m_val)}")
            
            # Check if there are any values in M19, M20 specifically
            print(f"\n  Specific M values for cladding:")
            for row in [19, 20]:
                m_val = sheet[f'M{row}'].value
                print(f"    M{row}: {repr(m_val)} (type: {type(m_val)})")
    
    # Check Base Costs sheet
    if 'Base Costs' in wb.sheetnames:
        base_costs = wb['Base Costs']
        print(f"\n📊 Base Costs sheet:")
        print(f"  Range A32:B37:")
        for row in range(32, 38):
            a_val = base_costs[f'A{row}'].value
            b_val = base_costs[f'B{row}'].value
            if a_val is not None or b_val is not None:
                print(f"    A{row}: {repr(a_val)}, B{row}: {repr(b_val)}")
    else:
        print(f"\n❌ 'Base Costs' sheet not found")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python check_formula_dependencies.py <excel_file>")
        sys.exit(1)
    
    check_formula_dependencies(sys.argv[1]) 